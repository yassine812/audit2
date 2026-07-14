from .utils.client_utils import get_base_url
from .fcm_utils import create_and_send_notification
from django.shortcuts import render, redirect, get_object_or_404, reverse
from django.http import JsonResponse, HttpResponseNotAllowed
from django.contrib.auth.decorators import login_required
from django.db.models.functions import Cast
from django.db.models import Q, Count, IntegerField, Avg
from django.contrib.staticfiles import finders
from .models import *
from django.db import transaction, IntegrityError
from django.urls import reverse
from django.core.validators import URLValidator
from django.core.exceptions import ValidationError
import re
import json
import dateparser
from .utils.openai_utils import generate_report
from .utils.prospect_ai import research_prospect
from django.utils.timesince import timesince
from django.utils import timezone
import logging
import pandas as pd
import requests
from rest_framework.views import APIView
from rest_framework.response import Response

from django.contrib.auth.forms import AuthenticationForm, PasswordResetForm, PasswordChangeForm
from django.contrib.auth.views import PasswordResetConfirmView, PasswordResetCompleteView
from django.contrib.auth import authenticate, login, logout, get_user_model, update_session_auth_hash
from django.contrib.auth.tokens import default_token_generator
from django.core.paginator import Paginator
from datetime import datetime, timedelta
from django.utils.timezone import make_aware
from django.views.decorators.http import require_GET, require_POST, require_http_methods
from django.utils import timezone
from .forms import *
from django.contrib.auth.hashers import make_password
from django.core.mail import send_mail, EmailMessage
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.template.loader import render_to_string
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.utils.html import strip_tags
from django.conf import settings
from django.contrib import messages
from django.utils.http import url_has_allowed_host_and_scheme
from django.db.models.functions import TruncMonth
from django.views.decorators.csrf import csrf_exempt
import pytz, logging, json, re, base64, unicodedata
from rest_framework.response import Response
from rest_framework.decorators import api_view
from rest_framework import permissions, serializers
from xhtml2pdf import pisa
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from django.http import HttpResponse
import os
from .ml.analysis import ProspectScorer, FrenchSentimentAnalyzer, MultilingualSentimentAnalyzer
from django.views import View
from django.utils.decorators import method_decorator
import requests

logger = logging.getLogger(__name__)

# ─── Alias de compatibilité ─────────────────────────────────────────────────
# Utilisateur et Societe sont désormais dans accounts ; on les re-expose ici
# pour que le code de views.py puisse continuer à les utiliser sans modification.
Utilisateur = get_user_model()
from accounts.models import Societe  # noqa: F401, E402


if '_redirect_back_or' not in globals():
    def _redirect_back_or(default_redirect_name, request):
        next_url = (request.POST.get('next') or request.GET.get('next') or '').strip()
        if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
            return redirect(next_url)
        return redirect(default_redirect_name)

# Dashboard
@method_decorator(login_required, name='dispatch')
class DashboardView(View):
    template_name = 'adminlte/sales/dashboard.html'

    def get_queryset_filters(self, request):
        user = request.user
        action_filter = Q()
        event_filter = Q()

        if user.is_superuser:
            return action_filter, event_filter

        if getattr(user, 'is_RC', False) and getattr(user, 'societe', None):
            action_filter &= Q(Q(created_by__societe=user.societe) | Q(pilote__societe=user.societe))
            event_filter &= Q(societe=user.societe)
        elif getattr(user, 'is_RO', False):
            ids = list(user.societes.values_list('id', flat=True))
            action_filter &= Q(Q(created_by__societe_id__in=ids) | Q(pilote__societe_id__in=ids))
            event_filter &= Q(societe_id__in=ids)
        else:
            action_filter &= Q(Q(created_by=user) | Q(pilote=user))
            event_filter &= Q(created_by=user)
        return action_filter, event_filter

    def get(self, request):
        from datetime import timedelta
        now = timezone.now()
        in_7_days = now + timedelta(days=7)
        last_30 = now - timedelta(days=30)
        prev_30 = now - timedelta(days=60)

        action_filter, event_filter = self.get_queryset_filters(request)

        # Counts
        actions_qs = Action.objects.filter(action_filter)
        counts = {
            'actions_total': actions_qs.count(),
            'calls': actions_qs.filter(is_Appel=True).count(),
            'emails': actions_qs.filter(is_Email=True).count(),
            'appointments': actions_qs.filter(is_RV=True).count(),
        }

        # Trends (last 30 days vs previous 30 days)
        def pct_change(current: int, previous: int) -> float:
            if previous <= 0:
                return 100.0 if current > 0 else 0.0
            return round(((current - previous) / previous) * 100.0, 1)

        window_qs = actions_qs.filter(date_heure_planifie__gte=last_30, date_heure_planifie__lte=now)
        prev_window_qs = actions_qs.filter(date_heure_planifie__gte=prev_30, date_heure_planifie__lt=last_30)

        cur_total = window_qs.count()
        prev_total = prev_window_qs.count()
        cur_calls = window_qs.filter(is_Appel=True).count()
        prev_calls = prev_window_qs.filter(is_Appel=True).count()
        cur_emails = window_qs.filter(is_Email=True).count()
        prev_emails = prev_window_qs.filter(is_Email=True).count()
        cur_rv = window_qs.filter(is_RV=True).count()
        prev_rv = prev_window_qs.filter(is_RV=True).count()

        trends = {
            'actions_total': {
                'pct': pct_change(cur_total, prev_total),
                'up': cur_total >= prev_total,
            },
            'calls': {
                'pct': pct_change(cur_calls, prev_calls),
                'up': cur_calls >= prev_calls,
            },
            'emails': {
                'pct': pct_change(cur_emails, prev_emails),
                'up': cur_emails >= prev_emails,
            },
            'appointments': {
                'pct': pct_change(cur_rv, prev_rv),
                'up': cur_rv >= prev_rv,
            },
        }

        # Upcoming events/actions (7 days)
        upcoming_events = (
            Evenement.objects.filter(event_filter, date_heure_planifie__gte=now)
            .order_by('date_heure_planifie')[:10]
        )
        upcoming_actions = (
            actions_qs.filter(date_heure_planifie__gte=now)
            .order_by('date_heure_planifie')[:10]
        )

        # Recent prospects (30 days) with role-based scoping
        entreprise_scope = Q()
        user = request.user
        if user.is_superuser:
            pass
        elif getattr(user, 'is_RC', False) and getattr(user, 'societe', None):
            entreprise_scope &= Q(societe=user.societe)
        elif getattr(user, 'is_RO', False):
            # Convert ManyRelatedManager to a list of IDs for the IN lookup
            societe_ids = list(user.societes.values_list('id', flat=True))
            entreprise_scope &= Q(societe_id__in=societe_ids) if societe_ids else Q(pk__isnull=True)
        else:
            if getattr(user, 'societe', None):
                entreprise_scope &= Q(societe=user.societe)
            else:
                # Aucun périmètre fiable sans filiale: ne rien retourner
                entreprise_scope &= Q(pk__isnull=True)

        recent_prospects = (
            Entreprise.objects.filter(Q(is_Prospect=True) & Q(date__gte=last_30) & entreprise_scope)
            .order_by('-date')[:10]
        )

        # Recent actions/events (latest 10)
        recent_actions = actions_qs.order_by('-date_heure_planifie')[:10]
        recent_events = Evenement.objects.filter(event_filter).order_by('-date_heure_planifie')[:10]

        context = {
            'counts': counts,
            'trends': trends,
            'upcoming_events': upcoming_events,
            'upcoming_actions': upcoming_actions,
            'recent_prospects': recent_prospects,
            'recent_actions': recent_actions,
            'recent_events': recent_events,
        }
        return render(request, self.template_name, context)

#appel
@login_required
def call_list(request):
    if request.user.is_superuser:
        calls = Action.objects.filter(is_Appel=True).order_by('-date_heure')
        societes = Societe.objects.all()
    elif getattr(request.user, 'is_RO', False):
        # Get the list of societe IDs for the admin
        societe_ids = list(request.user.societes.values_list('id', flat=True))
        if societe_ids:
            calls = Action.objects.filter(
                Q(is_Appel=True) & 
                (Q(societe_id__in=societe_ids) | 
                 Q(created_by__societe_id__in=societe_ids))).order_by('-date_heure')
            societes = Societe.objects.filter(id__in=societe_ids)
        else:
            calls = Action.objects.none()
            societes = Societe.objects.none()
    elif request.user.is_RC and request.user.societe:
        calls = Action.objects.filter(
            Q(is_Appel=True) & 
            (Q(societe=request.user.societe) | 
             Q(created_by__societe=request.user.societe))).order_by('-date_heure')
        societes = Societe.objects.filter(id=request.user.societe_id)
    else:
        calls = Action.objects.filter(is_Appel=True, created_by=request.user).order_by('-date_heure')
        societes = Societe.objects.filter(id=request.user.societe_id)
    
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    societe_filter = request.GET.get('societe', '')
    created_date_from = request.GET.get('created_from', '')
    created_date_to = request.GET.get('created_to', '')
    planned_date_from = request.GET.get('planned_from', '')
    planned_date_to = request.GET.get('planned_to', '')
    realized_date_from = request.GET.get('realized_from', '')
    realized_date_to = request.GET.get('realized_to', '')
    entreprise_filter = request.GET.get('entreprise', '')
    type_entreprise_filter = request.GET.get('type_entreprise', '')
    
    if search_query:
        calls = calls.filter(
            Q(sujet__icontains=search_query) |
            Q(compte_rendu__icontains=search_query) |
            Q(notes__icontains=search_query))
    
    if status_filter:
        calls = calls.filter(etat=status_filter)
    
    if societe_filter and (request.user.is_superuser or request.user.is_RC or request.user.is_RO):
        calls = calls.filter(societe__id=societe_filter)
        
    if type_entreprise_filter:
        if type_entreprise_filter == 'client':
            calls = calls.filter(entreprise__is_CLT=True)
        elif type_entreprise_filter == 'prospect':
            calls = calls.filter(entreprise__is_Prospect=True, entreprise__is_CLT=False)
        
    if entreprise_filter:
        calls = calls.filter(entreprise__id=entreprise_filter)
        
    entreprise_stats = calls.values('entreprise__id', 'entreprise__nom', 'entreprise__is_CLT') \
                       .filter(entreprise__is_Concurent=False, date_heure_realiser__isnull=False)\
                       .annotate(total_calls=Count('id')) \
                       .order_by('-entreprise__is_CLT', 'entreprise__nom')
                       
    clients_data = []
    prospects_data = []

    for stat in entreprise_stats:
        if stat['entreprise__is_CLT']:
            clients_data.append({
                'name': stat['entreprise__nom'],
                'count': stat['total_calls']
            })
        else:
            prospects_data.append({
                'name': stat['entreprise__nom'],
                'count': stat['total_calls']
            })
            
    monthly_stats = calls.exclude(date_heure_realiser__isnull=True) \
                    .filter(entreprise__is_Concurent=False) \
                    .annotate(month=TruncMonth('date_heure_realiser')) \
                    .values('month', 'entreprise__is_CLT', 'entreprise__is_Prospect') \
                    .annotate(total=Count('id')) \
                    .order_by('month')

    months = []
    client_counts = []
    prospect_counts = []
    total_counts = []

    for stat in monthly_stats:
        month_str = stat['month'].strftime("%Y-%m")
        if month_str not in months:
            months.append(month_str)
            client_counts.append(0)
            prospect_counts.append(0)
            total_counts.append(0)
        
        idx = months.index(month_str)
        if stat['entreprise__is_CLT']:
            client_counts[idx] = stat['total']
        elif stat['entreprise__is_Prospect']:
            prospect_counts[idx] = stat['total']
        
        total_counts[idx] += stat['total']
    
    if created_date_from:
        try:
            created_date_from = datetime.strptime(created_date_from, '%Y-%m-%d')
            calls = calls.filter(date_heure__gte=created_date_from)
        except ValueError:
            pass
    
    if created_date_to:
        try:
            created_date_to = datetime.strptime(created_date_to, '%Y-%m-%d')
            calls = calls.filter(date_heure__lte=created_date_to)
        except ValueError:
            pass
    
    if planned_date_from:
        try:
            planned_date_from = datetime.strptime(planned_date_from, '%Y-%m-%d')
            calls = calls.filter(date_heure_planifie__gte=planned_date_from)
        except ValueError:
            pass
    
    if planned_date_to:
        try:
            planned_date_to = datetime.strptime(planned_date_to, '%Y-%m-%d')
            calls = calls.filter(date_heure_planifie__lte=planned_date_to)
        except ValueError:
            pass
    
    if realized_date_from and realized_date_to:
        try:
            realized_date_from = datetime.strptime(realized_date_from, '%Y-%m-%d')
            realized_date_to = datetime.strptime(realized_date_to, '%Y-%m-%d')
            calls = calls.filter(
                date_heure_realiser__gte=realized_date_from,
                date_heure_realiser__lte=realized_date_to
            )
        except ValueError:
            pass
        
    entreprises = Entreprise.objects.filter(is_Concurent=False).order_by('nom')
    if not request.user.is_superuser:
        if getattr(request.user, 'is_RO', False):
            # Get the list of societe IDs for the admin
            societe_ids = list(request.user.societes.values_list('id', flat=True))
            if societe_ids:
                entreprises = entreprises.filter(societe_id__in=societe_ids)
            else:
                entreprises = entreprises.none()
        else:
            # For non-admin users, use their single societe
            societe = getattr(request.user, 'societe', None)
            if societe:
                entreprises = entreprises.filter(societe=societe)
            else:
                entreprises = entreprises.none()
    
    paginator = Paginator(calls, 100) 
    page_number = request.GET.get('page')
    calls = paginator.get_page(page_number)
    
    context = {
        'calls': calls,
        'is_superuser': request.user.is_superuser,
        'is_RC': request.user.is_RC,
        'is_RO': request.user.is_RO,
        'societes': societes,
        'status_choices': dict(Action.ETAT_CHOICES_APPEL),
        'search_query': search_query,
        'status_filter': status_filter,
        'societe_filter': societe_filter,
        'created_date_from': created_date_from,
        'created_date_to': created_date_to,
        'planned_date_from': planned_date_from,
        'planned_date_to': planned_date_to,
        'realized_date_from': realized_date_from,
        'realized_date_to': realized_date_to,
        'request': request,
        'clients_data': clients_data,
        'prospects_data': prospects_data,
        'entreprises': entreprises,
        'entreprise_filter': entreprise_filter,
        'type_entreprise_filter': type_entreprise_filter,
        'months': months,
        'client_counts': client_counts,
        'prospect_counts': prospect_counts,
        'total_counts': total_counts,
    }
    
    return render(request, 'adminlte/sales/sales/actions/appels.html', context)

@login_required
def get_all_filtered_calls(request):
    if request.user.is_superuser:
        calls = Action.objects.filter(is_Appel=True).order_by('-date_heure')
    elif getattr(request.user, 'is_RO', False):
        # Get the list of societe IDs for the admin
        societe_ids = list(request.user.societes.values_list('id', flat=True))
        if societe_ids:
            calls = Action.objects.filter(
                Q(is_Appel=True) & 
                (Q(societe_id__in=societe_ids) | 
                 Q(created_by__societe_id__in=societe_ids))).order_by('-date_heure')
        else:
            calls = Action.objects.none()
    elif getattr(request.user, 'is_RC', False) and hasattr(request.user, 'societe') and request.user.societe:
        calls = Action.objects.filter(
            Q(is_Appel=True) & 
            (Q(societe=request.user.societe) | 
             Q(created_by__societe=request.user.societe))).order_by('-date_heure')
    else:
        calls = Action.objects.filter(is_Appel=True, created_by=request.user).order_by('-date_heure')
    
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    societe_filter = request.GET.get('societe', '')
    created_date_from = request.GET.get('created_from', '')
    created_date_to = request.GET.get('created_to', '')
    planned_date_from = request.GET.get('planned_from', '')
    planned_date_to = request.GET.get('planned_to', '')
    realized_date_from = request.GET.get('realized_from', '')
    realized_date_to = request.GET.get('realized_to', '')
    
    if search_query:
        calls = calls.filter(
            Q(sujet__icontains=search_query) |
            Q(compte_rendu__icontains=search_query) |
            Q(notes__icontains=search_query))
    
    if status_filter:
        calls = calls.filter(etat=status_filter)
    
    if societe_filter and (request.user.is_superuser or request.user.is_RC or request.user.is_RO):
        calls = calls.filter(societe__id=societe_filter)
    
    if created_date_from:
        try:
            created_date_from = datetime.strptime(created_date_from, '%Y-%m-%d')
            calls = calls.filter(date_heure__gte=created_date_from)
        except ValueError:
            pass
    
    if created_date_to:
        try:
            created_date_to = datetime.strptime(created_date_to, '%Y-%m-%d')
            calls = calls.filter(date_heure__lte=created_date_to)
        except ValueError:
            pass
    
    if planned_date_from:
        try:
            planned_date_from = datetime.strptime(planned_date_from, '%Y-%m-%d')
            calls = calls.filter(date_heure_planifie__gte=planned_date_from)
        except ValueError:
            pass
    
    if planned_date_to:
        try:
            planned_date_to = datetime.strptime(planned_date_to, '%Y-%m-%d')
            calls = calls.filter(date_heure_planifie__lte=planned_date_to)
        except ValueError:
            pass
    
    if realized_date_from and realized_date_to:
        try:
            realized_date_from = datetime.strptime(realized_date_from, '%Y-%m-%d')
            realized_date_to = datetime.strptime(realized_date_to, '%Y-%m-%d')
            calls = calls.filter(
                date_heure_realiser__gte=realized_date_from,
                date_heure_realiser__lte=realized_date_to
            )
        except ValueError:
            pass
    
    return JsonResponse({
        'calls': [{
            'id': call.id,
            'sujet': call.sujet,
            'etat': call.etat,
            'date_heure_planifie': call.date_heure_planifie.isoformat() if call.date_heure_planifie else None,
            'date_heure_realiser': call.date_heure_realiser.isoformat() if call.date_heure_realiser else None,
            'notes': call.notes,
        } for call in calls]
    }, safe=False)

@require_GET
def get_calls_by_date(request):
    date_str = request.GET.get('date')
    date_type = request.GET.get('date_type', 'planned')
    
    try:
        date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return JsonResponse([], safe=False)
    
    # Base queryset with date constraint
    if date_type == 'planned':
        qs = Action.objects.filter(is_Appel=True, date_heure_planifie__date=date)
    else:
        qs = Action.objects.filter(is_Appel=True, date_heure_realiser__date=date)

    # Permission scoping
    if request.user.is_superuser:
        pass
    elif getattr(request.user, 'is_RO', False):
        # Get the list of societe IDs for the admin
        societe_ids = list(request.user.societes.values_list('id', flat=True))
        if societe_ids:
            qs = qs.filter(
                Q(societe_id__in=societe_ids) | 
                Q(created_by__societe_id__in=societe_ids)
            )
        else:
            qs = Action.objects.none()
    elif getattr(request.user, 'is_RC', False) and hasattr(request.user, 'societe') and request.user.societe:
        qs = qs.filter(
            Q(societe=request.user.societe) | 
            Q(created_by__societe=request.user.societe)
        )
    else:
        qs = qs.filter(created_by=request.user)

    calls_list = list(qs.values('id', 'sujet', 'date_heure_planifie', 'date_heure_realiser', 'etat', 'notes'))
    
    for call in calls_list:
        if 'date_heure_planifie' in call and call['date_heure_planifie']:
            call['date_heure_planifie'] = call['date_heure_planifie'].isoformat()
        if 'date_heure_realiser' in call and call['date_heure_realiser']:
            call['date_heure_realiser'] = call['date_heure_realiser'].isoformat()
    
    return JsonResponse(calls_list, safe=False)

@login_required
@require_POST
def update_action_date(request):
    try:
        action_id = request.POST.get('action_id')
        new_date_str = request.POST.get('new_date')
        date_type = request.POST.get('date_type')
        
        action = Action.objects.get(id=action_id)
        # Permission check: SU, RC (same societe), or creator
        if not (
            request.user.is_superuser or request.user.is_RO or
            (getattr(request.user, 'is_RC', False) and action.societe_id == getattr(request.user.societe, 'id', None)) or
            action.created_by_id == request.user.id
        ):
            return JsonResponse({'status': 'error', 'message': "Permission refusée"}, status=403)
        
        new_date = datetime.fromisoformat(new_date_str)
        
        if new_date.tzinfo is not None:
            new_date = new_date.astimezone(pytz.UTC)
            new_date = new_date.replace(tzinfo=None)  
            
        new_date = make_aware(new_date)
        
        if date_type == 'planned':
            action.date_heure_planifie = new_date
        elif date_type == 'realized':
            action.date_heure_realiser = new_date
        
        action.save()
        
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

def _fetch_client_from_api(numero_tiers: str, societe: Societe):
    if not numero_tiers or not societe:
        return None

    base_url = get_base_url(societe.id)
    if not base_url:
        logger.error("Base URL introuvable pour la société %s", societe.id if societe else None)
        return None

    url = f"{settings.SAGE_API_HOST}/WebServices100/{base_url}/TiersService/rest/Clients/{numero_tiers}"
    headers = {"Authorization": settings.SAGE_API_TOKEN, "Accept": "application/json"}

    try:
        r = requests.get(url, headers=headers, timeout=12)
        if r.status_code != 200:
            logger.error("API client %s -> %s : %s", numero_tiers, r.status_code, r.text[:300])
            return None
        data = r.json()
    except Exception as e:
        logger.exception("Erreur d'appel API client: %s", e)
        return None

    intitule = data.get("Intitule") or data.get("intitule") or f"Client {numero_tiers}"
    email = data.get("Email") or data.get("fact_email") or ""
    adresse = data.get("Adresse") or data.get("fact_adresse") or ""
    tel = data.get("Telephone") or data.get("Telephone1") or data.get("Telephone2") or data.get("fact_tel") or ""
    secteur = data.get("SecteurActivite") or ""

    return {
        "nom": intitule,
        "num_compte": data.get("NumeroTiers") or numero_tiers,
        "telephone": tel,
        "email": email,
        "adresse": adresse,
        "is_CLT": True,
        "is_Prospect": False,
        "is_Concurent": False,
        "societe": societe,
        "secteur_activite": secteur,
        "date": timezone.now().date(),
    }

def _get_or_create_client_by_numero(numero_tiers: str, societe: Societe):
    if not numero_tiers:
        return None
    existing = Entreprise.objects.filter(num_compte=numero_tiers, is_CLT=True).first()
    if existing:
        return existing
    payload = _fetch_client_from_api(numero_tiers, societe)
    if not payload:
        return None
    try:
        with transaction.atomic():
            return Entreprise.objects.create(**payload)
    except Exception as e:
        logger.exception("Erreur création client local depuis API: %s", e)
        return None

@login_required
def add_call(request):
    if request.method == 'POST':
        try:
            sujet = request.POST.get('sujet')
            compte_rendu = request.POST.get('compte_rendu', '')
            notes = request.POST.get('notes', '')
            etat = request.POST.get('etat', '')
            societe_id = request.POST.get('societe')
            date_heure_planifie = request.POST.get('date_heure_planifie')
            date_heure_realiser = request.POST.get('date_heure_realiser')
            entreprise_id = request.POST.get('entreprise')
            type_entreprise = request.POST.get('type_entreprise')
            
            errors = {}
            
            if not sujet:
                errors['sujet'] = 'Ce champ est obligatoire'
            if not date_heure_planifie:
                errors['date_heure_planifie'] = 'Ce champ est obligatoire'
            if etat and etat not in dict(Action.ETAT_CHOICES_APPEL):
                errors['etat'] = "L'état pour un appel doit être 'Réussi' ou 'Non Réussi'"

            if not entreprise_id:
                return JsonResponse({
                    'success': False,
                    'errors': {'entreprise': 'Veuillez sélectionner un prospect/client pour cet appel'}
                }, status=400)
                
            if not type_entreprise or type_entreprise not in ['client', 'prospect']:
                return JsonResponse({
                    'success': False,
                    'errors': {'type': 'Type d\'entreprise invalide. Doit être \'client\' ou \'prospect\''}
                }, status=400)
            
            if errors:
                return JsonResponse({'errors': errors}, status=400)
            
            # La filiale sera déterminée automatiquement en fonction du prospect/client choisi.
            societe = None
            if societe_id:
                try:
                    societe = Societe.objects.get(id=societe_id)
                except Societe.DoesNotExist:
                    societe = None
            if not societe and getattr(request.user, 'societe', None):
                societe = request.user.societe
            
            # Societe indiquée par l'autocomplete client (nom), si fournie
            client_societe_name = request.POST.get('client_societe_name', '').strip()
            societe_from_client_name = None
            if client_societe_name:
                societe_from_client_name = Societe.objects.filter(nom=client_societe_name).first()
                # Préférer la société du client pour la suite si présente
                if societe_from_client_name:
                    societe = societe_from_client_name
                
            entreprise = None
            if type_entreprise == 'prospect':
                if not entreprise_id:
                    return JsonResponse({'success': False, 'errors': {'entreprise': 'Veuillez sélectionner un prospect'}}, status=400)
                try:
                    entreprise = Entreprise.objects.get(id=entreprise_id)
                except (ValueError, Entreprise.DoesNotExist):
                    return JsonResponse({'success': False, 'errors': {'entreprise': 'Prospect sélectionné non valide'}}, status=400)
                    
            elif type_entreprise == 'client':
                if not entreprise_id:
                    return JsonResponse({'success': False, 'errors': {'entreprise': 'Numéro de compte client manquant'}}, status=400)
                try:
                    entreprise = Entreprise.objects.get(num_compte=entreprise_id, is_CLT=True)
                except Entreprise.DoesNotExist:
                    try:
                        # Déterminer la filiale à partir du client si possible
                        societe_client = None
                        try:
                            existing_client = (
                                Entreprise.objects
                                .filter(num_compte=entreprise_id)
                                .select_related('societe')
                                .first()
                            )
                            if existing_client and existing_client.societe_id:
                                societe_client = existing_client.societe
                        except Exception:
                            societe_client = None

                        # Si le nom de la société du client a été fourni par le front, on le privilégie
                        preferred_societe = societe_client or societe_from_client_name or societe or getattr(request.user, 'societe', None)
                        base_url = get_base_url(preferred_societe.id) if preferred_societe else None
                        if not base_url:
                            return JsonResponse(
                                {'success': False, 'errors': {'societe': 'Configuration de la société non trouvée'}}, 
                                status=400
                            )
                            
                        fact_url = f"{settings.SAGE_API_HOST}/WebServices100/{base_url}/TiersService/rest/Clients/{entreprise_id}"
                        headers = {
                            'Authorization': settings.SAGE_API_TOKEN,
                            'Accept': 'application/json'
                        }
                        
                        client_response = requests.get(fact_url, headers=headers, timeout=10)
                        if client_response.status_code != 200:
                            return JsonResponse(
                                {'success': False, 'errors': {'numTiers': 'Impossible de récupérer les informations du client'}}, 
                                status=400
                            )
                            
                        # Parsing JSON tolérant (réponses vides ou non-JSON)
                        try:
                            client_data = client_response.json() if client_response.content else {}
                        except ValueError:
                            client_data = {}
                        
                        # Déterminer la société préférée pour le client à créer
                        preferred_societe = societe_from_client_name or societe or getattr(request.user, 'societe', None)
                        # Créer le client directement
                        entreprise = Entreprise.objects.create(
                            nom=client_data.get('Intitule', f'Client {entreprise_id}'),
                            adresse=client_data.get('Adresse', ''),
                            telephone=client_data.get('Telephone', ''),
                            email=client_data.get('Email', ''),
                            num_compte=entreprise_id,
                            is_CLT=True,
                            is_Prospect=False,
                            is_Concurent=False,
                            societe=preferred_societe,
                            secteur_activite='',
                            date=timezone.now().date()
                        )
                        
                    except Exception as e:
                        import traceback
                        error_msg = f'Erreur lors de la création du client: {str(e)}'
                        logger.error(f"{error_msg}\n{traceback.format_exc()}")
                        return JsonResponse(
                            {'success': False, 'errors': {'numTiers': error_msg}}, 
                            status=400
                        )
            
            # Calcul final de la filiale de l'appel
            final_societe = None
            if entreprise and getattr(entreprise, 'societe', None):
                final_societe = entreprise.societe
            else:
                final_societe = societe or getattr(request.user, 'societe', None)

            pilote = None
            if final_societe:
                pilote = get_user_model().objects.filter(is_RC=True, societe=final_societe).first()
            
            Action.objects.create(
                sujet=sujet,
                compte_rendu=compte_rendu,
                notes=notes,
                etat=etat,
                societe=final_societe,
                created_by=request.user,
                pilote=pilote,
                is_Appel=True,
                entreprise= entreprise,
                date_heure_planifie=date_heure_planifie or None,
                date_heure_realiser=date_heure_realiser or None
            )
            
            return JsonResponse({
                'success': 'Appel ajouté avec succès',
                'redirect_url': reverse('prospection:call_list') 
            }, status=200)
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return render(request, 'adminlte/sales/sales/actions/appels.html')

def _normalize_numero_tiers(raw: str) -> str:
    """
    Reçoit une saisie venant du front:
      - "12345 - ACME SA"  -> "12345"
      - "  000012  "       -> "000012"
      - "C00123-Client X"  -> "C00123-Client X" (on ne supprime pas les lettres)
    On ne touche pas aux zéros de tête ni aux lettres éventuelles.
    """
    s = (raw or "").strip()
    # cas le plus courant "CODE - NOM"
    if " - " in s:
        s = s.split(" - ", 1)[0].strip()
    return s

def _resolve_contact(contact_raw, societe):
    """
    contact_raw peut être:
      - un ID local d'Entreprise (int/str d'int)
      - un NumeroTiers (str) ; ex "001234" ou "C00123"
      - éventuellement "001234 - Raison sociale"
    On tente ID local, sinon on tente NumeroTiers (création si absent).
    """
    if not contact_raw:
        return None

    # 1) Essai ID local
    try:
        as_int = int(str(contact_raw))
        ent = Entreprise.objects.filter(id=as_int).first()
        if ent:
            return ent
    except (ValueError, TypeError):
        pass

    # 2) NumeroTiers (nettoyé)
    numero_tiers = _normalize_numero_tiers(str(contact_raw))

    # Si on trouve déjà localement par num_compte, on renvoie
    ent = Entreprise.objects.filter(num_compte=numero_tiers).first()
    if ent:
        return ent

    # Sinon on crée depuis l'API distante
    return _get_or_create_client_by_numero(numero_tiers, societe)

@csrf_exempt
@login_required
def generate_action_report(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)

    try:
        data = json.loads(request.body or "{}")
    except json.JSONDecodeError as e:
        return JsonResponse({'error': f'JSON invalide: {str(e)}'}, status=400)

    # Champs reçus
    company_id = data.get('company')  # peut être None pour non-superuser
    contact_raw = data.get('contact')  # ID local OU NumeroTiers
    action_type = data.get('action_type')
    subject = data.get('subject', '')
    planned_date = data.get('planned_date', '')
    answers = data.get('answers', {})

    # Société
    try:
        if request.user.is_superuser or request.user.is_RO:
            if not company_id:
                return JsonResponse({'error': 'ID de filiale manquant'}, status=400)
            
            # Essayer d'abord par ID, puis par nom si ce n'est pas un entier
            try:
                # Si c'est un entier, rechercher par ID
                company = Societe.objects.get(id=int(company_id))
            except (ValueError, TypeError):
                # Si ce n'est pas un entier, rechercher par nom
                company = Societe.objects.get(nom=company_id)
        else:
            company = getattr(request.user, 'societe', None)
            if not company:
                return JsonResponse({'error': "Aucune filiale associée à l'utilisateur"}, status=400)
    except Societe.DoesNotExist:
        return JsonResponse({'error': "Filiale invalide"}, status=400)

    # Contact (entreprise)
    if not contact_raw:
        return JsonResponse({'error': 'Veuillez sélectionner une entreprise'}, status=400)

    contact = _resolve_contact(contact_raw, company)
    if not contact:
        return JsonResponse({'error': "Impossible de résoudre l'entreprise"}, status=400)

    # Validations
    if action_type not in ['call', 'email', 'appointment']:
        return JsonResponse({'error': "Type d'action invalide"}, status=400)
    if not subject:
        return JsonResponse({'error': 'Sujet manquant'}, status=400)

    # Templates
    templates = {
        'call': """
        Generate a concise professional call report in French (1 paragraph, 5-7 sentences) for quality control purposes.
        Include these elements:
        - Context: Call with the company {contact} from {company} regarding "{subject}" on {date}
        - Key discussion points: {q0}
        - Client's concerns or reactions: {q1}
        - Agreed actions or next steps: {q2}
        Requirements:
        - Neutral, factual tone
        - Highlight quality control aspects
        - Include sentiment indicators (positive/neutral/negative)
        - No bullet points or examples
        - Focus on factual observations
        - do not mark data other than what I give you, such as the number of participants
        - does not add additional data from you
        """,
        'email': """
        Generate a professional email summary in French (1 paragraph, 5-7 sentences) for quality control tracking.
        Include:
        - Context: Email to the company {contact} at {company} about "{subject}" sent on {date}
        - Main purpose and key content: {q0}
        - Sensitive or critical points addressed: {q1}
        - Required follow-up or response: {q2}
        Requirements:
        - Formal business style
        - Highlight quality/compliance aspects
        - Include tone assessment
        - No greetings or signatures
        - Pure factual summary
        - does not add additional data from you
        - do not mark data other than what I give you, such as the number of participants
        """,
        'appointment': """
        Generate a professional meeting report in French (1 paragraph, 5-7 sentences) for quality control records.
        Include:
        - Context: Meeting with the company {contact} from {company} about "{subject}" on {date}
        - Key discussion topics: {q0}
        - Participant engagement and reactions: {q1}
        - Decisions and action items: {q2}
        Requirements:
        - Professional but concise
        - Note any quality-related observations
        - Include engagement level (high/medium/low)
        - No opinions, only facts
        - Structured as single coherent paragraph
        - does not add additional data from you
        - do not mark data other than what I give you, such as the number of participants
        """
    }

    prompt = templates[action_type].format(
        contact=str(contact),  # __str__ de l’Entreprise
        company=str(company),
        subject=subject,
        date=planned_date,
        q0=answers.get('q0', ''),
        q1=answers.get('q1', ''),
        q2=answers.get('q2', ''),
    )

    # === ICI: appelle ta fonction de génération ===
    try:
        generated_text = generate_report(prompt)  # ta fonction existante
        clean_report = (generated_text or "").strip()
    except Exception as e:
        logger.exception("Erreur generate_report: %s", e)
        msg = str(e)
        if 'ResourceExhausted' in msg or '429' in msg or 'Quota exceeded' in msg:
            return JsonResponse(
                {'error': "Quota IA dépassé. Veuillez réessayer plus tard ou vérifier votre plan/facturation."},
                status=429,
            )
        if 'models/' in msg and ('not found' in msg.lower() or '404' in msg):
            return JsonResponse(
                {'error': "Modèle IA introuvable/non supporté. Vérifiez OPENAI_MODEL ou GEMINI_MODEL dans la configuration."},
                status=400,
            )
        return JsonResponse({'error': "Erreur lors de la génération du rapport"}, status=500)

    return JsonResponse({'report': clean_report, 'status': 'success'}, status=200)

@login_required
def edit_call(request, call_id):
    call = get_object_or_404(Action, id=call_id, is_Appel=True)
    
    if not (request.user.is_superuser or request.user.is_RO or
            (request.user.is_RC and call.societe == request.user.societe) or 
            call.created_by == request.user):
        return JsonResponse({'error': "Vous n'avez pas la permission de modifier cet appel"}, status=403)
    
    if request.method == 'POST':
        try:
            logger.info(f"Début de la modification de l'appel {call_id}")
            logger.debug(f"Données reçues: {dict(request.POST)}")
            
            # Récupération des données du formulaire
            call.sujet = request.POST.get('sujet', call.sujet)
            call.compte_rendu = request.POST.get('compte_rendu', call.compte_rendu)
            call.notes = request.POST.get('notes', call.notes)
            call.etat = request.POST.get('etat', call.etat)
            call.date_heure_planifie = request.POST.get('date_heure_planifie', call.date_heure_planifie)
            # Société du client (nom) transmise par l'autocomplete
            client_societe_name = (request.POST.get('client_societe_name') or '').strip()
            societe_from_client_name = Societe.objects.filter(nom=client_societe_name).first() if client_societe_name else None
            
            # Vérifier si c'est un client et si on a un numéro de compte
            type_entreprise = request.POST.get('type_entreprise')
            entreprise_id = request.POST.get('entreprise')
            
            logger.info(f"Type entreprise: {type_entreprise}, ID entreprise: {entreprise_id}")
            
            if type_entreprise == 'client' and entreprise_id:
                # Vérifier si c'est un ID numérique (existant) ou un numéro de compte (nouveau client)
                if not entreprise_id.isdigit():
                    # C'est un nouveau client, on doit le créer
                    try:
                        preferred_societe = societe_from_client_name or getattr(request.user, 'societe', None)
                        base_url = get_base_url(preferred_societe.id) if preferred_societe else None
                        if not base_url:
                            return JsonResponse(
                                {'error': 'Impossible de déterminer la base de données du client'}, 
                                status=400
                            )
                            
                        api_url = f"{settings.SAGE_API_HOST}/WebServices100/{base_url}/TiersService/rest/Clients/{entreprise_id}"
                        logger.info(f"Appel API pour les détails du client: {api_url}")
                        
                        response = requests.get(
                            api_url,
                            headers={
                                'Authorization': settings.SAGE_API_TOKEN,
                                'Accept': 'application/json'
                            },
                            timeout=10
                        )
                        
                        if response.status_code != 200:
                            error_msg = f"Échec de la récupération des données du client: {response.status_code} - {response.text}"
                            logger.error(error_msg)
                            return JsonResponse(
                                {'error': 'Impossible de récupérer les informations du client depuis le serveur distant'}, 
                                status=400
                            )
                        
                        try:
                            client_data = response.json()
                            logger.debug(f"Données client reçues: {client_data}")
                            
                            # Vérifier que les champs obligatoires sont présents
                            if not client_data.get('Intitule'):
                                raise ValueError("Le champ 'Intitule' est manquant dans la réponse de l'API")
                                
                            if not client_data.get('NumeroTiers'):
                                client_data['NumeroTiers'] = entreprise_id
                                
                            # Préparer les données pour la création
                            entreprise_data = {
                                'nom': client_data.get('Intitule', 'Nouveau Client'),
                                'num_compte': client_data.get('NumeroTiers', entreprise_id),
                                'telephone': client_data.get('Telephone1') or client_data.get('Telephone2', ''),
                                'email': client_data.get('Email', ''),
                                'adresse': client_data.get('Adresse', ''),
                                'is_CLT': True,
                                'is_Prospect': False,
                                'is_Concurent': False,
                                'societe': preferred_societe or request.user.societe,
                                'secteur_activite': client_data.get('SecteurActivite', ''),
                                'date': timezone.now().date()
                            }
                            
                            logger.info(f"Création du client avec les données: {entreprise_data}")
                            
                            # Créer le client dans la base de données
                            with transaction.atomic():
                                entreprise = Entreprise.objects.create(**entreprise_data)
                                logger.info(f"Client créé avec succès: ID={entreprise.id}, Nom={entreprise.nom}")
                                entreprise_id = str(entreprise.id)
                                
                        except json.JSONDecodeError as e:
                            error_msg = f"Erreur de décodage de la réponse JSON: {str(e)}"
                            logger.error(f"{error_msg} - Réponse: {response.text[:500]}")
                            return JsonResponse(
                                {'error': 'Erreur de format des données du client'}, 
                                status=400
                            )
                            
                        except Exception as e:
                            error_msg = f"Erreur lors du traitement des données du client: {str(e)}"
                            logger.error(error_msg, exc_info=True)
                            return JsonResponse(
                                {'error': 'Erreur lors du traitement des données du client'}, 
                                status=400
                            )
                            
                    except Exception as e:
                        logger.error(f"Erreur lors de la création du client: {str(e)}", exc_info=True)
                        return JsonResponse(
                            {'error': f'Erreur lors de la création du client: {str(e)}'}, 
                            status=400
                        )
            
            # Vérifier que l'entreprise existe avant de l'assigner
            try:
                logger.info(f"Recherche de l'entreprise avec le numéro de compte: {entreprise_id}")
                
                # D'abord essayer de trouver par numéro de compte
                entreprise = Entreprise.objects.filter(num_compte=entreprise_id).first()
                
                # Si pas trouvé, essayer par ID
                if not entreprise:
                    logger.info(f"Aucune entreprise trouvée avec le numéro de compte {entreprise_id}, recherche par ID...")
                    try:
                        entreprise = Entreprise.objects.get(id=entreprise_id)
                    except (ValueError, Entreprise.DoesNotExist):
                        pass
                
                if entreprise:
                    call.entreprise = entreprise
                    logger.info(f"Entreprise trouvée: {entreprise.nom} (ID: {entreprise.id}, Numéro: {entreprise.num_compte})")
                else:
                    logger.info(f"Aucune entreprise trouvée, création d'un nouveau client avec le numéro: {entreprise_id}")
                    try:
                        preferred_societe = societe_from_client_name or getattr(request.user, 'societe', None)
                        base_url = get_base_url(preferred_societe.id) if preferred_societe else None
                        if not base_url:
                            return JsonResponse(
                                {'error': 'Impossible de déterminer la base de données du client'}, 
                                status=400
                            )
                            
                        api_url = f"{settings.SAGE_API_HOST}/WebServices100/{base_url}/TiersService/rest/Clients/{entreprise_id}"
                        logger.info(f"Appel API pour les détails du client: {api_url}")
                        
                        response = requests.get(
                            api_url,
                            headers={
                                'Authorization': settings.SAGE_API_TOKEN,
                                'Accept': 'application/json'
                            },
                            timeout=10
                        )
                        
                        if response.status_code != 200:
                            error_msg = f"Échec de la récupération des données du client: {response.status_code} - {response.text}"
                            logger.error(error_msg)
                            return JsonResponse(
                                {'error': 'Impossible de récupérer les informations du client depuis le serveur distant'}, 
                                status=400
                            )
                        
                        client_data = response.json()
                        logger.debug(f"Données client reçues: {client_data}")
                        
                        # Vérifier que les champs obligatoires sont présents
                        if not client_data.get('Intitule'):
                            raise ValueError("Le champ 'Intitule' est manquant dans la réponse de l'API")
                            
                        if not client_data.get('NumeroTiers'):
                            client_data['NumeroTiers'] = entreprise_id
                            
                        # Préparer les données pour la création
                        entreprise_data = {
                            'nom': client_data.get('Intitule', 'Nouveau Client'),
                            'num_compte': client_data.get('NumeroTiers', entreprise_id),
                            'telephone': client_data.get('fact_tel', ''),
                            'email': client_data.get('fact_email', ''),
                            'adresse': client_data.get('fact_adresse', ''),
                            'is_CLT': True,
                            'is_Prospect': False,
                            'is_Concurent': False,
                            'societe': preferred_societe or request.user.societe,
                            'secteur_activite': client_data.get('SecteurActivite', ''),
                            'date': timezone.now().date()
                        }
                        
                        logger.info(f"Création du client avec les données: {entreprise_data}")
                        
                        # Créer le client dans la base de données
                        with transaction.atomic():
                            entreprise = Entreprise.objects.create(**entreprise_data)
                            logger.info(f"Client créé avec succès: ID={entreprise.id}, Nom={entreprise.nom}")
                            call.entreprise = entreprise
                            
                    except Exception as e:
                        error_msg = f"Erreur lors de la création du client: {str(e)}"
                        logger.error(error_msg, exc_info=True)
                        return JsonResponse(
                            {'error': f"Impossible de créer le client: {str(e)}"}, 
                            status=400
                        )
                    
            except Exception as e:
                error_msg = f"Erreur lors de la recherche de l'entreprise: {str(e)}"
                logger.error(error_msg, exc_info=True)
                return JsonResponse(
                    {'error': "Erreur lors de la recherche de l'entreprise"}, 
                    status=500
                )
            
            # Mettre à jour la société si nécessaire
            societe_id = request.POST.get('societe', call.societe_id)
            logger.info(f"Société ID: {societe_id} (actuelle: {call.societe_id})")
            
            # Mettre à jour la date de modification
            call.updated_at = timezone.now()
            call.updated_by = request.user
            
            # Validation des champs obligatoires
            errors = {}
            
            if not call.sujet:
                errors['sujet'] = 'Ce champ est obligatoire'
                logger.warning("Champ 'sujet' manquant")
                
            if not call.date_heure_planifie:
                errors['date_heure_planifie'] = 'Ce champ est obligatoire'
                logger.warning("Champ 'date_heure_planifie' manquant")
                
            # ne pas imposer la sélection manuelle d'une filiale
                
            if not call.entreprise_id:
                errors['entreprise'] = 'Ce champ est obligatoire'
                logger.warning("Aucune entreprise associée à l'appel")
            
            if errors:
                logger.error(f"Erreurs de validation: {errors}")
                return JsonResponse({'success': False, 'errors': errors}, status=400)
            
            try:
                # Déterminer automatiquement la filiale de l'appel en fonction de l'entreprise
                if call.entreprise and getattr(call.entreprise, 'societe', None):
                    call.societe = call.entreprise.societe
                elif getattr(request.user, 'societe', None):
                    call.societe = request.user.societe
                
                # Mise à jour de la date de réalisation si fournie
                date_heure_realiser = request.POST.get('date_heure_realiser')
                if date_heure_realiser:
                    logger.info(f"Mise à jour de la date de réalisation: {date_heure_realiser}")
                    call.date_heure_realiser = date_heure_realiser
                
                # Sauvegarde de l'appel
                call.save()
                logger.info(f"Appel {call_id} mis à jour avec succès")
                
                # Retourner une réponse JSON de succès
                return JsonResponse({
                    'success': True,
                    'message': 'Appel mis à jour avec succès',
                    'redirect_url': reverse('prospection:call_list')
                })
            
            except Societe.DoesNotExist:
                error_msg = f"La société avec l'ID {societe_id} n'existe pas"
                logger.error(error_msg)
                return JsonResponse(
                    {'error': error_msg}, 
                    status=400
                )
            except Exception as e:
                error_msg = f"Erreur lors de la mise à jour de l'appel: {str(e)}"
                logger.error(error_msg, exc_info=True)
                return JsonResponse({
                    'success': False,
                    'error': "Une erreur est survenue lors de la mise à jour de l'appel"
                }, status=500)
        
        except Exception as e:
            error_msg = f"Erreur inattendue: {str(e)}"
            logger.error(error_msg, exc_info=True)
            return JsonResponse(
                {'error': "Une erreur inattendue est survenue"}, 
                status=500
            )
    
    return render(request, 'adminlte/sales/sales/actions/appels.html')

@login_required
def delete_call(request, call_id):
    call = get_object_or_404(Action, id=call_id, is_Appel=True)
    
    if not (request.user.is_superuser or request.user.is_RO or
            (request.user.is_RC and call.societe == request.user.societe) or 
            call.created_by == request.user):
        return JsonResponse({'error': "Vous n'avez pas l'autorisation de supprimer cet appel"}, status=403)
    
    if request.method == 'POST':
        try:
            call.delete()
            return JsonResponse({'success': 'Appel supprimé avec succès!'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Méthode de demande invalide'}, status=405)

@login_required
def call_details(request, call_id):
    call = get_object_or_404(Action, id=call_id, is_Appel=True)
    
    if not (request.user.is_superuser or request.user.is_RO or
            (request.user.is_RC and call.societe == request.user.societe) or 
            (call.created_by and call.created_by == request.user)):
        return JsonResponse({'error': "Vous n'avez pas la permission de voir cet appel"}, status=403)
    
    data = {
        'id': call.id,
        'sujet': call.sujet,
        'compte_rendu': call.compte_rendu,
        'notes': call.notes,
        'etat': call.etat,
        'societe': call.societe.nom if call.societe else '',
        'entreprise': call.entreprise.nom if call.entreprise else '',
        'societe_id': call.societe.id if call.societe else None,
        'created_by': call.created_by.username if call.created_by else '',
        'pilote': call.pilote.username if call.pilote else '',
        'date_heure': call.date_heure.strftime('%Y-%m-%d %H:%M'),
        'date_heure_planifie': call.date_heure_planifie.strftime('%Y-%m-%d %H:%M') if call.date_heure_planifie else '',
        'date_heure_realiser': call.date_heure_realiser.strftime('%Y-%m-%d %H:%M') if call.date_heure_realiser else '',
    }
    
    return JsonResponse(data)

def send_call_details_email(request):
    if request.method == 'POST':
        try:
            logo_path = finders.find('dist/img/abserveLogo.png')
            
            if not logo_path:
                raise FileNotFoundError(
                    "Logo introuvable. Chemins vérifiés:\n" + 
                    "\n".join(finders.searched_locations)
                )
            
            with open(logo_path, "rb") as image_file:
                logo_base64 = base64.b64encode(image_file.read()).decode('utf-8')
            recipient_email = request.POST.get('recipient_email')
            subject = request.POST.get('subject')
            message = request.POST.get('message', '')
            call_details = {
                'sujet': request.POST.get('call_details[sujet]'),
                'entreprise': request.POST.get('call_details[entreprise]'),
                'societe': request.POST.get('call_details[societe]'),
                'createdBy': request.POST.get('call_details[createdBy]'),
                'pilote': request.POST.get('call_details[pilote]'),
                'dateHeure': request.POST.get('call_details[dateHeure]'),
                'datePlanifie': request.POST.get('call_details[datePlanifie]'),
                'dateRealiser': request.POST.get('call_details[dateRealiser]'),
                'etat': request.POST.get('call_details[etat]'),
                'compteRendu': request.POST.get('call_details[compteRendu]'),
                'additional_message': message,
                'logo_base64': logo_base64
            }
            
            html_string = render_to_string('adminlte/emails/call_details.html', call_details)
            
            pdf_file = BytesIO()
            pisa_status = pisa.CreatePDF(html_string, dest=pdf_file, encoding='UTF-8', link_callback=lambda uri, _: uri)
            if pisa_status.err:
                return JsonResponse({'success': False, 'error': 'Erreur de génération PDF'})
            pdf_file.seek(0)
            
            email = EmailMessage(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [recipient_email]
            )
            email.attach(f'details_appel.pdf', pdf_file.getvalue(), 'application/pdf')
            email.send()
            
            return JsonResponse({'success': True})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Méthode de demande invalide'})

#event
@login_required
def event_list(request, typeEvent):
    user = request.user
    if user.is_superuser:
        events = Evenement.objects.filter(type=typeEvent).order_by('-date')
        societes = Societe.objects.all()
    elif getattr(user, 'is_RO', False):
        ids = list(user.societes.values_list('id', flat=True))
        events = (
            Evenement.objects
            .filter(Q(type=typeEvent) & (Q(societe_id__in=ids) | Q(created_by__societe_id__in=ids)))
            .order_by('-date')
        )
        societes = user.societes.all()
    elif user.is_RC and user.societe:
        events = Evenement.objects.filter(
            Q(type=typeEvent) & (Q(societe=user.societe) | Q(created_by__societe=user.societe))
        ).order_by('-date')
        societes = Societe.objects.filter(id=user.societe_id)
    elif getattr(user, 'is_C', False):
        events = Evenement.objects.filter(
            Q(type=typeEvent) & (Q(societe=user.societe) | Q(created_by__societe=user.societe))
        ).order_by('-date')
        societes = Societe.objects.filter(id=user.societe_id)
    else:
        events = Evenement.objects.filter(type=typeEvent, created_by=user).order_by('-date')
        societes = Societe.objects.filter(id=user.societe_id)
    
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    societe_filter = request.GET.get('societe', '')
    created_date_from = request.GET.get('created_from', '')
    created_date_to = request.GET.get('created_to', '')
    planned_date_from = request.GET.get('planned_from', '')
    planned_date_to = request.GET.get('planned_to', '')
    realized_date_from = request.GET.get('realized_from', '')
    realized_date_to = request.GET.get('realized_to', '')
    entreprise_filter = request.GET.get('entreprise', '')
    categorie_filter = request.GET.get('categorie', '')
    
    # Données pour les événements planifiés par mois
    planned_data = (
        events.filter(type=typeEvent)
        .annotate(month=TruncMonth('date_heure_planifie'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )
    
    # Données pour les événements réalisés par mois
    realized_data = (
        events.filter(type=typeEvent, date_heure_realiser__isnull=False)
        .annotate(month=TruncMonth('date_heure_realiser'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )
    
    # Préparation des labels (mois) et datasets
    months = list(set(
        [item['month'].strftime("%Y-%m") for item in planned_data] +
        [item['month'].strftime("%Y-%m") for item in realized_data if item['month']]
    ))
    months.sort()
    
    planned_counts = {item['month'].strftime("%Y-%m"): item['count'] for item in planned_data}
    realized_counts = {item['month'].strftime("%Y-%m"): item['count'] for item in realized_data if item['month']}
    
    chart_data = {
        'labels': months,
        'datasets': [
            {
                'label': 'Planifié',
                'data': [planned_counts.get(month, 0) for month in months],
                'borderColor': '#4e73df',
                'backgroundColor': 'rgba(54, 162, 235, 0.2)',
                'tension': 0.1
            },
            {
                'label': 'Complété',
                'data': [realized_counts.get(month, 0) for month in months],
                'borderColor': '#1cc88a',
                'backgroundColor': 'rgba(75, 192, 192, 0.2)',
                'tension': 0.1
            }
        ]
    }
    
    sector_data = (
        events.filter(type=typeEvent, date_heure_realiser__isnull=False)
        .values('secteur_activite')
        .annotate(total=Count('id'))
        .order_by('-total')[:10]  # Top 10 secteurs
    )
    
    sector_chart_data = {
        'labels': [item['secteur_activite'] or 'Non spécifié' for item in sector_data],
        'datasets': [{
            'label': "Nombre d'événements",
            'data': [item['total'] for item in sector_data],
            'backgroundColor': [
                '#4e73df', '#1cc88a', '#36b9cc', '#f6c23e', 
                '#e74a3b', '#858796', '#6f42c1', '#fd7e14',
                '#20c997', '#e83e8c'
            ]
        }]
    }
    
    if search_query:
        events = events.filter(
            Q(nom__icontains=search_query) |
            Q(lieu__icontains=search_query) |
            Q(secteur_activite__icontains=search_query) |
            Q(notes__icontains=search_query))
    
    if status_filter:
        events = events.filter(etat=status_filter)
        
    if categorie_filter:
        events = events.filter(categorie=categorie_filter)
    
    if societe_filter and (request.user.is_superuser or request.user.is_RO):
        events = events.filter(societe__id=societe_filter)
        
    if entreprise_filter:
        events = events.filter(entreprise__id=entreprise_filter)
    
    if created_date_from:
        try:
            created_date_from = datetime.strptime(created_date_from, '%Y-%m-%d')
            events = events.filter(date__gte=created_date_from)
        except ValueError:
            pass
    
    if created_date_to:
        try:
            created_date_to = datetime.strptime(created_date_to, '%Y-%m-%d')
            events = events.filter(date__lte=created_date_to)
        except ValueError:
            pass
    
    if planned_date_from:
        try:
            planned_date_from = datetime.strptime(planned_date_from, '%Y-%m-%d')
            events = events.filter(date_heure_planifie__gte=planned_date_from)
        except ValueError:
            pass
    
    if planned_date_to:
        try:
            planned_date_to = datetime.strptime(planned_date_to, '%Y-%m-%d')
            events = events.filter(date_heure_planifie__lte=planned_date_to)
        except ValueError:
            pass
    
    if realized_date_from and realized_date_to:
        try:
            realized_date_from = datetime.strptime(realized_date_from, '%Y-%m-%d')
            realized_date_to = datetime.strptime(realized_date_to, '%Y-%m-%d')
            events = events.filter(
                date_heure_realiser__gte=realized_date_from,
                date_heure_realiser__lte=realized_date_to
            )
        except ValueError:
            pass
        
    entreprises = Entreprise.objects.all().order_by('nom')
    if typeEvent == 'interne':
        entreprises = entreprises.filter(is_Concurent=False, is_CLT=True)
    else:
        entreprises = entreprises.filter(is_Concurent=False, is_Prospect=True, is_CLT=False)
    if not request.user.is_superuser and request.user.societe:
        entreprises = entreprises.filter(societe=request.user.societe)
    
    paginator = Paginator(events, 100)
    page_number = request.GET.get('page')
    events = paginator.get_page(page_number)
    
    # Préparer les catégories en fonction du type d'événement
    if typeEvent == 'interne':
        # Pour les événements internes, exclure 'salon'
        categorie_choices = [choice for choice in Evenement.CATEGORIE_CHOICES if choice[0] != 'salon']
    else:
        # Pour les événements externes, inclure toutes les catégories
        categorie_choices = Evenement.CATEGORIE_CHOICES
    
    context = {
        'events': events,
        'is_superuser': request.user.is_superuser,
        'is_RC': request.user.is_RC,
        'is_RO': request.user.is_RO,
        'societes': societes,
        'status_choices': dict(Evenement.ETAT_CHOICES_EVENT),
        'search_query': search_query,
        'status_filter': status_filter,
        'societe_filter': societe_filter,
        'created_date_from': created_date_from,
        'created_date_to': created_date_to,
        'planned_date_from': planned_date_from,
        'planned_date_to': planned_date_to,
        'realized_date_from': realized_date_from,
        'realized_date_to': realized_date_to,
        'CATEGORIE_CHOICES': categorie_choices,
        'categorie_filter': categorie_filter,
        'request': request,
        'entreprises': entreprises,
        'entreprise_filter': entreprise_filter,
        'type_event': typeEvent,
        'chart_data': json.dumps(chart_data),
        'sector_chart_data': json.dumps(sector_chart_data)
    }
     
    if typeEvent == 'interne':
        return render(request, 'adminlte/sales/sales/events/interne.html', context)
    else:
        return render(request, 'adminlte/sales/sales/events/externe.html', context)

@login_required
def add_event(request, typeEvent):
    if request.method == 'POST':
        try:
            nom = request.POST.get('nom')
            notes = request.POST.get('notes', '')
            lieu = request.POST.get('lieu', '')
            secteur_activite = request.POST.get('secteur_activite', '')
            etat = request.POST.get('etat', '')
            societe_id = request.POST.get('societe')
            date_heure_planifie = request.POST.get('date_heure_planifie')
            date_heure_realiser = request.POST.get('date_heure_realiser')
            entreprise_id = request.POST.get('entreprise')
            categorie = request.POST.get('categorie')
            num_compte = request.POST.get('num_compte')
            client_societe_name = (request.POST.get('client_societe_name') or '').strip()
            errors = {}
            
            if not nom:
                errors['nom'] = 'Ce champ est obligatoire'
            if not date_heure_planifie:
                errors['date_heure_planifie'] = 'Ce champ est obligatoire'
            if not lieu:
                errors['lieu'] = 'Ce champ est obligatoire'
            if not secteur_activite:
                errors['secteur'] = 'Ce champ est obligatoire'
            # Le champ prospect est maintenant optionnel pour les événements externes
            
            if errors:
                return JsonResponse({'errors': errors}, status=400)
            
            # Résolution de la société: nom fourni (client_societe_name) > id fourni > user.societe
            societe = None
            if client_societe_name:
                societe = Societe.objects.filter(nom__iexact=client_societe_name).first()
            if not societe and societe_id:
                try:
                    societe = Societe.objects.get(id=societe_id)
                except Societe.DoesNotExist:
                    societe = None
            if not societe and request.user.societe:
                societe = request.user.societe
                
            entreprise = None
                
            if num_compte and num_compte != 'null' and num_compte != '':
                # Si c'est un événement interne, on vérifie d'abord par numéro de compte
                if typeEvent == 'interne':
                    # societe déjà résolue ci-dessus. Si null, tenter par nom du poste legacy societe-name
                    if not societe:
                        legacy_nom = (request.POST.get('societe-name') or '').strip()
                        if legacy_nom:
                            societe = Societe.objects.filter(nom__iexact=legacy_nom).first()
                    try:
                        # Essayer de trouver l'entreprise par numéro de compte
                        if societe:
                            entreprise = Entreprise.objects.get(num_compte=num_compte, is_CLT=True, societe=societe)
                        else:
                            entreprise = Entreprise.objects.get(num_compte=num_compte, is_CLT=True)
                    except Entreprise.DoesNotExist:
                        
                        try:
                            # Récupérer les informations du client depuis l'API
                            base_url = get_base_url(societe.id if societe else None)
                            if not base_url:
                                return JsonResponse(
                                    {'error': 'Configuration de la société non trouvée'}, 
                                    status=400
                                )
                                
                            fact_url = f"{settings.SAGE_API_HOST}/WebServices100/{base_url}/TiersService/rest/Clients/{num_compte}"
                            headers = {
                                'Authorization': settings.SAGE_API_TOKEN,
                                'Accept': 'application/json'
                            }
                            
                            client_response = requests.get(fact_url, headers=headers, timeout=15)
                            if client_response.status_code != 200:
                                return JsonResponse(
                                    {'error': 'Impossible de récupérer les informations du client'}, 
                                    status=400
                                )

                            # Parse JSON en toute sécurité
                            try:
                                client_data = client_response.json()
                            except ValueError:
                                client_data = {}
                            
                            entreprise = Entreprise.objects.create(
                                nom=client_data.get('Intitule', f'Client {num_compte}'),
                                adresse=client_data.get('Adresse', ''),
                                telephone=client_data.get('Telephone', ''),
                                email=client_data.get('Email', ''),
                                num_compte=num_compte,
                                is_CLT=True,
                                is_Prospect=False,
                                is_Concurent=False,
                                societe=societe,
                                secteur_activite='',
                                date=timezone.now().date()
                            )
                            
                        except Exception as e:
                            import traceback
                            logger.error(f"Error creating client: {str(e)}\n{traceback.format_exc()}")
                            return JsonResponse(
                                {'error': f'Erreur lors de la création du client: {str(e)}'}, 
                                status=400
                            )
                else:
                    # Pour les événements externes, on garde l'ancien comportement
                    try:
                        entreprise = Entreprise.objects.get(id=entreprise_id)
                    except (ValueError, Entreprise.DoesNotExist):
                        return JsonResponse(
                            {'error': 'Entreprise sélectionnée non valide'}, 
                            status=400
                        )
            elif entreprise_id and entreprise_id != '' and entreprise_id != 'null':
                # Si entreprise_id est fourni mais pas num_compte
                try:
                    entreprise = Entreprise.objects.get(id=entreprise_id)
                except (ValueError, Entreprise.DoesNotExist):
                    return JsonResponse(
                        {'error': 'Entreprise sélectionnée non valide'}, 
                        status=400
                    )
            # Sinon, entreprise reste None (optionnel)
            # Si l'entreprise résolue a une filiale, on l'utilise comme filiale d'événement
            if entreprise and getattr(entreprise, 'societe', None):
                societe = entreprise.societe

            pilote = None
            if societe:
                pilote = get_user_model().objects.filter(is_RC=True, societe=societe).first()
            
            event = Evenement.objects.create(
                nom=nom,
                lieu=lieu,
                secteur_activite=secteur_activite,
                type=typeEvent,
                notes=notes,
                etat=etat,
                societe=societe,
                created_by=request.user,
                pilote=pilote,
                categorie=categorie,
                entreprise= entreprise,
                date_heure_planifie=date_heure_planifie or None,
                date_heure_realiser=date_heure_realiser or None
            )
            
            return JsonResponse({
                'success': 'Événement ajouté avec succès',
                'redirect_url': reverse('prospection:event_list', kwargs={'typeEvent': typeEvent})
            }, status=200)
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    if typeEvent:
        return render(request, 'adminlte/sales/sales/events/interne.html')
    else:
        return render(request, 'adminlte/sales/sales/events/externe.html')

@login_required
def edit_event(request, event_id):
    event = get_object_or_404(Evenement, id=event_id)
    
    if not (request.user.is_superuser or request.user.is_RO or
            (request.user.is_RC and event.societe == request.user.societe) or 
            event.created_by == request.user):
        return JsonResponse({'error': "Vous n'avez pas la permission de modifier cet événement"}, status=403)

    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire sans modifier l'objet
            nom = request.POST.get('nom', event.nom)
            lieu = request.POST.get('lieu', event.lieu)
            secteur_activite = request.POST.get('secteur_activite', event.secteur_activite)
            notes = request.POST.get('notes', event.notes)  # Garder l'ancienne valeur par défaut
            etat = request.POST.get('etat', event.etat)  # Garder l'ancienne valeur par défaut
            date_heure_planifie = request.POST.get('date_heure_planifie', event.date_heure_planifie)
            # Ne pas mélanger l'ID entreprise (BD) avec le numéro de compte (ERP)
            entreprise_id = request.POST.get('entreprise')
            categorie = request.POST.get('categorie', event.categorie)  # Garder l'ancienne valeur par défaut
            num_compte = request.POST.get('num_compte')
            print(num_compte) 
            # Si num_compte absent mais une entreprise est choisie, essayer de le déduire
            if (not num_compte or num_compte in ['null', '']) and entreprise_id and event.type != 'interne':
                try:
                    _ent = Entreprise.objects.get(id=entreprise_id)
                    if getattr(_ent, 'num_compte', None):
                        num_compte = _ent.num_compte
                except (ValueError, Entreprise.DoesNotExist):
                    pass
            societe_id = request.POST.get('societe')
            client_societe_name = (request.POST.get('client_societe_name') or '').strip()
            errors = {}
            
            if not nom:
                errors['nom'] = 'Ce champ est obligatoire'
            if not date_heure_planifie:
                errors['date_heure_planifie'] = 'Ce champ est obligatoire'
            if not lieu:
                errors['lieu'] = 'Ce champ est obligatoire'
            if not secteur_activite:
                errors['secteur_activite'] = 'Ce champ est obligatoire'
            
            # Si des erreurs, on les retourne sans rien modifier
            if errors:
                return JsonResponse({'errors': errors}, status=400)

            # À ce stade, toutes les validations sont passées, on peut modifier l'objet
            event.nom = nom
            event.lieu = lieu
            event.secteur_activite = secteur_activite
            event.notes = notes
            event.etat = etat
            event.date_heure_planifie = date_heure_planifie
            event.categorie = categorie

            # Résolution de la société: nom fourni (client_societe_name) > id fourni > user.societe
            societe = None
            if client_societe_name:
                societe = Societe.objects.filter(nom__iexact=client_societe_name).first()
            if not societe and societe_id:
                try:
                    societe = Societe.objects.get(id=societe_id)
                except Societe.DoesNotExist:
                    societe = None
            if not societe and request.user.societe:
                societe = request.user.societe
                
            entreprise = None
            if entreprise_id and entreprise_id != '' and entreprise_id != 'null':
                # Si c'est un événement interne, on vérifie d'abord par numéro de compte
                if event.type == 'interne': 
                    if not societe: 
                        legacy_nom = (request.POST.get('societe-name') or '').strip()
                        if legacy_nom:
                            societe = Societe.objects.filter(nom=legacy_nom).first()
                    try:
                        if societe:
                            entreprise = Entreprise.objects.get(num_compte=entreprise_id, is_CLT=True, societe=societe)
                        event.entreprise = entreprise
                        if getattr(entreprise, 'societe', None):
                            event.societe = entreprise.societe
                    except Entreprise.DoesNotExist:
                        # Créer directement le client sans passer par le signal
                        try:
                            # Récupérer les informations du client depuis l'API
                            base_url = get_base_url((event.societe.id if event.societe else (societe.id if societe else None)))
                            if not base_url:
                                return JsonResponse(
                                    {'error': 'Configuration de la société non trouvée'}, 
                                    status=400
                                )
                                
                            fact_url = f"{settings.SAGE_API_HOST}/WebServices100/{base_url}/TiersService/rest/Clients/{num_compte}"
                            headers = {
                                'Authorization': settings.SAGE_API_TOKEN,
                                'Accept': 'application/json'
                            }
                            
                            client_response = requests.get(fact_url, headers=headers, timeout=15)
                            if client_response.status_code != 200:
                                return JsonResponse(
                                    {'error': 'Impossible de récupérer les informations du client'}, 
                                    status=400
                                )
                            # Sécuriser le parsing JSON
                            try:
                                client_data = client_response.json()
                            except ValueError:
                                client_data = {} 
                            
                            # Créer le client directement
                            entreprise = Entreprise.objects.create(
                                nom=client_data.get('Intitule', f'Client {entreprise_id}'),
                                adresse=client_data.get('Adresse', ''),
                                telephone=client_data.get('Telephone', ''),
                                email=client_data.get('Email', ''),
                                num_compte=entreprise_id,
                                is_CLT=True,
                                is_Prospect=False,
                                is_Concurent=False,
                                societe=event.societe,
                                secteur_activite='',
                                date=timezone.now().date()
                            )
                            event.entreprise = entreprise
                            if getattr(entreprise, 'societe', None):
                                event.societe = entreprise.societe
                            event.save()
                        
                        except Exception as e:
                            import traceback
                            logger.error(f"Error creating client: {str(e)}\n{traceback.format_exc()}")
                            return JsonResponse(
                                {'error': f'Erreur lors de la création du client: {str(e)}'}, 
                                status=400
                            )
                else:
                    # Pour les événements externes, on garde l'ancien comportement
                    try:
                        entreprise = Entreprise.objects.get(id=entreprise_id)
                        event.entreprise = entreprise
                        if getattr(entreprise, 'societe', None):
                            event.societe = entreprise.societe
                    except (ValueError, Entreprise.DoesNotExist):
                        return JsonResponse(
                            {'error': 'Entreprise sélectionnée non valide'}, 
                            status=400
                        )
            # Sinon, on ne modifie pas le client (garder la valeur existante)
            
            # Si après les opérations ci-dessus, la société de l'événement est encore vide mais résolue localement, l'appliquer
            if not getattr(event, 'societe', None) and societe:
                event.societe = societe
            
            # Gestion de la date de réalisation si fournie
            date_heure_realiser = request.POST.get('date_heure_realiser')
            if date_heure_realiser:
                event.date_heure_realiser = date_heure_realiser
            
            # Sauvegarde des modifications
            event.save()
            
            # Retourner une réponse JSON pour les requêtes AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': 'Événement mis à jour avec succès',
                    'redirect_url': reverse('prospection:event_list', kwargs={'typeEvent': event.type})
                })
            
            # Pour les requêtes normales (non-AJAX)
            return redirect('prospection:event_list', typeEvent=event.type)
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    if event.type:
        return render(request, 'adminlte/sales/sales/events/interne.html')
    else:
        return render(request, 'adminlte/sales/sales/events/externe.html')

@login_required
def delete_event(request, event_id):
    event = get_object_or_404(Evenement, id=event_id)
    
    if not (request.user.is_superuser or request.user.is_RO or
            (request.user.is_RC and event.societe == request.user.societe) or 
            event.created_by == request.user):
        return JsonResponse({'error': "Vous n'avez pas l'autorisation de supprimer cet événement"}, status=403)
    
    if request.method == 'POST':
        try:
            event.delete()
            return JsonResponse({'success': 'Événement supprimé avec succès!'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Méthode de demande invalide'}, status=405)

@login_required
def event_details(request, event_id):
    event = get_object_or_404(Evenement, id=event_id)
    
    if not (request.user.is_superuser or request.user.is_RO or
            (request.user.is_RC and event.societe == request.user.societe) or 
            event.created_by == request.user):
        return JsonResponse({'error': "Vous n'avez pas la permission de voir cet événement"}, status=403)
    
    data = {
        'id': event.id,
        'nom': event.nom,
        'notes': event.notes,
        'etat': event.etat,
        'lieu': event.lieu,
        'categorie': event.categorie,
        'secteur_activite': event.secteur_activite,
        'societe': event.societe.nom if event.societe else '',
        'entreprise': event.entreprise.nom if event.entreprise else '',
        'societe_id': event.societe.id if event.societe else None,
        'created_by': event.created_by.username,
        'pilote': event.pilote.username if event.pilote else '',
        'date': event.date.strftime('%Y-%m-%d %H:%M'),
        'date_heure_planifie': event.date_heure_planifie.strftime('%Y-%m-%d %H:%M') if event.date_heure_planifie else '',
        'date_heure_realiser': event.date_heure_realiser.strftime('%Y-%m-%d %H:%M') if event.date_heure_realiser else '',
    }
    
    return JsonResponse(data)

@login_required
def action_details(request, action_id):
    action = get_object_or_404(Action, id=action_id)
    
    if not (request.user.is_superuser or request.user.is_RO or
            (request.user.is_RC and action.societe == request.user.societe) or 
            (action.created_by and action.created_by == request.user)):
        return JsonResponse({'error': "Vous n'avez pas l'autorisation de voir cette action"}, status=403)
    
    action_type = None
    if action.is_Appel:
        action_type = 'Call'
    elif action.is_Email:
        action_type = 'Email'
    elif action.is_RV:
        action_type = 'Meeting'
    
    data = {
        'id': action.id,
        'type': action_type,
        'subject': action.sujet,
        'status': action.etat,
        'report': action.compte_rendu,
        'notes': action.notes,
        'company': action.societe.nom if action.societe else '',
        'enterprise': action.entreprise.nom if action.entreprise else '',
        'created_by': action.created_by.username if action.created_by else '',
        'pilot': action.pilote.username if action.pilote else '',
        'creation_date': action.date_heure.strftime('%Y-%m-%d %H:%M'),
        'planned_date': action.date_heure_planifie.strftime('%Y-%m-%d %H:%M') if action.date_heure_planifie else '',
        'completed_date': action.date_heure_realiser.strftime('%Y-%m-%d %H:%M') if action.date_heure_realiser else '',
    }
    
    return JsonResponse(data)

@require_GET
def get_events_by_date(request, typeEvent):
    date_str = request.GET.get('date')
    date_type = request.GET.get('date_type', 'planned')
    
    try:
        date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return JsonResponse([], safe=False)
    
    # Apply access filters for date-based fetch
    user = request.user
    base = Evenement.objects.filter(type=typeEvent)
    if user.is_superuser:
        base = base
    elif getattr(user, 'is_RO', False):
        ids = list(user.societes.values_list('id', flat=True))
        base = base.filter(Q(societe_id__in=ids) | Q(created_by__societe_id__in=ids))
    elif user.is_RC and user.societe:
        base = base.filter(Q(societe=user.societe) | Q(created_by__societe=user.societe))
    elif getattr(user, 'is_C', False):
        base = base.filter(Q(societe=user.societe) | Q(created_by__societe=user.societe))
    else:
        base = base.filter(created_by=user)

    if date_type == 'planned':
        events = base.filter(date_heure_planifie__date=date).values('id', 'nom', 'date_heure_planifie', 'etat', 'notes')
    else:
        events = base.filter(date_heure_realiser__date=date).values('id', 'nom', 'date_heure_realiser', 'etat', 'notes')
    
    events_list = list(events)
    
    for event in events_list:
        if 'date_heure_planifie' in event and event['date_heure_planifie']:
            event['date_heure_planifie'] = event['date_heure_planifie'].isoformat()
        if 'date_heure_realiser' in event and event['date_heure_realiser']:
            event['date_heure_realiser'] = event['date_heure_realiser'].isoformat()
    
    return JsonResponse(events_list, safe=False)

@login_required
def get_all_filtered_events(request, typeEvent):
    user = request.user
    if user.is_superuser:
        events = Evenement.objects.filter(type=typeEvent).order_by('-date')
    elif getattr(user, 'is_RO', False):
        ids = list(user.societes.values_list('id', flat=True))
        events = Evenement.objects.filter(
            Q(type=typeEvent) & (Q(societe_id__in=ids) | Q(created_by__societe_id__in=ids))
        ).order_by('-date')
    elif user.is_RC and user.societe:
        events = Evenement.objects.filter(
            Q(type=typeEvent) & (Q(societe=user.societe) | Q(created_by__societe=user.societe))
        )
    elif getattr(user, 'is_C', False):
        events = Evenement.objects.filter(
            Q(type=typeEvent) & (Q(societe=user.societe) | Q(created_by__societe=user.societe))
        )
    else:
        events = Evenement.objects.filter(type=typeEvent, created_by=user)
    
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    societe_filter = request.GET.get('societe', '')
    created_date_from = request.GET.get('created_from', '')
    created_date_to = request.GET.get('created_to', '')
    planned_date_from = request.GET.get('planned_from', '')
    planned_date_to = request.GET.get('planned_to', '')
    realized_date_from = request.GET.get('realized_from', '')
    realized_date_to = request.GET.get('realized_to', '')
    
    if search_query:
        events = events.filter(
            Q(nom__icontains=search_query) |
            Q(lieu__icontains=search_query) |
            Q(secteur_activite__icontains=search_query) |
            Q(notes__icontains=search_query))
    
    if status_filter:
        events = events.filter(etat=status_filter)
    
    if societe_filter and (request.user.is_superuser or request.user.is_RC or request.user.is_RO):
        events = events.filter(societe__id=societe_filter)
    
    if created_date_from:
        try:
            created_date_from = datetime.strptime(created_date_from, '%Y-%m-%d')
            events = events.filter(date__gte=created_date_from)
        except ValueError:
            pass
    
    if created_date_to:
        try:
            created_date_to = datetime.strptime(created_date_to, '%Y-%m-%d')
            events = events.filter(date__lte=created_date_to)
        except ValueError:
            pass
    
    if planned_date_from:
        try:
            planned_date_from = datetime.strptime(planned_date_from, '%Y-%m-%d')
            events = events.filter(date_heure_planifie__gte=planned_date_from)
        except ValueError:
            pass
    
    if planned_date_to:
        try:
            planned_date_to = datetime.strptime(planned_date_to, '%Y-%m-%d')
            events = events.filter(date_heure_planifie__lte=planned_date_to)
        except ValueError:
            pass
    
    if realized_date_from and realized_date_to:
        try:
            realized_date_from = datetime.strptime(realized_date_from, '%Y-%m-%d')
            realized_date_to = datetime.strptime(realized_date_to, '%Y-%m-%d')
            events = events.filter(
                date_heure_realiser__gte=realized_date_from,
                date_heure_realiser__lte=realized_date_to
            )
        except ValueError:
            pass
    
    return JsonResponse({
        'events': [{
            'id': event.id,
            'nom': event.nom,
            'etat': event.etat,
            'date_heure_planifie': event.date_heure_planifie.isoformat() if event.date_heure_planifie else None,
            'date_heure_realiser': event.date_heure_realiser.isoformat() if event.date_heure_realiser else None,
            'notes': event.notes,
        } for event in events]
    }, safe=False)

@login_required
@require_POST
def update_event_date(request):
    try:
        event_id = request.POST.get('event_id')
        new_date_str = request.POST.get('new_date')
        date_type = request.POST.get('date_type')
        
        event = Evenement.objects.get(id=event_id)
        # Permission check: SU, RO, RC (même société), ou créateur
        if not (
            request.user.is_superuser or getattr(request.user, 'is_RO', False) or
            (getattr(request.user, 'is_RC', False) and event.societe_id == getattr(request.user.societe, 'id', None)) or
            event.created_by_id == request.user.id
        ):
            return JsonResponse({'error': 'Accès refusé'}, status=403)
        
        new_date = datetime.fromisoformat(new_date_str)
        
        if new_date.tzinfo is not None:
            new_date = new_date.astimezone(pytz.UTC)
            new_date = new_date.replace(tzinfo=None)  
            
        new_date = make_aware(new_date)
        
        if date_type == 'planned':
            event.date_heure_planifie = new_date
        elif date_type == 'realized':
            event.date_heure_realiser = new_date
        
        event.save()
        
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

#email
@login_required
def email_list(request):
    if request.user.is_superuser:
        emails = Action.objects.filter(is_Email=True)
        societes = Societe.objects.all()
    elif getattr(request.user, 'is_RO', False):
        # Get the list of societe IDs for the admin
        societe_ids = list(request.user.societes.values_list('id', flat=True))
        if societe_ids:
            emails = Action.objects.filter(
                Q(is_Email=True) &
                (Q(societe_id__in=societe_ids) | 
                 Q(created_by__societe_id__in=societe_ids))
            )
            societes = Societe.objects.filter(id__in=societe_ids)
        else:
            emails = Action.objects.none()
            societes = Societe.objects.none()
    elif getattr(request.user, 'is_RC', False) and hasattr(request.user, 'societe') and request.user.societe:
        emails = Action.objects.filter(
            Q(is_Email=True) &
            (Q(societe=request.user.societe) | 
             Q(created_by__societe=request.user.societe))
        )
        societes = Societe.objects.filter(id=request.user.societe_id)
    else:
        emails = Action.objects.filter(is_Email=True, created_by=request.user)
        societes = Societe.objects.filter(id=request.user.societe_id)
    
    emails = emails.order_by('-date_heure')
    
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    societe_filter = request.GET.get('societe', '')
    created_date_from = request.GET.get('created_from', '')
    created_date_to = request.GET.get('created_to', '')
    planned_date_from = request.GET.get('planned_from', '')
    planned_date_to = request.GET.get('planned_to', '')
    realized_date_from = request.GET.get('realized_from', '')
    realized_date_to = request.GET.get('realized_to', '')
    entreprise_filter = request.GET.get('entreprise', '')
    type_entreprise_filter = request.GET.get('type_entreprise', '')
    
    if search_query:
        emails = emails.filter(
            Q(sujet__icontains=search_query) |
            Q(compte_rendu__icontains=search_query) |
            Q(notes__icontains=search_query))
    
    if status_filter:
        emails = emails.filter(etat=status_filter)
    
    if societe_filter and (request.user.is_superuser or request.user.is_RC or request.user.is_RO):
        emails = emails.filter(societe__id=societe_filter)
        
    if type_entreprise_filter:
        if type_entreprise_filter == 'client':
            emails = emails.filter(entreprise__is_CLT=True)
        elif type_entreprise_filter == 'prospect':
            emails = emails.filter(entreprise__is_Prospect=True, entreprise__is_CLT=False)
        
    if entreprise_filter:
        emails = emails.filter(entreprise__id=entreprise_filter)
        
    entreprise_stats = emails.values('entreprise__id', 'entreprise__nom', 'entreprise__is_CLT') \
                       .filter(entreprise__is_Concurent=False, date_heure_realiser__isnull=False)\
                       .annotate(total_emails=Count('id')) \
                       .order_by('-entreprise__is_CLT', 'entreprise__nom')
        
    clients_data = []
    prospects_data = []

    for stat in entreprise_stats:
        if stat['entreprise__is_CLT']:
            clients_data.append({
                'name': stat['entreprise__nom'],
                'count': stat['total_emails']
            })
        else:
            prospects_data.append({
                'name': stat['entreprise__nom'],
                'count': stat['total_emails']
            })
            
    monthly_stats = emails.exclude(date_heure_realiser__isnull=True) \
                    .filter(entreprise__is_Concurent=False) \
                    .annotate(month=TruncMonth('date_heure_realiser')) \
                    .values('month', 'entreprise__is_CLT', 'entreprise__is_Prospect') \
                    .annotate(total=Count('id')) \
                    .order_by('month')

    months = []
    client_counts = []
    prospect_counts = []
    total_counts = []

    for stat in monthly_stats:
        month_str = stat['month'].strftime("%Y-%m")
        if month_str not in months:
            months.append(month_str)
            client_counts.append(0)
            prospect_counts.append(0)
            total_counts.append(0)
        
        idx = months.index(month_str)
        if stat['entreprise__is_CLT']:
            client_counts[idx] = stat['total']
        elif stat['entreprise__is_Prospect']:
            prospect_counts[idx] = stat['total']
        
        total_counts[idx] += stat['total']
    
    if created_date_from:
        try:
            created_date_from = datetime.strptime(created_date_from, '%Y-%m-%d')
            emails = emails.filter(date_heure__gte=created_date_from)
        except ValueError:
            pass
    
    if created_date_to:
        try:
            created_date_to = datetime.strptime(created_date_to, '%Y-%m-%d')
            emails = emails.filter(date_heure__lte=created_date_to)
        except ValueError:
            pass
    
    if planned_date_from:
        try:
            planned_date_from = datetime.strptime(planned_date_from, '%Y-%m-%d')
            emails = emails.filter(date_heure_planifie__gte=planned_date_from)
        except ValueError:
            pass
    
    if planned_date_to:
        try:
            planned_date_to = datetime.strptime(planned_date_to, '%Y-%m-%d')
            emails = emails.filter(date_heure_planifie__lte=planned_date_to)
        except ValueError:
            pass
    
    if realized_date_from and realized_date_to:
        try:
            realized_date_from = datetime.strptime(realized_date_from, '%Y-%m-%d')
            realized_date_to = datetime.strptime(realized_date_to, '%Y-%m-%d')
            emails = emails.filter(
                date_heure_realiser__gte=realized_date_from,
                date_heure_realiser__lte=realized_date_to
            )
        except ValueError:
            pass
        
    entreprises = Entreprise.objects.filter(is_Concurent=False).order_by('nom')
    if not request.user.is_superuser:
        if getattr(request.user, 'is_RO', False):
            # Get the list of societe IDs for the admin
            societe_ids = list(request.user.societes.values_list('id', flat=True))
            if societe_ids:
                entreprises = entreprises.filter(societe_id__in=societe_ids)
            else:
                entreprises = Entreprise.objects.none()
        elif hasattr(request.user, 'societe') and request.user.societe:
            entreprises = entreprises.filter(societe=request.user.societe)
    
    paginator = Paginator(emails, 100)
    page_number = request.GET.get('page')
    emails = paginator.get_page(page_number)
    
    context = {
        'emails': emails,
        'is_superuser': request.user.is_superuser,
        'is_RC': request.user.is_RC,
        'is_RO': request.user.is_RO,
        'societes': societes,
        'status_choices': dict(Action.ETAT_CHOICES_EMAIL),
        'search_query': search_query,
        'status_filter': status_filter,
        'societe_filter': societe_filter,
        'created_date_from': created_date_from,
        'created_date_to': created_date_to,
        'planned_date_from': planned_date_from,
        'planned_date_to': planned_date_to,
        'realized_date_from': realized_date_from,
        'realized_date_to': realized_date_to,
        'request': request,
        'clients_data': clients_data,
        'prospects_data': prospects_data,
        'entreprises': entreprises,
        'entreprise_filter': entreprise_filter,
        'type_entreprise_filter': type_entreprise_filter,
        'months': months,
        'client_counts': client_counts,
        'prospect_counts': prospect_counts,
        'total_counts': total_counts,
    }
    
    return render(request, 'adminlte/sales/sales/actions/emails.html', context)

@login_required
def get_all_filtered_emails(request):
    if request.user.is_superuser:
        emails = Action.objects.filter(is_Email=True).order_by('-date_heure')
    elif getattr(request.user, 'is_RO', False):
        # Get the list of societe IDs for the admin
        societe_ids = list(request.user.societes.values_list('id', flat=True))
        if societe_ids:
            emails = Action.objects.filter(
                Q(is_Email=True) & 
                (Q(societe_id__in=societe_ids) | 
                 Q(created_by__societe_id__in=societe_ids))
            ).order_by('-date_heure')
        else:
            emails = Action.objects.none()
    elif getattr(request.user, 'is_RC', False) and hasattr(request.user, 'societe') and request.user.societe:
        emails = Action.objects.filter(
            Q(is_Email=True) & 
            (Q(societe=request.user.societe) | 
             Q(created_by__societe=request.user.societe))
        ).order_by('-date_heure')
    else:
        emails = Action.objects.filter(is_Email=True, created_by=request.user).order_by('-date_heure')
    
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    societe_filter = request.GET.get('societe', '')
    created_date_from = request.GET.get('created_from', '')
    created_date_to = request.GET.get('created_to', '')
    planned_date_from = request.GET.get('planned_from', '')
    planned_date_to = request.GET.get('planned_to', '')
    realized_date_from = request.GET.get('realized_from', '')
    realized_date_to = request.GET.get('realized_to', '')
    
    if search_query:
        emails = emails.filter(
            Q(sujet__icontains=search_query) |
            Q(compte_rendu__icontains=search_query) |
            Q(notes__icontains=search_query))
    
    if status_filter:
        emails = emails.filter(etat=status_filter)
    
    if societe_filter and (request.user.is_superuser or request.user.is_RC or request.user.is_RO):
        emails = emails.filter(societe__id=societe_filter)
    
    if created_date_from:
        try:
            created_date_from = datetime.strptime(created_date_from, '%Y-%m-%d')
            emails = emails.filter(date_heure__gte=created_date_from)
        except ValueError:
            pass
    
    if created_date_to:
        try:
            created_date_to = datetime.strptime(created_date_to, '%Y-%m-%d')
            emails = emails.filter(date_heure__lte=created_date_to)
        except ValueError:
            pass
    
    if planned_date_from:
        try:
            planned_date_from = datetime.strptime(planned_date_from, '%Y-%m-%d')
            emails = emails.filter(date_heure_planifie__gte=planned_date_from)
        except ValueError:
            pass
    
    if planned_date_to:
        try:
            planned_date_to = datetime.strptime(planned_date_to, '%Y-%m-%d')
            emails = emails.filter(date_heure_planifie__lte=planned_date_to)
        except ValueError:
            pass
    
    if realized_date_from and realized_date_to:
        try:
            realized_date_from = datetime.strptime(realized_date_from, '%Y-%m-%d')
            realized_date_to = datetime.strptime(realized_date_to, '%Y-%m-%d')
            emails = emails.filter(
                date_heure_realiser__gte=realized_date_from,
                date_heure_realiser__lte=realized_date_to
            )
        except ValueError:
            pass
    
    return JsonResponse({
        'emails': [{
            'id': email.id,
            'sujet': email.sujet,
            'etat': email.etat,
            'date_heure_planifie': email.date_heure_planifie.isoformat() if email.date_heure_planifie else None,
            'date_heure_realiser': email.date_heure_realiser.isoformat() if email.date_heure_realiser else None,
            'notes': email.notes,
        } for email in emails]
    }, safe=False)

@require_GET
def get_emails_by_date(request):
    date_str = request.GET.get('date')
    date_type = request.GET.get('date_type', 'planned')
    
    try:
        date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return JsonResponse([], safe=False)
    
    if date_type == 'planned':
        qs = Action.objects.filter(is_Email=True, date_heure_planifie__date=date)
    else:
        qs = Action.objects.filter(is_Email=True, date_heure_realiser__date=date)

    if request.user.is_superuser:
        pass
    elif getattr(request.user, 'is_RO', False):
        # Get the list of societe IDs for the admin
        societe_ids = list(request.user.societes.values_list('id', flat=True))
        if societe_ids:
            qs = qs.filter(
                Q(societe_id__in=societe_ids) | 
                Q(created_by__societe_id__in=societe_ids)
            )
        else:
            qs = Action.objects.none()
    elif getattr(request.user, 'is_RC', False) and hasattr(request.user, 'societe') and request.user.societe:
        qs = qs.filter(
            Q(societe=request.user.societe) | 
            Q(created_by__societe=request.user.societe)
        )
    else:
        qs = qs.filter(created_by=request.user)

    emails_list = list(qs.values('id', 'sujet', 'date_heure_planifie', 'date_heure_realiser', 'etat', 'notes'))
    
    for email in emails_list:
        if 'date_heure_planifie' in email and email['date_heure_planifie']:
            email['date_heure_planifie'] = email['date_heure_planifie'].isoformat()
        if 'date_heure_realiser' in email and email['date_heure_realiser']:
            email['date_heure_realiser'] = email['date_heure_realiser'].isoformat()
    
    return JsonResponse(emails_list, safe=False)

@login_required
@require_POST
def update_email_date(request):
    try:
        action_id = request.POST.get('action_id')
        new_date_str = request.POST.get('new_date')
        date_type = request.POST.get('date_type')
        
        action = Action.objects.get(id=action_id)
        # Permission check: SU, RC (same societe), or creator
        if not (
            request.user.is_superuser or request.user.is_RO or
            (getattr(request.user, 'is_RC', False) and action.societe_id == getattr(request.user.societe, 'id', None)) or
            action.created_by_id == request.user.id
        ):
            return JsonResponse({'status': 'error', 'message': "Permission refusée"}, status=403)
        
        new_date = datetime.fromisoformat(new_date_str)
        
        if new_date.tzinfo is not None:
            new_date = new_date.astimezone(pytz.UTC)
            new_date = new_date.replace(tzinfo=None)  
            
        new_date = make_aware(new_date)
        
        if date_type == 'planned':
            action.date_heure_planifie = new_date
        elif date_type == 'realized':
            action.date_heure_realiser = new_date
        
        action.save()
        
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

def _json_success_response(message, redirect_url=None, **extra):
    payload = {'success': True, 'message': message, **extra}
    if redirect_url:
        payload['redirect_url'] = redirect_url
    return JsonResponse(payload)

@login_required
def add_email(request):
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])

    errors = {}
    sujet = (request.POST.get('sujet') or '').strip()
    compte_rendu = (request.POST.get('compte_rendu') or '').strip()
    notes = (request.POST.get('notes') or '').strip()
    etat = (request.POST.get('etat') or '').strip()
    societe_id = (request.POST.get('societe') or '').strip()
    date_plan_str = (request.POST.get('date_heure_planifie') or '').strip()
    date_real_str = (request.POST.get('date_heure_realiser') or '').strip()
    type_entreprise = (request.POST.get('type_entreprise') or 'client').strip()
    entreprise_id = (request.POST.get('entreprise') or '').strip()       # pk prospect OU pk client local
    client_numero  = (request.POST.get('client_numero') or '').strip()   # NumeroTiers client
    prospect_id    = (request.POST.get('prospect_id') or '').strip()
    client_societe_name = (request.POST.get('client_societe_name') or '').strip()

    if not sujet:
        errors['sujet'] = "Le sujet est obligatoire"
    if etat and etat not in dict(Action.ETAT_CHOICES_EMAIL):
        errors['etat'] = "L'état pour un email doit être 'Lu' ou 'Non Lu'"
    # Détermination initiale de la société: par nom fourni par l'autocomplete, sinon par id, sinon user
    societe = None
    if client_societe_name:
        societe = Societe.objects.filter(nom=client_societe_name).first()
    if not societe and societe_id:
        try:
            societe = Societe.objects.get(id=societe_id)
        except Societe.DoesNotExist:
            societe = None
    if not societe:
        societe = getattr(request.user, 'societe', None)

    date_plan = _parse_dt_local_to_aware(date_plan_str, 'date_heure_planifie', errors)
    date_real = _parse_dt_local_to_aware(date_real_str, 'date_heure_realiser', {}) if date_real_str else None

    # entreprise (prospect / client)
    entreprise = None
    if type_entreprise == 'prospect':
        if not prospect_id:
            errors['prospect'] = "Veuillez sélectionner un prospect"
        else:
            try:
                entreprise = Entreprise.objects.get(id=prospect_id, is_CLT=False)
            except Entreprise.DoesNotExist:
                errors['prospect'] = "Prospect invalide"
    else:
        # client
        if not (client_numero or entreprise_id):
            errors['client'] = "Veuillez sélectionner ou rechercher un client"
        else:
            try:
                if client_numero:
                    # Utiliser la société du client si fournie pour choisir la bonne base
                    preferred_societe = Societe.objects.filter(nom=client_societe_name).first() if client_societe_name else societe
                    entreprise = _fetch_or_create_client(client_numero, preferred_societe)
                else:
                    entreprise = Entreprise.objects.get(id=entreprise_id, is_CLT=True)
            except Exception as e:
                errors['client'] = str(e)

    if errors:
        return JsonResponse({'success': False, 'errors': errors}, status=400)

    try:
        # Filiale finale de l'email: entreprise.societe si dispo, sinon societe calculée, sinon user
        final_societe = entreprise.societe if (entreprise and getattr(entreprise, 'societe', None)) else (societe or getattr(request.user, 'societe', None))
        email = Action.objects.create(
            sujet=sujet, compte_rendu=compte_rendu or '', notes=notes or '',
            etat=etat, societe=final_societe, created_by=request.user,
            pilote = None if final_societe is None else get_user_model().objects.filter(is_RC=True, societe=final_societe).first(),
            is_Email=True, entreprise=entreprise,
            date_heure_planifie=date_plan, date_heure_realiser=date_real
        )
        logger.info("Email %s créé par %s", email.id, request.user)
        return _json_success_response(message="E-mail créé avec succès", redirect_url=reverse('prospection:email_list'), email_id=email.id)
    except Exception:
        logger.exception("Erreur création email")
        return JsonResponse({'success': False, 'errors': {'__all__': 'Erreur serveur lors de la création'}}, status=500)


# ---------- MODIFICATION ----------
@login_required
def edit_email(request, email_id):
    email = get_object_or_404(Action, id=email_id, is_Email=True)

    # Permissions
    if not (
        request.user.is_superuser or request.user.is_RO or
        (getattr(request.user, 'is_RC', False) and email.societe_id == getattr(request.user.societe, 'id', None)) or
        email.created_by_id == request.user.id
    ):
        return JsonResponse({'error': "Vous n'avez pas la permission de modifier cet e-mail"}, status=403)

    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])

    errors = {}
    sujet = (request.POST.get('sujet') or email.sujet or '').strip()
    etat  = (request.POST.get('etat')  or email.etat  or '').strip() or email.etat
    notes = (request.POST.get('notes') or email.notes or '').strip()
    compte_rendu = (request.POST.get('compte_rendu') or email.compte_rendu or '').strip()
    date_plan_str = (request.POST.get('date_heure_planifie') or '').strip()
    date_real_str = (request.POST.get('date_heure_realiser') or '').strip()
    client_type = (request.POST.get('client_type') or '').strip()  # 'client'|'prospect'
    entreprise_param = (request.POST.get('entreprise') or '').strip()  # pk (prospect ou client local)
    numero_tiers = (request.POST.get('numero_tiers') or '').strip()    # NumeroTiers pour client
    client_societe_name = (request.POST.get('client_societe_name') or '').strip()

    if not sujet:
        errors['sujet'] = "Ce champ est obligatoire"
    email.sujet = sujet
    email.etat = etat
    email.notes = notes
    email.compte_rendu = compte_rendu

    # dates
    if date_plan_str:
        dtp = _parse_dt_local_to_aware(date_plan_str, 'date_heure_planifie', errors)
        if dtp: email.date_heure_planifie = dtp
    if date_real_str:
        dtr = _parse_dt_local_to_aware(date_real_str, 'date_heure_realiser', errors)
        if dtr: email.date_heure_realiser = dtr
    if not email.date_heure_planifie:
        errors['date_heure_planifie'] = "Ce champ est obligatoire"

    # Déterminer la filiale préférée à partir du nom transmis ou de l'utilisateur
    preferred_societe = None
    if client_societe_name:
        preferred_societe = Societe.objects.filter(nom=client_societe_name).first()
    if not preferred_societe:
        preferred_societe = email.societe or getattr(request.user, 'societe', None)

    # entreprise
    entreprise = email.entreprise
    try:
        if client_type == 'client':
            raw = numero_tiers or entreprise_param
            if not raw:
                errors['entreprise'] = "Veuillez sélectionner un client"
            else:
                # Vérifier si le client a changé en comparant avec l'existant
                if email.entreprise and email.entreprise.is_CLT:
                    # Vérifier si c'est le même client (par ID ou num_compte)
                    try:
                        if str(email.entreprise.id) == str(raw) or str(email.entreprise.num_compte) == str(raw):
                            # Pas de changement, garder le client existant
                            entreprise = email.entreprise
                        else:
                            # Client différent, résoudre le nouveau
                            entreprise = _resolve_contact(raw, preferred_societe)
                            if not entreprise or not entreprise.is_CLT:
                                errors['entreprise'] = "Client sélectionné non valide"
                    except:
                        # En cas d'erreur de comparaison, résoudre normalement
                        entreprise = _resolve_contact(raw, preferred_societe)
                        if not entreprise or not entreprise.is_CLT:
                            errors['entreprise'] = "Client sélectionné non valide"
                else:
                    # Pas de client existant ou changement de prospect vers client
                    entreprise = _resolve_contact(raw, preferred_societe)
                    if not entreprise or not entreprise.is_CLT:
                        errors['entreprise'] = "Client sélectionné non valide"
        elif client_type == 'prospect':
            if not entreprise_param:
                errors['entreprise'] = "Veuillez sélectionner un prospect"
            else:
                # Vérifier si c'est le même prospect
                if email.entreprise and not email.entreprise.is_CLT and str(email.entreprise.id) == str(entreprise_param):
                    # Pas de changement, garder le prospect existant
                    entreprise = email.entreprise
                else:
                    entreprise = Entreprise.objects.get(id=entreprise_param, is_CLT=False)
        # si pas de client_type, on conserve
    except Entreprise.DoesNotExist:
        errors['entreprise'] = "Prospect sélectionné non valide"
    except Exception as e:
        errors['entreprise'] = str(e)

    if errors:
        return JsonResponse({'errors': errors}, status=400)

    email.entreprise = entreprise
    # Ajuster automatiquement la filiale de l'email
    if entreprise and getattr(entreprise, 'societe', None):
        email.societe = entreprise.societe
    elif preferred_societe:
        email.societe = preferred_societe
    try:
        email.save()
        return _json_success("E-mail modifié avec succès")
    except Exception:
        logger.exception("Erreur sauvegarde email")
        return JsonResponse({'errors': {'__all__': 'Erreur serveur lors de la sauvegarde'}}, status=500)

@login_required
def delete_email(request, email_id):
    email = get_object_or_404(Action, id=email_id, is_Email=True)
    
    if not (request.user.is_superuser or request.uer.is_RO or
            (request.user.is_RC and email.societe == request.user.societe) or 
            email.created_by == request.user):
        return JsonResponse({'error': "Vous n'avez pas l'autorisation de supprimer cet e-mail"}, status=403)
    
    if request.method == 'POST':
        try:
            email.delete()
            return JsonResponse({'success': 'email supprimé avec succès!'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Méthode de demande invalide'}, status=405)

@login_required
def send_email_action(request):
    """Envoyer un email et créer une action Email"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)
    
    try:
        # Récupération des données
        recipients_json = request.POST.get('recipients', '[]')
        recipients = json.loads(recipients_json)
        
        if not recipients:
            return JsonResponse({'success': False, 'message': 'Aucun destinataire spécifié'}, status=400)
        
        subject = request.POST.get('send_subject', '').strip()
        message = request.POST.get('send_message', '').strip()
        date_planifie_str = request.POST.get('send_date_heure_planifie', '').strip()
        type_entreprise = request.POST.get('send_type_entreprise', 'client').strip()
        attachment = request.FILES.get('send_attachment')
        
        # Validation
        if not subject:
            return JsonResponse({'success': False, 'message': 'L\'objet est obligatoire'}, status=400)
        if not message:
            return JsonResponse({'success': False, 'message': 'Le message est obligatoire'}, status=400)
        
        # Parse date
        date_planifie = _parse_dt_local_to_aware(date_planifie_str, 'date_heure_planifie', {}) if date_planifie_str else timezone.now()
        
        # Déterminer l'entreprise
        entreprise = None
        societe = None
        
        if type_entreprise == 'prospect':
            prospect_id = request.POST.get('send_prospect_id', '').strip()
            if prospect_id:
                try:
                    entreprise = Entreprise.objects.get(id=prospect_id, is_CLT=False)
                    societe = entreprise.societe
                except Entreprise.DoesNotExist:
                    return JsonResponse({'success': False, 'message': 'Prospect invalide'}, status=400)
        else:  # client
            client_numero = request.POST.get('send_client_numero', '').strip()
            entreprise_id = request.POST.get('send_entreprise', '').strip()
            client_societe_name = request.POST.get('send_client_societe_name', '').strip()
            
            if client_numero:
                preferred_societe = Societe.objects.filter(nom=client_societe_name).first() if client_societe_name else None
                try:
                    entreprise = _fetch_or_create_client(client_numero, preferred_societe)
                    societe = entreprise.societe if entreprise else preferred_societe
                except Exception as e:
                    return JsonResponse({'success': False, 'message': f'Erreur client: {str(e)}'}, status=400)
            elif entreprise_id:
                try:
                    entreprise = Entreprise.objects.get(id=entreprise_id, is_CLT=True)
                    societe = entreprise.societe
                except Entreprise.DoesNotExist:
                    return JsonResponse({'success': False, 'message': 'Client invalide'}, status=400)
        
        if not entreprise:
            return JsonResponse({'success': False, 'message': 'Veuillez sélectionner un client ou prospect'}, status=400)
        
        # Société finale
        final_societe = societe or getattr(request.user, 'societe', None)
        
        # Pilote (RC de la société)
        pilote = get_user_model().objects.filter(is_RC=True, societe=final_societe).first() if final_societe else None
        
        # Envoi de l'email
        try:
            email_obj = EmailMessage(
                subject=subject,
                body=message,
                from_email=settings.EMAIL_HOST_USER,
                to=recipients,
            )
            
            if attachment:
                email_obj.attach(attachment.name, attachment.read(), attachment.content_type)
            
            email_obj.send(fail_silently=False)
            logger.info(f"Email envoyé à {recipients} par {request.user.username}")
        except Exception as e:
            logger.error(f"Erreur envoi email: {str(e)}")
            return JsonResponse({'success': False, 'message': f'Erreur lors de l\'envoi de l\'email: {str(e)}'}, status=500)
        
        # Créer l'action Email
        try:
            action = Action.objects.create(
                sujet=subject,
                notes=message,
                compte_rendu='',
                etat='non_lu',
                societe=final_societe,
                entreprise=entreprise,
                created_by=request.user,
                pilote=pilote,
                is_Email=True,
                date_heure_planifie=date_planifie,
            )
            logger.info(f"Action Email {action.id} créée par {request.user.username}")
        except Exception as e:
            logger.error(f"Erreur création action: {str(e)}")
            return JsonResponse({'success': False, 'message': f'Email envoyé mais erreur lors de la création de l\'action: {str(e)}'}, status=500)
        
        return JsonResponse({
            'success': True,
            'message': f'Email envoyé avec succès à {len(recipients)} destinataire(s)',
            'action_id': action.id
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Format des destinataires invalide'}, status=400)
    except Exception as e:
        logger.exception("Erreur envoi email action")
        return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'}, status=500)

@login_required
def send_reminder_action(request):
    """Envoyer un rappel d'email"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)
    
    try:
        # Récupération des données
        recipients_json = request.POST.get('recipients', '[]')
        recipients = json.loads(recipients_json)
        
        if not recipients:
            return JsonResponse({'success': False, 'message': 'Aucun destinataire spécifié'}, status=400)
        
        subject = request.POST.get('reminder_subject', '').strip()
        message = request.POST.get('reminder_message', '').strip()
        original_email_id = request.POST.get('original_email_id', '').strip()
        attachment = request.FILES.get('reminder_attachment')
        
        # Validation
        if not subject:
            return JsonResponse({'success': False, 'message': 'L\'objet est obligatoire'}, status=400)
        if not message:
            return JsonResponse({'success': False, 'message': 'Le message est obligatoire'}, status=400)
        
        # Récupérer l'email original pour les infos
        original_email = None
        if original_email_id:
            try:
                original_email = Action.objects.get(id=original_email_id, is_Email=True)
            except Action.DoesNotExist:
                pass
        
        # Envoi du rappel
        try:
            email_obj = EmailMessage(
                subject=subject,
                body=message,
                from_email=settings.EMAIL_HOST_USER,
                to=recipients,
            )
            
            if attachment:
                email_obj.attach(attachment.name, attachment.read(), attachment.content_type)
            
            email_obj.send(fail_silently=False)
            logger.info(f"Rappel envoyé à {recipients} par {request.user.username}")
        except Exception as e:
            logger.error(f"Erreur envoi rappel: {str(e)}")
            return JsonResponse({'success': False, 'message': f'Erreur lors de l\'envoi du rappel: {str(e)}'}, status=500)
        
        # Créer une nouvelle action Email pour le rappel
        if original_email:
            try:
                action = Action.objects.create(
                    sujet=subject,
                    notes=message,
                    compte_rendu='',
                    etat='non_lu',
                    societe=original_email.societe,
                    entreprise=original_email.entreprise,
                    created_by=request.user,
                    pilote=original_email.pilote,
                    is_Email=True,
                    date_heure_planifie=timezone.now(),
                )
                logger.info(f"Action Rappel {action.id} créée par {request.user.username}")
            except Exception as e:
                logger.error(f"Erreur création action rappel: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'message': f'Rappel envoyé avec succès à {len(recipients)} destinataire(s)'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Format des destinataires invalide'}, status=400)
    except Exception as e:
        logger.exception("Erreur envoi rappel")
        return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'}, status=500)

@login_required
def email_details(request, email_id):
    email = get_object_or_404(Action, id=email_id, is_Email=True)
    
    if not (request.user.is_superuser or request.user.is_RO or
            (request.user.is_RC and email.societe == request.user.societe) or 
            (email.created_by and email.created_by == request.user)):
        return JsonResponse({'error': "Vous n'avez pas l'autorisation de voir cet e-mail"}, status=403)
    
    data = {
        'id': email.id,
        'sujet': email.sujet,
        'compte_rendu': email.compte_rendu,
        'notes': email.notes,
        'etat': email.etat, 
        'entreprise': email.entreprise.nom if email.entreprise else '',
        'societe': email.societe.nom if email.societe else '',
        'societe_id': email.societe.id if email.societe else None,
        'created_by': email.created_by.username if email.created_by else '',
        'pilote': email.pilote.username if email.pilote else '',
        'date_heure': email.date_heure.strftime('%Y-%m-%d %H:%M'),
        'date_heure_planifie': email.date_heure_planifie.strftime('%Y-%m-%d %H:%M') if email.date_heure_planifie else '',
        'date_heure_realiser': email.date_heure_realiser.strftime('%Y-%m-%d %H:%M') if email.date_heure_realiser else '',
    }
    
    return JsonResponse(data)

def send_email_details_pdf(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode de demande invalide'})

    try:
        # 1) Logo
        logo_path = finders.find('dist/img/abserveLogo.png')
        logo_base64 = ''
        if logo_path:
            with open(logo_path, "rb") as f:
                logo_base64 = base64.b64encode(f.read()).decode('utf-8')

        # 2) Champs simples
        recipient_email = request.POST.get('recipient_email')
        subject = request.POST.get('subject')
        message = request.POST.get('message', '')

        # 3) Parser le JSON envoyé par le client
        raw_details = request.POST.get('email_details', '{}')
        try:
            details = json.loads(raw_details)  # <= ICI le parse !
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'email_details JSON invalide'})

        # 4) Construire le contexte du template
        email_details = {
            'sujet': details.get('sujet') or '',
            'entreprise': details.get('entreprise') or '',
            'societe': details.get('societe') or '',
            'createdBy': details.get('createdBy') or '',
            'pilote': details.get('pilote') or '',
            'dateHeure': details.get('dateHeure') or '',
            'datePlanifie': details.get('datePlanifie') or '',
            'dateRealiser': details.get('dateRealiser') or '',
            'etat': details.get('etat') or '',
            'compteRendu': details.get('compteRendu') or '',
            'additional_message': message or '',
            'logo_base64': logo_base64,
        }

        # 5) Génération PDF
        html_string = render_to_string('adminlte/emails/email_details.html', email_details)
        pdf_file = BytesIO()
        pisa_status = pisa.CreatePDF(html_string, dest=pdf_file, encoding='UTF-8', link_callback=lambda uri, _: uri)
        if pisa_status.err:
            return JsonResponse({'success': False, 'error': 'Erreur de génération PDF'})
        pdf_file.seek(0)

        # 6) Envoi
        email = EmailMessage(subject, message, settings.DEFAULT_FROM_EMAIL, [recipient_email])
        email.attach('details_email.pdf', pdf_file.getvalue(), 'application/pdf')
        email.send()

        return JsonResponse({'success': True})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

#entreprise
@login_required
def gestion_entreprises(request):
    user = request.user
    search_query = request.GET.get('search')
    societe_filter = request.GET.get('societe')
    sort_by_date = request.GET.get('sort_by_date')
    # Clients-only listing (includes ex-prospects since is_CLT=True)
    if user.is_superuser:
        entreprises_list = Entreprise.objects.filter(is_Concurent=False, is_CLT=True).order_by('nom')
    elif getattr(user, 'is_RO', False):
        # Get the list of societe IDs for the admin
        societe_ids = list(user.societes.values_list('id', flat=True))
        if societe_ids:
            entreprises_list = Entreprise.objects.filter(is_Concurent=False, is_CLT=True, societe_id__in=societe_ids).order_by('nom')
        else:
            entreprises_list = Entreprise.objects.none()
    elif getattr(user, 'is_RC', False) and getattr(user, 'societe', None):
        entreprises_list = Entreprise.objects.filter(is_Concurent=False, is_CLT=True, societe=user.societe).order_by('nom')
    elif getattr(user, 'is_C', False):
        entreprises_list = Entreprise.objects.filter(is_Concurent=False, is_CLT=True, societe=user.societe).order_by('nom')
    else:
        # Default case for other users with no specific access
        entreprises_list = Entreprise.objects.none()

    if search_query:
        entreprises_list = entreprises_list.filter(nom__icontains=search_query)

    societe = None
    if societe_filter and (user.is_superuser or user.is_RO):
        try:
            societe = Societe.objects.get(id=societe_filter)
            entreprises_list = entreprises_list.filter(societe=societe)
        except Societe.DoesNotExist:
            pass

    if sort_by_date == 'asc':
        entreprises_list = entreprises_list.order_by('date')
    else:
        # Par défaut ou 'desc', trier par date décroissante
        entreprises_list = entreprises_list.order_by('-date')

    paginator = Paginator(entreprises_list, 100)
    page_number = request.GET.get('page')
    entreprises = paginator.get_page(page_number)

    # Récupérer toutes les années disponibles dans les dates de conversion
    if user.is_superuser:
        # Pour les superusers, on prend toutes les conversions
        entreprises_conversion = Entreprise.objects.filter(
            is_CLT=True,
            is_Prospect=True,  # Maintenant un client (n'est plus un prospect)
            date_conversion__isnull=False
        ).order_by('nom')
    elif getattr(user, 'is_RO', False):
        # Get the list of societe IDs for the admin
        societe_ids = list(user.societes.values_list('id', flat=True))
        if societe_ids:
            entreprises_conversion = Entreprise.objects.filter(
                is_CLT=True,
                is_Prospect=True,
                date_conversion__isnull=False,
                societe_id__in=societe_ids
            ).order_by('nom')
        else:
            entreprises_conversion = Entreprise.objects.none()
    else:
        # Pour les autres utilisateurs, on filtre par société si elle existe
        societe = getattr(user, 'societe', None)
        if societe:
            entreprises_conversion = Entreprise.objects.filter(
                is_CLT=True,
                is_Prospect=True,
                date_conversion__isnull=False,
                societe=societe
            ).order_by('nom')
        else:
            entreprises_conversion = Entreprise.objects.none()

    if societe_filter and (user.is_superuser or user.is_RO):
        entreprises_conversion = entreprises_conversion.filter(societe=societe)

    # Récupérer toutes les années distinctes des conversions
    years = list(entreprises_conversion.dates('date_conversion', 'year'))
    years = [year.year for year in years]
    
    # Si aucune année n'est trouvée, on prend l'année courante
    if not years:
        years = [timezone.now().year]
    
    # Trier les années par ordre décroissant
    years = sorted(years, reverse=True)
    
    # Préparer les données pour le graphique
    # On va regrouper par année et par mois
    conversion_data = {}
    
    for year in years:
        # Initialiser tous les mois à 0 pour l'année courante
        conversion_data[year] = {month: 0 for month in range(1, 13)}
        
        # Récupérer les données réelles pour chaque mois
        for month in range(1, 13):
            start_date = timezone.datetime(year, month, 1, 0, 0, 0, tzinfo=timezone.get_current_timezone())
            if month == 12:
                end_date = timezone.datetime(year + 1, 1, 1, 0, 0, 0, tzinfo=timezone.get_current_timezone())
            else:
                end_date = timezone.datetime(year, month + 1, 1, 0, 0, 0, tzinfo=timezone.get_current_timezone())
            
            count = entreprises_conversion.filter(
                date_conversion__gte=start_date,
                date_conversion__lt=end_date
            ).count()
            
            conversion_data[year][month] = count
    
    # Préparer les données pour le template
    chart_data = []
    for year, months in conversion_data.items():
        for month, count in months.items():
            chart_data.append({
                'year': int(year),  # S'assurer que c'est un entier
                'month': int(month),  # S'assurer que c'est un entier
                'count': int(count),  # S'assurer que c'est un entier
                'month_year': f"{year}-{month:02d}"
            })
    
    if user.is_superuser:
        societes = Societe.objects.all().order_by('nom')
    elif getattr(user, 'is_RO', False):
        # Get the list of societe IDs for the admin
        societe_ids = list(user.societes.values_list('id', flat=True))
        societes = Societe.objects.filter(id__in=societe_ids).order_by('nom') if societe_ids else Societe.objects.none()
    else:
        societes = None
    
    # Récupérer les années triées par ordre décroissant
    years = sorted(conversion_data.keys(), reverse=True)
    
    # Convertir en JSON de manière sécurisée
    chart_data_json = json.dumps(chart_data, ensure_ascii=False)
    
    context = {
        'entreprises': entreprises,
        'search_query': search_query,
        'societes': societes,
        'selected_societe': societe, 
        'societe_filter': societe_filter,
        'sort_by_date': sort_by_date,
        'is_superuser': user.is_superuser,
        'is_responsable': user.is_RC,
        'is_RO': user.is_RO,
        'chart_data': chart_data_json,  # Données JSON pour le graphique
        'years': years,  # Liste des années disponibles
    }
        
    return render(request, 'adminlte/sales/users/entreprises/clients-prospects.html', context)
    
@login_required
def gestion_concurrents(request):
    user = request.user
    search_query = request.GET.get('search')
    societe_filter = request.GET.get('societe')
    sort_by_date = request.GET.get('sort_by_date')  

    if user.is_superuser:
        entreprises_list = Entreprise.objects.filter(is_Concurent=True).order_by('nom')
    elif getattr(user, 'is_RO', False):
        # Get the list of societe IDs for the admin
        societe_ids = list(user.societes.values_list('id', flat=True))
        if societe_ids:
            entreprises_list = Entreprise.objects.filter(
                is_Concurent=True, 
                societe_id__in=societe_ids
            ).order_by('nom')
        else:
            entreprises_list = Entreprise.objects.none()
    elif getattr(user, 'is_RC', False) and getattr(user, 'societe', None):
        entreprises_list = Entreprise.objects.filter(
            is_Concurent=True, 
            societe=user.societe
        ).order_by('nom')
    elif getattr(user, 'is_C', False):
        entreprises_list = Entreprise.objects.filter(
            is_Concurent=True, 
            societe=user.societe
        ).order_by('nom')                                                                                                                                                                                           
    else:
        # Default case for other users with no specific access
        entreprises_list = Entreprise.objects.none()

    if search_query:
        entreprises_list = entreprises_list.filter(nom__icontains=search_query)

    societe = None
    if societe_filter and (user.is_superuser or user.is_RO):
        try:
            societe = Societe.objects.get(id=societe_filter)
            entreprises_list = entreprises_list.filter(societe=societe)
        except Societe.DoesNotExist:
            pass

    if sort_by_date == 'asc':
        entreprises_list = entreprises_list.order_by('date')  
    elif sort_by_date == 'desc':
        entreprises_list = entreprises_list.order_by('-date')  
    else:
        entreprises_list = entreprises_list.order_by('-date')

    paginator = Paginator(entreprises_list, 100)
    page_number = request.GET.get('page')
    entreprises = paginator.get_page(page_number)

    if user.is_superuser:
        societes = Societe.objects.all().order_by('nom')
    elif getattr(user, 'is_RO', False):
        # Get the list of societe IDs for the admin
        societe_ids = list(user.societes.values_list('id', flat=True))
        societes = Societe.objects.filter(id__in=societe_ids).order_by('nom') if societe_ids else Societe.objects.none()
    else:
        societes = None

    return render(request, 'adminlte/sales/users/entreprises/concurrents.html', {
        'entreprises': entreprises,
        'societes': societes,
        'selected_societe': societe,
        'societe_filter': societe_filter,
        'search_query': search_query,
        'sort_by_date': sort_by_date,  
        'is_superuser': user.is_superuser,
        'is_RO': user.is_RO,
        'is_responsable': user.is_RC,
    })

@login_required
def add_entreprise(request, type):
    user = request.user
    if request.method == 'POST':
        form = EntrepriseForm(request.POST)
        if form.is_valid():
            entreprise = form.save(commit=False)
            if not (user.is_superuser or user.is_RO):
                entreprise.societe = user.societe
                
            # Gestion du type d'entreprise
            if type == 'client':
                entreprise.is_CLT = True
                # Vérifier si le numéro de compte est déjà utilisé pour cette société
                if entreprise.num_compte and Entreprise.objects.filter(
                    societe=entreprise.societe, 
                    num_compte=entreprise.num_compte
                ).exists():
                    error_msg = f"Le numéro de compte {entreprise.num_compte} est déjà utilisé pour cette société."
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'errors': {'num_compte': [error_msg]}}, status=400)
                    messages.error(request, error_msg)
                    if type == 'concurrent':
                        return redirect('prospection:gestion_concurrents')
                    else:
                        return redirect('prospection:gestion_entreprises')
            elif type == 'prospect':
                entreprise.is_Prospect = True
                entreprise.num_compte = None
            elif type == 'concurrent':
                entreprise.is_Concurent = True
                entreprise.num_compte = None
                
            entreprise.save()
            
            # Réponse AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Prospect ajouté avec succès',
                    'entreprise': {
                        'id': entreprise.id,
                        'nom': entreprise.nom,
                        'adresse': entreprise.adresse,
                        'secteur_activite': entreprise.secteur_activite,
                        'telephone': entreprise.telephone,
                        'email': entreprise.email,
                        'societe_id': entreprise.societe.id if entreprise.societe else None,
                        'societe_nom': entreprise.societe.nom if entreprise.societe else ''
                    }
                })
                
            # Redirection pour les requêtes normales
            if type == 'concurrent':
                return redirect('prospection:gestion_concurrents')
            elif type == 'prospect':
                return _redirect_back_or('prospection:prospects_list', request)
            else:
                return redirect('prospection:gestion_entreprises')
        else:
            # Gestion des erreurs de formulaire pour AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    else:
         if user.is_superuser or user.is_RO:
            form = EntrepriseForm()
         else:
            form = EntrepriseForm(initial={'societe': user.societe})
    
    # Si c'est une requête AJAX mais qu'il y a eu une erreur
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'errors': form.errors if 'form' in locals() else {}}, status=400)
        
    # Pour les requêtes normales, afficher le formulaire
    if type == 'concurrent':
        return render(request, f'adminlte/sales/users/entreprises/concurrents.html', {'form': form})
    else:
        return render(request, f'adminlte/sales/users/entreprises/clients-prospects.html', {'form': form})

@login_required
def convert_prospect_to_client(request, id):
    entreprise = get_object_or_404(Entreprise, id=id)
    if entreprise.is_Prospect:
        entreprise.is_CLT = True
        entreprise.is_Prospect = True
        if not entreprise.num_compte:  # Si le numéro de compte n'est pas déjà défini
            entreprise.num_compte = None
        entreprise.date_conversion = now()  
        entreprise.save()
    return _redirect_back_or('prospection:prospects_list', request)

@login_required
def edit_entreprise(request, id, type):
    entreprise = get_object_or_404(Entreprise, id=id)
    if request.method == 'POST':
        form = EntrepriseFormEdit(request.POST, instance=entreprise)
        
        if form.is_valid():
            # On prépare l'instance sans enregistrer pour pouvoir ajuster les champs
            instance = form.save(commit=False)

            if type == 'client':
                # Vérifier si le numéro de compte est modifié et déjà utilisé pour cette société
                new_num_compte = form.cleaned_data.get('num_compte')
                if new_num_compte and new_num_compte != entreprise.num_compte:
                    if Entreprise.objects.filter(
                        societe=entreprise.societe,
                        num_compte=new_num_compte
                    ).exclude(id=entreprise.id).exists():
                        messages.error(request, f"Le numéro de compte {new_num_compte} est déjà utilisé pour cette société.")
                        return redirect('prospection:gestion_entreprises')
            elif type in ('prospect', 'concurrent'):
                instance.num_compte = None

            instance.save()
            messages.success(request, "Entreprise mise à jour avec succès.")
            if type == 'concurrent':
                return redirect('prospection:gestion_concurrents')
            elif type == 'prospect':
                return _redirect_back_or('prospection:prospects_list', request)
            else:
                return redirect('prospection:gestion_entreprises')
        else:
            # Formulaire invalide: afficher un message et rediriger vers la liste appropriée
            try:
                error_text = form.errors.as_text()
            except Exception:
                error_text = ''
            if error_text:
                messages.error(request, f"Formulaire invalide. Veuillez vérifier les champs saisis. Détails: {error_text}")
            else:
                messages.error(request, "Formulaire invalide. Veuillez vérifier les champs saisis.")

            if type == 'concurrent':
                return redirect('prospection:gestion_concurrents')
            elif type == 'prospect':
                return redirect('prospection:prospects_list')
            else:
                return redirect('prospection:gestion_entreprises')
    else:
        form = EntrepriseFormEdit(instance=entreprise)
    if type == 'concurrent':
        return render(request, f'adminlte/sales/users/entreprises/concurrents.html', {'form': form})
    elif type == 'prospect':
        return render(request, f'adminlte/sales/users/entreprises/prospects.html', {'form': form})
    else:
        return render(request, f'adminlte/sales/users/entreprises/clients-prospects.html', {'form': form})

@login_required
def delete_entreprise(request, id, type):
    entreprise = get_object_or_404(Entreprise, id=id)
    if request.method == 'POST':
        entreprise.delete()
        if type == 'concurrent':
            return redirect('prospection:gestion_concurrents')
        elif type == 'prospect':
            return _redirect_back_or('prospection:prospects_list', request)
        else:
            return redirect('prospection:gestion_entreprises')
    if type == 'concurrent':
        return render(request, f'adminlte/sales/users/entreprises/concurrents.html', {'entreprise': entreprise})
    else:
        return render(request, f'adminlte/sales/users/entreprises/clients-prospects.html', {'entreprise': entreprise})
    
#index
@login_required
def index(request):
    return render(request, 'adminlte/index.html')

#rendez-vous
@login_required
def rv_list(request):
    if request.user.is_superuser:
        rvs = Action.objects.filter(is_RV=True).order_by('-date_heure')
        societes = Societe.objects.all()
    elif getattr(request.user, 'is_RO', False):
        # Get the list of societe IDs for the admin
        societe_ids = list(request.user.societes.values_list('id', flat=True))
        if societe_ids:
            rvs = Action.objects.filter(
                Q(is_RV=True) & 
                (Q(societe_id__in=societe_ids) | 
                 Q(created_by__societe_id__in=societe_ids))
            ).order_by('-date_heure')
            societes = Societe.objects.filter(id__in=societe_ids)
        else:
            rvs = Action.objects.none()
            societes = Societe.objects.none()
    elif getattr(request.user, 'is_RC', False) and hasattr(request.user, 'societe') and request.user.societe:
        rvs = Action.objects.filter(
            Q(is_RV=True) & 
            (Q(societe=request.user.societe) | 
             Q(created_by__societe=request.user.societe))
        ).order_by('-date_heure')
        societes = Societe.objects.filter(id=request.user.societe.id) if request.user.societe else Societe.objects.none()
    else:
        rvs = Action.objects.filter(is_RV=True, created_by=request.user).order_by('-date_heure')
        societes = Societe.objects.filter(id=request.user.societe.id) if hasattr(request.user, 'societe') and request.user.societe else Societe.objects.none()
    
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    societe_filter = request.GET.get('societe', '')
    created_date_from = request.GET.get('created_from', '')
    created_date_to = request.GET.get('created_to', '')
    planned_date_from = request.GET.get('planned_from', '')
    planned_date_to = request.GET.get('planned_to', '')
    realized_date_from = request.GET.get('realized_from', '')
    realized_date_to = request.GET.get('realized_to', '')
    entreprise_filter = request.GET.get('entreprise', '')
    type_entreprise_filter = request.GET.get('type_entreprise', '')
    
    if search_query:
        rvs = rvs.filter(
            Q(sujet__icontains=search_query) |
            Q(compte_rendu__icontains=search_query) |
            Q(notes__icontains=search_query))
    
    if status_filter:
        rvs = rvs.filter(etat=status_filter)
    
    if societe_filter and (request.user.is_superuser or request.user.is_RC or request.user.is_RO):
        rvs = rvs.filter(societe__id=societe_filter)
        
    if type_entreprise_filter:
        if type_entreprise_filter == 'client':
            rvs = rvs.filter(entreprise__is_CLT=True)
        elif type_entreprise_filter == 'prospect':
            rvs = rvs.filter(entreprise__is_Prospect=True, entreprise__is_CLT=False)
        
    if entreprise_filter:
        rvs = rvs.filter(entreprise__id=entreprise_filter)
        
    entreprise_stats = rvs.values('entreprise__id', 'entreprise__nom', 'entreprise__is_CLT') \
                       .filter(entreprise__is_Concurent=False, date_heure_realiser__isnull=False)\
                       .annotate(total_rvs=Count('id')) \
                       .order_by('-entreprise__is_CLT', 'entreprise__nom')
        
    clients_data = []
    prospects_data = []

    for stat in entreprise_stats:
        if stat['entreprise__is_CLT']:
            clients_data.append({
                'name': stat['entreprise__nom'],
                'count': stat['total_rvs']
            })
        else:
            prospects_data.append({
                'name': stat['entreprise__nom'],
                'count': stat['total_rvs']
            })
            
    monthly_stats = rvs.exclude(date_heure_realiser__isnull=True) \
                    .filter(entreprise__is_Concurent=False) \
                    .annotate(month=TruncMonth('date_heure_realiser')) \
                    .values('month', 'entreprise__is_CLT', 'entreprise__is_Prospect') \
                    .annotate(total=Count('id')) \
                    .order_by('month')

    months = []
    client_counts = []
    prospect_counts = []
    total_counts = []

    for stat in monthly_stats:
        month_str = stat['month'].strftime("%Y-%m")
        if month_str not in months:
            months.append(month_str)
            client_counts.append(0)
            prospect_counts.append(0)
            total_counts.append(0)
        
        idx = months.index(month_str)
        if stat['entreprise__is_CLT']:
            client_counts[idx] = stat['total']
        elif stat['entreprise__is_Prospect']:
            prospect_counts[idx] = stat['total']
        
        total_counts[idx] += stat['total']
    
    if created_date_from:
        try:
            created_date_from = datetime.strptime(created_date_from, '%Y-%m-%d')
            rvs = rvs.filter(date_heure__gte=created_date_from)
        except ValueError:
            pass
    
    if created_date_to:
        try:
            created_date_to = datetime.strptime(created_date_to, '%Y-%m-%d')
            rvs = rvs.filter(date_heure__lte=created_date_to)
        except ValueError:
            pass
    
    if planned_date_from:
        try:
            planned_date_from = datetime.strptime(planned_date_from, '%Y-%m-%d')
            rvs = rvs.filter(date_heure_planifie__gte=planned_date_from)
        except ValueError:
            pass
    
    if planned_date_to:
        try:
            planned_date_to = datetime.strptime(planned_date_to, '%Y-%m-%d')
            rvs = rvs.filter(date_heure_planifie__lte=planned_date_to)
        except ValueError:
            pass
    
    if realized_date_from and realized_date_to:
        try:
            realized_date_from = datetime.strptime(realized_date_from, '%Y-%m-%d')
            realized_date_to = datetime.strptime(realized_date_to, '%Y-%m-%d')
            rvs = rvs.filter(
                date_heure_realiser__gte=realized_date_from,
                date_heure_realiser__lte=realized_date_to
            )
        except ValueError:
            pass
    
    entreprises = Entreprise.objects.all().order_by('nom')
    entreprises = entreprises.filter(is_Concurent=False)
    if not request.user.is_superuser:
        if getattr(request.user, 'is_RO', False):
            # Récupérer les IDs des sociétés pour l'admin
            societe_ids = list(request.user.societes.values_list('id', flat=True))
            if societe_ids:
                entreprises = entreprises.filter(societe_id__in=societe_ids)
            else:
                entreprises = Entreprise.objects.none()
        elif hasattr(request.user, 'societe') and request.user.societe:
            entreprises = entreprises.filter(societe=request.user.societe)
        else:
            entreprises = Entreprise.objects.none()
    
    paginator = Paginator(rvs, 100)
    page_number = request.GET.get('page')
    rvs = paginator.get_page(page_number)
    
    context = {
        'rvs': rvs,
        'is_superuser': request.user.is_superuser,
        'is_RC': request.user.is_RC,
        'is_RO': request.user.is_RO,
        'societes': societes,
        'status_choices': dict(Action.ETAT_CHOICES_RENDEZ_VOUS),
        'search_query': search_query,
        'status_filter': status_filter,
        'societe_filter': societe_filter,
        'created_date_from': created_date_from,
        'created_date_to': created_date_to,
        'planned_date_from': planned_date_from,
        'planned_date_to': planned_date_to,
        'realized_date_from': realized_date_from,
        'realized_date_to': realized_date_to,
        'request': request,
        'clients_data': clients_data,
        'prospects_data': prospects_data,
        'entreprises': entreprises,
        'entreprise_filter': entreprise_filter,
        'type_entreprise_filter': type_entreprise_filter,
        'months': months,
        'client_counts': client_counts,
        'prospect_counts': prospect_counts,
        'total_counts': total_counts,
    }
    
    return render(request, 'adminlte/sales/sales/actions/rvs.html', context)

@login_required
def get_all_filtered_rvs(request):
    if request.user.is_superuser:
        rvs = Action.objects.filter(is_RV=True).order_by('-date_heure')
    elif getattr(request.user, 'is_RO', False):
        # Get the list of societe IDs for the admin
        societe_ids = list(request.user.societes.values_list('id', flat=True))
        if societe_ids:
            rvs = Action.objects.filter(
                Q(is_RV=True) & 
                (Q(societe_id__in=societe_ids) | 
                 Q(created_by__societe_id__in=societe_ids))
            ).order_by('-date_heure')
        else:
            rvs = Action.objects.none()
    elif getattr(request.user, 'is_RC', False) and hasattr(request.user, 'societe') and request.user.societe:
        rvs = Action.objects.filter(
            Q(is_RV=True) & 
            (Q(societe=request.user.societe) | 
             Q(created_by__societe=request.user.societe))
        ).order_by('-date_heure')
    else:
        rvs = Action.objects.filter(is_RV=True, created_by=request.user).order_by('-date_heure')
    
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    societe_filter = request.GET.get('societe', '')
    created_date_from = request.GET.get('created_from', '')
    created_date_to = request.GET.get('created_to', '')
    planned_date_from = request.GET.get('planned_from', '')
    planned_date_to = request.GET.get('planned_to', '')
    realized_date_from = request.GET.get('realized_from', '')
    realized_date_to = request.GET.get('realized_to', '')
    
    if search_query:
        rvs = rvs.filter(
            Q(sujet__icontains=search_query) |
            Q(compte_rendu__icontains=search_query) |
            Q(notes__icontains=search_query))
    
    if status_filter:
        rvs = rvs.filter(etat=status_filter)
    
    if societe_filter and (request.user.is_superuser or request.user.is_RC or request.user.is_RO):
        rvs = rvs.filter(societe__id=societe_filter)
    
    if created_date_from:
        try:
            created_date_from = datetime.strptime(created_date_from, '%Y-%m-%d')
            rvs = rvs.filter(date_heure__gte=created_date_from)
        except ValueError:
            pass
    
    if created_date_to:
        try:
            created_date_to = datetime.strptime(created_date_to, '%Y-%m-%d')
            rvs = rvs.filter(date_heure__lte=created_date_to)
        except ValueError:
            pass
    
    if planned_date_from:
        try:
            planned_date_from = datetime.strptime(planned_date_from, '%Y-%m-%d')
            rvs = rvs.filter(date_heure_planifie__gte=planned_date_from)
        except ValueError:
            pass
    
    if planned_date_to:
        try:
            planned_date_to = datetime.strptime(planned_date_to, '%Y-%m-%d')
            rvs = rvs.filter(date_heure_planifie__lte=planned_date_to)
        except ValueError:
            pass
    
    if realized_date_from and realized_date_to:
        try:
            realized_date_from = datetime.strptime(realized_date_from, '%Y-%m-%d')
            realized_date_to = datetime.strptime(realized_date_to, '%Y-%m-%d')
            rvs = rvs.filter(
                date_heure_realiser__gte=realized_date_from,
                date_heure_realiser__lte=realized_date_to
            )
        except ValueError:
            pass
    
    return JsonResponse({
        'rvs': [{
            'id': rv.id,
            'sujet': rv.sujet,
            'etat': rv.etat,
            'date_heure_planifie': rv.date_heure_planifie.isoformat() if rv.date_heure_planifie else None,
            'date_heure_realiser': rv.date_heure_realiser.isoformat() if rv.date_heure_realiser else None,
            'notes': rv.notes,
        } for rv in rvs]
    }, safe=False)

@require_GET
def get_rvs_by_date(request):
    date_str = request.GET.get('date')
    date_type = request.GET.get('date_type', 'planned')
    
    try:
        date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return JsonResponse([], safe=False)
    
    if date_type == 'planned':
        qs = Action.objects.filter(is_RV=True, date_heure_planifie__date=date)
    else:
        qs = Action.objects.filter(is_RV=True, date_heure_realiser__date=date)

    if not request.user.is_superuser:
        if getattr(request.user, 'is_RO', False):
            # Get the list of societe IDs for the admin
            societe_ids = list(request.user.societes.values_list('id', flat=True))
            if societe_ids:
                qs = qs.filter(
                    Q(societe_id__in=societe_ids) | 
                    Q(created_by__societe_id__in=societe_ids)
                )
            else:
                qs = qs.none()
        elif getattr(request.user, 'is_RC', False) and hasattr(request.user, 'societe') and request.user.societe:
            qs = qs.filter(
                Q(societe=request.user.societe) | 
                Q(created_by__societe=request.user.societe)
            )
        else:
            qs = qs.filter(created_by=request.user)

    rvs_list = list(qs.values('id', 'sujet', 'date_heure_planifie', 'date_heure_realiser', 'etat', 'notes'))
    
    for rv in rvs_list:
        if 'date_heure_planifie' in rv and rv['date_heure_planifie']:
            rv['date_heure_planifie'] = rv['date_heure_planifie'].isoformat()
        if 'date_heure_realiser' in rv and rv['date_heure_realiser']:
            rv['date_heure_realiser'] = rv['date_heure_realiser'].isoformat()
    
    return JsonResponse(rvs_list, safe=False)

@login_required
@require_POST
def update_rv_date(request):
    try:
        action_id = request.POST.get('action_id')
        new_date_str = request.POST.get('new_date')
        date_type = request.POST.get('date_type')
        
        action = Action.objects.get(id=action_id)
        # Permission check: SU, RC (same societe), or creator
        if not (
            request.user.is_superuser or request.user.is_RO or
            (getattr(request.user, 'is_RC', False) and action.societe_id == getattr(request.user.societe, 'id', None)) or
            action.created_by_id == request.user.id
        ):
            return JsonResponse({'status': 'error', 'message': "Permission refusée"}, status=403)
        
        new_date = datetime.fromisoformat(new_date_str)
        
        if new_date.tzinfo is not None:
            new_date = new_date.astimezone(pytz.UTC)
            new_date = new_date.replace(tzinfo=None)  
            
        new_date = make_aware(new_date)
        
        if date_type == 'planned':
            action.date_heure_planifie = new_date
        elif date_type == 'realized':
            action.date_heure_realiser = new_date
        
        action.save()
        
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

DATETIME_FMT = "%Y-%m-%dT%H:%M"

def _parse_dt_local_to_aware(dt_str, field_key, errors):
    """
    Parse 'YYYY-MM-DDTHH:MM' (input type="datetime-local") en datetime aware.
    Ajoute une erreur dans `errors` si invalide et renvoie None.
    """
    if not dt_str:
        return None
    try:
        naive = datetime.strptime(dt_str, DATETIME_FMT)
        return timezone.make_aware(naive)
    except ValueError as e:
        # Renvoie l'erreur sur la clé du champ cohérente avec le template
        errors[field_key] = f"Format de date invalide ({field_key})."
        return None

def _fetch_or_create_client(num_tiers, societe):
    """
    Récupère un client local par num_tiers (num_compte) et société ou le crée en se basant
    sur le webservice externe. Lève Exception en cas d'échec.
    """
    # Valider num_tiers
    num_tiers = (num_tiers or '').strip()
    if not num_tiers:
        raise Exception("Numéro de compte client manquant")

    try:
        # Vérifier d'abord si le client existe déjà pour cette société
        return Entreprise.objects.get(num_compte=num_tiers, societe=societe, is_CLT=True)
    except Entreprise.DoesNotExist:
        # Vérifier si le numéro de compte est déjà utilisé par une autre société
        if societe and Entreprise.objects.filter(num_compte=num_tiers, is_CLT=True).exclude(societe=societe).exists():
            raise Exception(f"Le numéro de compte {num_tiers} est déjà utilisé par une autre société.")

        # Appel WS externe pour récupérer les informations du client
        base_url = get_base_url(societe.id if societe else None)
        if not base_url:
            raise Exception("Configuration de la société non trouvée")

        fact_url = f"{settings.SAGE_API_HOST}/WebServices100/{base_url}/TiersService/rest/Clients/{num_tiers}"
        headers = {
            "Authorization": settings.SAGE_API_TOKEN,
            "Accept": "application/json",
        }
        try:
            resp = requests.get(fact_url, headers=headers, timeout=10)
        except requests.RequestException as e:
            raise Exception(f"Connexion au service client impossible: {e}")

        if resp.status_code != 200:
            raise Exception("Client introuvable dans Sage")

        data = resp.json() if resp.content else {}
        name = (data or {}).get("Intitule", "").strip()
        if not name:
            # Ne PAS créer de client local si Sage ne renvoie pas un nom valide
            raise Exception("Client introuvable dans Sage")

        # Création du client avec la société associée (seulement si Sage a renvoyé des infos valides)
        return Entreprise.objects.create(
            nom=name,
            adresse=(data.get("Adresse") or ""),
            telephone=(data.get("Telephone") or ""),
            email=(data.get("Email") or ""),
            num_compte=num_tiers,
            is_CLT=True,
            is_Prospect=False,
            is_Concurent=False,
            societe=societe,
            secteur_activite="",
            date=timezone.now().date(),
        )

def _json_success(message="Opération effectuée", redirect_name="rv_list"):
    return JsonResponse(
        {"success": message, "redirect_url": reverse(redirect_name)},
        status=200,
    )

def _json_success(msg, redirect_url=None):
    return JsonResponse({'success': msg, 'redirect_url': redirect_url or ''})

@login_required
def add_rv(request):
    # Toujours JSON pour l'appel AJAX. Si non-AJAX, on redirige.
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])

    errors = {}
    # Champs
    sujet = (request.POST.get('sujet') or '').strip()
    compte_rendu = (request.POST.get('compte_rendu') or '').strip()
    notes = (request.POST.get('notes') or '').strip()
    etat = (request.POST.get('etat') or '').strip()
    societe_id = (request.POST.get('societe') or '').strip()
    date_planifie_str = (request.POST.get('date_heure_planifie') or '').strip()
    date_realiser_str = (request.POST.get('date_heure_realiser') or '').strip()
    entreprise_id = (request.POST.get('entreprise') or '').strip()  # prospect id OU id client via autocomplete
    num_tiers = (request.POST.get('numTiers') or '').strip()        # client (num_compte)
    client_type = (request.POST.get('client_type') or '').strip()   # 'client' | 'prospect'
    client_societe_name = (request.POST.get('client_societe_name') or '').strip()
    # Validations
    if not sujet:
        errors['sujet'] = "Le sujet est obligatoire"
    if not date_planifie_str:
        errors['date_heure_planifie'] = "La date planifiée est obligatoire"
    if not client_type:
        errors['type'] = "Veuillez choisir Client ou Prospect"
    else:
        if client_type == 'prospect' and not entreprise_id:
            errors['prospect'] = "Veuillez sélectionner un prospect"
        if client_type == 'client' and not num_tiers:
            errors['client'] = "Veuillez sélectionner un client"

    # Société (inférée)
    societe = None
    # priorité: nom transmis par l'autocomplete -> id transmis -> user
    if client_societe_name:
        societe = Societe.objects.filter(nom=client_societe_name).first()
    if not societe and societe_id:
        try:
            societe = Societe.objects.get(id=societe_id)
        except Societe.DoesNotExist:
            societe = None
    if not societe:
        societe = getattr(request.user, 'societe', None)

    # Dates
    date_heure_planifie = _parse_dt_local_to_aware(date_planifie_str, 'date_heure_planifie', errors) if date_planifie_str else None
    date_heure_realiser = _parse_dt_local_to_aware(date_realiser_str, 'date_heure_realiser', errors) if date_realiser_str else None

    if errors:
        # AJAX
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'errors': errors}, status=400)
        # non-AJAX
        for m in errors.values():
            messages.error(request, m)
        return redirect('prospection:rv_list')

    # Entreprise
    try:
        if client_type == 'prospect':
            entreprise = Entreprise.objects.get(id=entreprise_id)
            # si le prospect a une filiale, on l'utilise comme filiale préférée
            if getattr(entreprise, 'societe', None):
                societe = entreprise.societe
        else:
            preferred_societe = Societe.objects.filter(nom=client_societe_name).first() if client_societe_name else societe
            entreprise = _fetch_or_create_client(num_tiers, preferred_societe)
    except Exception as e:
        key = 'prospect' if client_type == 'prospect' else 'client'
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'errors': {key: str(e)}}, status=400)
        messages.error(request, str(e))
        return redirect('prospection:rv_list')

    # Pilote
    pilote = None
    # Filiale finale = entreprise.societe si disponible, sinon societe inférée
    final_societe = entreprise.societe if (entreprise and getattr(entreprise, 'societe', None)) else societe
    if final_societe:
        pilote = get_user_model().objects.filter(is_RC=True, societe=final_societe).first()

    # Création
    try:
        rv = Action.objects.create(
            sujet=sujet,
            compte_rendu=compte_rendu or '',
            notes=notes or '',
            etat=etat,
            societe=final_societe,
            created_by=request.user,
            pilote=pilote,
            is_RV=True,
            entreprise=entreprise,
            date_heure_planifie=date_heure_planifie,
            date_heure_realiser=date_heure_realiser
        )
        logger.info("RV %s créé par %s", rv.id, request.user)
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return _json_success("Rendez-vous créé avec succès")
        messages.success(request, "Rendez-vous créé avec succès")
        return redirect('prospection:rv_list')
    except Exception:
        logger.exception("Erreur création RV")
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'errors': {'__all__': 'Erreur serveur lors de la création'}}, status=500)
        messages.error(request, "Erreur serveur lors de la création")
        return redirect('prospection:rv_list')


@login_required
def edit_rv(request, rv_id):
    rv = get_object_or_404(Action, id=rv_id, is_RV=True)
    errors = {}

    # Permissions
    if not (request.user.is_superuser or request.user.is_RO or 
            (getattr(request.user, 'is_RC', False) and rv.societe_id == getattr(request.user.societe, 'id', None)) or 
            rv.created_by_id == request.user.id):
        return JsonResponse({'error': "Vous n'avez pas la permission de modifier ce rendez-vous"}, status=403)

    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])
    # Champs de base
    rv.sujet = (request.POST.get('sujet') or rv.sujet or '').strip()
    rv.compte_rendu = (request.POST.get('compte_rendu') or rv.compte_rendu or '').strip()
    rv.notes = (request.POST.get('notes') or rv.notes or '').strip()
    rv.etat = (request.POST.get('etat') or rv.etat or '').strip() or rv.etat

    # Dates
    date_planifie_str = (request.POST.get('date_heure_planifie') or '').strip()
    date_realiser_str = (request.POST.get('date_heure_realiser') or '').strip()

    if date_planifie_str:
        dt = _parse_dt_local_to_aware(date_planifie_str, 'edit-date_heure_planifie', errors)
        if dt: rv.date_heure_planifie = dt
    if date_realiser_str:
        dt = _parse_dt_local_to_aware(date_realiser_str, 'edit-date_heure_realiser', errors)
        if dt: rv.date_heure_realiser = dt

    # Société (préférence depuis client_societe_name ou existant ou user)
    client_societe_name = (request.POST.get('client_societe_name') or '').strip()
    preferred_societe = None
    if client_societe_name:
        preferred_societe = Societe.objects.filter(nom=client_societe_name).first()
    if not preferred_societe:
        preferred_societe = rv.societe or getattr(request.user, 'societe', None)

    # Entreprise (client/prospect)
    client_type = (request.POST.get('client_type') or request.POST.get('client_type_selector') or '').strip()
    entreprise_id = (request.POST.get('entreprise_id') or request.POST.get('entreprise') or '').strip()
    
    # DEBUG
    logger.info(f"[edit_rv] client_type={client_type}, entreprise_id={entreprise_id}, client_societe_name={client_societe_name}")
    logger.info(f"[edit_rv] rv.entreprise actuel: ID={rv.entreprise.id if rv.entreprise else None}, is_CLT={rv.entreprise.is_CLT if rv.entreprise else None}")
    
    entreprise = rv.entreprise
    try:
        if client_type == 'client':
            if not entreprise_id:
                errors['edit-client-error'] = "Veuillez sélectionner un client"
            else:
                # D'abord vérifier si le client existe en base (par ID ou num_compte)
                existing_client = None
                try:
                    # Essayer par ID
                    logger.info(f"[edit_rv] Recherche client par ID: {entreprise_id}")
                    existing_client = Entreprise.objects.get(id=int(entreprise_id), is_CLT=True)
                    logger.info(f"[edit_rv] ✓ Client trouvé par ID: {existing_client.nom} (ID={existing_client.id})")
                except (ValueError, Entreprise.DoesNotExist):
                    # Essayer par num_compte
                    logger.info(f"[edit_rv] Recherche client par num_compte: {entreprise_id}")
                    existing_client = Entreprise.objects.filter(num_compte=entreprise_id, is_CLT=True).first()
                    if existing_client:
                        logger.info(f"[edit_rv] ✓ Client trouvé par num_compte: {existing_client.nom}")
                
                if existing_client:
                    # Client existe déjà, on le garde (PAS d'appel Sage)
                    entreprise = existing_client
                else:
                    # Client n'existe pas, on doit le créer depuis Sage
                    logger.info(f"[edit_rv] Client non trouvé localement, appel à _fetch_or_create_client")
                    entreprise = _fetch_or_create_client(entreprise_id, preferred_societe)
                    logger.info(f"[edit_rv] Client créé: {entreprise.nom} (ID={entreprise.id})")
                    
        elif client_type == 'prospect':
            if not entreprise_id:
                errors['entreprise'] = "Veuillez sélectionner un prospect"
            else:
                # Vérifier si c'est le même prospect
                if rv.entreprise and not rv.entreprise.is_CLT and str(rv.entreprise.id) == str(entreprise_id):
                    logger.info(f"[edit_rv] Même prospect, on le garde")
                    entreprise = rv.entreprise
                else:
                    entreprise = Entreprise.objects.get(id=entreprise_id, is_CLT=False)
        # si client_type vide, on ne change pas l'entreprise
    except Entreprise.DoesNotExist:
        errors['entreprise'] = "Prospect sélectionné non valide"
    except Exception as e:
        key = 'edit-client-error' if client_type == 'client' else 'entreprise'
        errors[key] = str(e)
        logger.error(f"[edit_rv] Erreur: {str(e)}", exc_info=True)

    # Validations minimales
    if not rv.sujet:
        errors['edit-sujet-error'] = "Ce champ est obligatoire"
    if not rv.date_heure_planifie:
        errors['edit-date_heure_planifie-error'] = "Ce champ est obligatoire"

    if errors:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'errors': errors}, status=400)
        for m in errors.values():
            messages.error(request, m)
        return redirect('prospection:rv_list')

    # Application finale
    rv.entreprise = entreprise
    if entreprise and getattr(entreprise, 'societe', None):
        rv.societe = entreprise.societe
    elif preferred_societe:
        rv.societe = preferred_societe

    try:
        rv.save()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return _json_success("Rendez-vous modifié avec succès")
        messages.success(request, "Rendez-vous modifié avec succès")
        return redirect('prospection:rv_list')
    except Exception:
        logger.exception("Erreur sauvegarde RV")
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'errors': {'__all__': 'Erreur serveur lors de la sauvegarde'}}, status=500)
        messages.error(request, "Erreur serveur lors de la sauvegarde")
        return redirect('prospection:rv_list')

@login_required
def delete_rv(request, rv_id):
    rv = get_object_or_404(Action, id=rv_id, is_RV=True)
    
    if not (request.user.is_superuser or request.user.is_admiin or
            (request.user.is_RC and rv.societe == request.user.societe) or 
            rv.created_by == request.user):
        return JsonResponse({'error': "Vous n'avez pas la permission de supprimer rendez-vous"}, status=403)
    
    if request.method == 'POST':
        try:
            rv.delete()
            return JsonResponse({'success': 'Rendez-vous supprimé avec succès!'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Méthode de demande invalide'}, status=405)

@login_required
def rv_details(request, rv_id):
    rv = get_object_or_404(Action, id=rv_id, is_RV=True)
    
    if not (request.user.is_superuser or request.user.is_RO or
            (request.user.is_RC and rv.societe == request.user.societe) or 
            (rv.created_by and rv.created_by == request.user)):
        return JsonResponse({'error': "Vous n'avez pas la permission de voir ce rendez-vous"}, status=403)
    
    data = {
        'id': rv.id,
        'sujet': rv.sujet,
        'compte_rendu': rv.compte_rendu,
        'notes': rv.notes,
        'etat': rv.etat,
        'entreprise': rv.entreprise.nom if rv.entreprise else '',
        'societe': rv.societe.nom if rv.societe else '',
        'societe_id': rv.societe.id if rv.societe else None,
        'created_by': rv.created_by.username if rv.created_by else '',
        'pilote': rv.pilote.username if rv.pilote else '',
        'date_heure': rv.date_heure.strftime('%Y-%m-%d %H:%M'),
        'date_heure_planifie': rv.date_heure_planifie.strftime('%Y-%m-%d %H:%M') if rv.date_heure_planifie else '',
        'date_heure_realiser': rv.date_heure_realiser.strftime('%Y-%m-%d %H:%M') if rv.date_heure_realiser else '',
    }
    
    return JsonResponse(data)

def send_appoitment_details_pdf(request):
    if request.method == 'POST':
        try:
            logo_path = finders.find('dist/img/abserveLogo.png')
            
            if not logo_path:
                raise FileNotFoundError(
                    "Logo introuvable. Chemins vérifiés.:\n" + 
                    "\n".join(finders.searched_locations)
                )
            
            with open(logo_path, "rb") as image_file:
                logo_base64 = base64.b64encode(image_file.read()).decode('utf-8') 
                
            recipient_email = request.POST.get('recipient_email')
            subject = request.POST.get('subject')
            message = request.POST.get('message', '')
            appoitment_details = {
                'sujet': request.POST.get('appoitment_details[sujet]'),
                'entreprise': request.POST.get('appoitment_details[entreprise]'),
                'societe': request.POST.get('appoitment_details[societe]'),
                'createdBy': request.POST.get('appoitment_details[createdBy]'),
                'pilote': request.POST.get('appoitment_details[pilote]'),
                'dateHeure': request.POST.get('appoitment_details[dateHeure]'),
                'datePlanifie': request.POST.get('appoitment_details[datePlanifie]'),
                'dateRealiser': request.POST.get('appoitment_details[dateRealiser]'),
                'etat': request.POST.get('appoitment_details[etat]'),
                'compteRendu': request.POST.get('appoitment_details[compteRendu]'),
                'additional_message': message,
                'logo_base64': logo_base64
            }
            
            html_string = render_to_string('adminlte/emails/appoitment_details.html', appoitment_details)
            
            pdf_file = BytesIO()
            pisa_status = pisa.CreatePDF(html_string, dest=pdf_file, encoding='UTF-8', link_callback=lambda uri, _: uri)
            if pisa_status.err:
                return JsonResponse({'success': False, 'error': 'Erreur de génération PDF'})
            pdf_file.seek(0)
            
            email = EmailMessage(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [recipient_email]
            )
            email.attach(f'details_rendez-vous.pdf', pdf_file.getvalue(), 'application/pdf')
            email.send()
            
            return JsonResponse({'success': True})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Méthode de demande invalide'})

#societe
@login_required
def societe(request):
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    societes_list = Societe.objects.all().order_by('nom')
    
    search_query = request.GET.get('search')
    if search_query:
        societes_list = societes_list.filter(nom__icontains=search_query)
    
    paginator = Paginator(societes_list, 5)
    page_number = request.GET.get('page', 1)
    societes = paginator.get_page(page_number)
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        data = {
            'societes': [{'id': s.id, 'nom': s.nom} for s in societes],
            'total': paginator.count,
            'page': societes.number,
            'pages': paginator.num_pages
        }
        return JsonResponse(data)
    return render(request, 'adminlte/sales/users/societe.html', {'societes': societes})

@login_required
def add_societe(request):
    if request.method == 'POST':
        try:
            nom = request.POST.get('nom')
            if not nom:
                return JsonResponse({'errors': {'nom': 'Le nom est requis'}}, status=400)
                
            if Societe.objects.filter(nom=nom).exists():
                return JsonResponse({'errors': {'nom': 'Cette filiale existe déjà'}}, status=400)
                
            societe = Societe.objects.create(nom=nom)
            return JsonResponse({
                'success': 'Filiale ajoutée avec succès',
                'societe': {'id': societe.id, 'nom': societe.nom}
            })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse({'error': 'Méthode non autorisée'}, status=405)

@login_required
def edit_societe(request, id):
    try:
        societe = Societe.objects.get(id=id)
        if request.method == 'POST':
            nom = request.POST.get('nom')
            if not nom:
                return JsonResponse({'errors': {'nom': 'Le nom est requis'}}, status=400)
                
            if Societe.objects.filter(nom=nom).exclude(id=id).exists():
                return JsonResponse({'errors': {'nom': 'Ce nom est déjà utilisé'}}, status=400)
                
            societe.nom = nom
            societe.save()
            return JsonResponse({
                'success': 'Filiale mise à jour avec succès',
                'societe': {'id': societe.id, 'nom': societe.nom}
            })
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    except Societe.DoesNotExist:
        return JsonResponse({'error': 'Filiale non trouvée'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def delete_societe(request, id):
    try:
        societe = Societe.objects.get(id=id)
        if request.method == 'POST':
            societe.delete()
            return JsonResponse({'success': 'Filiale supprimée avec succès'})
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    except Societe.DoesNotExist:
        return JsonResponse({'error': 'Filiale non trouvée'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
                                                                
#user
@login_required
def liste_utilisateurs(request):
    user = request.user

    if user.is_superuser:
        utilisateurs_list = Utilisateur.objects.all().order_by('-date_joined')
        societes = Societe.objects.all().order_by('nom')
    elif getattr(user, 'is_RO', False):
        # Get the list of societe IDs for the admin
        societe_ids = list(user.societes.values_list('id', flat=True))
        if societe_ids:
            utilisateurs_list = Utilisateur.objects.filter(societe_id__in=societe_ids).order_by('-date_joined')
            societes = Societe.objects.filter(id__in=societe_ids).order_by('nom')
        else:
            utilisateurs_list = Utilisateur.objects.none()
            societes = Societe.objects.none()
    elif getattr(user, 'is_RC', False) and getattr(user, 'societe', None):
        utilisateurs_list = Utilisateur.objects.filter(societe=user.societe).order_by('-date_joined')
        societes = Societe.objects.filter(id=user.societe.id).order_by('nom')
    else:
        utilisateurs_list = Utilisateur.objects.none()
        societes = Societe.objects.none()

    is_active_filter = request.GET.get('is_active', '').strip()
    role_filter = request.GET.get('role', '').strip()
    societe_filter = request.GET.get('societe', '').strip()
    search_query = request.GET.get('search', '').strip()

    if user.is_superuser or user.is_RO:
        if societe_filter:
            # Include users assigned directly to the filiale and admins linked via M2M societes
            utilisateurs_list = utilisateurs_list.filter(
                Q(societe_id=societe_filter) | (Q(is_RO=True) & Q(societes__id=societe_filter))
            ).distinct()
    
    if is_active_filter:
        if is_active_filter == 'activated':
            utilisateurs_list = utilisateurs_list.filter(is_active=True)
        elif is_active_filter == 'deactivated':
            utilisateurs_list = utilisateurs_list.filter(is_active=False)

    if role_filter:
        if role_filter == 'RC':
            utilisateurs_list = utilisateurs_list.filter(is_RC=True)
        elif role_filter == 'C':
            utilisateurs_list = utilisateurs_list.filter(is_C=True)
        elif role_filter == 'admin':
            utilisateurs_list = utilisateurs_list.filter(is_RO=True)

    if search_query:
        utilisateurs_list = utilisateurs_list.filter(
            Q(username__icontains=search_query) | Q(email__icontains=search_query)
        )
        
    user_counts = {
        'total_commercials': utilisateurs_list.filter(is_C=True, is_active=True).count(),
        'total_rc': utilisateurs_list.filter(is_RC=True, is_active=True).count(),
        'total_admins': utilisateurs_list.filter(is_RO=True, is_active=True).count(),
        'total_users': utilisateurs_list.filter(is_active=True).count()
    }

    paginator = Paginator(utilisateurs_list, 100)
    page_number = request.GET.get('page', 1)
    utilisateurs = paginator.get_page(page_number)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        html = render_to_string('adminlte/sales/users/user.html', {
            'utilisateurs': utilisateurs,
            'role_filter': role_filter,
            'societe_filter': societe_filter,
            'search_query': search_query,
            'is_active_filter': is_active_filter,
            'societes': societes,
            'is_superuser': request.user.is_superuser,
            'is_RO': request.user.is_RO,
        }, request=request)
        return JsonResponse({'html': html})
    
    return render(request, 'adminlte/sales/users/user.html', {
        'utilisateurs': utilisateurs,
        'role_filter': role_filter,
        'societe_filter': societe_filter,
        'search_query': search_query,
        'is_active_filter': is_active_filter,
        'societes': societes,
        'is_superuser': request.user.is_superuser,
        'is_RO': request.user.is_RO,
        'user_counts': user_counts,
    })
    
@login_required
def ajouter_utilisateur(request):
    if request.method == 'POST':
        errors = {}
        
        try:
            username = request.POST.get('username')
            email = request.POST.get('email')
            telephone = request.POST.get('telephone')
            societe_id = request.POST.get('societe')
            password1 = request.POST.get('password1')
            password2 = request.POST.get('password2')
            role = request.POST.get('role')
            is_RO = request.POST.get('is_RO') in ['1', 'true', 'on']
            societes_ids = request.POST.getlist('societes')
            
            if not username:
                errors['username'] = "Le nom d'utilisateur est requis"
            if not email:
                errors['email'] = "L'e-mail est requis"
            if not telephone:
                errors['telephone'] = 'Le numéro de téléphone est requis'
            # Require either single societe (non-admin) or at least one multi-societe (admin)
            if is_RO:
                if not societes_ids:
                    errors['societes'] = 'Au moins une filiale est requise pour un administrateur'
            else:
                if not societe_id:
                    errors['societe'] = 'La filiale est requise'
            if not password1:
                errors['password1'] = 'Le mot de passe est requis'
            if not password2:
                errors['password2'] = 'La confirmation du mot de passe est requise'
            if not role:
                errors['role'] = 'Le rôle est requis'
            if (not is_RO) and role == 'RC' and societe_id:
                try:
                    societe = Societe.objects.get(id=societe_id)
                    if Utilisateur.objects.filter(societe=societe, is_RC=True).exists():
                        errors['societe'] = "Cette filiale dispose déjà d'un responsable commercial. Une filiale ne peut avoir qu'un seul responsable commercial"
                except Societe.DoesNotExist:
                    errors['societe'] = 'Filiale introuvable'
                
            if username:
                if len(username) < 4:
                    errors['username'] = "Le nom d'utilisateur doit comporter au moins 4 caractères"
                elif len(username) > 30:
                    errors['username'] = "Le nom d'utilisateur ne peut pas dépasser 30 caractères"
                elif not re.match(r'^[a-zA-Z0-9_]+$', username):
                    errors['username'] = "Le nom d'utilisateur ne peut contenir que des lettres, des chiffres et des traits de soulignement"
                elif ' ' in username:
                    errors['username'] = "Le nom d'utilisateur ne peut pas contenir d'espaces"
                    
            if password1:
                if len(password1) < 8:
                    errors['password1'] = 'Le mot de passe doit comporter au moins 8 caractères'
                elif not any(char.isdigit() for char in password1):
                    errors['password1'] = 'Le mot de passe doit contenir au moins un chiffre'
                elif not any(char.isupper() for char in password1):
                    errors['password1'] = 'Le mot de passe doit contenir au moins une lettre majuscule'
                elif not any(char.islower() for char in password1):
                    errors['password1'] = 'Le mot de passe doit contenir au moins une lettre minuscule'
            if errors:
                return JsonResponse({'errors': errors}, status=200)

            # Validate selected societes
            societe = None
            if is_RO:
                # Ensure all ids exist
                valid_ids = list(Societe.objects.filter(id__in=societes_ids).values_list('id', flat=True)) if societes_ids else []
                missing = set(societes_ids) - set(str(x) for x in valid_ids)
                if missing:
                    errors['societes'] = 'Filiale(s) introuvable(s)'
            else:
                try:
                    societe = Societe.objects.get(id=societe_id)
                except Societe.DoesNotExist:
                    errors['societe'] = 'Filiale introuvable'

            if password1 != password2:
                errors['password2'] = 'Les mots de passe ne correspondent pas'

            User = get_user_model()
            if User.objects.filter(username=username).exists():
                errors['username'] = "Ce nom d'utilisateur est déjà utilisé"
            if User.objects.filter(email=email).exists():
                errors['email'] = 'Cet e-mail est déjà utilisé'
            if errors:
                return JsonResponse({'errors': errors}, status=200)

            utilisateur = Utilisateur.objects.create(
                username=username,
                email=email,
                telephone=telephone,
                societe=societe if not is_RO else None,
                password=make_password(password1),
                is_RC=(not is_RO and role == 'RC'),
                is_C=(not is_RO and role == 'C'),
                is_RO=is_RO
            )

            # Attach multiple societes for admin
            if is_RO and societes_ids:
                utilisateur.societes.set(Societe.objects.filter(id__in=societes_ids))

            subject = 'Votre nouveau compte a été créé'
            login_url = request.build_absolute_uri('/login/')
            context = {
                'username': username,
                'password': password1,
                'login_url': login_url,
            }
            html_message = render_to_string('adminlte/sales/users/welcome_email.html', context)
            plain_message = strip_tags(html_message) 
            from_email = settings.DEFAULT_FROM_EMAIL
            to_email = email

            send_mail(
                subject,
                plain_message,
                from_email,
                [to_email],
                html_message=html_message,
                fail_silently=False,
            )

            return JsonResponse({
                'success': 'Utilisateur créé avec succès',
                'user': {
                    'id': utilisateur.id,
                    'username': utilisateur.username,
                    'email': utilisateur.email,
                    'telephone': utilisateur.telephone,
                    'societe': (", ".join(utilisateur.societes.values_list('nom', flat=True)) if utilisateur.is_RO else (str(utilisateur.societe.nom) if utilisateur.societe else None)),
                    'date_joined': utilisateur.date_joined,
                    'is_active': utilisateur.is_active,
                    'role': ('Administrateur' if utilisateur.is_RO else ('Responsable Commercial' if utilisateur.is_RC else 'Commercial'))
                }
            }, status=201)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        form = UtilisateurCreationForm()
    return render(request, 'adminlte/sales/users/user.html', {'form': form, 'errors': {}})

@login_required
def modifier_utilisateur(request, id):
    utilisateur = get_object_or_404(Utilisateur, id=id)
    
    if request.method == 'POST':
        errors = {}
        
        username = request.POST.get('username')
        email = request.POST.get('email')
        telephone = request.POST.get('telephone')
        societe_id = request.POST.get('societe')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        role = request.POST.get('role')
        is_RO = request.POST.get('is_RO') in ['1','true','on'] or role == 'ADMIN'
        societes_ids = request.POST.getlist('societes')
        
        # Validate RC uniqueness only when not admin
        if (not is_RO) and role == 'RC' and societe_id:
                try:
                    societe = Societe.objects.get(id=societe_id)
                    if Utilisateur.objects.filter(societe=societe, is_RC=True).exclude(id=id).exists():
                        errors['role'] = "Cette filiale dispose déjà d'un responsable commercial. Une filiale ne peut avoir qu'un seul responsable commercial"
                except Societe.DoesNotExist:
                    errors['societe'] = 'Filiale introuvable'
        
        if username:
                if len(username) < 4:
                    errors['username'] = "Le nom d'utilisateur doit comporter au moins 4 caractères"
                elif len(username) > 30:
                    errors['username'] = "Le nom d'utilisateur ne peut pas dépasser 30 caractères"
                elif not re.match(r'^[a-zA-Z0-9_]+$', username):
                    errors['username'] = "Le nom d'utilisateur ne peut contenir que des lettres, des chiffres et des traits de soulignement"
                elif ' ' in username:
                    errors['username'] = "Le nom d'utilisateur ne peut pas contenir d'espaces"
                    
        if password1:
                if len(password1) < 8:
                    errors['password1'] = 'Le mot de passe doit comporter au moins 8 caractères'
                elif not any(char.isdigit() for char in password1):
                    errors['password1'] = 'Le mot de passe doit contenir au moins un chiffre'
                elif not any(char.isupper() for char in password1):
                    errors['password1'] = 'Le mot de passe doit contenir au moins une lettre majuscule'
                elif not any(char.islower() for char in password1):
                    errors['password1'] = 'Le mot de passe doit contenir au moins une lettre minuscule'
        
        User = get_user_model()
        if User.objects.filter(username=username).exclude(id=utilisateur.id).exists():
            errors['username'] = "Ce nom d'utilisateur est déjà utilisé"
        if User.objects.filter(email=email).exclude(id=utilisateur.id).exists():
            errors['email'] = 'Cet e-mail est déjà utilisé'
        if password1 and password1 != password2:
            errors['password2'] = 'Les mots de passe ne correspondent pas'
        # Societe requirements: admin needs at least one M2M; non-admin needs FK
        if is_RO:
            if not societes_ids:
                errors['societes'] = 'Au moins une filiale est requise pour un administrateur'
            else:
                valid_ids = list(Societe.objects.filter(id__in=societes_ids).values_list('id', flat=True))
                missing = set(societes_ids) - set(str(x) for x in valid_ids)
                if missing:
                    errors['societes'] = 'Filiale(s) introuvable(s)'
        else:
            if not societe_id:
                errors['societe'] = 'La filiale est requise'
        
        if errors:
            return JsonResponse({'errors': errors}, status=200)
        
        utilisateur.username = username
        utilisateur.email = email
        utilisateur.telephone = telephone
        
        # Assign companies according to admin flag
        if is_RO:
            utilisateur.societe = None
        else:
            try:
                utilisateur.societe = Societe.objects.get(id=societe_id)
            except Societe.DoesNotExist:
                errors['societe'] = 'Filiale introuvable'
                return JsonResponse({'errors': errors}, status=200)
        
        # Toggle roles
        if is_RO:
            utilisateur.is_RO = True
            utilisateur.is_RC = False
            utilisateur.is_C = False
        else:
            utilisateur.is_RO = False
            if role == 'RC':
                utilisateur.is_RC = True
                utilisateur.is_C = False
            elif role == 'C':
                utilisateur.is_RC = False
                utilisateur.is_C = True
        
        if password1:
            utilisateur.password = make_password(password1)
        
        utilisateur.save()
        # Update M2M societes for admin
        if is_RO:
            utilisateur.societes.set(Societe.objects.filter(id__in=societes_ids))
        else:
            utilisateur.societes.clear()
        
        return JsonResponse({
            'success': "L'utilisateur a été mis à jour avec succès",
            'user': {
                'id': utilisateur.id,
                'username': utilisateur.username,
                'email': utilisateur.email,
                'telephone': utilisateur.telephone,
                'societe': (", ".join(utilisateur.societes.values_list('nom', flat=True)) if utilisateur.is_RO else (str(utilisateur.societe) if utilisateur.societe else None)),
                'role': ('Administrateur' if utilisateur.is_RO else ('Responsable Commercial' if utilisateur.is_RC else 'Commercial'))
            }
        })
    
    return JsonResponse({'error': 'Méthode non autorisée'}, status=405)

@login_required
def activer_utilisateur(request, id):
    if request.method == 'POST':
        try:
            utilisateur = get_object_or_404(Utilisateur, id=id)
            username = utilisateur.username
            utilisateur.is_active = True
            utilisateur.save()
            
            return JsonResponse({
                'success': f"L'utilisateur {username} a été activé avec succès",
                'action': 'activated'
            })
        except Exception as e:
            return JsonResponse({
                'error': str(e)
            }, status=500)
    
    return JsonResponse({
        'error': 'Méthode non autorisée'
    }, status=405)

@login_required
def supprimer_utilisateur(request, id):
    if request.method == 'POST':
        try:
            utilisateur = get_object_or_404(Utilisateur, id=id)
            username = utilisateur.username
            utilisateur.is_active = False
            utilisateur.save()
            
            return JsonResponse({
                'success': f"L'utilisateur {username} a été désactivé avec succès",
                'action': 'deactivated'
            })
        except Exception as e:
            return JsonResponse({
                'error': str(e)
            }, status=500)
    
    return JsonResponse({
        'error': 'Méthode non autorisée'
    }, status=405)

def home_view(request):
    return render(request, 'adminlte/sales/login.html')

def notfound_view(request):
    return render(request, 'adminlte/sales/404.html')

@csrf_exempt
def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                remember_me = request.POST.get('remember_me') == 'on'
                redirect_to = 'prospection:index'
                response = redirect(redirect_to)
                if remember_me:
                    response.set_cookie('remember_me', 'true', max_age=1209600)  # 2 weeks
                else:
                    response.delete_cookie('remember_me')
                return response
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    if field == 'username' and error == "Ce champ est obligatoire":
                        messages.error(request, "Le nom d'utilisateur est requis")
                    elif field == 'password' and error == "Ce champ est obligatoire":
                        messages.error(request, "Un mot de passe est requis")
                    elif field == '__all__' and error == "Veuillez saisir un nom d'utilisateur et un mot de passe corrects. Notez que les deux champs peuvent être sensibles à la casse":
                        messages.error(request, "Nom d'utilisateur ou mot de passe incorrect")
                    else:
                        messages.error(request, f'{field}: {error}')
    else:
        form = AuthenticationForm()
    remember_me = request.COOKIES.get('remember_me') == 'true'
    return render(request, 'adminlte/sales/login.html', {'form': form, 'remember_me': remember_me})

@login_required
def logout_view(request):
    logout(request)
    return redirect('prospection:login')

def password_reset_request(request):
    if request.method == 'POST':
        form = PasswordResetForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            user = Utilisateur.objects.filter(email=email).first()
            if user:
                subject = "Réinitialisation du mot de passe demandée"
                email_template_name = "adminlte/sales/password_reset_email.html"
                context = {
                    "email": user.email,
                    'domain': '127.0.0.1:8000',
                    'site_name': 'AB Serve',
                    "uid": urlsafe_base64_encode(force_bytes(user.pk)),
                    "user": user,
                    'token': default_token_generator.make_token(user),
                    'protocol': 'http',
                }
                email_content = render_to_string(email_template_name, context)
                send_mail(subject, None, settings.DEFAULT_FROM_EMAIL, [user.email], html_message=email_content,fail_silently=False)
                messages.success(request, 'Un e-mail a été envoyé pour réinitialiser votre mot de passe.')
                return redirect('prospection:login')
            else:
                messages.error(request, 'Aucun utilisateur trouvé avec cette adresse e-mail. Veuillez vérifier et réessayer.')
    else:
        form = PasswordResetForm()
    return render(request, 'adminlte/sales/password_reset.html', {'form': form})
    
class CustomPasswordResetConfirmView(PasswordResetConfirmView):
    template_name = 'adminlte/sales/password_reset_confirm.html'

class CustomPasswordResetCompleteView(PasswordResetCompleteView):
    template_name = 'adminlte/sales/password_reset_complete.html'
    
@login_required
def get_user_info(request):
    user = request.user
    return JsonResponse({
        'username': user.username,
        'email': user.email,
        'first_name': user.first_name,
        'last_name': user.last_name,
        'is_authenticated': user.is_authenticated
    })
    
@login_required
def profile_view(request):
    user = request.user
    return render(request, 'adminlte/sales/profile.html', {'user': user})
    
@login_required
def update_profile(request):
    if request.method == 'POST':
        user = request.user
        new_email = request.POST.get('email')
        new_telephone = request.POST.get('telephone')
        
        if Utilisateur.objects.filter(email=new_email).exclude(pk=user.pk).exists():
            return JsonResponse({
                'success': False,
                'errors': {'email': ['Cet e-mail est déjà utilisé par un autre compte.']}
            })
        
        user.email = new_email
        user.telephone = new_telephone
        
        try:
            user.save()
            return JsonResponse({
                'success': True,
                'message': 'Votre profil a été mis à jour avec succès.',
                'email': user.email,
                'telephone': user.telephone
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'errors': {'__all__': ["Une erreur s'est produite lors de la mise à jour de votre profil."]}
            })
    
    return JsonResponse({'success': False, 'errors': {'__all__': ['Demande invalide']}})

@login_required
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            return JsonResponse({
                'success': True,
                'message': 'Votre mot de passe a été mis à jour avec succès.'
            })
        else:
            errors = {}
            for field in form.errors:
                errors[field] = form.errors[field]
            return JsonResponse({
                'success': False,
                'errors': errors
            })
    return JsonResponse({'success': False, 'errors': {'__all__': ['Demande invalide']}})

@csrf_exempt
@require_POST
@login_required
def register_fcm_device(request):
    try:
        if request.content_type != 'application/json':
            return JsonResponse({'error': 'Content-Type must be application/json'}, status=400)
        
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        
        token = data.get('token')
        if not token:
            return JsonResponse({'error': 'Token is required'}, status=400)

        device, created = FCMDevice.objects.update_or_create(
            user=request.user,
            registration_id=token,
            defaults={
                'active': True,
                'date_created': timezone.now()
            }
        )
        
        return JsonResponse({
            'status': 'success',
            'created': created,
            'device_id': device.id,
            'token': device.registration_id
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc() 
        return JsonResponse({
            'error': 'Internal server error',
            'details': str(e)
        }, status=500)

@login_required
def all_notifications(request):
    notifications = NotificationUtilisateur.objects.filter(
        utilisateur=request.user
    ).select_related('notification').order_by('-notification__date_heure')
    
    if request.method == 'GET':
        unread_ids = [n.id for n in notifications if not n.est_lu]
        if unread_ids:
            NotificationUtilisateur.objects.filter(
                id__in=unread_ids
            ).update(est_lu=True)
            
    paginator = Paginator(notifications, 4)
    page_number = request.GET.get('page')
    notifications = paginator.get_page(page_number)
    
    return render(request, 'adminlte/notifications/all.html', {
        'notifications': notifications
    })

@api_view(['GET'])
@login_required
def unread_notifications_count(request):
    count = NotificationUtilisateur.objects.filter(
        utilisateur=request.user,
        est_lu=False
    ).count()
    return Response({'count': count})

@api_view(['POST'])
@login_required
def mark_notifications_as_read(request):
    notification_ids = request.data.get('notification_ids', [])
    
    if not notification_ids:
        return Response({'status': 'error', 'message': 'Aucun identifiant de notification fourni'}, status=400)
    
    try:
        updated = NotificationUtilisateur.objects.filter(
            id__in=notification_ids,
            utilisateur=request.user,
            est_lu=False
        ).update(est_lu=True)
        
        return Response({
            'status': 'success',
            'marked': updated,
            'message': f'{updated} notification(s) marquée(s) comme lue(s)'
        })
    except Exception as e:
        return Response({'status': 'error', 'message': str(e)}, status=500)

@api_view(['POST'])
@login_required
def mark_all_notifications_as_read(request):
    updated = NotificationUtilisateur.objects.filter(
        utilisateur=request.user,
        est_lu=False
    ).update(est_lu=True)
    
    return Response({'status': 'success', 'marked': updated})

@require_GET
@login_required
def notifications_api(request):
    try:
        user_notifications = request.user.notificationutilisateur_set.select_related(
            'notification'
        ).order_by('-notification__date_heure')[:10]

        data = {
            'unread_count': request.user.notificationutilisateur_set.filter(est_lu=False).count(),
            'notifications': [
                {
                    'id': nu.id,
                    'notification_id': nu.notification.id,
                    'message': nu.notification.message,
                    'type': nu.notification.type,
                    'details_url': nu.notification.lien_id,
                    'unread': not nu.est_lu,
                    'time_ago': timesince(nu.notification.date_heure, timezone.now()),
                    'url': '#',
                }
                for nu in user_notifications
            ]
        }
        return JsonResponse(data)
    except Exception as e:
        logger.exception("notifications_api error")
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@require_POST
def mark_notification_as_read(request, pk):
    notification = get_object_or_404(NotificationUtilisateur, pk=pk, utilisateur=request.user)
    notification.est_lu = True
    notification.save()
    return JsonResponse({'status': 'success'})

@login_required
def swot_list(request):
    type_filter = request.GET.get('type', '')
    axe_filter = request.GET.get('axe', '')
    societe_filter = request.GET.get('societe', '')
    entreprise_filter = request.GET.get('entreprise', '')
    created_from = request.GET.get('created_from', '')
    created_to = request.GET.get('created_to', '')
    search_query = request.GET.get('search', '')
    
    radar_data = {}
    
    if request.user.is_superuser:
        swots = Swot.objects.all().order_by('-date')
        radar_swots = Swot.objects.all()
        societes = Societe.objects.all()
    elif hasattr(request.user, 'is_RO') and request.user.is_RO:
        societe_ids = list(request.user.societes.values_list('id', flat=True))
        swots = Swot.objects.filter(societe_id__in=societe_ids).order_by('-date')
        radar_swots = Swot.objects.filter(societe_id__in=societe_ids)
        societes = Societe.objects.filter(id__in=societe_ids)
    elif hasattr(request.user, 'is_RC') and request.user.is_RC and hasattr(request.user, 'societe') and request.user.societe:
        swots = Swot.objects.filter(societe=request.user.societe).order_by('-date')
        radar_swots = Swot.objects.filter(societe=request.user.societe)
        societes = Societe.objects.filter(id=request.user.societe_id)
    
    if type_filter:
        swots = swots.filter(type=type_filter)
    if axe_filter:
        swots = swots.filter(axe=axe_filter)
    if societe_filter and (request.user.is_superuser or request.user.is_RO):
        swots = swots.filter(societe_id=societe_filter)
    if entreprise_filter:
        swots = swots.filter(entreprise_id=entreprise_filter)
    if created_from:
        swots = swots.filter(date__gte=created_from)
    if created_to:
        swots = swots.filter(date__lte=created_to)
    if search_query:
        swots = swots.filter(description__icontains=search_query)
    
    if not (request.user.is_superuser or request.user.is_RO) and hasattr(request.user, 'societe'):
        swots = swots.filter(societe=request.user.societe)
        radar_swots = radar_swots.filter(societe=request.user.societe)
        
    if societe_filter and (request.user.is_superuser or request.user.id_admin):
        radar_swots = radar_swots.filter(societe_id=societe_filter)
        
    type_colors = {
        'force': 'rgba(40, 167, 69, 0.6)',    # Vert
        'faiblesse': 'rgba(255, 193, 7, 0.6)', # Jaune
        'opportunite': 'rgba(23, 162, 184, 0.6)', # Bleu
        'menace': 'rgba(220, 53, 69, 0.6)'     # Rouge
    }
    
    for swot in radar_swots:
        type_key = swot.get_type_display()
        if type_key not in radar_data:
            radar_data[type_key] = {
                'axes': {},
                'color': type_colors.get(swot.type, 'rgba(0, 0, 0, 0.6)')
            }
        
        if swot.get_axe_display() not in radar_data[type_key]['axes']:
            radar_data[type_key]['axes'][swot.get_axe_display()] = 0
        radar_data[type_key]['axes'][swot.get_axe_display()] += 1
        
    radar_datasets = []
    radar_labels = [axe[1] for axe in Swot.AXE_CHOICES]
    
    for type_swot, data in radar_data.items():
        dataset = {
            'label': type_swot,
            'data': [data['axes'].get(axe, 0) for axe in radar_labels],
            'backgroundColor': data['color'],
            'borderColor': data['color'].replace('0.6', '1'),
            'borderWidth': 2,
            'pointBackgroundColor': data['color'].replace('0.6', '1'),
            'pointBorderColor': '#fff',
            'pointHoverBackgroundColor': '#fff',
            'pointHoverBorderColor': data['color'].replace('0.6', '1')
        }
        radar_datasets.append(dataset)
    
    paginator = Paginator(swots, 100)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistiques pour les graphiques
    type_distribution = swots.values('type').annotate(count=Count('id')).order_by('-count')
    axe_distribution = swots.values('axe').annotate(count=Count('id')).order_by('-count')
    
    context = {
        'swots': page_obj,
        'type_choices': dict(Swot.TYPE_CHOICES),
        'axe_choices': dict(Swot.AXE_CHOICES),
        'societes': societes,
        'entreprises': Entreprise.objects.filter(
            is_Concurent=True,
            societe_id__in=(societe.id for societe in societes) if not request.user.is_superuser else None
        ) if not request.user.is_superuser else Entreprise.objects.filter(is_Concurent=True),
        'type_filter': type_filter,
        'axe_filter': axe_filter,
        'societe_filter': societe_filter,
        'entreprise_filter': entreprise_filter,
        'created_date_from': created_from,
        'created_date_to': created_to,
        'search_query': search_query,
        'type_distribution': type_distribution,
        'axe_distribution': axe_distribution,
        'is_superuser': request.user.is_superuser,
        'is_RO': request.user.is_RO,
        'radar_labels': radar_labels,
        'radar_datasets': radar_datasets,
        'is_RC': request.user.is_RC,
    }
    
    return render(request, 'adminlte/sales/sales/swot.html', context)

@csrf_exempt
@login_required
def add_swot_web(request):
    if request.method == 'POST':
        try:
            data = request.POST
            swot = Swot(
                type=data.get('type'),
                description=data.get('description'),
                axe=data.get('axe'),
                societe=request.user.societe if not (request.user.is_superuser or request.user.is_RO) else Societe.objects.filter(id=data.get('societe')).first(),
                entreprise_id=data.get('entreprise'),
                created_by=request.user
            )
            swot.save()
            return JsonResponse({'success': True, 'message': 'SWOT ajouté avec succès!'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'})

@login_required
def edit_swot(request, swot_id):
    swot = get_object_or_404(Swot, id=swot_id)
    if request.method == 'POST':
        try:
            data = request.POST
            swot.type = data.get('type')
            swot.description = data.get('description')
            swot.axe = data.get('axe')
            if request.user.is_superuser or request.user.is_RO:
                swot.societe_id = data.get('societe') if data.get('societe') else None
            else:
                swot.societe = request.user.societe
            swot.entreprise_id = data.get('entreprise')
            swot.save()
            return JsonResponse({'success': True, 'message': 'SWOT modifié avec succès!'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'})

@login_required
def delete_swot(request, swot_id):
    swot = get_object_or_404(Swot, id=swot_id)
    if request.method == 'POST':
        try:
            swot.delete()
            return JsonResponse({'success': True, 'message': 'SWOT supprimé avec succès!'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'})

@login_required
def get_swot_details(request, swot_id):
    swot = get_object_or_404(Swot, id=swot_id)
    data = {
        'type': swot.type,
        'type_display': swot.get_type_display(),
        'axe': swot.axe,
        'axe_display': swot.get_axe_display(),
        'description': swot.description if swot.description else '--',
        'societe': swot.societe.nom if swot.societe else 'Non spécifié',
        'entreprise': swot.entreprise.nom,
        'created_by': swot.created_by.username if swot.created_by else 'Utilisateur inconnu',
        'date': swot.date.strftime('%d/%m/%Y'),
    }
    return JsonResponse(data)

class ProspectAnalysisView(View):
    template_name = 'adminlte/sales/sales/actions/analyse.html'
    
    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_monthly_action_counts(self, prospects):
        """
        Récupère le nombre d'actions par mois et par type pour les prospects donnés.
        Retourne un dictionnaire avec les années disponibles et les données mensuelles.
        """
        # Obtenir toutes les actions pour les prospects
        actions = Action.objects.filter(
            entreprise__in=prospects
        ).annotate(
            month=TruncMonth('date_heure')
        ).values('month', 'is_Appel', 'is_Email', 'is_RV').annotate(
            count=Count('id')
        ).order_by('month')
        
        # Initialiser les structures de données
        monthly_data = {}
        years = set()
        
        # Remplir les données mensuelles
        for action in actions:
            month = action['month']
            year = month.year
            month_key = month.strftime('%Y-%m')
            years.add(year)
            
            if month_key not in monthly_data:
                monthly_data[month_key] = {
                    'month': month.strftime('%B %Y'),
                    'call': 0,
                    'email': 0,
                    'meeting': 0
                }
            
            # Compter les actions par type
            if action['is_Appel']:
                monthly_data[month_key]['call'] += action['count']
            elif action['is_Email']:
                monthly_data[month_key]['email'] += action['count']
            elif action['is_RV']:
                monthly_data[month_key]['meeting'] += action['count']
        
        # Trier les mois par ordre chronologique
        sorted_months = sorted(monthly_data.items(), key=lambda x: x[0])
        
        # Convertir en format adapté pour le graphique
        chart_data = {
            'months': [data['month'] for _, data in sorted_months],
            'calls': [data['call'] for _, data in sorted_months],
            'emails': [data['email'] for _, data in sorted_months],
            'meetings': [data['meeting'] for _, data in sorted_months],
            'years': [str(year) for year in sorted(years, reverse=True)]  # Convert years to strings and sort descending
        }
        
        # Convert the data to JSON-serializable format
        chart_data['months'] = json.dumps(chart_data['months'])
        chart_data['calls'] = json.dumps(chart_data['calls'])
        chart_data['emails'] = json.dumps(chart_data['emails'])
        chart_data['meetings'] = json.dumps(chart_data['meetings'])
        chart_data['years'] = json.dumps(chart_data['years'])
        
        return chart_data

    def get(self, request):
        societe_filter = request.GET.get('societe')
        search_query = request.GET.get('search')
        min_score = int(request.GET.get('min_score', 0))
        
        prospects = Entreprise.objects.filter(is_Prospect=True).prefetch_related('actions')
        
        if not (request.user.is_superuser or request.user.is_RO):
            prospects = prospects.filter(societe=request.user.societe)
        
        if request.user.is_RO:
            prospects = prospects.filter(societe__in=request.user.societes.all())
        
        if societe_filter:
            prospects = prospects.filter(societe_id=societe_filter)
        
        if search_query:
            prospects = prospects.filter(
                Q(secteur_activite__icontains=search_query) |
                Q(nom__icontains=search_query))
        
        scorer = ProspectScorer()
        
        # Option pour forcer le réentraînement (admin seulement)
        force_retrain = 'retrain' in request.GET and request.user.is_superuser
        if force_retrain:
            scorer.train_model(prospects, force_retrain=True)
        
        # Analyser chaque prospect
        prospects_data = []
        high_probability = medium_probability = low_probability = 0
        
        for prospect in prospects:
            actions = prospect.actions.all()
            analysis = scorer.calculate_prospect_score(prospect, actions)
            
            # Classer les prospects par score
            if analysis['score'] >= 75:
                high_probability += 1
            elif analysis['score'] >= 50:
                medium_probability += 1
            else:
                low_probability += 1
            
            # Détails des actions
            actions_analysis = analysis['actions_analysis']
            total_calls = actions_analysis.get('call', {}).get('count', 0)
            total_emails = actions_analysis.get('email', {}).get('count', 0)
            total_meetings = actions_analysis.get('meeting', {}).get('count', 0)
            
            last_action = prospect.actions.order_by('-date_heure').first()
            
            sentiment_score = analysis['details']['sentiment']
            if sentiment_score >= 60:
                sentiment_level = 'positive'
                sentiment_text = 'Positif'
            elif sentiment_score >= 30:
                sentiment_level = 'neutral'
                sentiment_text = 'Neutre'
            else:
                sentiment_level = 'negative'
                sentiment_text = 'Négatif'
            
            prospects_data.append({
                'prospect': prospect,
                'score': analysis['score'],
                'details': analysis['details'],
                'actions_analysis': actions_analysis,
                'total_actions': len(actions),
                'total_calls': total_calls,
                'total_emails': total_emails,
                'total_meetings': total_meetings,
                'last_action': last_action.date_heure if last_action else None,
                'sentiment_level': sentiment_level,
                'sentiment_text': sentiment_text,
                'details_json': json.dumps(analysis['details']),  
                'actions_analysis_json': json.dumps(actions_analysis)
            })
        
        # Trier par score décroissant et filtrer par score minimum
        prospects_data = [p for p in prospects_data if p['score'] >= min_score]
        prospects_data.sort(key=lambda x: x['score'], reverse=True)
        
        total_prospects = len(prospects_data)
        high_percentage = (high_probability / total_prospects * 100) if total_prospects > 0 else 0
        medium_percentage = (medium_probability / total_prospects * 100) if total_prospects > 0 else 0
        low_percentage = (low_probability / total_prospects * 100) if total_prospects > 0 else 0
        
        context = {
            'prospects_data': prospects_data,
            'current_date': timezone.now().date(),
            'high_probability': high_probability,
            'medium_probability': medium_probability,
            'low_probability': low_probability,
            'high_percentage': high_percentage,
            'medium_percentage': medium_percentage,
            'low_percentage': low_percentage,
            'societes': Societe.objects.all() if request.user.is_superuser else 
                       (request.user.societes.all() if getattr(request.user, 'is_RO', False) else 
                       Societe.objects.filter(id=request.user.societe.id) if hasattr(request.user, 'societe') and request.user.societe else 
                       Societe.objects.none()),
            'societe_filter': societe_filter,
            'search_query': search_query,
            'min_score': min_score,
            'is_superuser': request.user.is_superuser,
            'is_RO': request.user.is_RO,
            'monthly_action_data': self.get_monthly_action_counts(prospects) if prospects.exists() else None
        }
        
        return render(request, self.template_name, context)
    
    def post(self, request):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # C'est une requête AJAX
            if request.POST.get('action') == 'train_model' and request.user.is_superuser:
                try:
                    prospects = Entreprise.objects.filter(is_Prospect=True).prefetch_related('actions')
                    
                    scorer = ProspectScorer()
                    scorer.train_model(prospects, force_retrain=True)
                    
                    return JsonResponse({'status': 'success'})
                except Exception as e:
                    return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
        
        return JsonResponse({'status': 'error', 'message': 'Requête invalide'}, status=400)

@login_required
def questions(request):
    questions = Question.objects.all()
    q = request.GET.get('filter_question')
    t = request.GET.get('filter_type')
    o = request.GET.get('filter_obligatoire')
    if q:
        questions = questions.filter(question_fr__icontains=q)
    if t:
        questions = questions.filter(type=t)
    if o:
        questions = questions.filter(obligatoire=o)
    paginator = Paginator(questions, 100)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Gestion des permissions pour les sociétés
    if not (request.user.is_superuser or getattr(request.user, 'is_RO', False)):
        # Utilisateur standard - ne voit que sa société
        if hasattr(request.user, 'societe') and request.user.societe:
            societes = Societe.objects.filter(id=request.user.societe_id)
        else:
            societes = Societe.objects.none()
    elif hasattr(request.user, 'is_RO') and request.user.is_RO:
        # Admin - voit ses sociétés
        societe_ids = list(request.user.societes.values_list('id', flat=True))
        societes = Societe.objects.filter(id__in=societe_ids)
    else:
        # Superuser - voit tout
        societes = Societe.objects.all()
        
    qs = Question.objects.all()
    # Calcul des statistiques
    stats = {
        'closed': qs.filter(type='closed').count(),
        'ouinon': qs.filter(type='ouinon').count(),
        'open': qs.filter(type='open').count(),
        'note': qs.filter(type='note').count(),
    }
    
    # Gestion des permissions
    if not (request.user.is_superuser or getattr(request.user, 'is_RO', False)):
        # Utilisateur standard - ne voit que les utilisateurs de sa société
        if hasattr(request.user, 'societe') and request.user.societe:
            users = Utilisateur.objects.filter(societe=request.user.societe)
            clients = Entreprise.objects.filter(societe=request.user.societe, is_CLT=True).order_by('nom')
        else:
            users = Utilisateur.objects.none()
            clients = Entreprise.objects.none()
            
    elif hasattr(request.user, 'is_RO') and request.user.is_RO:
        # Admin - voit les utilisateurs de ses sociétés
        societe_ids = list(request.user.societes.values_list('id', flat=True))
        users = Utilisateur.objects.filter(societe_id__in=societe_ids)
        clients = Entreprise.objects.filter(societe_id__in=societe_ids, is_CLT=True).order_by('nom')
    else:
        # Superuser - voit tout
        users = Utilisateur.objects.all()
        clients = Entreprise.objects.filter(is_CLT=True).order_by('nom')
    context = {'questions': page_obj, 'page_obj': page_obj, 'filter_question': q, 'filter_type': t, 'filter_obligatoire': o, 'stats': stats, 'users': users, 'clients': clients, 'societes': societes}
    return render(request, 'adminlte/sales/sales/enquetes/questions.html', context)

def question_stats(request, question_id):
    try:
        question = Question.objects.get(id=question_id)
    except Question.DoesNotExist:
        return JsonResponse({'error': 'Question non trouvée'}, status=404)
    
    # Scope enquetes et réponses selon les permissions de l'utilisateur
    enq_qs = Enquete.objects.all()
    user = request.user
    
    if user.is_superuser:
        # Les superusers voient toutes les enquêtes (pas de filtre supplémentaire)
        pass
    elif hasattr(user, 'is_RO') and user.is_RO:
        # Admin - voit les enquêtes de toutes ses sociétés
        societe_ids = list(user.societes.values_list('id', flat=True))
        if societe_ids:
            enq_qs = enq_qs.filter(client__societe_id__in=societe_ids)
        else:
            # Si l'admin n'a pas de sociétés, ne rien afficher
            enq_qs = enq_qs.none()
    else:
        # Utilisateur standard - voit les enquêtes de sa société
        if hasattr(user, 'societe') and user.societe:
            enq_qs = enq_qs.filter(client__societe=user.societe)
        else:
            # Si l'utilisateur n'a pas de société, ne rien afficher
            enq_qs = enq_qs.none()

    # Apply optional filters from modal
    client_id = request.GET.get('client_id')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    created_by_id = request.GET.get('created_by_id')
    societe_id = request.GET.get('societe_id')

    if client_id:
        try:
            enq_qs = enq_qs.filter(client_id=int(client_id))
        except (TypeError, ValueError):
            pass
    if created_by_id:
        try:
            enq_qs = enq_qs.filter(created_by_id=int(created_by_id))
        except (TypeError, ValueError):
            pass
    if date_from:
        # date_reponse is a DateField; filter directly with __gte / __lte
        enq_qs = enq_qs.filter(date_reponse__gte=date_from)
    if date_to:
        enq_qs = enq_qs.filter(date_reponse__lte=date_to)

    if societe_id:
        try:
            enq_qs = enq_qs.filter(client__societe_id=int(societe_id))
        except (TypeError, ValueError):
            pass
    
    responses = Reponse.objects.filter(question=question, enquete__in=enq_qs)

    stats = []
    comments = []
    comment_sentiment = None
    average = None
    sentiment_avg = None

    if question.type == 'closed':
        stats = [
            responses.filter(reponse='Très satisfait').count(),
            responses.filter(reponse='Satisfait').count(),
            responses.filter(reponse='Peu satisfait').count(),
            responses.filter(reponse='Pas du tout satisfait').count()
        ]
        comments = [
            {
                'text': r.commentaire,
                'client_name': r.client.nom if getattr(r, 'client', None) else None,
                'date': r.enquete.date_reponse.strftime('%d/%m/%Y') if getattr(r.enquete, 'date_reponse', None) else None
            }
            for r in responses.exclude(commentaire__isnull=True).exclude(commentaire='')
        ]
        if comments:
            analyzer = MultilingualSentimentAnalyzer()
            pos = neu = neg = 0
            total = 0
            for c in comments:
                res = analyzer.analyze(c.get('text') or '')
                s = res.get('score', 0) or 0
                total += 1
                if s > 0.1:
                    pos += 1
                elif s < -0.1:
                    neg += 1
                else:
                    neu += 1
            comment_sentiment = {
                'pos': pos,
                'neu': neu,
                'neg': neg,
                'total': total,
                'pos_pct': round((pos / total) * 100) if total else 0,
                'neg_pct': round((neg / total) * 100) if total else 0,
            }

    elif question.type == 'ouinon':
        stats = [
            responses.filter(reponse='Oui').count(),
            responses.filter(reponse='Non').count(),
            responses.filter(reponse='Peut-être').count()
        ]
        comments = [
            {
                'text': r.commentaire,
                'client_name': r.client.nom if getattr(r, 'client', None) else None,
                'date': r.enquete.date_reponse.strftime('%d/%m/%Y') if getattr(r.enquete, 'date_reponse', None) else None
            }
            for r in responses.exclude(commentaire__isnull=True).exclude(commentaire='')
        ]
        if comments:
            analyzer = FrenchSentimentAnalyzer()
            pos = neu = neg = 0
            total = 0
            for c in comments:
                res = analyzer.analyze(c.get('text') or '')
                s = res.get('score', 0) or 0
                total += 1
                if s > 0.1:
                    pos += 1
                elif s < -0.1:
                    neg += 1
                else:
                    neu += 1
            comment_sentiment = {
                'pos': pos,
                'neu': neu,
                'neg': neg,
                'total': total,
                'pos_pct': round((pos / total) * 100) if total else 0,
                'neg_pct': round((neg / total) * 100) if total else 0,
            }

    elif question.type == 'note':
        # Build histogram and average in Python to be tolerant of formats like '10 ', '10.0'
        stats = [0] * 10
        count = 0
        sum_notes = 0.0
        # Only consider responses with non-empty value
        for r in responses.exclude(reponse__isnull=True).exclude(reponse=''):
            raw = str(r.reponse).strip()
            if not raw:
                continue
            # Accept integers and decimals by parsing as float, then rounding to nearest int
            try:
                # Replace commas with dots for locales if needed
                raw_norm = raw.replace(',', '.')
                val = float(raw_norm)
                note = int(round(val))
            except (ValueError, TypeError):
                continue
            if 1 <= note <= 10:
                stats[note - 1] += 1
                sum_notes += note
                count += 1

        average = (sum_notes / count) if count > 0 else 0
        comments = [
            {
                'text': r.commentaire,
                'client_name': r.client.nom if getattr(r, 'client', None) else None,
                'date': r.enquete.date_reponse.strftime('%d/%m/%Y') if getattr(r.enquete, 'date_reponse', None) else None
            }
            for r in responses.exclude(commentaire__isnull=True).exclude(commentaire='')
        ]
        if comments:
            analyzer = FrenchSentimentAnalyzer()
            pos = neu = neg = 0
            total = 0
            for c in comments:
                res = analyzer.analyze(c.get('text') or '')
                s = res.get('score', 0) or 0
                total += 1
                if s > 0.1:
                    pos += 1
                elif s < -0.1:
                    neg += 1
                else:
                    neu += 1
            comment_sentiment = {
                'pos': pos,
                'neu': neu,
                'neg': neg,
                'total': total,
                'pos_pct': round((pos / total) * 100) if total else 0,
                'neg_pct': round((neg / total) * 100) if total else 0,
            }

    elif question.type == 'open':
        analyzer = MultilingualSentimentAnalyzer()
        texts_qs = responses.exclude(reponse='').values_list('reponse', flat=True)
        pos = neu = neg = 0
        total = 0
        score_sum = 0.0
        for txt in texts_qs:
            res = analyzer.analyze(txt or '')
            score = res.get('score', 0) or 0
            score_sum += score
            total += 1
            if score > 0.1:
                pos += 1
            elif score < -0.1:
                neg += 1
            else:
                neu += 1
        stats = [pos, neu, neg]
        sentiment_avg = (score_sum / total) if total > 0 else 0
        comments = [
            {
                'text': r.reponse,
                'client_name': r.client.nom if getattr(r, 'client', None) else None,
                'date': r.enquete.date_reponse.strftime('%d/%m/%Y') if getattr(r.enquete, 'date_reponse', None) else None
            }
            for r in responses.exclude(reponse='')
        ]

    # Get the language from the request parameters or default to French
    language = request.GET.get('lang', 'fr').lower()
    
    # Get the question text in the requested language, defaulting to French
    question_text = question.question_fr  # Default to French
    if language == 'en' and question.question_en:
        question_text = question.question_en
    elif language == 'de' and question.question_de:
        question_text = question.question_de
    
    return JsonResponse({
        'question_id': question.id,
        'question_text': question_text,
        'question_text_fr': question.question_fr,
        'question_text_en': question.question_en,
        'question_text_de': question.question_de,
        'question_type': question.type,
        'stats': stats,
        'comments': comments,
        'total_responses': responses.count(),
        'average': average if question.type == 'note' else None,
        'sentiment_avg': sentiment_avg if question.type == 'open' else None,
        'comment_sentiment': comment_sentiment,
        'language': language
    })
    
def add_question(request):
    if request.method == 'POST':
        try:
            question_fr = request.POST.get('question_fr')
            question_en = request.POST.get('question_en')
            question_de = request.POST.get('question_de', '')  # Champ optionnel
            question_type = request.POST.get('type')
            obligatoire = request.POST.get('obligatoire', '0') == '1'
            questions = Question.objects.all()
            paginator = Paginator(questions, 5)  
            last_page = paginator.num_pages
            
            # Validate required fields
            if not question_fr:
                return JsonResponse({'errors': {'question_fr': 'La question en français est requise'}}, status=400)
            
            if not question_en:
                return JsonResponse({'errors': {'question_en': 'The English question is required'}}, status=400)
            
            if not question_type:
                return JsonResponse({'errors': {'type': 'Le type de question est requis'}}, status=400)
                
            valid_types = [choice[0] for choice in Question.Type.choices]
            if question_type not in valid_types:
                return JsonResponse({'errors': {'type': 'Type de question invalide'}}, status=400)
                
            # Check if either French or English version already exists
            if (Question.objects.filter(question_fr=question_fr).exists() or 
                Question.objects.filter(question_en=question_en).exists()):
                return JsonResponse({
                    'errors': {
                        'question': 'Une question avec ce texte existe déjà (en français ou en anglais)'
                    }
                }, status=400)
                
            question = Question.objects.create(
                question_fr=question_fr,
                question_en=question_en,
                question_de=question_de,
                type=question_type,
                obligatoire=obligatoire
            )

            qs = Question.objects.all()
            stats = {
                'closed': qs.filter(type='closed').count(),
                'ouinon': qs.filter(type='ouinon').count(),
                'open': qs.filter(type='open').count(),
                'note': qs.filter(type='note').count(),
            }
            
            return JsonResponse({
                'success': 'Question ajoutée avec succès',
                'question': {
                    'id': question.id, 
                    'question_fr': question.question_fr,
                    'question_en': question.question_en,
                    'question_de': question.question_de,
                    'type': question.type,
                    'obligatoire': question.obligatoire
                },
                'last_page': last_page,
                'stats': stats
            })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse({'error': 'Méthode non autorisée'}, status=405)

@login_required
def edit_question(request, question_id):
    question = get_object_or_404(Question, id=question_id)
    if request.method == 'POST':
        try:
            question_fr = request.POST.get('question_fr')
            question_en = request.POST.get('question_en')
            question_de = request.POST.get('question_de', '')  # Champ optionnel
            
            if not question_fr:
                return JsonResponse({'errors': {'question_fr': 'La question en français est requise'}}, status=400)
            
            if not question_en:
                return JsonResponse({'errors': {'question_en': 'The English question is required'}}, status=400)
            
            # Check if another question with the same text exists
            if (Question.objects.filter(question_fr=question_fr).exclude(id=question_id).exists() or 
                Question.objects.filter(question_en=question_en).exclude(id=question_id).exists()):
                return JsonResponse({
                    'errors': {
                        'question': 'Une question avec ce texte existe déjà (en français ou en anglais)'
                    }
                }, status=400)
            
            question.question_fr = question_fr
            question.question_en = question_en
            question.question_de = question_de
            question.type = request.POST.get('type')
            question.obligatoire = request.POST.get('obligatoire') == '1'
            question.save()
            
            return JsonResponse({
                'success': 'Question modifiée avec succès',
                'question': {
                    'id': question.id,
                    'question_fr': question.question_fr,
                    'question_en': question.question_en,
                    'type': question.type,
                    'obligatoire': question.obligatoire
                },
            })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse({'error': 'Méthode non autorisée'}, status=405)

@login_required
def delete_question(request, question_id):
    question = get_object_or_404(Question, id=question_id)
    if request.method == 'POST':
        try:
            question.delete()
            return JsonResponse({'success': 'Question supprimée avec succès'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse({'error': 'Méthode non autorisée'}, status=405)

def _filters_q(request):
    """Construit le Q() à partir des mêmes filtres que la page + renvoie un dict meta utile."""
    q = Q()
    g = request.GET

    # Champs texte
    titre = g.get('titre')
    if titre:
        q &= Q(titre__icontains=titre)

    # Dates
    def _rng(field):
        vmin = g.get(f'{field}_min')
        vmax = g.get(f'{field}_max')
        if vmin:
            q_nonlocal = Q(**{f'{field}__gte': vmin})
        else:
            q_nonlocal = Q()
        if vmax:
            q_nonlocal &= Q(**{f'{field}__lte': vmax})
        return q_nonlocal

    q &= _rng('date_creation')
    q &= _rng('date_modification')
    q &= _rng('date_envoi')
    q &= _rng('date_rappel')
    q &= _rng('date_reponse')

    # Booléens via présence de dates
    envoyee = g.get('envoyee')
    if envoyee == '1':
        q &= Q(date_envoi__isnull=False)
    elif envoyee == '0':
        q &= Q(date_envoi__isnull=True)

    rappellee = g.get('rappellee')
    if rappellee == '1':
        q &= Q(date_rappel__isnull=False)
    elif rappellee == '0':
        q &= Q(date_rappel__isnull=True)

    repondue = g.get('repondue')
    if repondue == '1':
        q &= Q(date_reponse__isnull=False)
    elif repondue == '0':
        q &= Q(date_reponse__isnull=True)

    # Sélecteurs
    client = g.get('client')
    if client:
        q &= Q(client_id=client)

    # Restriction de périmètre
    if request.user.is_superuser or request.user.is_RO:
        societe = g.get('societe')
        if societe:
            q &= Q(client__societe_id=societe)
        created_by = g.get('created_by')
        if created_by:
            q &= Q(created_by_id=created_by)
    else:
        # Non superuser : force la société de l'utilisateur
        q &= Q(client__societe=request.user.societe)

    return q

def _base_queryset(request):
    qs = Enquete.objects.all()
    if not request.user.is_superuser:
        if hasattr(request.user, 'is_RO') and request.user.is_RO:
            # Pour les admins, on filtre par leurs sociétés
            societe_ids = list(request.user.societes.values_list('id', flat=True))
            qs = qs.filter(
                Q(client__societe_id__in=societe_ids) |
                Q(created_by=request.user) |
                Q(created_by__societe_id__in=societe_ids)
            )
        else:
            # Pour les utilisateurs standards, on filtre par leur société uniquement
            if hasattr(request.user, 'societe') and request.user.societe:
                qs = qs.filter(
                    Q(client__societe=request.user.societe) |
                    Q(created_by=request.user) |
                    Q(created_by__societe=request.user.societe)
                )
            else:
                # Si l'utilisateur n'a pas de société, on ne retourne que ce qu'il a créé
                qs = qs.filter(created_by=request.user)
    return qs

def _list_context(request, page_obj):
    """Données annexes pour le template + la QS filtrée servlet friendly."""
    if request.user.is_superuser:
        clients = Entreprise.objects.filter(is_CLT=True).select_related('societe').order_by('nom')
        societes = Societe.objects.all() or []
        users = Utilisateur.objects.all()
    elif hasattr(request.user, 'is_RO') and request.user.is_RO:
        # Pour les admins, on récupère les sociétés associées
        societe_ids = list(request.user.societes.values_list('id', flat=True))
        clients = Entreprise.objects.filter(is_CLT=True, societe_id__in=societe_ids).order_by('nom')
        societes = Societe.objects.filter(id__in=societe_ids) or []
        users = Utilisateur.objects.filter(societe_id__in=societe_ids)
    else:
        # Pour les utilisateurs standards
        if hasattr(request.user, 'societe') and request.user.societe:
            clients = Entreprise.objects.filter(is_CLT=True, societe=request.user.societe).order_by('nom')
            societes = Societe.objects.filter(id=request.user.societe_id) or []
            users = Utilisateur.objects.filter(societe=request.user.societe)
        else:
            # Si l'utilisateur n'a pas de société, on ne retourne que lui-même
            clients = Entreprise.objects.none()
            societes = Societe.objects.none()
            users = Utilisateur.objects.filter(id=request.user.id)
    
    questions_non_obl = Question.objects.filter(obligatoire=False)
    all_questions = Question.objects.all().order_by('-id')

    # Préparer QS sans le param page pour la pagination propre
    qs_params = request.GET.copy()
    qs_params.pop('page', True)
    qs_string = qs_params.urlencode()

    _client_id = request.GET.get('client', '')
    _societe_id = request.GET.get('societe', '')
    _created_by_id = request.GET.get('created_by', '')
    _envoyee = request.GET.get('envoyee', '')
    _rappellee = request.GET.get('rappellee', '')
    _repondue = request.GET.get('repondue', '')

    return {
        'page_obj': page_obj,
        'clients': clients,
        'societes': societes,
        'users': users,
        'questions_non_obl': questions_non_obl,
        'all_questions': all_questions,
        'qs': qs_string,  # à réutiliser dans les liens de pagination
        'clients_opts': [(cl, str(cl.id) == _client_id) for cl in clients],
        'societes_opts': [(s, str(s.id) == _societe_id) for s in societes],
        'users_opts': [(u, str(u.id) == _created_by_id) for u in users],
        'filter_envoyee_oui': _envoyee == '1',
        'filter_envoyee_non': _envoyee == '0',
        'filter_rappellee_oui': _rappellee == '1',
        'filter_rappellee_non': _rappellee == '0',
        'filter_repondue_oui': _repondue == '1',
        'filter_repondue_non': _repondue == '0',
    }

@login_required
def enquetes_list(request):
    """Liste + filtres + modals + charts (charts chargés via endpoint JSON)"""
    q = _filters_q(request)
    enquetes = _base_queryset(request).filter(q).order_by('-date_creation', '-id')

    paginator = Paginator(enquetes, 100)
    page_number = request.GET.get('page') or 1
    page_obj = paginator.get_page(page_number)

    ctx = _list_context(request, page_obj)
    return render(request, 'adminlte/sales/sales/enquetes/enquetes.html', ctx)

def ensure_client_from_form(request, field_name, societe=None):
    """
    Retourne une instance Entreprise à partir du champ 'client' du formulaire.
    - Si 'client' est un entier => on le prend comme ID interne.
    - Sinon => on le traite comme Numéro de compte Sage et on crée (ou retrouve) l'entreprise via l'API.
    Lève ValueError avec message utilisateur si échec.
    """
    if not societe:
        raise ValueError("Une société est requise pour trouver un client.")
    raw = (request.POST.get(field_name) or '').strip()
    if not raw:
        raise ValueError("Veuillez sélectionner un client.")

    # 1) Essai ID interne
    try:
        cid = int(raw)
        return Entreprise.objects.get(id=cid, societe=societe)
    except (ValueError, Entreprise.DoesNotExist):
        pass  # Ce n'est pas un ID interne, on traite comme num_compte Sage

    num_compte = raw  # alphanum acceptable

    # Déjà en base avec ce num_compte ?
    existing = Entreprise.objects.filter(num_compte=num_compte, societe=societe).first()
    if existing:
        return existing

    # Sinon: récupérer via l'API Sage et créer
    from .utils.client_utils import get_base_url
    base_url = get_base_url(societe.id)
    if not base_url:
        raise ValueError("Impossible de déterminer la société pour interroger Sage.")
    url = f"{settings.SAGE_API_HOST}/WebServices100/{base_url}/TiersService/rest/Clients/{num_compte}"
    headers = {
        'Authorization': settings.SAGE_API_TOKEN,
        'Accept': 'application/json'
    }

    try:
        resp = requests.get(url, headers=headers, timeout=10)
    except requests.RequestException:
        raise ValueError("Échec de communication avec Sage (timeout).")
    
    if resp.status_code != 200:
        raise ValueError(f"Client {num_compte} introuvable dans Sage (HTTP {resp.status_code}).")
    
    data = resp.json() or {}

    # Champs potentiels côté Sage (adapte les clés si besoin)
    nom = data.get('Intitule', f"Client {num_compte}")
    adresse = data.get('Adresse', '') or data.get('fact_adresse', '')
    email = data.get('EMail', '') or data.get('fact_email', '')
    telephone = data.get('Telephone', '') or data.get('fact_tel', '')

    # NE PAS utiliser transaction.atomic() ici car déjà dans une transaction parente
    # Cela causerait "database is locked" avec SQLite
    try:
        created, was_created = Entreprise.objects.get_or_create(
            num_compte=num_compte,
            societe=societe,
            defaults={
                'nom': nom,
                'adresse': adresse,
                'email': email or '',
                'telephone': telephone,
                'is_CLT': True,
                'is_Prospect': False,
                'is_Concurent': False,
                'date': timezone.now().date(),
                'secteur_activite': 'Non spécifié'
            }
        )
        
        # Si l'entreprise existait déjà, mettre à jour avec les données Sage
        if not was_created:
            created.nom = nom
            created.adresse = adresse
            if email:
                created.email = email
            created.telephone = telephone
            created.is_CLT = True
            created.save()
        
        return created
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise ValueError(f"Erreur lors de la création/récupération du client: {str(e)}")

@login_required
def add_enquete(request):
    if request.method != 'POST':
        # fallback GET => revenir à la liste
        return redirect('prospection:enquetes')

    description = (request.POST.get('description') or '').strip()
    questions_ids = request.POST.getlist('questions')
    societe_nom = request.POST.get('societe')
    societe = None
    
    if societe_nom:
        societe = Societe.objects.filter(nom=societe_nom).first()
    errors = {}

    try:
        client = ensure_client_from_form(request, 'client', societe)
    except ValueError as e:
        errors['client'] = str(e)

    if not description:
        errors['description'] = "La description est obligatoire."

    if errors:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'errors': errors}, status=400)
        for v in errors.values():
            messages.error(request, v)
        return redirect('prospection:enquetes')

    # Date de création
    date_creation_str = request.POST.get('date_creation')
    if date_creation_str:
        try:
            date_creation = timezone.datetime.strptime(date_creation_str, '%Y-%m-%d').date()
        except Exception:
            date_creation = timezone.now().date()
    else:
        date_creation = timezone.now().date()

    try:
        with transaction.atomic():
            enquete = Enquete.objects.create(
                description=description,
                date_creation=date_creation,
                created_by=request.user,
                client=client,
            )
            # titre convention
            enquete.titre = f"{enquete.date_creation}_{client.nom}"
            enquete.save()

            # Questions : obligatoires + cochées
            questions_obligatoires = Question.objects.filter(obligatoire=True)
            questions_selectionnees = Question.objects.filter(id__in=questions_ids)
            enquete.questions.set(list(questions_obligatoires) + list(questions_selectionnees))
    except Exception as e:
        logger.exception("Erreur création enquête")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'errors': {'__all__': str(e)}}, status=400)
        messages.error(request, f"Erreur lors de la création : {e}")
        return redirect('prospection:enquetes')

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # renvoyer la dernière page
        paginator = Paginator(_base_queryset(request).filter(_filters_q(request)).order_by('-date_creation', '-id'), 20)
        last_page = paginator.num_pages or 1
        return JsonResponse({"success": "Enquête créée", "last_page": last_page})

    return redirect('prospection:enquetes')

@login_required
@require_POST
def edit_enquete(request, enquete_id):
    enquete = get_object_or_404(Enquete, id=enquete_id)
    description = (request.POST.get('description') or '').strip()
    questions_ids = request.POST.getlist('questions')
    societe_nom = request.POST.get('societe')
    societe = None
    
    if societe_nom:
        societe = Societe.objects.filter(nom=societe_nom).first()

    # Validation simple
    if not description:
        msg = "La description est obligatoire."
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'error': msg}, status=400)
        messages.error(request, msg)
        return redirect('prospection:enquetes')

    # Maj client autorisée si non envoyé (parité backend/frontend)
    if not getattr(enquete, 'date_envoi', None):
        raw_client = (request.POST.get('client') or '').strip()
        if raw_client:
            try:
                new_client = ensure_client_from_form(request, 'client', societe)
                enquete.client = new_client
                enquete.titre = f"{enquete.date_creation}_{new_client.nom}"
            except ValueError as e:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'error': str(e)}, status=400)
                messages.error(request, str(e))
                return redirect('prospection:enquetes')

    enquete.description = description
    enquete.date_modification = timezone.now().date()

    # Toujours conserver les obligatoires
    obligatory_ids = list(Question.objects.filter(obligatoire=True).values_list('id', flat=True))
    new_set = set(obligatory_ids)
    new_set.update([int(qid) for qid in questions_ids if qid])

    try:
        with transaction.atomic():
            enquete.save()
            enquete.questions.set(new_set)
    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'error': f"Erreur de sauvegarde : {e}"}, status=400)
        messages.error(request, f"Erreur de sauvegarde : {e}")
        return redirect('prospection:enquetes')

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True, 'message': "Enquête mise à jour avec succès"})

    return redirect('prospection:enquetes')

@login_required
@require_POST
def delete_enquete(request, enquete_id):
    enquete = get_object_or_404(Enquete, id=enquete_id)
    if enquete.date_envoi:
        return JsonResponse({'error': "Impossible de supprimer une enquête déjà envoyée."}, status=400)
    titre = enquete.titre
    try:
        enquete.delete()
        return JsonResponse({'success': f"L'enquête « {titre} » a été supprimée avec succès"})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def get_enquete_data(enquete):
    """Récupère les données structurées d'une enquête"""
    questions = enquete.questions.all()
    reponses = {r.question_id: r for r in enquete.reponses.all()}
    
    return {
        'id': enquete.id,
        'titre': enquete.titre,
        'description': enquete.description,
        'client': str(enquete.client) if enquete.client else None,
        'created_by': str(enquete.created_by) if enquete.created_by else None,
        'date_creation': enquete.date_creation.strftime('%d/%m/%Y') if enquete.date_creation else None,
        'date_modification': enquete.date_modification.strftime('%d/%m/%Y') if enquete.date_modification else None,
        'date_envoi': enquete.date_envoi.strftime('%d/%m/%Y') if enquete.date_envoi else None,
        'date_rappel': enquete.date_rappel.strftime('%d/%m/%Y') if enquete.date_rappel else None,
        'date_reponse': enquete.date_reponse.strftime('%d/%m/%Y') if enquete.date_reponse else None,
        'is_completed': enquete.is_completed,
        'questions': [
            {
                'id': q.id,
                'question': q.question_fr,
                'type': q.get_type_display(),
                'type_code': q.type,
                'obligatoire': q.obligatoire,
                'reponse': reponses.get(q.id).reponse if reponses.get(q.id) else None,
                'commentaire': reponses.get(q.id).commentaire if reponses.get(q.id) else None,
            } for q in questions
        ]
    }

def detail_enquete(request, enquete_id):
    e = get_object_or_404(Enquete, id=enquete_id)
    data = {
        'id': e.id,
        'titre': e.titre,
        'client': getattr(e.client, 'nom', ''),
        'created_by': getattr(e.created_by, 'get_full_name', lambda: e.created_by.username)(),
        'description': e.description,
        'date_creation': e.date_creation.isoformat() if e.date_creation else '',
        'date_modification': e.date_modification.isoformat() if e.date_modification else '',
        'date_envoi': e.date_envoi.isoformat() if e.date_envoi else '',
        'date_reponse': e.date_reponse.isoformat() if e.date_reponse else '',
        'is_completed': getattr(e, 'is_completed', False),
        'questions': [],
    }
    # joindre réponses
    rep_map = {
        (r.question_id): {'reponse': r.reponse, 'commentaire': r.commentaire}
        for r in Reponse.objects.filter(enquete=e)
    }
    for q in e.questions.all():
        meta = rep_map.get(q.id, {'reponse': None, 'commentaire': None})
        data['questions'].append({
            'id': q.id,
            'question': q.question_fr,
            'type': q.type,
            'obligatoire': q.obligatoire,
            'reponse': meta['reponse'],
            'commentaire': meta['commentaire'],
        })
    return JsonResponse(data)

def download_enquete_pdf(request, enquete_id):
    """Génère et télécharge un PDF de l'enquête (sans LibreOffice)"""
    enquete = get_object_or_404(Enquete, id=enquete_id)
    data = get_enquete_data(enquete)

    # Regrouper les questions par type pour imiter la structure du modèle
    type_map = {
        'closed': 'Échelle',
        'ouinon': 'Oui - Non - Peut-être',
        'open': 'Commentaire',
        'note': 'Note (1-10)'
    }
    grouped = {label: [] for label in type_map.values()}
    for q in data['questions']:
        label = type_map.get(q['type_code'], q['type'])
        grouped.setdefault(label, []).append(q)

    # Resolve logo absolute path and convert to file URI (xhtml2pdf friendly)
    _logo_abs = (finders.find('dist/img/abserveLogo.png') or finders.find('dist/img/logo.png'))
    _logo_uri = None
    _logo_data_uri = None
    if _logo_abs:
        # Build a proper file URI for xhtml2pdf (file:///C:/... with forward slashes)
        _logo_uri = 'file:///' + _logo_abs.replace('\\', '/')
        try:
            with open(_logo_abs, 'rb') as _f:
                _b64 = base64.b64encode(_f.read()).decode('ascii')
                ext = os.path.splitext(_logo_abs)[1].lower().lstrip('.')
                mime = 'png' if ext == 'png' else 'jpeg'
                _logo_data_uri = f"data:image/{mime};base64,{_b64}"
        except Exception:
            _logo_data_uri = None

    context = {
        'enquete': enquete,
        'client': data['client'],
        'created_by': data['created_by'],
        'date_creation': data['date_creation'],
        'date_envoi': data['date_envoi'],
        'date_reponse': data['date_reponse'],
        'grouped': grouped,
        'now': timezone.now().strftime('%d/%m/%Y %H:%M'),
        # Absolute filesystem path and URI for xhtml2pdf image loading
        'logo_path': _logo_abs,
        'logo_uri': _logo_uri,
        'logo_data_uri': _logo_data_uri,
    }

    html = render_to_string('adminlte/sales/sales/enquetes/pdf_template.html', context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode('utf-8')), result, encoding='utf-8')
    if pdf.err:
        return HttpResponse('Une erreur est survenue lors de la génération du PDF', status=500)
    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="enquete_{enquete.id}.pdf"'
    return response

def download_enquete_excel(request, enquete_id):
    """Génère et télécharge un fichier Excel qui MIMIQUE la page PDF (une seule feuille)."""
    enquete = get_object_or_404(Enquete, id=enquete_id)
    data = get_enquete_data(enquete)

    # Regrouper les questions par type comme pour le PDF
    type_map = {
        'closed': 'Échelle',
        'ouinon': 'Oui - Non - Peut-être',
        'open': 'Commentaire',
        'note': 'Note (1-10)'
    }
    grouped = {label: [] for label in type_map.values()}
    for q in data['questions']:
        label = type_map.get(q['type_code'], q['type'])
        grouped.setdefault(label, []).append(q)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Enquête'

    # Largeurs colonnes (approx. pour ressembler à la mise en page PDF)
    col_widths = {
        'A': 8,   # N° / logo zone
        'B': 38,  # Titre / infos
        'C': 30,  # Contact
        'D': 18,  # Espace / Type
        'E': 54,  # Question
        'F': 18,  # Réponse
        'G': 30,  # Commentaire
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # Styles de base
    thin = Side(style='thin', color='999999')
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    # Blue theme and bold title
    title_font = Font(name='Calibri', size=18, bold=True, color='1E88E5')
    bold_font = Font(bold=True)
    # Light blue fill for section headers
    header_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

    # Bande verte fine (comme PDF)
    ws.merge_cells('A1:G1')
    # Top thin band in blue
    ws['A1'].fill = PatternFill(start_color='1E88E5', end_color='1E88E5', fill_type='solid')
    ws.row_dimensions[1].height = 5

    row = 2
    # Ligne d'entête: logo | titre | coordonnées
    # Logo
    try:
        _logo_abs = (finders.find('dist/img/abserveLogo.png') or finders.find('dist/img/logo.png'))
        if _logo_abs:
            from openpyxl.drawing.image import Image as XLImage  # local import to avoid hard dependency
            img = XLImage(_logo_abs)
            img.width = 90
            img.height = 90
            ws.add_image(img, f'A{row}')
            ws.row_dimensions[row].height = 70
    except Exception:
        pass

    # Centrer horizontalement l'impression de la page
    ws.print_options.horizontalCentered = True

    # Titre centré au milieu (B..E) et coordonnées à gauche (F..G)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)  # B:E
    cell_title = ws.cell(row=row, column=2, value='Questionnaire Satisfaction Client')
    cell_title.font = title_font
    cell_title.alignment = center

    # Coordonnées (gauche) dans F..G
    contact = (
        "AB Serve\n"
        "53 route de Rombas | 57140 Woippy | France\n"
        "T +33(0)3 87 58 98 98 | contact@ab-serve.com\n"
        "www.ab-serve.com"
    )
    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)  # F:G
    cell_contact = ws.cell(row=row, column=6, value=contact)
    cell_contact.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    row += 2

    # Bloc métadonnées (labels en gras largeur fixe, valeurs normales)
    # Colonnes fixes: A:B (label gauche), C:D (valeur gauche), E:F (label droite), G:G (valeur droite)
    meta_pairs = [
        ("Date :", data.get('date_creation') or '', "Client :", data.get('client') or ''),
        ("Service :", "Commercial", "", ""),
        ("Interlocuteur :", "", "Service :", ""),
        ("Date de la prestation AB Serve :", data.get('date_envoi') or '', "1ère prestation AB Serve :", "OUI     NON"),
    ]
    for l1, v1, l2, v2 in meta_pairs:
        # Merges
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)  # A:B label
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)  # C:D value
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)  # E:F label
        # Write left label/value
        c_label_left = ws.cell(row=row, column=1, value=l1)
        c_label_left.font = bold_font
        c_label_left.alignment = left
        c_value_left = ws.cell(row=row, column=3, value=v1)
        c_value_left.alignment = left
        # Write right label/value
        c_label_right = ws.cell(row=row, column=5, value=l2)
        # Only set bold font when there is a label to avoid assigning StyleProxy
        if l2:
            c_label_right.font = bold_font
        c_label_right.alignment = left
        c_value_right = ws.cell(row=row, column=7, value=v2)
        c_value_right.alignment = left
        row += 1

    row += 1  # espace

    # Sections par type (mimique PDF)
    first_header_row = None
    # Suivi des longueurs max pour ajuster les largeurs de colonnes
    max_q_len = 0  # Question (col B)
    max_cmt_len = 0  # Commentaire (col F)
    for section_title, items in grouped.items():
        # Ne rien afficher si aucune question pour ce type
        if not items:
            continue
        # Titre de section fond vert clair
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        sec = ws.cell(row=row, column=1, value=section_title)
        sec.font = bold_font
        sec.fill = header_fill
        sec.alignment = left
        row += 1

        # En-têtes du tableau (coller la colonne Question juste après # pour éviter un grand espace)
        # Tableau structuré: # | Question | Type | Obligatoire | Réponse | Commentaire
        headers = ['#', 'Question', 'Type', 'Obligatoire', 'Réponse', 'Commentaire']
        header_cols = [1, 2, 3, 4, 5, 6]  # A..F
        header_row = row
        for h, col_idx in zip(headers, header_cols):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border_all

        row += 1
        # Figer l'en-tête la première fois
        if not ws.freeze_panes:
            ws.freeze_panes = f"A{row}"
        if first_header_row is None:
            first_header_row = header_row

        # Lignes
        for idx, q in enumerate(items, start=1):
            # Col A: index
            c = ws.cell(row=row, column=1, value=idx)
            c.alignment = center
            c.border = border_all
            # Col B: question (wrap)
            q_text = q.get('question') or ''
            c = ws.cell(row=row, column=2, value=q_text)
            c.alignment = left
            c.border = border_all
            if isinstance(q_text, str):
                max_q_len = max(max_q_len, len(q_text))
            # Col C: type (libellé)
            c = ws.cell(row=row, column=3, value=type_map.get(q.get('type_code'), q.get('type')))
            c.alignment = center
            c.border = border_all
            # Col D: obligatoire (Oui/Non)
            c = ws.cell(row=row, column=4, value='Oui' if q.get('obligatoire') else 'Non')
            c.alignment = center
            c.border = border_all
            # Col E: reponse
            c = ws.cell(row=row, column=5, value=q.get('reponse') or '')
            c.alignment = center
            c.border = border_all
            # Col F: commentaire (wrap)
            cmt_text = q.get('commentaire') or ''
            c = ws.cell(row=row, column=6, value=cmt_text)
            c.alignment = left
            c.border = border_all
            if isinstance(cmt_text, str):
                max_cmt_len = max(max_cmt_len, len(cmt_text))
            row += 1

        row += 1  # espace entre sections

    # Auto-filter sur l'ensemble du tableau si disponible
    try:
        if first_header_row is not None and row > first_header_row:
            ws.auto_filter.ref = f"A{first_header_row}:F{row-1}"
    except Exception:
        pass

    # Ajustement dynamique des largeurs en fonction du contenu réel
    try:
        # Heuristique: approx 1 unité de largeur ~ 1 char (Calibri). Limiter à un maximum pour éviter un fichier trop large.
        if max_q_len:
            ws.column_dimensions['B'].width = min(100, max(ws.column_dimensions['B'].width or 38, max_q_len * 0.9))
        if max_cmt_len:
            ws.column_dimensions['F'].width = min(80, max(ws.column_dimensions['F'].width or 18, max_cmt_len * 0.75))
    except Exception:
        pass

    # Retourner la réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="enquete_{enquete.id}.xlsx"'
    wb.save(response)
    return response

def export_enquete_pdf(request, enquete_id):
    # Wrapper pour compatibilité avec les URLs existantes
    return download_enquete_pdf(request, enquete_id)

def export_enquete_excel(request, enquete_id):
    # Wrapper pour compatibilité avec les URLs existantes
    return download_enquete_excel(request, enquete_id)

def repondre_enquete(request, token):
    token_obj = get_object_or_404(EnqueteToken, token=token)
    enquete = token_obj.enquete
    client = token_obj.client
    
    # Get language from token or default to French
    language = token_obj.language
    
    if not token_obj.is_valid():
        return render(request, 'adminlte/sales/sales/enquetes/token_expire.html', {
            'enquete': enquete,
            'language': language
        })

    questions = enquete.questions.all().order_by('id')

    existing_responses = {
        r.question.id: {
            'response': r.reponse,
            'comment': r.commentaire
        } for r in enquete.reponses.filter(client=client)
    }

    last_answered_id = max(existing_responses.keys(), default=0)
    current_question_num = 1
    
    if existing_responses:
        for i, question in enumerate(questions, start=1):
            if question.id not in existing_responses:
                current_question_num = i
                break
        else:
            current_question_num = len(questions)

    if request.method == 'POST':
        if 'submit_final' in request.POST:
            enquete.date_reponse = timezone.now().date()
            enquete.is_completed = True
            enquete.save()
            token_obj.used = True
            token_obj.save()
            
            User = get_user_model()
            # Collect the two relevant societes: client's and creator's
            target_societes = set()
            if enquete.client and enquete.client.societe:
                target_societes.add(enquete.client.societe.pk)
            if enquete.created_by and enquete.created_by.societe:
                target_societes.add(enquete.created_by.societe.pk)
            # Admins matched via ForeignKey societe OR ManyToMany societes
            if target_societes:
                admin_recipients = User.objects.filter(
                    is_RO=True
                ).filter(
                    models.Q(societe__in=target_societes) |
                    models.Q(societes__in=target_societes)
                )
            else:
                admin_recipients = User.objects.none()
            superusers = User.objects.filter(is_superuser=True)
            recipients = (admin_recipients | superusers).distinct()
            
            message = f"L'enquête '{enquete.titre}' a reçu une nouvelle réponse de {client.nom}."
            
            create_and_send_notification(
                message=message,
                users=recipients,
                type='enquete',
                id=enquete.id
            )
            return redirect(f"{reverse('prospection:reponse_ok')}?language={language}")
        
        question_id = request.POST.get('question_id')
        if question_id:
            question = get_object_or_404(Question, id=question_id)
            main_response = request.POST.get('response', '').strip()
            comment = request.POST.get('comment', '').strip()
            
            if main_response:
                Reponse.objects.update_or_create(
                    enquete=enquete,
                    question=question,
                    client=client,
                    defaults={
                        'reponse': main_response,
                        'commentaire': comment if comment else None
                    }
                )
                existing_responses[question.id] = {
                    'response': main_response,
                    'comment': comment if comment else None
                }

    return render(request, 'adminlte/sales/sales/enquetes/repondre.html', {
        'enquete': enquete,
        'client': client,
        'questions': questions,
        'current_question_num': current_question_num,
        'existing_responses': json.dumps(existing_responses),
        'no_required': True,
        'is_completed': enquete.is_completed,
        'token': token,
        'language': language  # Pass the language to the template
    })

@csrf_exempt
def save_response_ajax(request, token):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée'})
    
    try:
        token_obj = get_object_or_404(EnqueteToken, token=token)
        enquete = token_obj.enquete
        client = token_obj.client
        
        if not token_obj.is_valid():
            return JsonResponse({'status': 'error', 'message': 'Token expiré'})
        
        data = json.loads(request.body)
        question_id = data.get('question_id')
        response = data.get('response', '') or ''
        comment = data.get('comment', '') or ''
        
        response = response.strip() if response else ''
        comment = comment.strip() if comment else ''
        
        if not question_id:
            return JsonResponse({'status': 'error', 'message': 'ID de question manquant'})
        
        question = get_object_or_404(Question, id=question_id)

        # Pour les questions ouvertes: ne jamais persister une réponse vide.
        # Si l'utilisateur efface sa réponse (ou passe sans répondre), on supprime la ligne existante.
        if question.type == 'open' and not response:
            Reponse.objects.filter(
                enquete=enquete,
                question=question,
                client=client,
            ).delete()
            return JsonResponse({'status': 'success', 'message': 'Réponse sauvegardée'})

        if response or comment:
            Reponse.objects.update_or_create(
                enquete=enquete,
                question=question,
                client=client,
                defaults={
                    'reponse': response if response else '',
                    'commentaire': comment if comment else None
                }
            )
        
        return JsonResponse({'status': 'success', 'message': 'Réponse sauvegardée'})
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

def reponse_ok(request):
    language = request.GET.get('language', 'fr')  # Récupère la langue depuis les paramètres GET ou utilise 'fr' par défaut
    return render(request, 'adminlte/sales/sales/enquetes/reponse_ok.html', {
        'language': language
    })

@csrf_exempt
@require_POST
def send_email_enquete(request):
    """
    Vue pour envoyer un email d'enquête ou de rappel avec personnalisation
    """
    try:
        enquete_id = request.POST.get('enquete_id')
        action_type = request.POST.get('action_type')  # 'send' ou 'rappel'
        recipient_email = request.POST.get('recipient_email')
        email_subject = request.POST.get('email_subject')
        email_body = request.POST.get('email_body')
        language = request.POST.get('language', 'fr')
        client_type = request.POST.get('client_type', 'froid')
        mission_name = request.POST.get('mission_name', '')
        
        if not all([enquete_id, action_type, recipient_email, email_subject, email_body]):
            return JsonResponse({
                'status': 'error',
                'error': 'Tous les champs requis doivent être remplis.'
            }, status=400)
        
        email_pattern = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
        if not re.match(email_pattern, recipient_email):
            return JsonResponse({
                'status': 'error',
                'error': 'Format d\'email invalide.'
            }, status=400)
        
        try:
            enquete = Enquete.objects.get(id=enquete_id)
        except Enquete.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'error': 'Enquête introuvable.'
            }, status=404)
        
        if client_type == 'chaud' and not mission_name.strip():
            return JsonResponse({
                'status': 'error',
                'error': 'Le nom de la mission est requis pour un client chaud.'
            }, status=400)
        
        client_name = enquete.client.nom if enquete.client else 'Client'
        # societe_id is REQUIRED and overrides any client company for the subject
        chosen_societe_id = (request.POST.get('societe_id') or '').strip()
        if not chosen_societe_id:
            return JsonResponse({
                'status': 'error',
                'error': "Le champ 'Société pour l'objet' est obligatoire."
            }, status=400)
        try:
            s_obj = Societe.objects.get(pk=chosen_societe_id)
        except Societe.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'error': "La société sélectionnée est introuvable."
            }, status=400)
        societe_name = (s_obj.nom or '').strip()
        # Subject format: RE: Enquête de satisfaction <societe> / <client>
        if societe_name:
            final_subject = f"RE: Enquête de satisfaction {societe_name} / {client_name}"
        else:
            final_subject = f"RE: Enquête de satisfaction {client_name}"
        final_body = email_body
        
        if client_type == 'chaud' and mission_name:
            final_body = final_body.replace('{mission}', mission_name)

        # Génération du token et de l'URL
        EnqueteToken.objects.filter(enquete=enquete).delete()
        token_obj = EnqueteToken.objects.create(
            enquete=enquete, 
            client=enquete.client,
            language=language  # Save the selected language with the token
        )
        token = token_obj.token
        # Build HTTPS URL explicitly to avoid mixed http scheme in emails
        url = 'https://' + request.get_host() + reverse('prospection:repondre_enquete', args=[token])
        
        if language == 'fr':
            final_body += f"\n\nVeuillez cliquer sur le lien suivant pour accéder à l'enquête :\n{url}\n\nMerci pour votre temps."
        else:
            final_body += f"\n\nPlease click on the following link to access the survey:\n{url}\n\nThank you for your time."
        
        try: 
            send_mail(
                subject=final_subject,
                message=final_body,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[recipient_email],
                fail_silently=False,
            )
            
            now = timezone.now().date()
            # Save recipient email for future reminders/defaults
            try:
                enquete.email = recipient_email
            except Exception:
                pass

            if action_type == 'send':
                enquete.date_envoi = now
                success_message = 'Enquête envoyée avec succès !'
                
            elif action_type == 'rappel':
                enquete.date_rappel = now
                success_message = 'Rappel envoyé avec succès !'
            
            enquete.save()
            
            return JsonResponse({
                'status': 'success',
                'message': success_message
            })
            
        except Exception as email_error:
            return JsonResponse({
                'status': 'error',
                'error': f'Erreur lors de l\'envoi de l\'email : {str(email_error)}'
            }, status=500)
    
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'error': f'Une erreur inattendue s\'est produite : {str(e)}'
        }, status=500)

def get_question_stats(question, enq_qs):
    """Get statistics for a specific question across all surveys"""
    responses = Reponse.objects.filter(
        question=question, enquete__in=enq_qs, reponse__isnull=False
    ).values_list('reponse', flat=True)

    # "safe" keys for templates (no spaces/accents)
    response_counts_safe = {
        'tres_satisfait': 0,
        'satisfait': 0,
        'peu_satisfait': 0,
        'non_satisfait': 0
    }
    for r in responses:
        x = str(r).strip().lower()
        if 'très' in x and 'satisfait' in x:
            response_counts_safe['tres_satisfait'] += 1
        elif 'peu' in x and 'satisfait' in x:
            response_counts_safe['peu_satisfait'] += 1
        elif ('non' in x or 'pas' in x) and 'satisfait' in x:
            response_counts_safe['non_satisfait'] += 1
        elif 'satisfait' in x:
            response_counts_safe['satisfait'] += 1

    total = sum(response_counts_safe.values())
    if not total:
        return None

    # Backward-compatible keys for existing JS on enquetes.html
    response_counts = {
        'très satisfait': response_counts_safe['tres_satisfait'],
        'satisfait': response_counts_safe['satisfait'],
        'peu satisfait': response_counts_safe['peu_satisfait'],
        'non satisfait': response_counts_safe['non_satisfait'],
    }

    response_percentages_safe = {k: round(v * 100.0 / total, 1) for k, v in response_counts_safe.items()}
    response_percentages = {
        'très satisfait': response_percentages_safe['tres_satisfait'],
        'satisfait': response_percentages_safe['satisfait'],
        'peu satisfait': response_percentages_safe['peu_satisfait'],
        'non satisfait': response_percentages_safe['non_satisfait'],
    }

    return {
        'question_id': question.id,
        'question_text': question.question_fr or f"Question {question.id}",
        'type': question.type,
        'labels': {
            'tres_satisfait': 'Très satisfait',
            'satisfait': 'Satisfait',
            'peu_satisfait': 'Peu satisfait',
            'non_satisfait': 'Pas du tout satisfait',
        },
        # For JS (existing)
        'response_counts': response_counts,
        'response_percentages': response_percentages,
        # For PDF templates (safe)
        'response_counts_safe': response_counts_safe,
        'response_percentages_safe': response_percentages_safe,
        'total_responses': total
    }


def _logo_context_for_pdf():
    # Resolve logo absolute path and convert to file URI (xhtml2pdf friendly)
    _logo_abs = (finders.find('dist/img/abserveLogo.png') or finders.find('dist/img/logo.png'))
    _logo_uri = None
    _logo_data_uri = None
    if _logo_abs:
        _logo_uri = 'file:///' + _logo_abs.replace('\\', '/')
        try:
            with open(_logo_abs, 'rb') as _f:
                _b64 = base64.b64encode(_f.read()).decode('ascii')
                ext = os.path.splitext(_logo_abs)[1].lower().lstrip('.')
                mime = 'png' if ext == 'png' else 'jpeg'
                _logo_data_uri = f"data:image/{mime};base64,{_b64}"
        except Exception:
            _logo_data_uri = None
    return {
        'logo_path': _logo_abs,
        'logo_uri': _logo_uri,
        'logo_data_uri': _logo_data_uri,
    }

@login_required
def enquetes_analytics(request):
    try:
        q = _filters_q(request)
        enq_qs = _base_queryset(request).filter(q)

        total_enq = enq_qs.count()
        responded_enq = enq_qs.filter(reponses__isnull=False).distinct().count()
        not_responded_enq = max(total_enq - responded_enq, 0)

        response_rate_pct = (responded_enq * 100.0 / total_enq) if total_enq else 0.0

        # Note moyenne
        notes_qs = Reponse.objects.filter(enquete__in=enq_qs, question__type='note') \
                                  .annotate(note_int=Cast('reponse', IntegerField()))
        average_note = notes_qs.aggregate(avg=Avg('note_int'))['avg']

        # Sentiment commentaires
        # Important: compter les commentaires vides comme neutres pour éviter de biaiser l'analyse.
        # On analyse d'abord `commentaire`, et pour les questions ouvertes on peut aussi analyser `reponse`.
        comments_qs = Reponse.objects.filter(enquete__in=enq_qs)
        analyzer = MultilingualSentimentAnalyzer()
        pos = neu = neg = 0
        total_comments = 0
        avg_accum = 0.0
        for r in comments_qs.iterator():
            txt = (r.commentaire or '').strip()
            if not txt and getattr(r, 'question', None) and getattr(r.question, 'type', None) == 'open':
                txt = (r.reponse or '').strip()

            # Commentaire vide => neutre (mais compté dans le total)
            if not txt:
                total_comments += 1
                neu += 1
                continue
            analysis = analyzer.analyze(txt) or {}
            sc = analysis.get('score', 0.0) or 0.0
            lab = (analysis.get('label') or '').lower()
            total_comments += 1
            avg_accum += sc
            if lab == 'positive':
                pos += 1
            elif lab == 'negative':
                neg += 1
            elif lab == 'neutral':
                neu += 1
            else:
                # Fallback (anciens seuils) si label indisponible
                if sc > 0.1:
                    pos += 1
                elif sc < -0.1:
                    neg += 1
                else:
                    neu += 1

        sentiment = {
            'pos': pos, 'neu': neu, 'neg': neg, 'total': total_comments,
            'pos_pct': (pos * 100.0 / total_comments) if total_comments else 0.0,
            'neu_pct': (neu * 100.0 / total_comments) if total_comments else 0.0,
            'neg_pct': (neg * 100.0 / total_comments) if total_comments else 0.0,
            'avg': (avg_accum / total_comments) if total_comments else 0.0
        }

        # Compteurs comparables
        sent_count = enq_qs.filter(date_envoi__isnull=False).count()
        reminder_count = enq_qs.filter(date_rappel__isnull=False).count()
        return_count = enq_qs.filter(date_reponse__isnull=False, is_completed=True).count()

        # Questions fermées
        closed_questions = Question.objects.filter(type='closed', reponses__enquete__in=enq_qs).distinct()
        question_stats = []
        for qobj in closed_questions:
            st = get_question_stats(qobj, enq_qs)
            if st and st.get('total_responses', 0) > 0:
                question_stats.append(st)

        return JsonResponse({
            'response_stats': {
                'total': total_enq,
                'responded': responded_enq,
                'not_responded': not_responded_enq,
                'response_rate_pct': response_rate_pct,
            },
            'average_note': average_note,
            'sentiment': sentiment,
            'question_stats': question_stats,
            'enquete_counts': {
                'sent': sent_count,
                'reminders': reminder_count,
                'returns': return_count,
            }
        })
    except Exception as e:
        logger.exception('Error computing enquetes analytics')
        return JsonResponse({'error': str(e)}, status=500)


def _min_mm_segments(segments, total_mm, min_mm=3.0):
    """Ensure non-zero segments are visible in PDF by enforcing a minimum width.
    segments: dict(name -> mm). Returns new dict with sum exactly = total_mm."""
    seg = {k: float(v or 0.0) for k, v in (segments or {}).items()}
    non_zero = [k for k, v in seg.items() if v > 0.0]
    if not non_zero:
        return seg

    for k in non_zero:
        if seg[k] < min_mm:
            seg[k] = min_mm
    s = sum(seg.values())
    if s <= 0.0:
        return seg
    if s > total_mm:
        scale = total_mm / s
        for k in seg:
            seg[k] = round(seg[k] * scale, 2)
    drift = round(total_mm - sum(seg.values()), 2)
    if abs(drift) > 0.01:
        kmax = max(seg.keys(), key=lambda k: seg[k])
        seg[kmax] = max(0.0, round(seg[kmax] + drift, 2))
    return seg


def _png_data_uri_from_drawing(drawing):
    try:
        from reportlab.graphics import renderPM
        png_bytes = renderPM.drawToString(drawing, fmt='PNG')
        if not png_bytes:
            return None
        import base64
        b64 = base64.b64encode(png_bytes).decode('ascii')
        return f'data:image/png;base64,{b64}'
    except Exception:
        return None


def _make_pie_chart_data_uri(labels, values, colors_hex, width=520, height=220):
    try:
        from reportlab.graphics.shapes import Drawing, String, Rect
        from reportlab.graphics.charts.piecharts import Pie
        from reportlab.lib import colors

        d = Drawing(width, height)
        pie = Pie()
        pie.x = 20
        pie.y = 10
        pie.width = 200
        pie.height = 200
        pie.data = [max(0.0, float(v or 0.0)) for v in (values or [])]
        pie.labels = [str(x) for x in (labels or [])]
        pie.slices.strokeWidth = 0.0
        for i, c in enumerate(colors_hex or []):
            try:
                pie.slices[i].fillColor = colors.HexColor(c)
            except Exception:
                pass
        pie.sideLabels = True
        pie.simpleLabels = False
        d.add(pie)

        lx = 250
        ly = height - 30
        for i, lab in enumerate(pie.labels):
            try:
                col = colors.HexColor((colors_hex or [])[i])
            except Exception:
                col = colors.black
            d.add(Rect(lx, ly - 10 - i * 18, 10, 10, fillColor=col, strokeColor=colors.HexColor('#94a3b8'), strokeWidth=0.5))
            d.add(String(lx + 16, ly - i * 18, lab, fontSize=10, fillColor=colors.HexColor('#334155')))

        return _png_data_uri_from_drawing(d)
    except Exception:
        return None


def _make_hbar_chart_data_uri(labels, values, colors_hex, width=520, height=220, max_value=None):
    try:
        from reportlab.graphics.shapes import Drawing, Rect, String
        from reportlab.lib import colors

        d = Drawing(width, height)
        pad_l = 140
        pad_r = 20
        pad_t = 20
        pad_b = 20
        bar_h = 18
        gap = 10
        plot_w = max(10, width - pad_l - pad_r)

        vals = [max(0.0, float(v or 0.0)) for v in (values or [])]
        m = float(max_value) if max_value is not None else (max(vals) if vals else 1.0)
        if m <= 0:
            m = 1.0

        for i, lab in enumerate(labels or []):
            y = height - pad_t - (i + 1) * (bar_h + gap)
            if y < pad_b:
                break
            d.add(String(10, y + 4, str(lab), fontSize=10, fillColor=colors.HexColor('#334155')))
            d.add(Rect(pad_l, y, plot_w, bar_h, fillColor=colors.white, strokeColor=colors.HexColor('#cbd5e1'), strokeWidth=0.8))
            w = plot_w * (vals[i] / m) if i < len(vals) else 0
            try:
                col = colors.HexColor((colors_hex or [])[i])
            except Exception:
                col = colors.HexColor('#42A5F5')
            d.add(Rect(pad_l, y, max(0, w), bar_h, fillColor=col, strokeColor=None))
            d.add(String(pad_l + plot_w + 6, y + 4, str(int(vals[i])) if vals[i].is_integer() else f'{vals[i]:.1f}', fontSize=10, fillColor=colors.HexColor('#334155')))

        return _png_data_uri_from_drawing(d)
    except Exception:
        return None


def _min_pct_segments(segments, min_pct=2.0):
    seg = {k: float(v or 0.0) for k, v in (segments or {}).items()}
    non_zero = [k for k, v in seg.items() if v > 0.0]
    if not non_zero:
        return seg
    for k in non_zero:
        if seg[k] < min_pct:
            seg[k] = min_pct
    s = sum(seg.values())
    if s <= 0.0:
        return seg
    if s != 100.0:
        scale = 100.0 / s
        for k in seg:
            seg[k] = round(seg[k] * scale, 2)
    drift = round(100.0 - sum(seg.values()), 2)
    if abs(drift) > 0.01:
        kmax = max(seg.keys(), key=lambda k: seg[k])
        seg[kmax] = max(0.0, round(seg[kmax] + drift, 2))
    return seg


@login_required
def download_enquetes_analytics_global_pdf(request):
    """Génère et télécharge le PDF de la section Analytique globale (filtres = querystring)."""
    q = _filters_q(request)
    enq_qs = _base_queryset(request).filter(q)

    total_enq = enq_qs.count()
    responded_enq = enq_qs.filter(reponses__isnull=False).distinct().count()
    not_responded_enq = max(total_enq - responded_enq, 0)
    response_rate_pct = (responded_enq * 100.0 / total_enq) if total_enq else 0.0
    responded_pct = round(response_rate_pct, 1)
    not_responded_pct = round(max(0.0, 100.0 - responded_pct), 1)

    # xhtml2pdf rendering is more reliable with absolute widths (mm) than % + floats.
    BAR_W_MM = 600.0

    resp_mm = {
        'responded': round(BAR_W_MM * responded_pct / 100.0, 2),
        'not_responded': round(BAR_W_MM * not_responded_pct / 100.0, 2),
    }
    resp_mm = _min_mm_segments(resp_mm, BAR_W_MM, min_mm=3.0)
    responded_mm = resp_mm.get('responded', 0.0)
    not_responded_mm = resp_mm.get('not_responded', 0.0)

    notes_qs = Reponse.objects.filter(enquete__in=enq_qs, question__type='note') \
                              .annotate(note_int=Cast('reponse', IntegerField()))
    average_note = notes_qs.aggregate(avg=Avg('note_int'))['avg']

    comments_qs = Reponse.objects.filter(enquete__in=enq_qs)
    analyzer = MultilingualSentimentAnalyzer()
    pos = neu = neg = 0
    total_comments = 0
    avg_accum = 0.0
    for r in comments_qs.iterator():
        txt = (r.commentaire or '').strip()
        if not txt and getattr(r, 'question', None) and getattr(r.question, 'type', None) == 'open':
            txt = (r.reponse or '').strip()

        # Commentaire vide => neutre (mais compté dans le total)
        if not txt:
            total_comments += 1
            neu += 1
            continue
        analysis = analyzer.analyze(txt) or {}
        sc = analysis.get('score', 0.0) or 0.0
        lab = (analysis.get('label') or '').lower()
        total_comments += 1
        avg_accum += sc
        if lab == 'positive':
            pos += 1
        elif lab == 'negative':
            neg += 1
        elif lab == 'neutral':
            neu += 1
        else:
            # Fallback (anciens seuils) si label indisponible
            if sc > 0.1:
                pos += 1
            elif sc < -0.1:
                neg += 1
            else:
                neu += 1

    sentiment = {
        'pos': pos,
        'neu': neu,
        'neg': neg,
        'total': total_comments,
        'pos_pct': (pos * 100.0 / total_comments) if total_comments else 0.0,
        'neu_pct': (neu * 100.0 / total_comments) if total_comments else 0.0,
        'neg_pct': (neg * 100.0 / total_comments) if total_comments else 0.0,
        'avg': (avg_accum / total_comments) if total_comments else 0.0,
    }

    sent_count = enq_qs.filter(date_envoi__isnull=False).count()
    reminder_count = enq_qs.filter(date_rappel__isnull=False).count()
    return_count = enq_qs.filter(date_reponse__isnull=False, is_completed=True).count()

    total_counts = sent_count + reminder_count + return_count
    comparatif_pct = {
        'sent': round((sent_count * 100.0 / total_counts), 1) if total_counts else 0.0,
        'reminders': round((reminder_count * 100.0 / total_counts), 1) if total_counts else 0.0,
        'returns': round((return_count * 100.0 / total_counts), 1) if total_counts else 0.0,
    }

    comparatif_mm = {
        'sent': round(BAR_W_MM * comparatif_pct['sent'] / 100.0, 2),
        'reminders': round(BAR_W_MM * comparatif_pct['reminders'] / 100.0, 2),
        'returns': round(BAR_W_MM * comparatif_pct['returns'] / 100.0, 2),
    }
    comparatif_mm = _min_mm_segments(comparatif_mm, BAR_W_MM, min_mm=3.0)

    sentiment_mm = {
        'pos': round(BAR_W_MM * float(sentiment.get('pos_pct') or 0.0) / 100.0, 2),
        'neu': round(BAR_W_MM * float(sentiment.get('neu_pct') or 0.0) / 100.0, 2),
        'neg': round(BAR_W_MM * float(sentiment.get('neg_pct') or 0.0) / 100.0, 2),
    }
    sentiment_mm = _min_mm_segments(sentiment_mm, BAR_W_MM, min_mm=3.0)

    sentiment_mm_rest = {
        'pos': max(0.0, round(BAR_W_MM - sentiment_mm['pos'], 2)),
        'neu': max(0.0, round(BAR_W_MM - sentiment_mm['neu'], 2)),
        'neg': max(0.0, round(BAR_W_MM - sentiment_mm['neg'], 2)),
    }

    charts = {
        'response_rate': None,
        'comparatif': None,
        'sentiment': None,
    }
    try:
        charts['response_rate'] = _make_pie_chart_data_uri(
            labels=['Répondu', 'Non répondu'],
            values=[responded_enq, not_responded_enq],
            colors_hex=['#66BB6A', '#EF5350'],
            width=520,
            height=220,
        )
        charts['comparatif'] = _make_hbar_chart_data_uri(
            labels=['Envois', 'Rappels', 'Retours'],
            values=[sent_count, reminder_count, return_count],
            colors_hex=['#42A5F5', '#FFB74D', '#66BB6A'],
            width=520,
            height=220,
            max_value=max(sent_count, reminder_count, return_count, 1),
        )
        charts['sentiment'] = _make_pie_chart_data_uri(
            labels=['Positif', 'Neutre', 'Négatif'],
            values=[sentiment['pos'], sentiment['neu'], sentiment['neg']],
            colors_hex=['#66BB6A', '#9E9E9E', '#EF5350'],
            width=520,
            height=220,
        )
    except Exception:
        charts = {'response_rate': None, 'comparatif': None, 'sentiment': None}

    context = {
        'now': timezone.now().strftime('%d/%m/%Y %H:%M'),
        'response_stats': {
            'total': total_enq,
            'responded': responded_enq,
            'not_responded': not_responded_enq,
            'response_rate_pct': responded_pct,
            'responded_pct': responded_pct,
            'not_responded_pct': not_responded_pct,
            'responded_mm': responded_mm,
            'not_responded_mm': not_responded_mm,
        },
        'average_note': average_note,
        'sentiment': sentiment,
        'enquete_counts': {
            'sent': sent_count,
            'reminders': reminder_count,
            'returns': return_count,
        },
        'comparatif_pct': comparatif_pct,
        'comparatif_mm': comparatif_mm,
        'sentiment_mm': sentiment_mm,
        'sentiment_mm_rest': sentiment_mm_rest,
        'charts': charts,
    }
    context.update(_logo_context_for_pdf())

    html = render_to_string('adminlte/sales/sales/enquetes/analytics_global_pdf.html', context)

    if (request.GET.get('format') or '').lower() == 'html':
        return HttpResponse(html, content_type='text/html; charset=utf-8')

    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode('utf-8')), result, encoding='utf-8')
    if pdf.err:
        return HttpResponse('Une erreur est survenue lors de la génération du PDF', status=500)
    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="analytique_globale_enquetes.pdf"'
    return response


@login_required
def download_enquetes_analytics_questions_pdf(request):
    """Génère et télécharge le PDF de la section Réponses détaillées par question (filtres = querystring)."""
    q = _filters_q(request)
    enq_qs = _base_queryset(request).filter(q)

    closed_questions = Question.objects.filter(type='closed', reponses__enquete__in=enq_qs).distinct()
    question_stats = []
    BAR_W_MM = 600.0
    for qobj in closed_questions:
        st = get_question_stats(qobj, enq_qs)
        if st and st.get('total_responses', 0) > 0:
            pct = st.get('response_percentages_safe') or {}
            mm = {
                'tres_satisfait': round(BAR_W_MM * float(pct.get('tres_satisfait') or 0.0) / 100.0, 2),
                'satisfait': round(BAR_W_MM * float(pct.get('satisfait') or 0.0) / 100.0, 2),
                'peu_satisfait': round(BAR_W_MM * float(pct.get('peu_satisfait') or 0.0) / 100.0, 2),
                'non_satisfait': round(BAR_W_MM * float(pct.get('non_satisfait') or 0.0) / 100.0, 2),
            }
            mm = _min_mm_segments(mm, BAR_W_MM, min_mm=3.0)
            st['bar_w_mm'] = BAR_W_MM
            st['bar_mm'] = mm
            question_stats.append(st)

    context = {
        'now': timezone.now().strftime('%d/%m/%Y %H:%M'),
        'question_stats': question_stats,
    }
    context.update(_logo_context_for_pdf())

    html = render_to_string('adminlte/sales/sales/enquetes/analytics_questions_pdf.html', context)

    if (request.GET.get('format') or '').lower() == 'html':
        return HttpResponse(html, content_type='text/html; charset=utf-8')

    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode('utf-8')), result, encoding='utf-8')
    if pdf.err:
        return HttpResponse('Une erreur est survenue lors de la génération du PDF', status=500)
    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reponses_detaillees_questions.pdf"'
    return response

@login_required
def get_client_details(request, client_id):
    """Fetch detailed information for a specific client including facturation details"""
    headers = {
        'Authorization': settings.SAGE_API_TOKEN,
        'Accept': 'application/json'
    }
    try:
        societe = request.GET.get('societe')
        base_candidates = []
        if societe:
            societe_id = Societe.objects.filter(nom=societe).first()
            base_candidates = [get_base_url(societe_id.id)]
        else:
            return JsonResponse({"error": "Société non reconnue"}, status=400)

        client_response = None
        for base in base_candidates:
            fact_url = f"{settings.SAGE_API_HOST}/WebServices100/{base}/TiersService/rest/Clients/{client_id}"
            try:
                r = requests.get(fact_url, headers=headers, timeout=10)
            except Exception:
                r = None
            if r is not None and r.status_code == 200:
                client_response = r
                break
        if client_response is None:
            return JsonResponse({"error": "Erreur lors de la récupération des informations du client"}, status=400)

        # Parsing JSON tolérant (correction Expecting value: line 1 column 1)
        try:
            client_data = client_response.json() if client_response.content else {}
        except ValueError:
            client_data = {}

        # Fallback local: chercher l'entreprise
        ent = None
        cid_raw = str(client_id or '').strip()
        # 1) Essayer par PK si client_id est un entier
        try:
            ent_pk = int(cid_raw)
            ent = Entreprise.objects.filter(pk=ent_pk, societe_id=societe_id.id).first()
        except (TypeError, ValueError):
            ent = None
        # 2) Variantes num_compte: exact, iexact, sans zéros de tête, icontains
        if not ent and cid_raw:
            cid_nz = cid_raw.lstrip('0') or cid_raw
            ent = (
                Entreprise.objects.filter(num_compte__iexact=cid_raw, societe_id=societe_id.id).first() or
                Entreprise.objects.filter(num_compte__iexact=cid_nz, societe_id=societe_id.id).first() or
                Entreprise.objects.filter(num_compte__icontains=cid_raw, societe_id=societe_id.id).first()
            )
        # 3) Par nom (fallback)
        if not ent and cid_raw:
            ent = Entreprise.objects.filter(nom__iexact=cid_raw, societe_id=societe_id.id).first()

        def coalesce(*vals):
            for v in vals:
                if v is None:
                    continue
                s = str(v).strip()
                if s:
                    return s
            return ""

        # Extraire champs SAGE avec clés alternatives
        numero_tiers = coalesce(client_data.get("NumeroTiers"), getattr(ent, 'num_compte', None))
        intitule = coalesce(client_data.get("Intitule"), getattr(ent, 'nom', None))

        # Adresse: composer si besoin
        adr = coalesce(
            client_data.get("Adresse"),
            " ".join(filter(None, [
                str(client_data.get("AdresseLigne1", "")),
                str(client_data.get("AdresseLigne2", "")),
                str(client_data.get("CodePostal", "")),
                str(client_data.get("Ville", ""))
            ])).strip(),
            getattr(ent, 'adresse', None)
        )

        email = coalesce(client_data.get("Email"), getattr(ent, 'email', None))
        tel = coalesce(
            client_data.get("Telephone"), client_data.get("Telephone1"), client_data.get("Telephone2"),
            getattr(ent, 'telephone', None)
        )

        result = {
            "NumeroTiers": numero_tiers,
            "Intitule": intitule,
            "fact_adresse": adr,
            "fact_email": email,
            "fact_tel": tel,
        }
        # Ajouter un warning si tout est vide pour informer le front
        if not any([numero_tiers, intitule, adr, email, tel]):
            result["warning"] = "Aucune donnée trouvée pour ce client (SAGE/local)."
        return JsonResponse(result)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@login_required
def get_active_customers(request):
    try:
        search_term = request.GET.get("q", "").strip().lower()
        if len(search_term) < 2:
            return JsonResponse({"results": []})

        # Build list of societes to search (id, name, base)
        societes_to_search = []
        show_societe = bool(getattr(request.user, 'is_RO', False) or getattr(request.user, 'is_superuser', False))
        if getattr(request.user, 'is_superuser', False):
            societes_to_search = [(s.id, s.nom, get_base_url(s.id)) 
                                for s in Societe.objects.all() 
                                if get_base_url(s.id)]
        elif getattr(request.user, 'is_RO', False):
            societes_to_search = [(s.id, s.nom, get_base_url(s.id)) 
                                for s in request.user.societes.all() 
                                if get_base_url(s.id)]
        else:
            if not request.user.societe_id:
                return JsonResponse({"error": "Société non reconnue"}, status=400)
            base = get_base_url(request.user.societe_id)
            if not base:
                return JsonResponse({"error": "Base non configurée pour cette société"}, status=400)
            societes_to_search = [(request.user.societe_id, 
                                 getattr(request.user.societe, 'nom', ''), 
                                 base)]

        # 1. Recherche dans la base de données locale
        local_results = []
        societe_ids = [sid for sid, _, _ in societes_to_search]
        local_filters = (Q(nom__icontains=search_term) | 
                        Q(num_compte__icontains=search_term) |
                        Q(num_compte__istartswith=search_term))
        
        clients_local = Entreprise.objects.filter(
            local_filters, 
            is_CLT=True,
            societe_id__in=societe_ids
        ).select_related('societe')[:20]  # Limiter les résultats

        for client in clients_local:
            txt = f"{client.num_compte} - {client.nom}" if client.num_compte else client.nom
            societe_nom = getattr(client.societe, 'nom', '')
            if show_societe and societe_nom:
                txt = f"{txt} - {societe_nom}"
            
            local_results.append({
                'id': client.num_compte or str(client.id),
                'text': txt,
                'nom': client.nom,
                'email': client.email,
                'telephone': client.telephone,
                'societe': societe_nom,
                'source': 'local'
            })

        # 2. Recherche dans Sage
        sage_results = []
        seen_sage_ids = set()
        
        for sid, sname, base in societes_to_search:
            try:
                url = f"{settings.SAGE_API_HOST}/WebServices100/{base}/TiersService/rest/Clients"
                headers = {'Authorization': settings.SAGE_API_TOKEN, 'Accept': 'application/json'}
                
                response = requests.get(url, headers=headers, timeout=10)
                if response.status_code != 200:
                    continue
                    
                clients = response.json() or []
                for client in clients:
                    if not isinstance(client, dict):
                        continue
                        
                    numero_tiers = str(client.get("NumeroTiers", "")).lower()
                    intitule = str(client.get("Intitule", "")).lower()
                    
                    if (search_term in numero_tiers or search_term in intitule):
                        sage_id = f"{sid}:{client.get('NumeroTiers')}"
                        if sage_id in seen_sage_ids:
                            continue
                            
                        seen_sage_ids.add(sage_id)
                        txt = f"{client.get('NumeroTiers')} - {client.get('Intitule')}"
                        if show_societe and sname:
                            txt = f"{txt} - {sname}"
                            
                        sage_results.append({
                            'id': client.get("NumeroTiers"),
                            'text': txt,
                            'nom': client.get("Intitule"),
                            'email': client.get("Email"),
                            'societe': sname,
                            'telephone': client.get("Telephone1") or client.get("Telephone2"),
                            'NumeroTiers': client.get("NumeroTiers"),
                            'Intitule': client.get("Intitule"),
                            'Email': client.get("Email"),
                            'Telephone': client.get("Telephone1") or client.get("Telephone2"),
                            'source': 'sage'
                        })
                        
            except Exception as e:
                logger.error(f"Erreur API Sage ({sname}): {str(e)}")
                continue

        # 3. Fusionner et dédupliquer les résultats
        seen = set()
        final_results = []
        
        # Ajouter d'abord les résultats locaux
        for item in local_results:
            key = item.get('id')
            if key not in seen:
                seen.add(key)
                final_results.append(item)
                
        # Puis les résultats Sage qui ne sont pas déjà présents
        for item in sage_results:
            key = item.get('id')
            if key not in seen:
                seen.add(key)
                final_results.append(item)

        # Trier par nom
        final_results.sort(key=lambda x: x.get('nom', '').lower())
        
        # Limiter à 20 résultats au total
        final_results = final_results[:20]

        return JsonResponse({
            "results": final_results,
        })

    except Exception as e:
        logger.error(f"Erreur inattendue dans get_active_customers: {str(e)}", exc_info=True)
        return JsonResponse({"error": "Une erreur est survenue lors de la recherche des clients."}, status=500)


# ----------------- Importation Enquêtes -----------------

NORM_TXT = lambda s: " ".join(
    "".join(c for c in unicodedata.normalize("NFD", str(s or "").strip()) if unicodedata.category(c) != "Mn"
).split())
LOWER = lambda s: NORM_TXT(s).lower()

def PARSE_DATE(v):
    if not v or not str(v).strip():
        return None
    parsed = dateparser.parse(str(v).strip())
    return parsed.date() if parsed else None

def SAFE_EMAIL(v):
    v = (v or "").strip().lower()
    try:
        validate_email(v); return v
    except ValidationError:
        return ""

SAT_MAP = {"pas satisfait":"1","peu satisfait":"2","satisfait":"3","tres satisfait":"4","très satisfait":"4"}
YNM = {"oui":"oui","non":"non","peut etre":"peut_etre","peut-être":"peut_etre"}

SAGE_AUTH_HEADER = {"Authorization": settings.SAGE_API_TOKEN, "Accept": "application/json"}

def fetch_client_from_sage(societe_id: int, client_id: str) -> dict | None:
    base = get_base_url(societe_id)
    if not base:
        return None
    url = f"{settings.SAGE_API_HOST}/WebServices100/{base}/TiersService/rest/Clients/{client_id}"
    try:
        r = requests.get(url, headers=SAGE_AUTH_HEADER, timeout=20)
        if r.status_code != 200:
            return None
        d = r.json() or {}
        return {
            "NumeroTiers": d.get("NumeroTiers",""),
            "Intitule": d.get("Intitule",""),
            "Adresse": d.get("Adresse",""),
            "Email": d.get("Email",""),
            "Telephone": d.get("Telephone","")
        }
    except Exception:
        return None


# =========================
# Serializers
# =========================

class FetchSageSerializer(serializers.Serializer):
    societe_id = serializers.IntegerField()
    client_id  = serializers.CharField()


class ImportSatisfactionSerializer(serializers.Serializer):
    file = serializers.FileField()
    societe_id = serializers.IntegerField(required=False)
    societe_nom = serializers.CharField(required=False, allow_blank=True)


# =========================
# Vue 1 : Fetch client Sage
# =========================

class FetchClientFromSageView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        ser = FetchSageSerializer(data=request.query_params)
        ser.is_valid(raise_exception=True)
        societe_id = ser.validated_data["societe_id"]
        client_id  = ser.validated_data["client_id"]

        data = fetch_client_from_sage(societe_id, client_id)
        if not data:
            return Response({"found": False, "data": {}}, status=404)
        return Response({"found": True, "data": data}, status=200)


# =========================
# Fonctions utilitaires pour l'importation
# =========================

def parse_date_safe(date_str, default=None):
    """Parse une date de manière sécurisée avec gestion des erreurs"""
    if not date_str or pd.isna(date_str):
        return default
    try:
        return pd.to_datetime(date_str, errors='coerce').date()
    except (ValueError, TypeError):
        return default

def get_safe_value(row, cmap, *names, default=""):
    """Récupère une valeur de manière sécurisée avec une valeur par défaut"""
    if not hasattr(row, 'get'):
        return default
    val = next((row[cmap[LOWER(n)]] for n in names if LOWER(n) in cmap), None)
    return default if val is None or (isinstance(val, (str, float)) and not str(val).strip()) else val

def pick_flag(row_dict, keys):
    """Vérifie si une des clés est présente et a une valeur positive"""
    for k in keys:
        v = row_dict.get(k, "")
        if LOWER(str(v)) in ("1", "x", "true", "vrai", "ok", "oui"):
            return k
    return None

# =========================
# Vue 2 : Import Excel Satisfaction
# =========================
def _has_value(x):
    if x is None or pd.isna(x):
        return False
    s = str(x).strip().lower()
    return s not in ("", "nan", "n/a", "na", "-", "null", "none", "non renseigné")

# views.py
class ImportSatisfactionExcelView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    SHEETS_SEND_ALIASES = ["Données envoi", "Donnees envoi", "Donnees d'envoi", "Donnees_envoi"]
    SHEETS_REP_ALIASES  = ["Réponses clients", "Reponses clients", "Reponses_clients", "Reponses"]

    def get_or_create_enterprise(self, societe, num_compte, nom_cli, email_cli, excel_row, skipped):
        """Gère la création ou la récupération d'une entreprise"""
        ent = None
        if num_compte:
            ent = Entreprise.objects.filter(societe=societe, num_compte=num_compte).first()
            if ent:
                return ent, False
        
        if not ent and email_cli:
            ent = Entreprise.objects.filter(societe=societe, email__iexact=email_cli).first()
        
        if not ent and nom_cli:
            qs = Entreprise.objects.filter(societe=societe, nom__iexact=nom_cli)
            ent = qs.first() if qs.count() == 1 else None

        created = False
        if not ent:
            sage = fetch_client_from_sage(societe.id, num_compte or nom_cli) or {}
            ent = Entreprise.objects.create(
                num_compte = num_compte or sage.get("NumeroTiers") or None,
                nom = nom_cli or sage.get("Intitule") or "Inconnu",
                email = email_cli or SAFE_EMAIL(sage.get("Email")),
                adresse = sage.get("Adresse", "") or "",
                telephone = sage.get("Telephone", "") or "",
                secteur_activite = "",
                societe = societe,
                is_CLT=True, is_Prospect=False, is_Concurent=False
            )
            created = True
        
        return ent, created

    def get_or_create_enquete(self, ent, date_envoi, date_rappel, date_rappel2, excel_row, skipped):
        """Gère la création ou la récupération d'une enquête"""
        enq = Enquete.objects.filter(
            entreprise=ent,
            date_envoi=date_envoi
        ).order_by('-date_creation').first()

        if enq:
            return enq, False

        enq = Enquete.objects.create(
            entreprise=ent,
            date_envoi=date_envoi,
            date_rappel=date_rappel,
            date_rappel2=date_rappel2,
            statut='envoyee',
            type_enquete='satisfaction',
            token=Enquete.generate_token()
        )
        return enq, True

    def process_response(self, enq, question_text, reponse_value, excel_row, skipped):
        """Traitement d'une réponse à une question"""
        if not question_text:
            return None
            
        # Récupération du commentaire de manière sécurisée
        comment = excel_row.get("Comment")
        comment = str(comment)[:500] if _has_value(comment) else ""

        # Recherche de la question par son texte
        question = Question.objects.filter(
            Q(texte=question_text) |
            Q(texte__iexact=question_text) |
            Q(texte__icontains=question_text[:15])
        ).first()

        if not question:
            return None

        # Pour les questions de type "open" (commentaire libre)
        if question.type == 'open':
            # Pour les questions ouvertes, on utilise le commentaire comme réponse
            reponse_value = comment if _has_value(comment) else ""
            
            reponse, created = Reponse.objects.update_or_create(
                enquete=enq,
                question=question,
                defaults={
                    'reponse': reponse_value,  # Le commentaire est la réponse pour les questions ouvertes
                    'commentaire': None,  # On ne met rien dans le commentaire
                    'client': enq.entreprise
                }
            )
            return reponse
        else:
            # Pour les autres types de questions
            reponse_str = str(reponse_value)[:500] if _has_value(reponse_value) else ""
            
            # Création ou mise à jour de la réponse dans tous les cas
            reponse, created = Reponse.objects.update_or_create(
                enquete=enq,
                question=question,
                defaults={
                    'reponse': reponse_str,
                    'commentaire': comment if _has_value(comment) else "",
                    'client': enq.entreprise
                }
            )
            return reponse

    def post(self, request):
        ser = ImportSatisfactionSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        f = ser.validated_data["file"]

        # ---- Société
        societe = None
        if "societe_id" in ser.validated_data:
            societe = Societe.objects.filter(id=ser.validated_data["societe_id"]).first()
        if not societe and ser.validated_data.get("societe_nom"):
            societe = Societe.objects.filter(nom__iexact=ser.validated_data["societe_nom"]).first()
        if not societe:
            return Response({"error": "Passe 'societe_id' ou 'societe_nom'."}, status=400)

        # ---- Ouvre l’xlsx
        try:
            xl = pd.ExcelFile(f, engine="openpyxl")
        except Exception as e:
            return Response({"error": f"Fichier Excel invalide: {e}"}, status=400)

        # ---- Pick des bonnes feuilles
        norm_names = {LOWER(n): n for n in xl.sheet_names}
        find_sheet = lambda aliases: next((norm_names[LOWER(a)] for a in aliases if LOWER(a) in norm_names), None)
        send_sheet = find_sheet(self.SHEETS_SEND_ALIASES)
        rep_sheet  = find_sheet(self.SHEETS_REP_ALIASES)
        if not send_sheet:
            return Response({"error": "Feuille 'Données envoi' introuvable", "disponibles": xl.sheet_names}, status=400)
        if not rep_sheet:
            return Response({"error": "Feuille 'Réponses clients' introuvable", "disponibles": xl.sheet_names}, status=400)

        # ---- Lecture avec détection d’entêtes (annotation supprimée pour compat 3.8+)
        def read_with_header_detection(sheet, required_aliases):
            head = xl.parse(sheet, header=None, nrows=25, dtype=str)
            best_idx, best_score = None, -1
            for i in range(len(head)):
                row = [LOWER(x) for x in list(head.iloc[i].fillna("").values)]
                score = sum(any(LOWER(a) in row for a in aliases) for aliases in required_aliases) / len(required_aliases)
                if score > best_score:
                    best_idx, best_score = i, score
                if score >= 0.6:
                    return xl.parse(sheet, header=i, dtype=str)
            return xl.parse(sheet, header=(best_idx if best_idx is not None else 0), dtype=str)

        SEND_REQ = [
            ["N° Compte tiers","No Compte tiers","Numero de compte","Numero client","Code client","Numero tiers","Numerotiers","N compte"],
            ["Client","Intitule","Raison sociale","Nom client"],
            ["Contact (email)","Contact","Email","E-mail","Adresse email","Adresse e-mail","Mail"],
            ["envoi (ok ou non)","envoi","envoye","envoyé"],
            ["Date d'envoi","Date envoi","date_envoi"],
            ["Rappel","Date rappel","date_rappel"],
            ["Retour (ok ou non)","Retour","repondu"],
            ["Date retour","Date_retour","date_retour"],
            ["Note Attribué","Note attribue","Note","Score"],
        ]
        REP_REQ = [
            ["Client","code client","numero tiers","n° compte tiers","numerotiers","id client"],
            ["Date retour","date retour","date_retour"],
            ["Question","intitule question","intitulé question"],
            ["Pas satisfait","pas_satisfait"],
            ["Peu satisfait","peu_satisfait"],
            ["Satisfait"],
            ["Très satisfait","tres satisfait","tres_satisfait"],
            ["Oui"],["Non"],["Peut-etre","peut-être","peut_etre"],
            ["Comment","commentaire","remarque"],
            ["Note","score","rating"],
        ]

        df_send = read_with_header_detection(send_sheet, SEND_REQ)
        df_rep  = read_with_header_detection(rep_sheet,  REP_REQ)

        send_map = {LOWER(c): c for c in df_send.columns}
        rep_map  = {LOWER(c): c for c in df_rep.columns}

        GET = lambda row, cmap, *names: next((row[cmap[LOWER(n)]] for n in names if LOWER(n) in cmap), None)

        # ---- Index de réconciliation à partir de "Données envoi"
        def norm_key(x):
            return LOWER(NORM_TXT(x or ""))

        send_name_to_num = {}
        send_email_to_num = {}
        send_nums = set()

        for _, rr in df_send.iterrows():
            ncomp = (GET(rr, send_map, "N° Compte tiers","Numero de compte","Numero client","Code client","Numero tiers","Numerotiers","N compte") or "").strip()
            cname = (GET(rr, send_map, "Client","Raison sociale","Intitule","Nom client") or "").strip()
            email_value = GET(rr, send_map, "Contact (email)","Contact","Email","E-mail","Adresse email","Adresse e-mail","Mail")
            ceml  = (str(email_value).strip().lower() if email_value is not None else "")

            if ncomp:
                send_nums.add(ncomp)
            if cname:
                nk = norm_key(cname)
                send_name_to_num.setdefault(nk, set()).add(ncomp or "")
            if ceml:
                if ceml not in send_email_to_num or (ncomp and not send_email_to_num.get(ceml)):
                    send_email_to_num[ceml] = ncomp

        # ---- Vérif en-têtes minimales
        essentials_send = ["n° compte tiers","client","date d'envoi"]
        miss = [k for k in essentials_send if k not in send_map]
        if miss:
            return Response({"error": "En-têtes minimales manquantes pour 'Données envoi'", "detected": list(df_send.columns), "missing": miss}, status=400)
        essentials_rep = ["client","question"]
        miss2 = [k for k in essentials_rep if k not in rep_map]
        if miss2:
            return Response({"error": "En-têtes minimales manquantes pour 'Réponses clients'", "detected": list(df_rep.columns), "missing": miss2}, status=400)

        # ---- Compteurs & journaux
        created = {"entreprises": 0, "enquetes": 0, "reponses": 0}
        updated = {"entreprises": 0, "enquetes": 0, "reponses": 0}
        skipped = {"send_rows": 0, "rep_rows": 0}
        skipped_details = {"send": [], "rep": []}

        # ---- Caches
        cache_ent_by_num  = {}
        cache_ent_by_mail = {}
        cache_ent_by_name = {}
        cache_enq = {}

        # ---- Map questions (clé: question_fr normalisée)
        cache_q = {LOWER(NORM_TXT(q.question_fr)): q for q in Question.objects.all()}

        def pick_flag(row_dict, keys):
            for k in keys:
                v = row_dict.get(k, "")
                if LOWER(v) in ("1","x","true","vrai","ok","oui"):
                    return k
            return None

        return_details = request.query_params.get("return_details") in ("1", "true", "yes")
        touched_enquetes = set()
        created_reponses_ids = []

        @transaction.atomic
        def import_core():
            nonlocal touched_enquetes, created_reponses_ids
            enq_max_ret = {}

            SAT_LABEL = {
                "pas satisfait": "Pas satisfait",
                "peu satisfait": "Peu satisfait",
                "satisfait": "Satisfait",
                "très satisfait": "Très satisfait",
                "tres satisfait": "Très satisfait",
            }
            YNM_LABEL = {
                "oui": "Oui",
                "non": "Non",
                "peut etre": "Peut-être",
                "peut-être": "Peut-être",
            }

            # ---------- 1) ENTREPRISES + ENQUÊTES
            for _, r in df_send.iterrows():
                excel_row = int(r.name) + 2
                num_compte = (GET(r, send_map, "N° Compte tiers","Numero de compte","Numero client","Code client","Numero tiers") or "").strip()
                nom_cli    = (GET(r, send_map, "Client","Raison sociale","Intitule","Nom client") or "").strip()
                email_val  = GET(r, send_map, "Contact (email)","Contact","Email","E-mail","Adresse email","Adresse e-mail")
                if _has_value(email_val):
                    email_cli = str(email_val).strip().lower()
                    if email_cli == "nan":
                        email_cli = ""
                else:
                    email_cli = ""
                d_envoi    = PARSE_DATE(GET(r, send_map, "Date d'envoi","Date envoi","date_envoi") or "")
                raw_rappel = GET(r, send_map, "Rappel","Date rappel","date_rappel")
                d_rappel = PARSE_DATE(raw_rappel) if _has_value(raw_rappel) else None
                raw_retour = GET(r, send_map, "Date retour","Date_retour","date_retour")
                d_retour = PARSE_DATE(raw_retour) if _has_value(raw_retour) else None

                if not (num_compte or nom_cli):
                    skipped["send_rows"] += 1
                    skipped_details["send"].append({"row": excel_row, "reason": "Ni num_compte ni nom client"})
                    continue
                if not d_envoi:
                    skipped["send_rows"] += 1
                    skipped_details["send"].append({"row": excel_row, "reason": "Date d'envoi manquante/illisible"})
                    continue

                # Résolution Entreprise (priorité num_compte au sein de la même société)
                ent = None
                if num_compte:
                    ent = cache_ent_by_num.get((societe.id, num_compte)) \
                        or Entreprise.objects.filter(societe=societe, num_compte=num_compte).first()
                else:
                    if email_cli:
                        ent = cache_ent_by_mail.get((societe.id, email_cli)) \
                            or Entreprise.objects.filter(societe=societe, email__iexact=email_cli).first()
                    if not ent and nom_cli:
                        qs = Entreprise.objects.filter(societe=societe, nom__iexact=nom_cli)
                        ent = qs.first() if qs.count() == 1 else None

                if not ent:
                    # Vérifier si le numéro de compte est déjà utilisé dans cette société
                    if num_compte and Entreprise.objects.filter(societe=societe, num_compte=num_compte).exists():
                        # Si le numéro de compte est déjà utilisé, on crée sans numéro de compte
                        num_compte = None
                        skipped["send_rows"] += 1
                        skipped_details["send"].append({
                            "row": excel_row, 
                            "reason": f"Le numéro de compte {num_compte} est déjà utilisé dans cette société",
                            "action": "Création sans numéro de compte"
                        })
                    
                    sage = fetch_client_from_sage(societe.id, (num_compte or nom_cli) or "") or {}
                    ent = Entreprise.objects.create(
                        num_compte = (sage.get("NumeroTiers") or (num_compte or None)),
                        nom        = (sage.get("Intitule") or ""),
                        email      = (SAFE_EMAIL(sage.get("Email")) or ""),
                        adresse    = (sage.get("Adresse", "") or ""),
                        telephone  = (sage.get("Telephone", "") or ""),
                        secteur_activite = "",
                        societe    = societe,
                        is_CLT=True, is_Prospect=False, is_Concurent=False
                    )
                    created["entreprises"] += 1
                else:
                    # Entreprise trouvée: forcer l'alignement sur les données SAGE si disponibles
                    if not ent.societe_id:
                        ent.societe = societe
                        ent.save(update_fields=["societe"])
                    sage = fetch_client_from_sage(societe.id, ent.num_compte or nom_cli or "") or {}
                    fields_to_update = []
                    if sage.get("Intitule") and ent.nom != sage.get("Intitule"):
                        ent.nom = sage.get("Intitule")
                        fields_to_update.append("nom")
                    if sage.get("Email") is not None:
                        new_email = SAFE_EMAIL(sage.get("Email"))
                        if new_email and ent.email != new_email:
                            ent.email = new_email
                            fields_to_update.append("email")
                    if sage.get("Adresse") is not None and ent.adresse != (sage.get("Adresse") or ""):
                        ent.adresse = sage.get("Adresse") or ""
                        fields_to_update.append("adresse")
                    if sage.get("Telephone") is not None and ent.telephone != (sage.get("Telephone") or ""):
                        ent.telephone = sage.get("Telephone") or ""
                        fields_to_update.append("telephone")
                    # si num_compte absent en base, essayer de le compléter depuis SAGE
                    if not ent.num_compte and sage.get("NumeroTiers"):
                        ent.num_compte = sage.get("NumeroTiers")
                        fields_to_update.append("num_compte")
                    if fields_to_update:
                        ent.save(update_fields=fields_to_update)
                    updated["entreprises"] += 1

                # Caches
                if num_compte:
                    cache_ent_by_num[(societe.id, num_compte)] = ent
                if email_cli:
                    cache_ent_by_mail[(societe.id, email_cli)] = ent
                if nom_cli:
                    cache_ent_by_name[(societe.id, nom_cli)] = ent

                # Enquête (unique par client + date_envoi)
                enq_key = (ent.id, d_envoi.isoformat())
                if enq_key not in cache_enq:
                    enq, was_created = Enquete.objects.update_or_create(
                        client=ent, 
                        date_envoi=d_envoi,
                        defaults={
                            "date_reponse": d_retour if d_retour else None,
                            "date_creation": d_envoi,
                            "description": f"{d_envoi}_{ent.nom}",
                            "created_by": request.user if request.user.is_authenticated else None,
                        }
                    )
                    # Si une date de retour est fournie, marquer l'enquête comme complétée
                    if d_retour and not enq.is_completed:
                        enq.is_completed = True
                        enq.save(update_fields=["is_completed"])
                    if d_rappel:
                        enq.date_rappel = d_rappel
                        enq.save(update_fields=["date_rappel"])
                    cache_enq[enq_key] = enq
                    (created if was_created else updated)["enquetes"] += 1
                else:
                    enq = cache_enq[enq_key]
                    if d_rappel and enq.date_rappel != d_rappel:
                        enq.date_rappel = d_rappel
                        enq.save(update_fields=["date_rappel"])
                    
                    # Mettre à jour la date de réponse: si pas de date_retour, la vider
                    new_date_reponse = d_retour if d_retour else None
                    if enq.date_reponse != new_date_reponse:
                        enq.date_reponse = new_date_reponse
                        enq.save(update_fields=["date_reponse"])
                        # Si une date de retour est définie suite à la mise à jour, marquer comme complétée
                        if new_date_reponse and not enq.is_completed:
                            enq.is_completed = True
                            enq.save(update_fields=["is_completed"])

            # ---------- 2) RÉPONSES
            for _, r in df_rep.iterrows():
                excel_row = int(r.name) + 2
                cli_val_raw = (GET(r, rep_map, "client","code client","numero tiers","n° compte tiers","numerotiers","id client") or "").strip()
                q_txt = LOWER(NORM_TXT(GET(r, rep_map, "question","intitule question","intitulé question") or ""))

                if not cli_val_raw or not q_txt:
                    skipped["rep_rows"] += 1
                    skipped_details["rep"].append({"row": excel_row, "reason": "Client ou question manquant"})
                    continue

                # Résolution du client à partir de "Réponses" via l’index "envoi"
                cli_num = None
                if cli_val_raw in send_nums:
                    cli_num = cli_val_raw
                else:
                    cand = send_email_to_num.get(cli_val_raw.lower())
                    if cand:
                        cli_num = cand
                    if not cli_num:
                        nk = norm_key(cli_val_raw)
                        nums = send_name_to_num.get(nk, set())
                        nums = {x for x in nums if x}
                        if len(nums) == 1:
                            cli_num = next(iter(nums))

                # Recherche Entreprise (priorité num_compte)
                ent = None
                if cli_num:
                    ent = cache_ent_by_num.get((societe.id, cli_num)) \
                          or Entreprise.objects.filter(societe=societe, num_compte=cli_num).first()
                if not ent:
                    ent = cache_ent_by_mail.get((societe.id, cli_val_raw.lower())) \
                          or Entreprise.objects.filter(societe=societe, email__iexact=cli_val_raw).first()
                if not ent:
                    qs = Entreprise.objects.filter(societe=societe, nom__iexact=cli_val_raw)
                    ent = qs.first() if qs.count() == 1 else None

                if not ent:
                    skipped["rep_rows"] += 1
                    skipped_details["rep"].append({
                        "row": excel_row,
                        "reason": f"Client introuvable: '{cli_val_raw}' (aucune correspondance num_compte/email/nom univoque via 'Données envoi')"
                    })
                    continue

                d_ret = PARSE_DATE(GET(r, rep_map, "date-retour","date retour","date_retour"))

                # Trouver l’enquête correspondante
                # 1) Priorité: faire correspondre la Date retour de la feuille Réponses avec Enquete.date_reponse
                # 2) Sinon, prendre l’enquête la plus proche avant ou égale à cette date par date_envoi
                # 3) Sinon, fallback à la dernière enquête (par date_envoi)
                enq = (Enquete.objects.filter(client=ent, date_reponse=d_ret).first() if d_ret else None) \
                      or (Enquete.objects.filter(client=ent, date_envoi__lte=d_ret).order_by('-date_envoi').first() if d_ret else None) \
                      or Enquete.objects.filter(client=ent).order_by('-date_envoi').first()
                if not enq:
                    skipped["rep_rows"] += 1
                    skipped_details["rep"].append({"row": excel_row, "reason": "Aucune enquête correspondante"})
                    continue

                q = cache_q.get(q_txt)
                if not q:
                    skipped["rep_rows"] += 1
                    skipped_details["rep"].append({"row": excel_row, "reason": f"Question non reconnue: '{q_txt}'"})
                    continue

                row_vals = {
                    "pas satisfait": GET(r, rep_map, "pas satisfait","pas_satisfait"),
                    "peu satisfait": GET(r, rep_map, "peu satisfait","peu_satisfait"),
                    "satisfait":     GET(r, rep_map, "satisfait"),
                    "très satisfait":GET(r, rep_map, "très satisfait","tres satisfait","tres_satisfait"),
                    "oui":           GET(r, rep_map, "oui"),
                    "non":           GET(r, rep_map, "non"),
                    "peut etre":     GET(r, rep_map, "peut etre","peut-être","peut_etre"),
                    "comment":       GET(r, rep_map, "comment","commentaire","remarque"),
                    "note":          GET(r, rep_map, "note","score","rating"),
                }

                # Valeur de réponse selon le type
                rep_value = ""
                if q.type == q.Type.CLOSED:
                    key = pick_flag(row_vals, ["pas satisfait","peu satisfait","satisfait","très satisfait","tres satisfait"])
                    if key: rep_value = SAT_LABEL.get(key) or ""
                elif q.type == q.Type.OUINON:
                    key = pick_flag(row_vals, ["oui","non","peut etre"])
                    if key: rep_value = YNM_LABEL.get(key) or ""
                elif q.type == q.Type.OPEN:
                    rep_value = str(row_vals.get("comment") or "").strip()
                elif q.type == q.Type.NOTE:
                    rep_value = str(row_vals.get("note") or "").strip()

                # Préparer le commentaire brut
                raw_cmt = row_vals.get("comment")
                cmt_str = "" if (raw_cmt is None or (isinstance(raw_cmt, float) and str(raw_cmt) == "nan")) else str(raw_cmt).strip()

                # Pour type OPEN: la colonne Comment devient la réponse, commentaire reste vide
                if q.type == q.Type.OPEN:
                    rep_value = cmt_str
                    commentaire = ""
                else:
                    # Pour les autres types: on peut stocker le commentaire même si la réponse est vide
                    commentaire = cmt_str

                # Si aucune réponse et aucun commentaire: ne pas créer de Réponse, mais associer la question
                if not rep_value and not commentaire:
                    skipped["rep_rows"] += 1
                    skipped_details["rep"].append({"row": excel_row, "reason": "Valeur de réponse vide et aucun commentaire"})
                    enq.questions.add(q)
                    touched_enquetes.add(enq.id)
                    continue

                # Lier et enregistrer
                enq.questions.add(q)
                touched_enquetes.add(enq.id)

                if d_ret:
                    prev = enq_max_ret.get(enq.id)
                    enq_max_ret[enq.id] = max(prev, d_ret) if prev else d_ret

                obj, _ = Reponse.objects.update_or_create(
                    enquete=enq, client=ent, question=q,
                    defaults={"reponse": str(rep_value), "commentaire": commentaire}
                )
                created_reponses_ids.append(obj.id)
                created["reponses"] += 1

            # Questions obligatoires si rien n’a été accroché
            mandatory_qs = list(Question.objects.filter(obligatoire=True))
            if mandatory_qs:
                for enq in cache_enq.values():
                    has_q = enq.questions.exists()
                    has_r = Reponse.objects.filter(enquete=enq).exists()
                    if not has_q and not has_r:
                        enq.questions.add(*mandatory_qs)

            # Finalisation (progress, dates)
            for enq_id in touched_enquetes:
                enq = Enquete.objects.select_for_update().get(id=enq_id)
                fields = []
                if not enq.is_completed:
                    enq.is_completed = True
                    fields.append("is_completed")
                max_ret = enq_max_ret.get(enq_id)
                if max_ret and enq.date_reponse != max_ret:
                    enq.date_reponse = max_ret
                    fields.append("date_reponse")
                if fields:
                    enq.date_modification = timezone.now().date()
                    fields.append("date_modification")
                    enq.save(update_fields=fields)

        import_core()

        payload = {
            "societe": {"id": societe.id, "nom": societe.nom},
            "created": created, "updated": updated, "skipped": skipped,
            "skipped_details": skipped_details
        }

        # Optionnel : renvoyer le détail des enquêtes & réponses touchées
        if return_details:
            enquetes = list(
                Enquete.objects.filter(id__in=touched_enquetes)
                .values("id", "client__nom", "client__num_compte", "date_envoi", "date_reponse", "is_completed")
            )
            reponses = list(
                Reponse.objects.filter(id__in=created_reponses_ids)
                .values("id", "enquete_id", "client_id", "question_id", "reponse", "commentaire")
            )
            payload["touched_enquetes"] = enquetes
            payload["created_reponses"] = reponses

        return Response(payload, status=200)


# =============================
# Prospects management (consolidated)
# =============================

@method_decorator(login_required, name='dispatch')
class ProspectsListView(View):
    def get(self, request):
        q = (request.GET.get('search') or '').strip()
        sort_by_date = request.GET.get('sort_by_date')
        societe_filter = request.GET.get('societe')

        qs = Entreprise.objects.filter(is_Prospect=True, is_CLT=False).order_by('nom')

        if request.user.is_superuser:
            pass
        elif getattr(request.user, 'is_RO', False):
            societe_ids = list(request.user.societes.values_list('id', flat=True))
            if societe_ids:
                qs = qs.filter(societe_id__in=societe_ids)
            else:
                qs = qs.none()
        else:
            societe = getattr(request.user, 'societe', None)
            if societe:
                qs = qs.filter(societe=societe)
            else:
                qs = qs.none()

        if q:
            qs = qs.filter(
                Q(nom__icontains=q) |
                Q(email__icontains=q) |
                Q(telephone__icontains=q) |
                Q(secteur_activite__icontains=q) |
                Q(num_compte__icontains=q)
            )
        if societe_filter:
            qs = qs.filter(societe_id=societe_filter)
        if sort_by_date == 'asc':
            qs = qs.order_by('date')
        elif sort_by_date == 'desc':
            qs = qs.order_by('-date')
        else:
            qs = qs.order_by('-date')

        paginator = Paginator(qs, 100)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        ctx = {
            'entreprises': page_obj,
            'is_superuser': request.user.is_superuser,
            'is_RO': request.user.is_RO,
            'is_responsable': getattr(request.user, 'is_RC', False),
            'search_query': q,
            'sort_by_date': sort_by_date or '',
            'societes': Societe.objects.all() if request.user.is_superuser else (Societe.objects.filter(id__in=request.user.societes.all()) if getattr(request.user, 'is_RO', False) else []),
            'societe_filter': societe_filter or '',
        }
        return render(request, 'adminlte/sales/users/entreprises/prospects.html', ctx)


@method_decorator(login_required, name='dispatch')
class ProspectDetailView(View):
    def get(self, request, entreprise_id: int):
        entreprise = get_object_or_404(Entreprise, pk=entreprise_id, is_Prospect=True)
        researches = entreprise.prospect_researches.select_related('created_by').all().order_by('-created_at')
        infos = entreprise.prospect_infos.select_related('research').all().order_by('-created_at')
        last_research = researches.first()
        counts = {
            'emails': infos.filter(type=ProspectInfo.InfoType.EMAIL).count(),
            'phones': infos.filter(type=ProspectInfo.InfoType.PHONE).count(),
            'websites': infos.filter(type=ProspectInfo.InfoType.WEBSITE).count(),
            'socials': infos.filter(type=ProspectInfo.InfoType.SOCIAL).count(),
            'addresses': infos.filter(type=ProspectInfo.InfoType.ADDRESS).count(),
            'notes': infos.filter(type=ProspectInfo.InfoType.NOTE).count(),
        }
        # Group infos by research (date)
        grouped_infos = []
        history_labels = []
        history_totals = []
        for r in researches:
            r_infos = infos.filter(research=r)
            g = {
                'research': r,
                'summary': r.summary,
                'emails': list(r_infos.filter(type=ProspectInfo.InfoType.EMAIL)),
                'phones': list(r_infos.filter(type=ProspectInfo.InfoType.PHONE)),
                'websites': list(r_infos.filter(type=ProspectInfo.InfoType.WEBSITE)),
                'socials': list(r_infos.filter(type=ProspectInfo.InfoType.SOCIAL)),
                'addresses': list(r_infos.filter(type=ProspectInfo.InfoType.ADDRESS)),
                'notes': list(r_infos.filter(type=ProspectInfo.InfoType.NOTE)),
            }
            total_this = sum(len(g[k]) for k in ['emails','phones','websites','socials','addresses','notes'])
            history_labels.append(r.created_at.strftime('%d/%m/%Y %H:%M'))
            history_totals.append(total_this)
            grouped_infos.append(g)
        ctx = {
            'entreprise': entreprise,
            'researches': researches,
            'infos': infos,
            'last_research': last_research,
            'info_counts': counts,
            'grouped_infos': grouped_infos,
            'history_labels': history_labels,
            'history_totals': history_totals,
            'info_counts_json': json.dumps(counts),
            'history_labels_json': json.dumps(history_labels),
            'history_totals_json': json.dumps(history_totals),
            'is_superuser': request.user.is_superuser,
            'is_RO': request.user.is_RO,
            'is_responsable': getattr(request.user, 'is_RC', False),
        }
        return render(request, 'adminlte/sales/users/entreprises/prospect_detail.html', ctx)


@login_required
def trigger_prospect_research(request, entreprise_id: int):
    if request.method != 'POST':
        return redirect('prospection:prospect_detail', entreprise_id=entreprise_id)

    entreprise = get_object_or_404(Entreprise, pk=entreprise_id, is_Prospect=True)
    country = (request.POST.get('country') or '').strip()
    language = (request.POST.get('language') or '').strip()
    notes = (request.POST.get('notes') or '').strip()
    parts = []
    if country:
        parts.append(f"Pays: {country}")
    if language:
        parts.append(f"Langue: {language}")
    if notes:
        parts.append(notes)
    extra_query = "; ".join(parts)

    try:
        result = research_prospect(entreprise.nom, entreprise.secteur_activite or '', extra_query)
    except Exception as e:
        messages.error(request, f"Erreur lors de la recherche AI: {e}")
        return redirect('prospection:prospect_detail', entreprise_id=entreprise.id)

    created_total = 0
    duplicates_total = 0  # Compteur de doublons ignorés
    created_by_type = {t: 0 for t in [
        ProspectInfo.InfoType.EMAIL,
        ProspectInfo.InfoType.PHONE,
        ProspectInfo.InfoType.WEBSITE,
        ProspectInfo.InfoType.SOCIAL,
        ProspectInfo.InfoType.ADDRESS,
        ProspectInfo.InfoType.NOTE,
    ]}
    duplicates_by_type = {t: 0 for t in [
        ProspectInfo.InfoType.EMAIL,
        ProspectInfo.InfoType.PHONE,
        ProspectInfo.InfoType.WEBSITE,
        ProspectInfo.InfoType.SOCIAL,
        ProspectInfo.InfoType.ADDRESS,
        ProspectInfo.InfoType.NOTE,
    ]}

    with transaction.atomic():
        research = ProspectResearch.objects.create(
            entreprise=entreprise,
            created_by=request.user,
            query=extra_query,
            summary=result.get('summary') or '',
            raw_result=result.get('raw') or '',
            confidence=result.get('confidence', 'medium')
        )
        data = (result.get('data') or {})

        # Build existing normalized sets per type to avoid duplicates
        existing_raw = entreprise.prospect_infos.all().values('type', 'value')
        existing_sets = {
            ProspectInfo.InfoType.EMAIL: set(),
            ProspectInfo.InfoType.PHONE: set(),
            ProspectInfo.InfoType.WEBSITE: set(),
            ProspectInfo.InfoType.SOCIAL: set(),
            ProspectInfo.InfoType.ADDRESS: set(),
        }
        def _norm_email(v: str) -> str:
            return (v or '').strip().lower()
        def _norm_phone(v: str) -> str:
            s = (v or '').strip()
            s = re.sub(r"[^0-9+]+", "", s)
            return s
        def _norm_url(v: str) -> str:
            s = (v or '').strip()
            if not s:
                return ''
            s = s.lower()
            if not s.startswith('http://') and not s.startswith('https://'):
                s = 'https://' + s
            # Strip trailing slash
            if len(s) > 1 and s.endswith('/'):
                s = s[:-1]
            return s
        def _norm_addr(v: str) -> str:
            return (v or '').strip()
        for row in existing_raw:
            t = row['type']
            val = row['value'] or ''
            if t == ProspectInfo.InfoType.EMAIL:
                existing_sets[t].add(_norm_email(val))
            elif t == ProspectInfo.InfoType.PHONE:
                existing_sets[t].add(_norm_phone(val))
            elif t in (ProspectInfo.InfoType.WEBSITE, ProspectInfo.InfoType.SOCIAL):
                existing_sets[t].add(_norm_url(val))
            elif t == ProspectInfo.InfoType.ADDRESS:
                existing_sets[t].add(_norm_addr(val))

        url_validator = URLValidator(schemes=['http', 'https'])

        # Country-specific rules (extend as needed)
        COUNTRY_RULES = {
            'fr': {
                'aliases': ['france', 'fr'],
                'tlds': ['.fr'],
                'phones': ['+33'],
            },
            'tn': {
                'aliases': ['tunisie', 'tunisia', 'tn'],
                'tlds': ['.tn'],
                'phones': ['+216'],
            },
            'de': {
                'aliases': ['allemagne', 'germany', 'deutschland', 'de'],
                'tlds': ['.de'],
                'phones': ['+49'],
            },
            'es': {
                'aliases': ['espagne', 'spain', 'es'],
                'tlds': ['.es'],
                'phones': ['+34'],
            },
            'it': {
                'aliases': ['italie', 'italy', 'it'],
                'tlds': ['.it'],
                'phones': ['+39'],
            },
            'ma': {
                'aliases': ['maroc', 'morocco', 'ma'],
                'tlds': ['.ma'],
                'phones': ['+212'],
            },
            'dz': {
                'aliases': ['algerie', 'algeria', 'dz'],
                'tlds': ['.dz'],
                'phones': ['+213'],
            },
        }
        sel = (country or '').strip().lower()
        selected_country_key = None
        if sel:
            for key, rule in COUNTRY_RULES.items():
                if sel == key or any(a in sel for a in rule['aliases']):
                    selected_country_key = key
                    break
        selected_rules = COUNTRY_RULES.get(selected_country_key) if selected_country_key else None

        def _domain_tld(u: str) -> str:
            try:
                from urllib.parse import urlparse
                host = urlparse(u).hostname or ''
                if '.' in host:
                    return '.' + host.split('.')[-1]
                return ''
            except Exception:
                return ''

        def _enforce_country_for_email(email: str) -> bool:
            if not selected_rules:
                return True
            try:
                domain = email.split('@', 1)[1]
                if not domain:
                    return False
                if '.' not in domain:
                    return False
                tld = '.' + domain.split('.')[-1].lower()
                return tld in selected_rules['tlds']
            except Exception:
                return False

        def _enforce_country_for_url(u: str) -> bool:
            if not selected_rules:
                return True
            tld = _domain_tld(u).lower()
            return bool(tld) and tld in selected_rules['tlds']

        def _enforce_country_for_phone(p: str) -> bool:
            if not selected_rules:
                return True
            return any(p.startswith(pref) for pref in selected_rules['phones'])

        def _enforce_country_for_address(a: str) -> bool:
            if not selected_rules:
                return True
            a2 = (a or '').lower()
            # Accept if address mentions one of the country aliases
            return any(alias in a2 for alias in selected_rules['aliases'])

        PLACEHOLDER_PATTERNS = (
            'non disponible', 'not available', 'n/a', 'introuvable', 'infructueuse'
        )

        def _is_placeholder(s: str) -> bool:
            s2 = (s or '').strip().lower()
            if not s2:
                return True
            return any(p in s2 for p in PLACEHOLDER_PATTERNS)

        def _is_valid_email(s: str) -> bool:
            return bool(re.match(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$", s or ''))

        def _is_valid_phone(s: str) -> bool:
            digits = re.sub(r"\D+", "", s or '')
            return len(digits) >= 7

        def _is_valid_url(s: str) -> bool:
            try:
                url_validator(s)
                return True
            except ValidationError:
                return False

        def _url_reachable(s: str) -> bool:
            try:
                r = requests.head(s, allow_redirects=True, timeout=4)
                if r.status_code >= 400:
                    # fallback to GET in case HEAD is blocked
                    r = requests.get(s, allow_redirects=True, timeout=6)
                return 200 <= r.status_code < 400
            except Exception:
                return False

        def _save_items(items, info_type):
            nonlocal created_total, duplicates_total
            for it in items or []:
                raw_value = (it.get('value') or '').strip()
                if not raw_value or _is_placeholder(raw_value):
                    continue
                # Normalize + validate depending on type
                if info_type == ProspectInfo.InfoType.EMAIL:
                    value = _norm_email(raw_value)
                    if not _is_valid_email(value):
                        continue
                    if not _enforce_country_for_email(value):
                        continue
                    if value in existing_sets[ProspectInfo.InfoType.EMAIL]:
                        duplicates_total += 1
                        duplicates_by_type[info_type] += 1
                        logger.info(f"Doublon email ignoré: {value}")
                        continue
                elif info_type == ProspectInfo.InfoType.PHONE:
                    value = _norm_phone(raw_value)
                    if not _is_valid_phone(value):
                        continue
                    if not _enforce_country_for_phone(value):
                        continue
                    if value in existing_sets[ProspectInfo.InfoType.PHONE]:
                        duplicates_total += 1
                        duplicates_by_type[info_type] += 1
                        logger.info(f"Doublon téléphone ignoré: {value}")
                        continue
                elif info_type in (ProspectInfo.InfoType.WEBSITE, ProspectInfo.InfoType.SOCIAL):
                    value = _norm_url(raw_value)
                    if not _is_valid_url(value):
                        continue
                    if value in existing_sets[info_type]:
                        duplicates_total += 1
                        duplicates_by_type[info_type] += 1
                        logger.info(f"Doublon {info_type} ignoré: {value}")
                        continue
                    # Reachability check to avoid 404/not found entries
                    if not _url_reachable(value):
                        continue
                    # Country TLD constraint for websites; socials are global, but still apply TLD if present
                    if info_type == ProspectInfo.InfoType.WEBSITE and not _enforce_country_for_url(value):
                        continue
                elif info_type == ProspectInfo.InfoType.ADDRESS:
                    value = _norm_addr(raw_value)
                    if len(value) < 5:
                        continue
                    if not _enforce_country_for_address(value):
                        continue
                    if value in existing_sets[ProspectInfo.InfoType.ADDRESS]:
                        duplicates_total += 1
                        duplicates_by_type[info_type] += 1
                        logger.info(f"Doublon adresse ignorée: {value}")
                        continue
                else:
                    value = raw_value

                try:
                    ProspectInfo.objects.create(
                        entreprise=entreprise,
                        research=research,
                        type=info_type,
                        value=value,
                        label=it.get('label'),
                        source_url=it.get('source_url'),
                        extra=None
                    )
                    created_total += 1
                    created_by_type[info_type] += 1
                    # add to existing set to prevent duplicates in same run
                    if info_type in existing_sets:
                        existing_sets[info_type].add(value)
                except IntegrityError as e:
                    # Doublon détecté au niveau base de données (contrainte unique)
                    duplicates_total += 1
                    duplicates_by_type[info_type] += 1
                    logger.warning(f"Doublon base de données pour {info_type}: {value} - {str(e)}")
                    continue

        _save_items(data.get('emails'), ProspectInfo.InfoType.EMAIL)
        _save_items(data.get('phones'), ProspectInfo.InfoType.PHONE)
        _save_items(data.get('websites'), ProspectInfo.InfoType.WEBSITE)
        _save_items(data.get('socials'), ProspectInfo.InfoType.SOCIAL)
        _save_items(data.get('addresses'), ProspectInfo.InfoType.ADDRESS)
        # Skip saving raw notes; we only display AI summary per research

    # Affichage des résultats avec le niveau de confidence
    confidence = result.get('confidence', 'medium')
    confidence_label = {
        'high': '✓ Fiabilité élevée',
        'medium': '⚠ Fiabilité moyenne',
        'low': '⚠ Fiabilité faible',
        'none': '✗ Échec de la recherche'
    }.get(confidence, '⚠ Fiabilité moyenne')
    
    # Messages détaillés pour l'utilisateur
    if created_total > 0:
        msg_parts = [
            f"Recherche AI effectuée ({confidence_label}).",
            f"✅ {created_total} nouvelle(s) information(s) enregistrée(s):"
        ]
        
        # Détails par type
        details = []
        if created_by_type[ProspectInfo.InfoType.EMAIL] > 0:
            details.append(f"Emails: {created_by_type[ProspectInfo.InfoType.EMAIL]}")
        if created_by_type[ProspectInfo.InfoType.PHONE] > 0:
            details.append(f"Téléphones: {created_by_type[ProspectInfo.InfoType.PHONE]}")
        if created_by_type[ProspectInfo.InfoType.WEBSITE] > 0:
            details.append(f"Sites: {created_by_type[ProspectInfo.InfoType.WEBSITE]}")
        if created_by_type[ProspectInfo.InfoType.SOCIAL] > 0:
            details.append(f"Réseaux: {created_by_type[ProspectInfo.InfoType.SOCIAL]}")
        if created_by_type[ProspectInfo.InfoType.ADDRESS] > 0:
            details.append(f"Adresses: {created_by_type[ProspectInfo.InfoType.ADDRESS]}")
        
        msg_parts.append(" • ".join(details))
        
        # Informations sur les doublons ignorés
        if duplicates_total > 0:
            dup_details = []
            if duplicates_by_type[ProspectInfo.InfoType.EMAIL] > 0:
                dup_details.append(f"Emails: {duplicates_by_type[ProspectInfo.InfoType.EMAIL]}")
            if duplicates_by_type[ProspectInfo.InfoType.PHONE] > 0:
                dup_details.append(f"Téléphones: {duplicates_by_type[ProspectInfo.InfoType.PHONE]}")
            if duplicates_by_type[ProspectInfo.InfoType.WEBSITE] > 0:
                dup_details.append(f"Sites: {duplicates_by_type[ProspectInfo.InfoType.WEBSITE]}")
            if duplicates_by_type[ProspectInfo.InfoType.SOCIAL] > 0:
                dup_details.append(f"Réseaux: {duplicates_by_type[ProspectInfo.InfoType.SOCIAL]}")
            if duplicates_by_type[ProspectInfo.InfoType.ADDRESS] > 0:
                dup_details.append(f"Adresses: {duplicates_by_type[ProspectInfo.InfoType.ADDRESS]}")
            
            msg_parts.append(f" | ⚠️ {duplicates_total} doublon(s) ignoré(s): {', '.join(dup_details)}")
        
        messages.success(request, " ".join(msg_parts))
        logger.info(f"Recherche terminée - Créés: {created_total}, Doublons: {duplicates_total}")
        
    elif confidence == 'none':
        messages.error(
            request, 
            f"❌ Erreur lors de la recherche AI. {result.get('summary', 'Veuillez réessayer ultérieurement.')}"
        )
        logger.error(f"Recherche échouée pour {entreprise.nom}")
        
    else:
        # Aucune nouvelle information mais peut-être des doublons
        if duplicates_total > 0:
            dup_details = []
            if duplicates_by_type[ProspectInfo.InfoType.EMAIL] > 0:
                dup_details.append(f"Emails: {duplicates_by_type[ProspectInfo.InfoType.EMAIL]}")
            if duplicates_by_type[ProspectInfo.InfoType.PHONE] > 0:
                dup_details.append(f"Téléphones: {duplicates_by_type[ProspectInfo.InfoType.PHONE]}")
            if duplicates_by_type[ProspectInfo.InfoType.WEBSITE] > 0:
                dup_details.append(f"Sites: {duplicates_by_type[ProspectInfo.InfoType.WEBSITE]}")
            if duplicates_by_type[ProspectInfo.InfoType.SOCIAL] > 0:
                dup_details.append(f"Réseaux: {duplicates_by_type[ProspectInfo.InfoType.SOCIAL]}")
            if duplicates_by_type[ProspectInfo.InfoType.ADDRESS] > 0:
                dup_details.append(f"Adresses: {duplicates_by_type[ProspectInfo.InfoType.ADDRESS]}")
            
            messages.info(
                request,
                f"Recherche AI effectuée ({confidence_label}). "
                f"Aucune nouvelle information. {duplicates_total} doublon(s) détecté(s) et ignoré(s): {', '.join(dup_details)}"
            )
        else:
            messages.info(
                request, 
                f"Recherche AI effectuée ({confidence_label}). Aucune nouvelle information trouvée."
            )
        logger.info(f"Recherche sans nouveauté - Doublons ignorés: {duplicates_total}")

    return redirect('prospection:prospect_detail', entreprise_id=entreprise.id)


@require_http_methods(["GET"])
@csrf_exempt
def loadSwotData(request):
    try:
        concurrent_id = request.GET.get('concurrent_id')
        
        # Si aucun concurrent_id n'est fourni, retourner tous les SWOTs avec leurs concurrents
        if not concurrent_id:
            # Récupérer tous les SWOTs
            swot_items = Swot.objects.select_related('entreprise').all()
            
            # Construire dynamiquement les groupes par type avec le nom du concurrent
            result = {}
            for item in swot_items:
                key = item.type  # ex: 'force', 'faiblesse', 'opportunite', 'menace', ...
                if key not in result:
                    result[key] = []
                
                # Créer un objet avec description et nom du concurrent
                item_data = {
                    'description': f"{item.description} - {item.axe}" if item.axe else item.description,
                    'concurrent': item.entreprise.nom if item.entreprise else 'N/A'
                }
                result[key].append(item_data)
            
            
            return JsonResponse({
                'status': 'success',
                'data': result
            })

        # Vérifier que le concurrent existe
        try:
            concurrent = Entreprise.objects.get(id=concurrent_id, is_Concurent=True)
        except Entreprise.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': 'Concurrent non trouvé'
            }, status=404)

        # Récupérer les données SWOT pour le concurrent spécifié
        swot_items = Swot.objects.filter(entreprise_id=concurrent_id)

        # Construire dynamiquement les groupes par type (supporte automatiquement de nouveaux types)
        result = {}
        for item in swot_items:
            item_text = f"{item.description} - {item.axe}" if item.axe else item.description
            key = item.type  # ex: 'force', 'faiblesse', 'opportunite', 'menace', ...
            # Utiliser une clé telle quelle pour permettre au frontend de l'afficher dynamiquement
            if key not in result:
                result[key] = []
            result[key].append(item_text)
        
        return JsonResponse({
            'status': 'success',
            'data': result
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({
            'status': 'error',
            'message': f'Erreur serveur: {str(e)}',
            'traceback': traceback.format_exc()
        }, status=500)

@login_required
def enquete_details_ajax(request, enquete_id):
    """Retourne les détails d'une enquête en JSON pour AJAX"""
    try:
        enquete = Enquete.objects.select_related('client', 'created_by').prefetch_related('questions').get(id=enquete_id)

        # Récupérer les questions et la dernière réponse du client (si définie) pour chaque question
        questions_payload = []
        client = enquete.client
        from .models import Reponse  # import local pour éviter cycles

        # Base questions: prefer M2M; if empty, derive from responses linked to this enquete
        questions_qs = enquete.questions.all()
        if not questions_qs.exists():
            from .models import Question, Reponse as _Resp
            qids = list(_Resp.objects.filter(enquete=enquete).values_list('question_id', flat=True).distinct())
            questions_qs = Question.objects.filter(id__in=qids)

        for q in questions_qs:
            # Filtrer les réponses pour cette enquête et question; si client est défini, privilégier celles du client
            base_qs = Reponse.objects.filter(enquete=enquete, question=q)
            qs = base_qs
            if client:
                qs = qs.filter(client=client)
            # Dernière réponse (par id DESC)
            last_resp = qs.order_by('-id').first()
            # Fallback: si aucune réponse pour ce client, prendre la dernière réponse globale
            if not last_resp:
                last_resp = base_qs.order_by('-id').first()

            questions_payload.append({
                'question': q.get_question_text('fr') if hasattr(q, 'get_question_text') else getattr(q, 'question_fr', ''),
                'type': q.type,
                'obligatoire': bool(getattr(q, 'obligatoire', False)),
                'reponse': last_resp.reponse if last_resp else '',
                'commentaire': last_resp.commentaire if last_resp and last_resp.commentaire else ''
            })

        return JsonResponse({
            'id': enquete.id,
            'titre': enquete.titre,
            'description': enquete.description,
            'date_creation': enquete.date_creation.strftime('%Y-%m-%d') if enquete.date_creation else None,
            'date_modification': enquete.date_modification.strftime('%Y-%m-%d') if enquete.date_modification else None,
            'date_envoi': enquete.date_envoi.strftime('%Y-%m-%d') if enquete.date_envoi else None,
            'date_rappel': enquete.date_rappel.strftime('%Y-%m-%d') if enquete.date_rappel else None,
            'date_reponse': enquete.date_reponse.strftime('%Y-%m-%d') if enquete.date_reponse else None,
            'client_email': enquete.client.email if enquete.client else None,
            'client_nom': enquete.client.nom if enquete.client else None,
            'created_by': (enquete.created_by.get_full_name() or enquete.created_by.username) if getattr(enquete, 'created_by', None) else '',
            'is_completed': bool(enquete.is_completed),
            'questions': questions_payload,
        })
    except Enquete.DoesNotExist:
        return JsonResponse({'error': 'Enquête introuvable'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def send_response_pdf_email(request):
    if request.method == 'POST':
        try:
            enquete_id = request.POST.get('enquete_id')
            recipient_emails = request.POST.getlist('recipient_emails[]')
            email_subject = request.POST.get('email_subject')
            email_message = request.POST.get('email_message', '')
            
            try:
                enquete = Enquete.objects.get(id=enquete_id)
            except Enquete.DoesNotExist:
                return JsonResponse({'error': 'Enquête introuvable'}, status=404)
            
            if not enquete.date_reponse:
                return JsonResponse({'error': 'Cette enquête n\'a pas encore de réponse'}, status=400)
            
            if not recipient_emails:
                return JsonResponse({'error': 'Au moins un destinataire est requis'}, status=400)

            try:
                from django.test import RequestFactory
                rf = RequestFactory()
                internal_request = rf.get(f'/enquetes/export/pdf/{enquete.id}/')
                internal_request.user = request.user
                export_response = export_enquete_pdf(internal_request, enquete.id)
                pdf_bytes = bytes(export_response.content)
            except Exception as e:
                return JsonResponse({'error': f"Erreur lors de la génération du PDF via l'endpoint existant: {str(e)}"}, status=500)
            
            pdf_filename = f"Reponses_Enquete_{enquete.client.nom.replace(' ', '_')}_{enquete.date_reponse.strftime('%Y%m%d') if enquete.date_reponse else 'inconnu'}.pdf"
            
            from django.core.mail import EmailMessage
            
            email = EmailMessage(
                subject=email_subject,
                body=email_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=recipient_emails,
            )
            
            email.attach(pdf_filename, pdf_bytes, 'application/pdf')
            email.send(fail_silently=False)
            
            return JsonResponse({
                'message': f'Email envoyé avec succès à {len(recipient_emails)} destinataire(s)',
                'recipients': recipient_emails
            })
            
        except Exception as e:
            import traceback
            return JsonResponse({
                'error': f'Erreur lors de l\'envoi de l\'email: {str(e)}',
                'traceback': traceback.format_exc()
            }, status=500)
    
    return JsonResponse({'error': 'Méthode non autorisée'}, status=405)

def parse_excel_date(date_value):
    """Parse une date depuis Excel en format datetime"""
    if pd.isna(date_value):
        return None
    try:
        if isinstance(date_value, str):
            return pd.to_datetime(date_value).to_pydatetime()
        return pd.to_datetime(date_value).to_pydatetime()
    except Exception as e:
        logger.warning(f"Erreur parsing date: {date_value}, erreur: {e}")
        return None


@login_required
def import_rvs(request):
    """Importe des rendez-vous depuis un fichier Excel"""
    if not (request.user.is_superuser or request.user.is_RO):
        messages.error(request, "Vous n'avez pas les permissions nécessaires.")
        return redirect('prospection:rv_list')
    
    if request.method != 'POST':
        return redirect('prospection:rv_list')
    
    excel_file = request.FILES.get('excel_file')
    societe_id = request.POST.get('societe_id')
    
    if not excel_file or not societe_id:
        messages.error(request, "Fichier Excel et filiale sont requis.")
        return redirect('prospection:rv_list')
    
    try:
        societe = Societe.objects.get(id=societe_id)
    except Societe.DoesNotExist:
        messages.error(request, "Filiale introuvable.")
        return redirect('prospection:rv_list')
    
    pilote_rc = Utilisateur.objects.filter(is_RC=True, societe=societe).first()
    
    success_count = 0
    error_count = 0
    errors = []
    
    try:
        df = pd.read_excel(excel_file, sheet_name='Rendez-vous', header=1, 
                          dtype={'N° compte tiers': str})
        
        required_cols = ['Date', 'Prise par', 'N° compte tiers', 'Société', 'Nom contact', 
                        'Poste', 'Tel', 'Mail', 'Volet', 'Etat du rendez-vous', 'Sujet']
        
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            messages.error(request, f"Colonnes manquantes dans le fichier: {', '.join(missing_cols)}")
            return redirect('prospection:rv_list')

        for idx, row in df.iterrows():
            try:
                date_planifie = parse_excel_date(row.get('Date')) or None
                
                volet_check = row.get('Volet')
                if pd.isna(volet_check) or not str(volet_check).strip():
                    error_count += 1
                    errors.append(f"Ligne {idx+3}: Volet manquant")
                    continue
                
                created_by = None
                prise_par = str(row.get('Prise par', '')).strip()
                if prise_par and not pd.isna(row.get('Prise par')):
                    created_by = Utilisateur.objects.filter(username=prise_par).first()
                
                entreprise = None
                volet = str(row.get('Volet', '')).strip()
                num_compte_raw = row.get('N° compte tiers')
                num_compte = str(num_compte_raw).strip() if not pd.isna(num_compte_raw) and str(num_compte_raw).strip() != 'nan' else ''
                nom_societe = str(row.get('Société', '')).strip()
                
                notes_parts = []
                if not pd.isna(row.get('Nom contact')):
                    notes_parts.append(f"Contact: {row.get('Nom contact')}")
                if not pd.isna(row.get('Poste')):
                    notes_parts.append(f"Poste: {row.get('Poste')}")
                if not pd.isna(row.get('Tel')):
                    tel_raw = row.get('Tel')
                    try:
                        tel_format = str(int(float(tel_raw)))
                    except (ValueError, TypeError):
                        tel_format = str(tel_raw)
                    notes_parts.append(f"Tel: {tel_format}")
                if not pd.isna(row.get('Mail')):
                    notes_parts.append(f"Mail: {row.get('Mail')}")
                notes = " | ".join(notes_parts)
                
                if 'Developpement' in volet or 'Développement' in volet:
                    if num_compte:
                        entreprise = Entreprise.objects.filter(
                            num_compte=num_compte,
                            societe=societe,
                            is_CLT=True
                        ).first()
                        
                        if not entreprise:
                            try:
                                sage = fetch_client_from_sage(societe.id, num_compte) or {}
                                entreprise = Entreprise.objects.create(
                                    num_compte=(sage.get("NumeroTiers") or (num_compte or None)),
                                    nom=sage.get("Intitule") or "Inconnu",
                                    email=SAFE_EMAIL(sage.get("Email")),
                                    adresse=sage.get("Adresse", "") or "",
                                    telephone=sage.get("Telephone", "") or "",
                                    secteur_activite="",
                                    societe=societe,
                                    is_CLT=True,
                                    is_Prospect=False,
                                    is_Concurent=False
                                )
                            except Exception as e:
                                logger.error(f"Erreur création client ligne {idx+3}: {e}")
                
                elif 'Prospection' in volet:
                    if nom_societe:
                        entreprise = Entreprise.objects.filter(
                            nom__iexact=nom_societe,
                            societe=societe,
                            is_Prospect=True,
                            is_CLT=False
                        ).first()
                        
                        if not entreprise:
                            email_prospect = str(row.get('Mail', '')).strip() if not pd.isna(row.get('Mail')) else ''
                            email_prospect = email_prospect.strip("'\"")
                            
                            tel_prospect_raw = row.get('Tel')
                            if pd.isna(tel_prospect_raw):
                                tel_prospect = ''
                            else:
                                try:
                                    tel_prospect = str(int(float(tel_prospect_raw)))
                                except (ValueError, TypeError):
                                    tel_prospect = str(tel_prospect_raw).strip()
                            
                            entreprise = Entreprise.objects.create(
                                nom=nom_societe,
                                telephone=tel_prospect,
                                date=timezone.now().date(),
                                is_Prospect=True,
                                is_CLT=False,
                                is_Concurent=False,
                                email=email_prospect or "",
                                societe=societe,
                                adresse="",
                                secteur_activite=""
                            )
                
                etat = 'planifie' 
                date_realiser = None
                etat_rv_raw = row.get('Etat du rendez-vous')
                if not pd.isna(etat_rv_raw):
                    etat_rv = str(etat_rv_raw).strip()
                    if 'Non Realiser' in etat_rv or 'Non Réaliser' in etat_rv:
                        etat = 'annule'
                    elif 'Realiser' in etat_rv or 'Réaliser' in etat_rv:
                        etat = 'termine'
                        date_realiser = date_planifie or None
                
                compte_rendu = str(row.get('Sujet', '')).strip() if not pd.isna(row.get('Sujet')) else ''
                
                client_name = entreprise.nom if entreprise else "Inconnu"
                sujet = f"Rendez-vous {client_name} {date_planifie.strftime('%d/%m/%Y') if date_planifie else ''}"
                
                Action.objects.create(
                    date_heure=timezone.now(),
                    date_heure_planifie=date_planifie,
                    date_heure_realiser=date_realiser,
                    compte_rendu=compte_rendu,
                    notes=notes,
                    sujet=sujet,
                    is_Appel=False,
                    is_Email=False,
                    is_RV=True,
                    etat=etat,
                    societe=societe,
                    entreprise=entreprise,
                    created_by=created_by,
                    pilote=pilote_rc
                )
                
                success_count += 1
                
            except Exception as e:
                error_count += 1
                errors.append(f"Ligne {idx+3}: {str(e)}")
                logger.error(f"Erreur importation RV ligne {idx+3}: {e}")
        
        if success_count > 0:
            messages.success(request, f"{success_count} rendez-vous importés avec succès.")
        if error_count > 0:
            messages.warning(request, f"{error_count} erreurs lors de l'importation. Détails: {'; '.join(errors[:5])}")
    
    except Exception as e:
        logger.error(f"Erreur générale importation RV: {e}")
        messages.error(request, f"Erreur lors de l'importation: {str(e)}")
    
    return redirect('prospection:rv_list')


@login_required
def import_calls(request):
    """Importe des appels depuis un fichier Excel"""
    if not (request.user.is_superuser or request.user.is_RO):
        messages.error(request, "Vous n'avez pas les permissions nécessaires.")
        return redirect('prospection:call_list')
    
    if request.method != 'POST':
        return redirect('prospection:call_list')
    
    excel_file = request.FILES.get('excel_file')
    societe_id = request.POST.get('societe_id')
    
    if not excel_file or not societe_id:
        messages.error(request, "Fichier Excel et filiale sont requis.")
        return redirect('prospection:call_list')
    
    try:
        societe = Societe.objects.get(id=societe_id)
    except Societe.DoesNotExist:
        messages.error(request, "Filiale introuvable.")
        return redirect('prospection:call_list')
    
    pilote_rc = Utilisateur.objects.filter(is_RC=True, societe=societe).first()
    
    success_count = 0
    error_count = 0
    errors = []
    
    try:
        df = pd.read_excel(excel_file, sheet_name='Appel', header=1,
                          dtype={'N° Compte tiers': str})
        
        required_cols = ['Date', 'Appel', 'Appeler par', 'N° Compte tiers', 'Société', 
                        'Nom contact', 'Poste', 'Tel', 'Mail', 'Reponse', 'Volet', 'Feedback']
        
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            messages.error(request, f"Colonnes manquantes dans le fichier: {', '.join(missing_cols)}")
            return redirect('prospection:call_list')
        
        for idx, row in df.iterrows():
            try:
                date_planifie = parse_excel_date(row.get('Date')) or None
                
                volet_check = row.get('Volet')
                if pd.isna(volet_check) or not str(volet_check).strip():
                    error_count += 1
                    errors.append(f"Ligne {idx+3}: Volet manquant")
                    continue
                
                created_by = None
                appeler_par = str(row.get('Appeler par', '')).strip()
                if appeler_par and not pd.isna(row.get('Appeler par')):
                    created_by = Utilisateur.objects.filter(username=appeler_par).first()
                
                entreprise = None
                volet = str(row.get('Volet', '')).strip()
                num_compte_raw = row.get('N° Compte tiers')
                num_compte = str(num_compte_raw).strip() if not pd.isna(num_compte_raw) and str(num_compte_raw).strip() != 'nan' else ''
                nom_societe = str(row.get('Société', '')).strip()
                
                notes_parts = []
                if not pd.isna(row.get('Nom contact')):
                    notes_parts.append(f"Contact: {row.get('Nom contact')}")
                if not pd.isna(row.get('Poste')):
                    notes_parts.append(f"Poste: {row.get('Poste')}")
                if not pd.isna(row.get('Tel')):
                    tel_raw = row.get('Tel')
                    try:
                        tel_format = str(int(float(tel_raw)))
                    except (ValueError, TypeError):
                        tel_format = str(tel_raw)
                    notes_parts.append(f"Tel: {tel_format}")
                if not pd.isna(row.get('Mail')):
                    notes_parts.append(f"Mail: {row.get('Mail')}")
                notes = " | ".join(notes_parts)
                
                if 'Developpement' in volet or 'Développement' in volet:
                    if num_compte:
                        entreprise = Entreprise.objects.filter(
                            num_compte=num_compte,
                            societe=societe,
                            is_CLT=True
                        ).first()
                        
                        if not entreprise:
                            try:
                                sage = fetch_client_from_sage(societe.id, num_compte) or {}
                                
                                entreprise = Entreprise.objects.create(
                                    num_compte=(sage.get("NumeroTiers") or (num_compte or None)),
                                    nom=sage.get("Intitule") or "Inconnu",
                                    email=SAFE_EMAIL(sage.get("Email")),
                                    adresse=sage.get("Adresse", "") or "",
                                    telephone=sage.get("Telephone", "") or "",
                                    secteur_activite="",
                                    societe=societe,
                                    is_CLT=True,
                                    is_Prospect=False,
                                    is_Concurent=False
                                )
                            except Exception as e:
                                logger.error(f"Erreur création client ligne {idx+3}: {e}")
                
                elif 'Prospection' in volet:
                    if nom_societe:
                        entreprise = Entreprise.objects.filter(
                            nom__iexact=nom_societe,
                            societe=societe,
                            is_Prospect=True,
                            is_CLT=False
                        ).first()
                        
                        if not entreprise:
                            email_prospect = str(row.get('Mail', '')).strip() if not pd.isna(row.get('Mail')) else ''
                            email_prospect = email_prospect.strip("'\"")
                            
                            tel_prospect_raw = row.get('Tel')
                            if pd.isna(tel_prospect_raw):
                                tel_prospect = ''
                            else:
                                try:
                                    tel_prospect = str(int(float(tel_prospect_raw)))
                                except (ValueError, TypeError):
                                    tel_prospect = str(tel_prospect_raw).strip()
                            
                            entreprise = Entreprise.objects.create(
                                nom=nom_societe,
                                telephone=tel_prospect,
                                date=timezone.now().date(),
                                is_Prospect=True,
                                is_CLT=False,
                                is_Concurent=False,
                                email=email_prospect or "",
                                societe=societe,
                                adresse="",
                                secteur_activite=""
                            )
                
                etat = 'planifie'
                date_realiser = None
                reponse_raw = row.get('Reponse')
                if not pd.isna(reponse_raw):
                    reponse = str(reponse_raw).strip().lower()
                    if 'oui' in reponse:
                        etat = 'reussi'
                        date_realiser = date_planifie or None
                    else:
                        etat = 'non_reussi'
                
                compte_rendu = str(row.get('Feedback', '')).strip() if not pd.isna(row.get('Feedback')) else ''
                
                client_name = entreprise.nom if entreprise else "Inconnu"
                sujet = f"Prise de rendez-vous {client_name}"
                
                Action.objects.create(
                    date_heure=timezone.now(),
                    date_heure_planifie=date_planifie,
                    date_heure_realiser=date_realiser,
                    compte_rendu=compte_rendu,
                    notes=notes,
                    sujet=sujet,
                    is_Appel=True,
                    is_Email=False,
                    is_RV=False,
                    etat=etat,
                    societe=societe,
                    entreprise=entreprise,
                    created_by=created_by,
                    pilote=pilote_rc
                )
                
                success_count += 1
                
            except Exception as e:
                error_count += 1
                errors.append(f"Ligne {idx+3}: {str(e)}")
                logger.error(f"Erreur importation appel ligne {idx+3}: {e}")
        
        if success_count > 0:
            messages.success(request, f"{success_count} appels importés avec succès.")
        if error_count > 0:
            messages.warning(request, f"{error_count} erreurs lors de l'importation. Détails: {'; '.join(errors[:5])}")
    
    except Exception as e:
        logger.error(f"Erreur générale importation appels: {e}")
        messages.error(request, f"Erreur lors de l'importation: {str(e)}")
    
    return redirect('prospection:call_list')
