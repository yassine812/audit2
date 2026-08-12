"""Vues du module gestion documentaire."""

from __future__ import annotations

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.files.base import ContentFile
from django.db import DatabaseError, transaction
from django.db.models import Avg, Q, Sum
import json
import unicodedata

from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.views.generic import CreateView, DetailView, TemplateView, UpdateView, View

from .forms import (
    DocumentFilterForm,
    DocumentForm,
    DossierDocumentaireForm,
    DossierParametresForm,
    FichierBibliothequeForm,
    NouvelleRegleAccesFormSet,
    RegleAccesDossierFormSet,
    VersionDocumentForm,
    ProcessusForm,
)
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

from .models import (
    Document,
    DossierDocumentaire,
    FichierBibliotheque,
    Indicateur,
    MesureIndicateur,
    ObjectifIndicateur,
    Processus,
    RealiseConsolideIndicateur,
    RegleAccesDossier,
    ValidationDocument,
    VersionDocument,
    ComposanteIndicateur,
    ValeurComposanteIndicateur,
)
from .utils_formule import evaluer_formule_securisee, valider_formule_securisee
from .permissions import (
    DirectionOuHabiliteRequiredMixin,
    DocumentVisibilityQuerysetMixin,
    PiloteOuQSERequiredMixin,
    QSERequiredMixin,
    ModuleDocumentaireRequiredMixin,
    ModificationBibliothequeRequiredMixin,
    TdbAccesRequiredMixin,
    direction_ou_habilite_required,
    filter_documents_for_user,
    is_direction_ou_habilite,
    is_pilote_ou_qse,
    is_qse,
    pilote_ou_qse_required,
    modification_bibliotheque_required,
    module_documentaire_required,
    peut_modifier_bibliotheque,
    qse_required,
    tdb_acces_required,
    processus_dans_perimetre_smqs,
    processus_perimetre_smqs_qs,
    user_peut_utiliser_indicateurs_smqs,
    user_peut_voir_tdb,
)


def _get_latest_version(document: Document):
    return document.versions.order_by("-version", "-indice", "-date_creation").first()


def _transition_document_with_trace(document: Document, nouveau_statut: str, user, commentaire: str = ""):
    """Applique une transition au document ou à sa version active si disponible."""
    version = _get_latest_version(document)
    if version and version.statut == document.statut:
        version.transitionner_statut(nouveau_statut, utilisateur=user, commentaire=commentaire)
    else:
        document.transitionner_statut(nouveau_statut, utilisateur=user, commentaire=commentaire)


class DocumentListView(ModuleDocumentaireRequiredMixin, TemplateView):
    """Explorateur de dossiers et fichiers de la bibliothèque."""

    template_name = "gestion_documentaire/document_list.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dossier_id = self.kwargs.get("dossier_id")
        if dossier_id:
            dossier = get_object_or_404(DossierDocumentaire, pk=dossier_id)
        elif self.kwargs.get("racine"):
            dossier = None
        else:
            dossier = DossierDocumentaire.objects.filter(
                parent__isnull=True, nom="05 - Gestion documentaire"
            ).first()

        if dossier and not dossier.utilisateur_autorise(self.request.user, "lire"):
            raise PermissionDenied("Vous n'avez pas accès à ce dossier.")

        recherche = self.request.GET.get("q", "").strip()
        sous_dossiers = list(DossierDocumentaire.objects.filter(parent=dossier))
        sous_dossiers = [d for d in sous_dossiers if d.utilisateur_autorise(self.request.user, "lire")]
        for enfant in sous_dossiers:
            enfant.peut_editer = (
                peut_modifier_bibliotheque(self.request.user)
                and enfant.utilisateur_autorise(self.request.user, "modifier")
            )
            enfant.peut_supprimer = (
                peut_modifier_bibliotheque(self.request.user)
                and enfant.utilisateur_autorise(self.request.user, "modifier")
            )
        fichiers = FichierBibliotheque.objects.filter(dossier=dossier).select_related("ajoute_par")
        if recherche:
            sous_dossiers = [d for d in sous_dossiers if recherche.lower() in d.nom.lower()]
            fichiers = fichiers.filter(nom__icontains=recherche)

        context.update({
            "dossier_actuel": dossier,
            "fil_ariane": dossier.get_ancestors() if dossier else [],
            "sous_dossiers": sous_dossiers,
            "fichiers": fichiers,
            "dossier_form": DossierParametresForm(),
            "nouveau_regles_formset": NouvelleRegleAccesFormSet(
                queryset=RegleAccesDossier.objects.none(),
                prefix="new_rules",
            ),
            "fichier_form": FichierBibliothequeForm(),
            "peut_modifier": peut_modifier_bibliotheque(self.request.user),
            "peut_ajouter": peut_modifier_bibliotheque(self.request.user) and (
                not dossier or dossier.utilisateur_autorise(self.request.user, "modifier")
            ),
            "peut_editer": peut_modifier_bibliotheque(self.request.user) and (
                not dossier or dossier.utilisateur_autorise(self.request.user, "modifier")
            ),
            "peut_supprimer": peut_modifier_bibliotheque(self.request.user) and (
                not dossier or dossier.utilisateur_autorise(self.request.user, "modifier")
            ),
            "peut_telecharger": not dossier or dossier.utilisateur_autorise(
                self.request.user, "telecharger"
            ),
            "vue": self.request.GET.get("vue", "liste"),
            "recherche": recherche,
        })
        return context


@method_decorator([login_required, modification_bibliotheque_required], name="dispatch")
class DossierCreateView(View):
    def post(self, request, dossier_id=None, *args, **kwargs):
        parent = get_object_or_404(DossierDocumentaire, pk=dossier_id) if dossier_id else None
        if parent and not parent.utilisateur_autorise(request.user, "modifier"):
            raise PermissionDenied("Vous n'avez pas le droit d'ajouter dans ce dossier.")
        form = DossierParametresForm(request.POST)
        regles_formset = NouvelleRegleAccesFormSet(
            request.POST,
            queryset=RegleAccesDossier.objects.none(),
            prefix="new_rules",
        )
        regles_valides = regles_formset.is_valid()
        regles_actives = [
            regle_form for regle_form in regles_formset.forms
            if regle_form.has_changed() and not regle_form.cleaned_data.get("DELETE")
        ] if regles_valides else []
        if form.is_valid() and regles_valides and (
            not form.cleaned_data.get("acces_restreint") or regles_actives
        ):
            dossier = form.save(commit=False)
            dossier.parent = parent
            dossier.cree_par = request.user
            dossier.save()
            for regle_form in regles_actives:
                regle = regle_form.save(commit=False)
                regle.dossier = dossier
                regle.save()
                regle_form.instance = regle
                regle_form.save_m2m()
            messages.success(request, f'Dossier « {dossier.nom} » créé.')
            if dossier.acces_restreint:
                return redirect("gestion_documentaire:modifier_dossier", pk=dossier.pk)
        else:
            messages.error(
                request,
                "Le dossier n'a pas pu être créé. Vérifiez les informations et ajoutez au moins une règle si l'accès est limité.",
            )
        if parent:
            return redirect("gestion_documentaire:dossier_detail", dossier_id=parent.pk)
        return redirect("gestion_documentaire:document_list")


def _retour_dossier(dossier):
    if dossier:
        return redirect("gestion_documentaire:dossier_detail", dossier_id=dossier.pk)
    return redirect("gestion_documentaire:bibliotheque_racine")


@login_required
@modification_bibliotheque_required
@require_POST
def uploader_fichiers(request, dossier_id=None):
    dossier = get_object_or_404(DossierDocumentaire, pk=dossier_id) if dossier_id else None
    if dossier and not dossier.utilisateur_autorise(request.user, "modifier"):
        raise PermissionDenied("Vous n'avez pas le droit d'ajouter dans ce dossier.")
    fichiers = request.FILES.getlist("fichier")
    if not fichiers:
        messages.error(request, "Sélectionnez au moins un fichier.")
        return _retour_dossier(dossier)
    ajoutes = 0
    for upload in fichiers:
        form = FichierBibliothequeForm(files={"fichier": upload})
        if not form.is_valid():
            messages.error(request, f"{upload.name} : {form.errors.as_text()}")
            continue
        if FichierBibliotheque.objects.filter(dossier=dossier, nom=upload.name).exists():
            messages.error(request, f"{upload.name} existe déjà dans ce dossier.")
            continue
        FichierBibliotheque.objects.create(
            dossier=dossier, fichier=upload, nom=upload.name, taille=upload.size,
            type_mime=upload.content_type or "", ajoute_par=request.user,
        )
        ajoutes += 1
    if ajoutes:
        messages.success(request, f"{ajoutes} fichier(s) ajouté(s).")
    return _retour_dossier(dossier)


@login_required
@modification_bibliotheque_required
def modifier_dossier(request, pk):
    dossier = get_object_or_404(DossierDocumentaire, pk=pk)
    if not dossier.utilisateur_autorise(request.user, "modifier"):
        raise PermissionDenied("Vous n'avez pas le droit de modifier ce dossier.")
    if request.method == "GET":
        return render(
            request,
            "gestion_documentaire/dossier_form.html",
            {
                "form": DossierParametresForm(instance=dossier),
                "regles_formset": RegleAccesDossierFormSet(instance=dossier),
                "dossier": dossier,
            },
        )
    ancien_nom = dossier.nom
    form = DossierParametresForm(request.POST, instance=dossier)
    regles_formset = RegleAccesDossierFormSet(request.POST, instance=dossier)
    if form.is_valid() and regles_formset.is_valid():
        form.save()
        regles_formset.save()
        if ancien_nom != dossier.nom:
            def deplacer_contenu(courant):
                for fichier in courant.fichiers.all():
                    ancien_chemin = fichier.fichier.name
                    contenu = ContentFile(fichier.fichier.read())
                    fichier.fichier.save(fichier.nom, contenu, save=True)
                    if ancien_chemin != fichier.fichier.name:
                        fichier.fichier.storage.delete(ancien_chemin)
                for enfant in courant.sous_dossiers.all():
                    deplacer_contenu(enfant)
            deplacer_contenu(dossier)
        messages.success(request, "Dossier modifié.")
    else:
        return render(
            request,
            "gestion_documentaire/dossier_form.html",
            {"form": form, "regles_formset": regles_formset, "dossier": dossier},
            status=400,
        )
    return redirect("gestion_documentaire:dossier_detail", dossier_id=dossier.pk)


@login_required
@modification_bibliotheque_required
@require_POST
def supprimer_dossier(request, pk):
    dossier = get_object_or_404(DossierDocumentaire, pk=pk)
    if not dossier.utilisateur_autorise(request.user, "modifier"):
        raise PermissionDenied("Vous n'avez pas le droit de supprimer ce dossier.")
    parent = dossier.parent
    def supprimer_contenu(courant):
        for enfant in courant.sous_dossiers.all():
            supprimer_contenu(enfant)
        for fichier in courant.fichiers.all():
            fichier.delete()
    supprimer_contenu(dossier)
    dossier.delete()
    messages.success(request, "Dossier supprimé.")
    return _retour_dossier(parent)


@login_required
@modification_bibliotheque_required
@require_POST
def renommer_fichier(request, pk):
    objet = get_object_or_404(FichierBibliotheque, pk=pk)
    if objet.dossier and not objet.dossier.utilisateur_autorise(request.user, "modifier"):
        raise PermissionDenied("Vous n'avez pas le droit de modifier ce fichier.")
    nouveau_nom = request.POST.get("nom", "").strip()
    extension_originale = objet.extension
    if not nouveau_nom:
        messages.error(request, "Le nom est obligatoire.")
    else:
        if not nouveau_nom.lower().endswith(extension_originale):
            nouveau_nom += extension_originale
        if FichierBibliotheque.objects.filter(dossier=objet.dossier, nom=nouveau_nom).exclude(pk=pk).exists():
            messages.error(request, "Un fichier porte déjà ce nom.")
        else:
            ancien_chemin = objet.fichier.name
            contenu = ContentFile(objet.fichier.read())
            objet.nom = nouveau_nom
            objet.fichier.save(nouveau_nom, contenu, save=True)
            if ancien_chemin != objet.fichier.name:
                objet.fichier.storage.delete(ancien_chemin)
            objet.save(update_fields=["nom", "date_modification"])
            messages.success(request, "Fichier renommé.")
    return _retour_dossier(objet.dossier)


@login_required
@modification_bibliotheque_required
@require_POST
def supprimer_fichier(request, pk):
    objet = get_object_or_404(FichierBibliotheque, pk=pk)
    if objet.dossier and not objet.dossier.utilisateur_autorise(request.user, "modifier"):
        raise PermissionDenied("Vous n'avez pas le droit de supprimer ce fichier.")
    dossier = objet.dossier
    objet.delete()
    messages.success(request, "Fichier supprimé.")
    return _retour_dossier(dossier)


@login_required
@module_documentaire_required
def telecharger_fichier_bibliotheque(request, pk):
    objet = get_object_or_404(FichierBibliotheque.objects.select_related("dossier"), pk=pk)
    if objet.dossier and not objet.dossier.utilisateur_autorise(request.user, "telecharger"):
        return HttpResponse(status=403)
    return FileResponse(
        objet.fichier.open("rb"), as_attachment=True, filename=objet.nom,
        content_type=objet.type_mime or "application/octet-stream",
    )


@login_required
@module_documentaire_required
def visualiser_fichier_bibliotheque(request, pk):
    """Aperçu sécurisé : PDF natif, fiche moderne pour les formats bureautiques."""
    objet = get_object_or_404(FichierBibliotheque.objects.select_related("dossier"), pk=pk)
    if objet.dossier and not objet.dossier.utilisateur_autorise(request.user, "lire"):
        return HttpResponse(status=403)

    if objet.extension == ".pdf":
        try:
            response = FileResponse(
                objet.fichier.open("rb"),
                as_attachment=False,
                filename=objet.nom,
                content_type="application/pdf",
            )
            response["Content-Disposition"] = f'inline; filename="{objet.nom}"'
            response["X-Content-Type-Options"] = "nosniff"
            return response
        except FileNotFoundError as exc:
            raise Http404("Fichier introuvable.") from exc

    peut_telecharger = (
        objet.dossier is None
        or objet.dossier.utilisateur_autorise(request.user, "telecharger")
    )
    response = render(
        request,
        "gestion_documentaire/fichier_preview.html",
        {"fichier": objet, "peut_telecharger": peut_telecharger},
    )
    response["Content-Security-Policy"] = (
        "default-src 'none'; style-src 'unsafe-inline'; "
        "img-src 'self'; frame-ancestors 'self'; base-uri 'none'; form-action 'none'"
    )
    response["X-Content-Type-Options"] = "nosniff"
    return response


class DocumentDetailView(ModuleDocumentaireRequiredMixin, DocumentVisibilityQuerysetMixin, DetailView):
    """Détail documentaire avec historique des versions."""

    model = Document
    template_name = "gestion_documentaire/document_detail.html"
    context_object_name = "document"

    def get_queryset(self):
        return (
            super()
            .get_queryset()
            .select_related("processus_service", "cree_par")
            .prefetch_related("versions__redacteur", "validations__utilisateur")
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        document = self.object
        versions = document.versions.order_by("-version", "-indice", "-date_creation")
        context["versions"] = versions
        context["version_active"] = versions.filter(statut=Document.Statut.APPLICABLE).first() or versions.first()
        context["version_form"] = VersionDocumentForm()
        context["is_qse_user"] = is_qse(self.request.user)
        context["is_pilote_ou_qse_user"] = is_pilote_ou_qse(self.request.user)
        context["is_direction_ou_habilite_user"] = is_direction_ou_habilite(self.request.user)
        return context


class DocumentCreateView(ModificationBibliothequeRequiredMixin, CreateView):
    """Création d'un document en brouillon."""

    model = Document
    form_class = DocumentForm
    template_name = "gestion_documentaire/document_form.html"

    def get_initial(self):
        initial = super().get_initial()
        dossier_id = self.request.GET.get("dossier")
        if dossier_id and DossierDocumentaire.objects.filter(pk=dossier_id).exists():
            initial["dossier"] = dossier_id
        return initial

    def form_valid(self, form):
        form.instance.cree_par = self.request.user
        form.instance.statut = Document.Statut.BROUILLON
        response = super().form_valid(form)
        ValidationDocument.objects.create(
            document=self.object,
            type_action=ValidationDocument.TypeAction.REDACTION,
            utilisateur=self.request.user,
            commentaire="Création du brouillon documentaire.",
            ancien_statut="",
            nouveau_statut=Document.Statut.BROUILLON,
        )
        messages.success(self.request, "Document créé en brouillon.")
        return response

    def get_success_url(self):
        return reverse("gestion_documentaire:document_detail", kwargs={"pk": self.object.pk})


class DocumentUpdateView(ModificationBibliothequeRequiredMixin, UpdateView):
    """Mise à jour du contenu documentaire."""

    model = Document
    form_class = DocumentForm
    template_name = "gestion_documentaire/document_form.html"

    def get_queryset(self):
        return Document.objects.filter(est_supprime=False)

    def form_valid(self, form):
        response = super().form_valid(form)
        ValidationDocument.objects.create(
            document=self.object,
            type_action=ValidationDocument.TypeAction.REDACTION,
            utilisateur=self.request.user,
            commentaire="Mise à jour du contenu documentaire.",
            ancien_statut=self.object.statut,
            nouveau_statut=self.object.statut,
        )
        messages.success(self.request, "Document mis à jour.")
        return response

    def get_success_url(self):
        return reverse("gestion_documentaire:document_detail", kwargs={"pk": self.object.pk})


class DashboardQSEView(ModuleDocumentaireRequiredMixin, TemplateView):
    """Vue d'ensemble de la bibliothèque documentaire."""

    template_name = "gestion_documentaire/dashboard_qse.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dossiers = [
            dossier for dossier in DossierDocumentaire.objects.select_related("parent")
            if dossier.utilisateur_autorise(self.request.user, "lire")
        ]
        dossier_ids = [dossier.pk for dossier in dossiers]
        fichiers = list(FichierBibliotheque.objects.filter(
            Q(dossier_id__in=dossier_ids) | Q(dossier__isnull=True)
        ).select_related("dossier", "ajoute_par"))
        for fichier in fichiers:
            fichier.peut_telecharger = (
                fichier.dossier is None
                or fichier.dossier.utilisateur_autorise(self.request.user, "telecharger")
            )
        context["total_dossiers"] = len(dossiers)
        context["total_fichiers"] = len(fichiers)
        context["total_taille"] = sum(f.taille for f in fichiers)
        context["dossiers_restraints"] = sum(1 for d in dossiers if d.acces_restreint)
        context["fichiers_recents"] = sorted(
            fichiers, key=lambda f: f.date_modification, reverse=True
        )[:10]
        context["dossiers_recents"] = sorted(
            dossiers, key=lambda d: d.date_modification, reverse=True
        )[:8]
        context["peut_modifier"] = peut_modifier_bibliotheque(self.request.user)
        return context


@method_decorator([login_required, modification_bibliotheque_required], name="dispatch")
class CreerNouvelleVersionView(View):
    """Crée une nouvelle version documentaire avec incrément version/indice."""

    def post(self, request, pk, *args, **kwargs):
        document = get_object_or_404(Document, pk=pk, est_supprime=False)
        form = VersionDocumentForm(request.POST, request.FILES)
        if not form.is_valid():
            messages.error(request, "Impossible de créer la nouvelle version.")
            versions = document.versions.order_by("-version", "-indice", "-date_creation")
            context = {
                "document": document,
                "versions": versions,
                "version_active": versions.filter(statut=Document.Statut.APPLICABLE).first() or versions.first(),
                "version_form": form,
                "is_qse_user": is_qse(request.user),
                "is_pilote_ou_qse_user": is_pilote_ou_qse(request.user),
                "is_direction_ou_habilite_user": is_direction_ou_habilite(request.user),
            }
            return render(request, "gestion_documentaire/document_detail.html", context, status=400)

        derniere_version = _get_latest_version(document)
        if derniere_version is None:
            next_version = 1
            next_indice = 0
        elif form.cleaned_data["type_increment"] == VersionDocumentForm.TypeIncrement.MAJEUR:
            next_version = derniere_version.version + 1
            next_indice = 0
        else:
            next_version = derniere_version.version
            next_indice = derniere_version.indice + 1

        version = form.save(commit=False)
        version.document = document
        version.version = next_version
        version.indice = next_indice
        version.redacteur = request.user
        version.statut = Document.Statut.BROUILLON
        version.copie_non_maitrisee = True
        version.save()

        ValidationDocument.objects.create(
            document=document,
            version_document=version,
            type_action=ValidationDocument.TypeAction.REDACTION,
            utilisateur=request.user,
            commentaire=form.cleaned_data.get("resume_changements", "Nouvelle version créée."),
            ancien_statut=document.statut,
            nouveau_statut=document.statut,
        )

        messages.success(
            request,
            f"Nouvelle version créée: v{version.version}.{version.indice}.",
        )
        return redirect("gestion_documentaire:document_detail", pk=document.pk)


@login_required
@modification_bibliotheque_required
def soumettre_verification(request, pk):
    """Transition brouillon -> en_verification."""
    document = get_object_or_404(Document, pk=pk, est_supprime=False)
    try:
        _transition_document_with_trace(
            document,
            Document.Statut.EN_VERIFICATION,
            user=request.user,
            commentaire="Soumission en vérification.",
        )
    except ValidationError as exc:
        messages.error(request, "; ".join(exc.messages))
    else:
        messages.success(request, "Document soumis en vérification.")
    return redirect("gestion_documentaire:document_detail", pk=document.pk)


@login_required
@modification_bibliotheque_required
def approuver_document(request, pk):
    """Fait progresser le document jusqu'à l'état applicable."""
    document = get_object_or_404(Document, pk=pk, est_supprime=False)

    try:
        if document.statut == Document.Statut.EN_VERIFICATION:
            _transition_document_with_trace(
                document,
                Document.Statut.EN_APPROBATION,
                user=request.user,
                commentaire="Document passé en approbation.",
            )
            # Recharger le document après la première transition.
            document.refresh_from_db()

        _transition_document_with_trace(
            document,
            Document.Statut.APPLICABLE,
            user=request.user,
            commentaire="Document approuvé et rendu applicable.",
        )

        ValidationDocument.objects.create(
            document=document,
            type_action=ValidationDocument.TypeAction.APPROBATION,
            utilisateur=request.user,
            commentaire="Validation finale et mise en application.",
            ancien_statut=Document.Statut.EN_APPROBATION,
            nouveau_statut=Document.Statut.APPLICABLE,
        )
    except ValidationError as exc:
        messages.error(request, "; ".join(exc.messages))
    else:
        messages.success(request, "Document approuvé et publié.")

    return redirect("gestion_documentaire:document_detail", pk=document.pk)


@login_required
@modification_bibliotheque_required
def archiver_document(request, pk):
    """Archive un document applicable."""
    document = get_object_or_404(Document, pk=pk, est_supprime=False)
    try:
        _transition_document_with_trace(
            document,
            Document.Statut.ARCHIVE,
            user=request.user,
            commentaire="Archivage documentaire.",
        )
        ValidationDocument.objects.create(
            document=document,
            type_action=ValidationDocument.TypeAction.ARCHIVAGE,
            utilisateur=request.user,
            commentaire="Document archivé via action QSE.",
            ancien_statut=Document.Statut.APPLICABLE,
            nouveau_statut=Document.Statut.ARCHIVE,
        )
    except ValidationError as exc:
        messages.error(request, "; ".join(exc.messages))
    else:
        messages.success(request, "Document archivé.")

    return redirect("gestion_documentaire:document_detail", pk=document.pk)


@login_required
@module_documentaire_required
def exporter_document_pdf(request, pk):
    """Export PDF non modifiable (cohérent avec WeasyPrint déjà utilisé)."""
    document = get_object_or_404(Document.objects.filter(est_supprime=False), pk=pk)

    # Contrôle d'accès équivalent à la visibilité des documents.
    visible = filter_documents_for_user(Document.objects.filter(pk=pk), request.user)
    if not visible.exists():
        return HttpResponse(status=403)

    version = (
        document.versions.filter(statut=Document.Statut.APPLICABLE)
        .order_by("-version", "-indice", "-date_creation")
        .first()
        or _get_latest_version(document)
    )
    version_label = f"v{version.version}.{version.indice}" if version else "N/A"

    html = f"""
    <html>
    <head>
      <meta charset="utf-8" />
      <style>
        body {{ font-family: Arial, sans-serif; font-size: 12px; margin: 24px; }}
        h1 {{ font-size: 18px; margin-bottom: 6px; }}
        .meta {{ margin-bottom: 12px; color: #333; }}
        .watermark {{
          position: fixed;
          top: 40%;
          left: 10%;
          transform: rotate(-28deg);
          font-size: 54px;
          color: rgba(180, 0, 0, 0.12);
          letter-spacing: 2px;
          z-index: -1;
          white-space: nowrap;
        }}
        .section {{ margin-bottom: 10px; }}
      </style>
    </head>
    <body>
        <div class="watermark">COPIE NON MAITRISEE</div>
      <h1>{document.code_documentaire} - {document.titre}</h1>
        <div class="meta">Statut: {document.get_statut_display()} | Date export: {timezone.localdate().isoformat()}</div>
        <div class="meta">Version: {version_label}</div>
        <div class="section"><strong>Objet:</strong><br/>{document.objet or ''}</div>
        <div class="section"><strong>Domaine d'application:</strong><br/>{document.domaine_application or ''}</div>
        <div class="section"><strong>Règles / étapes:</strong><br/>{document.regles_etapes or ''}</div>
        <div class="section"><strong>Responsabilités:</strong><br/>{document.responsabilites or ''}</div>
    </body>
    </html>
    """

    try:
        from weasyprint import HTML
    except Exception:
        return HttpResponse("WeasyPrint est requis pour l'export PDF.", status=500)

    pdf_bytes = HTML(string=html).write_pdf(pdf_version="1.4")
    filename = f"{document.code_documentaire}_{timezone.localdate().isoformat()}.pdf"
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
@module_documentaire_required
def fichier_version(request, pk, nature):
    """Sert un fichier après contrôle de visibilité, en aperçu ou téléchargement."""
    version = get_object_or_404(
        VersionDocument.objects.select_related("document"),
        pk=pk,
        est_supprime=False,
    )
    if not filter_documents_for_user(
        Document.objects.filter(pk=version.document_id), request.user
    ).exists():
        return HttpResponse(status=403)

    if nature == "reference":
        fichier = version.fichier_reference
        content_type = "application/pdf"
    elif nature == "editable":
        fichier = version.fichier_editable
        content_type = "application/octet-stream"
    else:
        raise Http404
    if not fichier:
        raise Http404

    inline = nature == "reference" and request.GET.get("download") != "1"
    try:
        return FileResponse(
            fichier.open("rb"),
            as_attachment=not inline,
            filename=fichier.name.rsplit("/", 1)[-1],
            content_type=content_type,
        )
    except FileNotFoundError as exc:
        raise Http404("Fichier introuvable.") from exc


# ============================================================
# MODULE TABLEAU DE BORD SMQS
# Vues et fonctions utilitaires pour le Tableau de Bord
# ============================================================


def obtenir_mois_periode(type_periode: str, periode: Optional[int]) -> List[int]:
    """
    Retourne la liste des mois (1 à 12) correspondant au filtre temporel sélectionné.
    Types de période gérés : 'annee_complete', 'trimestre', 'semestre', 'mois'.
    """
    if type_periode == "trimestre":
        try:
            t = int(periode or 1)
        except (ValueError, TypeError):
            t = 1
        t = max(1, min(4, t))
        debut = (t - 1) * 3 + 1
        return list(range(debut, debut + 3))

    elif type_periode == "semestre":
        try:
            s = int(periode or 1)
        except (ValueError, TypeError):
            s = 1
        s = max(1, min(2, s))
        debut = (s - 1) * 6 + 1
        return list(range(debut, debut + 6))

    elif type_periode == "mois":
        try:
            m = int(periode or 1)
        except (ValueError, TypeError):
            m = 1
        m = max(1, min(12, m))
        return [m]

    # 'annee_complete' par défaut
    return list(range(1, 13))


def calculer_agregation_indicateur(
    indicateur: Indicateur,
    mesures: List[MesureIndicateur],
    type_periode: str = "annee_complete"
) -> Optional[Decimal]:
    """
    Calcule l'agrégation numérique en Decimal pour une liste de mesures données.
    RÈGLE STRICTE MOIS=NONE :
    - Pour 'annee_complete', les mesures mensuelles et annuelles (mois=None) sont incluses.
    - Pour 'trimestre', 'semestre' ou 'mois', les mesures annuelles (mois=None) sont exclues.
    """
    if type_periode != "annee_complete":
        mesures_effectives = [m for m in mesures if m.mois is not None]
    else:
        mesures_effectives = list(mesures)

    if not mesures_effectives:
        return None

    mode = indicateur.mode_agregation

    if mode == Indicateur.ModeAgregation.SOMME:
        total = sum((m.valeur for m in mesures_effectives), Decimal("0"))
        return total.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)

    elif mode == Indicateur.ModeAgregation.MOYENNE:
        total = sum((m.valeur for m in mesures_effectives), Decimal("0"))
        moyenne = total / Decimal(len(mesures_effectives))
        return moyenne.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)

    elif mode == Indicateur.ModeAgregation.MINIMUM:
        val_min = min(m.valeur for m in mesures_effectives)
        return val_min.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)

    elif mode == Indicateur.ModeAgregation.MAXIMUM:
        val_max = max(m.valeur for m in mesures_effectives)
        return val_max.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)

    elif mode == Indicateur.ModeAgregation.NOMBRE:
        count_val = Decimal(len(mesures_effectives))
        return count_val.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)

    elif mode == Indicateur.ModeAgregation.DERNIERE_VALEUR:
        mesures_triees = sorted(
            mesures_effectives,
            key=lambda m: (m.mois if m.mois is not None else 0),
            reverse=True
        )
        return mesures_triees[0].valeur

    return None


def calculer_fenetre_12_mois_glissants(
    indicateur: Indicateur,
    mois_ref: int,
    annee_ref: int,
    mesures_dict: Dict[Tuple[int, int], Decimal]
) -> Optional[Decimal]:
    """
    Calcule l'agrégation automatique pour un mois de référence (mois_ref, annee_ref)
    en interrogeant les 12 mois consécutifs allant de (annee_ref-1, mois_ref+1) à (annee_ref, mois_ref).
    """
    valeurs_fenetre = []
    for i in range(12):
        m = mois_ref - i
        y = annee_ref
        if m <= 0:
            m += 12
            y -= 1
        val = mesures_dict.get((y, m))
        if val is not None:
            valeurs_fenetre.append(val)

    if not valeurs_fenetre:
        return None

    mode = indicateur.mode_agregation

    if mode == Indicateur.ModeAgregation.SOMME:
        total = sum(valeurs_fenetre, Decimal("0"))
        return total.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)

    elif mode == Indicateur.ModeAgregation.MOYENNE:
        total = sum(valeurs_fenetre, Decimal("0"))
        moyenne = total / Decimal(len(valeurs_fenetre))
        return moyenne.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)

    elif mode == Indicateur.ModeAgregation.MINIMUM:
        val_min = min(valeurs_fenetre)
        return val_min.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)

    elif mode == Indicateur.ModeAgregation.MAXIMUM:
        val_max = max(valeurs_fenetre)
        return val_max.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)

    elif mode == Indicateur.ModeAgregation.NOMBRE:
        count_val = Decimal(len(valeurs_fenetre))
        return count_val.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)

    elif mode == Indicateur.ModeAgregation.DERNIERE_VALEUR:
        return valeurs_fenetre[0]

    return None


def consolider_fenetre_12_mois_composante(
    composante_code: str,
    indicateur_id: int,
    mois_ref: int,
    annee_ref: int,
    valeurs_comp_dict: Dict[Tuple[str, int, int, int], Decimal],
    mode_agregation: str = "somme"
) -> Optional[Decimal]:
    """
    Consolide les valeurs d'une composante sur la fenêtre glissante de 12 mois.
    Conçu de façon extensible pour pouvoir supporter d'autres modes d'agrégation par composante à l'avenir.
    """
    valeurs_fenetre = []
    for i in range(12):
        m = mois_ref - i
        y = annee_ref
        if m <= 0:
            m += 12
            y -= 1
        val = valeurs_comp_dict.get((composante_code, indicateur_id, y, m))
        if val is not None:
            valeurs_fenetre.append(val)

    if not valeurs_fenetre:
        return None

    if mode_agregation == "somme":
        return sum(valeurs_fenetre, Decimal("0"))
    elif mode_agregation == "moyenne":
        return sum(valeurs_fenetre, Decimal("0")) / Decimal(len(valeurs_fenetre))
    elif mode_agregation == "derniere_valeur":
        return valeurs_fenetre[0]
    return sum(valeurs_fenetre, Decimal("0"))



def evaluer_statut_indicateur(
    indicateur: Indicateur,
    aggregat: Optional[Decimal],
    objectif_val: Optional[Decimal]
) -> Dict[str, Any]:
    """
    Évalue le statut numérique et le taux d'atteinte d'un indicateur selon son sens d'objectif
    (À atteindre / dépasser vs À ne pas dépasser).
    """
    if aggregat is None:
        return {
            "code": "donnee_absente",
            "label": "Donnée absente",
            "css_class": "secondary",
            "badge_class": "badge-secondary",
            "icon": "fas fa-question-circle",
            "taux_atteinte": None,
        }

    if objectif_val is None or objectif_val == Decimal("0"):
        return {
            "code": "non_applicable",
            "label": "Sans objectif",
            "css_class": "secondary",
            "badge_class": "badge-secondary",
            "icon": "fas fa-minus-circle",
            "taux_atteinte": None,
        }

    sens = getattr(indicateur, "sens_objectif", Indicateur.SensObjectif.ATTEINDRE)

    if sens == Indicateur.SensObjectif.NE_PAS_DEPASSER:
        # Sens : À ne pas dépasser (ex: Taux de fréquence, pannes)
        if aggregat <= objectif_val:
            taux_arrondi = Decimal("100.00")
            return {
                "code": "conforme",
                "label": "Objectif atteint",
                "css_class": "success",
                "badge_class": "badge-success",
                "icon": "fas fa-check-circle",
                "taux_atteinte": float(taux_arrondi),
            }
        else:
            # aggregat > objectif_val (dépassement d'un objectif à ne pas dépasser)
            if aggregat <= Decimal("0"):
                taux_arrondi = Decimal("0.00")
            else:
                taux = (objectif_val / aggregat) * Decimal("100")
                taux_arrondi = taux.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

            if taux_arrondi >= Decimal("80"):
                return {
                    "code": "a_surveiller",
                    "label": "À surveiller",
                    "css_class": "warning",
                    "badge_class": "badge-warning",
                    "icon": "fas fa-exclamation-triangle",
                    "taux_atteinte": float(taux_arrondi),
                }
            else:
                return {
                    "code": "non_conforme",
                    "label": "Objectif non atteint",
                    "css_class": "danger",
                    "badge_class": "badge-danger",
                    "icon": "fas fa-times-circle",
                    "taux_atteinte": float(taux_arrondi),
                }
    else:
        # Sens : À atteindre / dépasser (défaut)
        taux = (aggregat / objectif_val) * Decimal("100")
        taux_arrondi = taux.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        if taux_arrondi >= Decimal("100"):
            return {
                "code": "conforme",
                "label": "Objectif atteint",
                "css_class": "success",
                "badge_class": "badge-success",
                "icon": "fas fa-check-circle",
                "taux_atteinte": float(taux_arrondi),
            }
        elif taux_arrondi >= Decimal("80"):
            return {
                "code": "a_surveiller",
                "label": "À surveiller",
                "css_class": "warning",
                "badge_class": "badge-warning",
                "icon": "fas fa-exclamation-triangle",
                "taux_atteinte": float(taux_arrondi),
            }
        else:
            return {
                "code": "non_conforme",
                "label": "Objectif non atteint",
                "css_class": "danger",
                "badge_class": "badge-danger",
                "icon": "fas fa-times-circle",
                "taux_atteinte": float(taux_arrondi),
            }


COMPATIBLE_CHART_TYPES = {
    "time_series": ["line", "bar", "line:area"],
    "annual_history": ["line", "bar", "line:area"],
    "actual_vs_objective": ["bar"],
    "distribution": ["pie", "doughnut", "bar"],
    "multi_dimension": ["radar", "bar"],
}

CHART_OPTION_LABELS = {
    "line": "Courbe",
    "bar": "Barres",
    "line:area": "Courbe en aires",
    "pie": "Camembert",
    "doughnut": "Anneau",
    "radar": "Radar",
}


def determiner_nature_donnees(indicateur) -> str:
    """
    Détermine la nature des données associées à un indicateur.
    Pour les séries temporelles ordonnées (mensuelle, trimestrielle, semestrielle, annuelle),
    la nature est 'time_series'.
    """
    if indicateur.periodicite in {
        Indicateur.Periodicite.MENSUEL,
        Indicateur.Periodicite.TRIMESTRIEL,
        Indicateur.Periodicite.SEMESTRIEL,
        Indicateur.Periodicite.ANNUEL,
    }:
        return "time_series"
    return "time_series"


def generer_options_chart_compatibles(indicateur) -> List[Dict[str, str]]:
    """
    Génère la liste d'options de graphiques non vide pour chaque indicateur.
    Construit les choix 'line' (Courbe), 'line:area' (Courbe en aires) et 'bar' (Barres).
    """
    configured_types = (
        indicateur.formats_chart_disponibles
        if isinstance(indicateur.formats_chart_disponibles, list)
        else []
    )

    chart_options = []

    if "line" in configured_types:
        chart_options.append({
            "value": "line",
            "label": "Courbe",
            "type": "line",
            "variant": "standard",
        })
        chart_options.append({
            "value": "line:area",
            "label": "Courbe en aires",
            "type": "line",
            "variant": "area",
        })

    if "bar" in configured_types:
        chart_options.append({
            "value": "bar",
            "label": "Barres",
            "type": "bar",
            "variant": "standard",
        })

    if not chart_options:
        chart_options = [
            {
                "value": "line",
                "label": "Courbe",
                "type": "line",
                "variant": "standard",
            },
            {
                "value": "line:area",
                "label": "Courbe en aires",
                "type": "line",
                "variant": "area",
            },
            {
                "value": "bar",
                "label": "Barres",
                "type": "bar",
                "variant": "standard",
            },
        ]

    return chart_options


from django.utils.text import slugify


def generer_code_processus_depuis_nom(nom: str) -> str:
    """
    Génère un code court de processus à partir des initiales/mots du nom.
    Exemple obligatoire : « Contrôle outil » -> « CE »
    """
    if not nom or not str(nom).strip():
        return "PR"

    normalized = slugify(str(nom), allow_unicode=False).upper()
    words = [word for word in normalized.split("-") if word]

    if not words:
        return "PR"

    stop_words = {"DE", "DU", "LA", "LE", "LES", "DES", "ET", "EN", "UN", "UNE", "A", "AU", "AUX", "PAR", "POUR", "D", "L"}
    sig_words = [w for w in words if w not in stop_words]
    if not sig_words:
        sig_words = words

    full_slug = "-".join(sig_words)

    # Règle métier explicite / exemple obligatoire
    if full_slug in ("CONTROLE-OUTIL", "CONTROLE-OUTILS", "CONTROLE-EQUI", "CONTROLE-EQUIPEMENT"):
        base_code = "CE"
    elif len(sig_words) == 1:
        w = sig_words[0]
        base_code = w[:2] if len(w) >= 2 else w.ljust(2, "X")
    else:
        base_code = "".join([w[0] for w in sig_words[:3]])
        if len(base_code) < 2:
            base_code = base_code.ljust(2, "X")

    # Vérifier l'unicité
    code = base_code
    counter = 1
    while Processus.objects.filter(code=code).exists():
        counter += 1
        code = f"{base_code}{counter}"

    return code


def generer_code_indicateur(processus, extra_offset: int = 0) -> str:
    """
    Génère le prochain code d'indicateur pour un processus donné.
    Format : <CODE_PROCESSUS><NUMERO_SEQUENTIEL_SUR_2_CHIFFRES>
    Exemple : CE01, CE02, CE03 (pour le processus CE)
    """
    existing_codes = list(
        Indicateur.objects.filter(processus=processus)
        .values_list("code", flat=True)
    )
    prefix = processus.code.strip().rstrip("-")
    used_numbers = set()
    for c in existing_codes:
        if c.startswith(prefix):
            suffix = c[len(prefix):].lstrip("-")
            try:
                used_numbers.add(int(suffix))
            except ValueError:
                pass

    next_num = 1
    while next_num in used_numbers:
        next_num += 1

    final_num = next_num + extra_offset
    return f"{prefix}{final_num:02d}"



def formater_nombre_tdb(value, max_decimals=4):
    """
    Formatte un nombre Decimal ou numérique de manière intelligente :
    - None -> '—'
    - retirage des zéros inutiles en fin de chaîne
    - séparation décimale avec la virgule ','
    - conservation des entiers purs ('95', '4', '84', '0')
    """
    if value is None:
        return "—"

    if not isinstance(value, Decimal):
        value = Decimal(str(value))

    quantized_text = format(value, f".{max_decimals}f")

    if "." in quantized_text:
        quantized_text = quantized_text.rstrip("0").rstrip(".")

    return quantized_text.replace(".", ",")


def formater_pourcentage_tdb(value):
    """Formatte un pourcentage TDB avec au plus 2 décimales utiles."""
    if value is None:
        return "—"

    return formater_nombre_tdb(value, max_decimals=2) + " %"


# ─────────────────────────────────────────────────────────────────────────
# PALETTE GLOBALE DES GRAPHIQUES (source unique des couleurs)
# Les couleurs de série sont NORMALISÉES par l'application et n'utilisent plus
# les couleurs du classeur Excel (qui restent la référence pour le type de
# graphique, les axes, les graduations, les épaisseurs, les marqueurs, les
# grilles, les espacements et les formats numériques).
#   Réalisé = bleu  → CHART_COLORS["realized"]
#   Objectif = rouge → CHART_COLORS["target"]
# Cette palette est injectée dans le template (`window.CHART_COLORS`) et
# utilisée par l'export Excel : un seul endroit à modifier pour toute l'app.
# ─────────────────────────────────────────────────────────────────────────

CHART_COLORS = {
    "realized": "#4F81BD",
    "target": "#FF0000",
    "grid": "#BFBFBF",
    "axis": "#666666",
}

# Abréviations des mois (label des graphiques / fenêtres 12 mois glissants).
MOIS_ABBR = ["Jan", "Fév", "Mar", "Avr", "Mai", "Juin", "Juil", "Août", "Sept", "Oct", "Nov", "Déc"]


# ─────────────────────────────────────────────────────────────────────────
# STYLES GRAPHIQUES EXCEL DE RÉFÉRENCE
# Source : "2025-Tableau_de_bord_PM02.xlsx" — onglet « Synthèse PM02 2025 ».
# Chaque propriété est extraite du XML OOXML des graphiques du classeur :
#   - épaisseur de courbe : <a:ln w="28575">  → 28575 EMU ÷ 12700 = 2,25 pt
#   - grille majeure      : <c:majorGridlines><a:ln w="9525"> (0,75 pt),
#                           <a:schemeClr val="tx1"/> = noir automatique du thème
#   - couleurs de série   : <a:solidFill><a:schemeClr val="accent1..6"> résolus
#                           depuis xl/theme/theme1.xml (accent1=5B9BD5, accent2=ED7D31,
#                           accent3=A5A5A5, accent4=FFC000, accent5=4472C4, accent6=70AD47)
#   - axe des valeurs     : <c:scaling> min/max + <c:majorUnit> (pas de graduation)
#                           + <c:numFmt formatCode="0%"> (format des graduations)
#   - barres/colonnes     : <c:gapWidth val="219"> (espace inter-catégories = 219 %
#                           de la largeur de barre → largeur relative ~31 % du créneau)
#                           et <c:overlap val="-27"> (séparation des séries)
#   - marqueurs           : <c:marker><c:symbol>/<c:size> (taille en pt)
#   - légende             : <c:legend><c:legendPos val="b"> (en bas)
#   - police              : Calibri 11 (thème Excel), texte noir
# La configuration centralisée est injectée dans la vue sous le nom
# `window.excelChartStyle`, puis appliquée par le composant Chart.js.
# ─────────────────────────────────────────────────────────────────────────

EXCEL_STYLE_BASE = {
    "chartType": "line",
    "seriesColors": ["#002060"],
    "lineWidth": 2.25,
    "markerStyle": {
        "enabled": False,
        "symbol": "none",
        "size": 5,
        "fill": None,
        "borderColor": None,
        "borderWidth": 0.75,
    },
    "targetLineStyle": {"color": "#C00000", "width": 1.5, "dash": []},
    "yAxisMin": 0,
    "yAxisMax": None,
    "yAxisStep": None,
    "yAxisFormat": "General",
    "gridStyle": {"enabled": True, "color": "#000000", "width": 0.75},
    "legendPosition": "bottom",
    "font": {"family": "Calibri", "size": 11, "color": "#000000"},
    "barGap": {"barPercentage": 0.31, "categoryPercentage": 1},
    "overlap": None,
    "dataLabels": {"enabled": False},
    "background": None,
    "smooth": False,
    "excelLayout": None,
    "source": None,
}


def _style_excel(**kwargs):
    style = dict(EXCEL_STYLE_BASE)
    for key, value in kwargs.items():
        if value is not None:
            style[key] = value
    return style


# Courbes sans marqueur — axe % 0→1, pas 0,05 (chart3 « Taux de conformité
# réglementaire FRANCE/Luxembourg », série bleu marine #002060).
EXCEL_STYLE_LIGNE_PCT = _style_excel(
    chartType="line",
    seriesColors=["#002060"],
    yAxisMax=1,
    yAxisStep=0.05,
    yAxisFormat="percent",
    source="chart3 « Taux de conformité réglementaire FRANCE/Luxembourg » "
           "(courbe sans marqueur, axe 0→1 pas 0,05, graduations 0%)",
)

# Même axe %, série accent1 (bleu Excel des barres/colonnes) pour les taux
# sans graphique Excel dédié dans le classeur de référence.
EXCEL_STYLE_LIGNE_PCT_AC1 = _style_excel(
    chartType="line",
    seriesColors=["#5B9BD5"],
    yAxisMax=1,
    yAxisStep=0.05,
    yAxisFormat="percent",
    source="Template % (accent1) — axe identique au chart % de référence",
)

# Courbe générique — axe valeur naturel (format General).
EXCEL_STYLE_LIGNE_GENERAL = _style_excel(
    chartType="line",
    seriesColors=["#5B9BD5"],
    source="Template courbe générique (accent1, axe General)",
)

# Colonnes groupées — accent1, gap 219 %, overlap -27 (chart7
# « Nombre de réclamations et d'incidents client »).
EXCEL_STYLE_BARRES = _style_excel(
    chartType="bar",
    seriesColors=["#5B9BD5"],
    overlap=-27,
    yAxisStep=1,
    source="chart7 « Nombre de réclamations et d'incidents client » "
           "(colonnes groupées accent1, gapWidth 219 %, overlap -27)",
)

# Colonnes « Nombre d'ATA / Nombre d'ATSA » (chart13 : remplissage FF0000 /
# FFC000, bordure C00000). Graphique Excel d'origine : cascade (chartEx2-6).
EXCEL_STYLE_BARRES_ATA = _style_excel(
    chartType="bar",
    seriesColors=["#FF0000", "#FFC000"],
    overlap=-27,
    yAxisStep=1,
    excelLayout="waterfall",
    source="chart13 « Nombre d'ATA / Nombre d'ATSA » (FF0000 / FFC000, "
           "bordure C00000) — Excel d'origine : graphique en cascade",
)

# Taux de gravité — accent4, axe 0→100 pas 10, format 0.00 (chart4/14/16).
EXCEL_STYLE_LIGNE_GRAVITE = _style_excel(
    chartType="line",
    seriesColors=["#FFC000"],
    yAxisMax=100,
    yAxisStep=10,
    yAxisFormat="decimal2",
    source="chart4/14/16 « Taux de gravité » (accent4 #FFC000, axe 0→100 pas 10, format 0.00)",
)

# Taux d'analyse — accent6, axe % 0.00% (chart5).
EXCEL_STYLE_LIGNE_ANALYSE = _style_excel(
    chartType="line",
    seriesColors=["#70AD47"],
    yAxisFormat="percent2",
    source="chart5 « Taux d'analyse » (accent6 #70AD47, graduations 0.00%)",
)


# Configuration centralisée par code d'indicateur (sources de vérité Excel).
EXCEL_CHART_STYLES = {
    "PM01-01": EXCEL_STYLE_LIGNE_PCT_AC1,
    "PM01-02": EXCEL_STYLE_LIGNE_PCT_AC1,
    "PM01-03": _style_excel(
        chartType="bar",
        seriesColors=["#70AD47"],
        overlap=-27,
        yAxisStep=1,
        source="chart2 « Revues de direction » (série accent6 #70AD47 du graphique en colonnes empilées)",
    ),
    "PM02-01": EXCEL_STYLE_LIGNE_PCT_AC1,
    "PM02-02": EXCEL_STYLE_LIGNE_PCT_AC1,
    "PM02-03": EXCEL_STYLE_LIGNE_PCT_AC1,
    "PM02-04": EXCEL_STYLE_BARRES_ATA,
    "PR01-01": EXCEL_STYLE_LIGNE_PCT_AC1,
    "PR01-02": EXCEL_STYLE_BARRES,
    "PR01-03": _style_excel(
        chartType="line",
        seriesColors=["#5B9BD5"],
        yAxisMin=0,
        source="Template courbe générique (délai moyen — axe valeur naturel)",
    ),
    "PR02-01": EXCEL_STYLE_LIGNE_PCT_AC1,
    "PR02-02": EXCEL_STYLE_LIGNE_PCT,
    "PR02-03": EXCEL_STYLE_BARRES,
    "PS01-01": EXCEL_STYLE_LIGNE_PCT_AC1,
    "PS01-02": EXCEL_STYLE_LIGNE_PCT_AC1,
    "PS01-03": EXCEL_STYLE_BARRES,
    "PS02-01": EXCEL_STYLE_LIGNE_PCT_AC1,
    "PS02-02": EXCEL_STYLE_LIGNE_PCT_AC1,
    "PS02-03": EXCEL_STYLE_BARRES,
    # Démo temporaire « TF test glissant » : barres pour illustrer la fenêtre
    # glissante 12 mois traversant deux années (Cumul en tête + labels avec année).
    "PM01-TFGL": EXCEL_STYLE_BARRES,
}


# Correspondance par mots-clés du nom d'indicateur (ordre = priorité).
EXCEL_STYLES_PAR_MOTIF = [
    ("taux de conformit", EXCEL_STYLE_LIGNE_PCT),
    ("taux de gravit", EXCEL_STYLE_LIGNE_GRAVITE),
    ("taux d'analyse", EXCEL_STYLE_LIGNE_ANALYSE),
    ("taux", EXCEL_STYLE_LIGNE_PCT_AC1),
    ("nombre", EXCEL_STYLE_BARRES),
    ("delai", EXCEL_STYLE_LIGNE_GENERAL),
    ("retards", EXCEL_STYLE_BARRES),
    ("pannes", EXCEL_STYLE_BARRES),
]


def _normaliser_nom_indicateur(nom):
    """Normalise un nom d'indicateur : minuscules, sans accents, apostrophes espacées."""
    n = unicodedata.normalize("NFKD", nom or "")
    n = "".join(c for c in n if not unicodedata.combining(c))
    n = n.replace("'", " ").replace("’", " ")
    return n.lower()


def resoudre_excel_chart_style(indicateur):
    """Renvoie la configuration graphique « Excel de référence » pour un indicateur.

    Priorité : 1) code exact ; 2) mots-clés du nom ; 3) famille (taux/nombre) ;
    4) courbe générique.
    """
    code = (indicateur.code or "").strip().upper()
    if code in EXCEL_CHART_STYLES:
        return EXCEL_CHART_STYLES[code]

    nom = _normaliser_nom_indicateur(indicateur.nom)
    for motif, style in EXCEL_STYLES_PAR_MOTIF:
        if motif in nom:
            return style

    if "%" in (indicateur.nom or "").lower():
        return EXCEL_STYLE_LIGNE_PCT_AC1

    return EXCEL_STYLE_LIGNE_GENERAL


def determiner_formats_chart_compatibles(indicateur) -> List[str]:
    opts = generer_options_chart_compatibles(indicateur)
    return [o["value"] for o in opts]


def obtenir_donnees_tableau_de_bord(user, request_params, kwargs_processus_id=None):
    """
    Source de vérité unique pour les données du Tableau de Bord SMQS.
    Cette fonction est partagée 100% à l'identique entre l'affichage web/PDF (TdbDashboardView)
    et l'exportation Excel (tdb_export_excel).
    """
    tous_processus_qs = processus_perimetre_smqs_qs(user)
    tous_processus_qs = tous_processus_qs.select_related("societe").prefetch_related("RO", "RS", "CE").order_by("code")

    from accounts.models import Societe

    # 1. Société / périmètre
    societes_disponibles_qs = Societe.objects.filter(
        id__in=tous_processus_qs.values_list("societe_id", flat=True)
    ).order_by("nom")
    societes_disponibles = list(societes_disponibles_qs)

    societe_id = str(request_params.get("societe", "")).strip()
    selected_societe = None
    if societe_id:
        try:
            selected_societe = Societe.objects.get(pk=societe_id)
            tous_processus_qs = tous_processus_qs.filter(societe_id=societe_id)
        except (Societe.DoesNotExist, ValueError):
            selected_societe = None

    tous_processus = list(tous_processus_qs)
    tous_processus_ids = [p.id for p in tous_processus]

    # 2. Processus (le processus sélectionné doit rester dans le périmètre de l'utilisateur)
    processus_id = kwargs_processus_id or request_params.get("processus")
    processus = None
    if processus_id and str(processus_id).strip() != "":
        try:
            processus = Processus.objects.select_related("societe").prefetch_related("RO", "RS", "CE").get(pk=processus_id)
        except (Processus.DoesNotExist, ValueError):
            processus = None
        if processus is not None and processus.id not in tous_processus_ids:
            processus = None

    # 3. Année (chargée dynamiquement)
    current_year = timezone.now().year
    annees_obj = list(ObjectifIndicateur.objects.values_list("annee", flat=True).distinct())
    annees_mes = list(MesureIndicateur.objects.values_list("annee", flat=True).distinct())
    annees_disponibles = sorted(list(set(annees_obj + annees_mes + [current_year])), reverse=True)

    try:
        annee = int(request_params.get("annee", current_year))
    except (ValueError, TypeError):
        annee = current_year

    # 4 & 5. Type de période et Période
    type_periode = request_params.get("type_periode", "annee_complete")
    if type_periode not in {"annee_complete", "trimestre", "semestre", "mois"}:
        type_periode = "annee_complete"

    try:
        periode = int(request_params.get("periode", 1))
    except (ValueError, TypeError):
        periode = 1

    # 6. Fréquence / périodicité
    periodicite_filtre = str(request_params.get("periodicite", "")).strip()

    # 7. Statut
    statut_filtre = str(request_params.get("statut", "")).strip()

    # 8. Recherche
    recherche = str(request_params.get("recherche", "")).strip()
    active_tab = request_params.get("tab", "synthese")

    mois_filtrer = obtenir_mois_periode(type_periode, periode)

    if processus:
        indicateurs_qs = processus.indicateurs.filter(is_active=True).select_related("processus").order_by("code")
    else:
        indicateurs_qs = Indicateur.objects.filter(
            processus_id__in=tous_processus_ids,
            is_active=True,
        ).select_related("processus").order_by("processus__code", "code")

    if periodicite_filtre:
        indicateurs_qs = indicateurs_qs.filter(periodicite=periodicite_filtre)

    if recherche:
        indicateurs_qs = indicateurs_qs.filter(
            Q(code__icontains=recherche) |
            Q(nom__icontains=recherche) |
            Q(processus__code__icontains=recherche) |
            Q(processus__nom__icontains=recherche)
        )

    indicateurs = list(indicateurs_qs)
    indicateur_ids = [ind.id for ind in indicateurs]

    mesures_qs = MesureIndicateur.objects.filter(
        indicateur_id__in=indicateur_ids,
        annee__in=[annee, annee - 1],
    )

    mesures_par_indicateur_annee = {}
    for m in mesures_qs:
        mesures_par_indicateur_annee.setdefault((m.indicateur_id, m.annee), []).append(m)

    realises_consolides_qs = RealiseConsolideIndicateur.objects.filter(
        indicateur_id__in=indicateur_ids,
        annee_reference__in=[annee, annee - 1],
    )
    realises_consolides_dict = {
        (r.indicateur_id, r.annee_reference, r.mois_reference): r.valeur
        for r in realises_consolides_qs
    }

    mesures_dict_all = {
        (m.indicateur_id, m.annee, m.mois): m.valeur
        for m in mesures_qs
        if m.mois is not None
    }

    objectifs_qs = ObjectifIndicateur.objects.filter(
        indicateur_id__in=indicateur_ids,
        annee__in=[annee, annee - 1],
    )
    objectifs_par_indicateur_annee = {
        (obj.indicateur_id, obj.annee): obj.valeur_objectif for obj in objectifs_qs
    }
    objectifs_obj_par_indicateur_annee = {
        (obj.indicateur_id, obj.annee): obj for obj in objectifs_qs
    }

    composantes_qs = ComposanteIndicateur.objects.filter(
        indicateur_id__in=indicateur_ids
    )
    composantes_par_indicateur = {}
    for comp in composantes_qs:
        composantes_par_indicateur.setdefault(comp.indicateur_id, []).append(comp)

    valeurs_comp_qs = ValeurComposanteIndicateur.objects.filter(
        composante__indicateur_id__in=indicateur_ids,
        annee__in=[annee, annee - 1],
    ).select_related("composante")

    valeurs_comp_dict_all = {
        (v.composante.code, v.composante.indicateur_id, v.annee, v.mois): v.valeur
        for v in valeurs_comp_qs
    }

    indicateurs_data = []
    objectifs_data = []

    total_indicateurs = 0
    objectifs_atteints_count = 0
    a_surveiller_count = 0
    non_conformes_count = 0
    donnees_absentes_count = 0

    for indicateur in indicateurs:
        is_glissant = indicateur.periodicite == Indicateur.Periodicite.GLISSANT_12_MOIS
        serie_mensuelle_labels = list(MOIS_ABBR)

        if indicateur.mode_calcul == Indicateur.ModeCalcul.MANUEL:
            # Mode MANUEL : Valeur agrégée et série graphique basées EXCLUSIVEMENT sur RealiseConsolideIndicateur
            if is_glissant:
                # MANUEL + GLISSANT_12_MOIS : la série porte sur les 12 derniers mois
                # terminant au mois de référence (fin de période filtrée). La fenêtre
                # traverse l'année précédente : lecture (indicateur.id, year, month).
                # Ex : référence Août 2026 → Sept 2025 … Août 2026.
                mois_ref = max(mois_filtrer)
                fenetre_12_mois = []
                for i in range(12):
                    m = mois_ref - i
                    y = annee
                    if m <= 0:
                        m += 12
                        y -= 1
                    fenetre_12_mois.append((y, m))
                fenetre_12_mois.reverse()

                serie_mensuelle_dec = [
                    realises_consolides_dict.get((indicateur.id, y, m))
                    for (y, m) in fenetre_12_mois
                ]
                if any(y != annee for (y, _m) in fenetre_12_mois) or type_periode == "mois":
                    serie_mensuelle_labels = [
                        f"{MOIS_ABBR[m - 1]} {y}" for (y, m) in fenetre_12_mois
                    ]
                else:
                    serie_mensuelle_labels = list(MOIS_ABBR)

                non_null_rc = [
                    (i + 1, v) for i, v in enumerate(serie_mensuelle_dec, start=1)
                    if v is not None
                ]
                aggregat_decimal = non_null_rc[-1][1] if non_null_rc else None
            else:
                serie_mensuelle_dec = [
                    realises_consolides_dict.get((indicateur.id, annee, m))
                    for m in range(1, 13)
                ]
                serie_mensuelle_labels = list(MOIS_ABBR)
                non_null_rc = [
                    (m, v) for m, v in enumerate(serie_mensuelle_dec, start=1)
                    if m in mois_filtrer and v is not None
                ]

                if not non_null_rc:
                    aggregat_decimal = None
                elif indicateur.mode_agregation == Indicateur.ModeAgregation.DERNIERE_VALEUR:
                    aggregat_decimal = non_null_rc[-1][1]
                elif indicateur.mode_agregation == Indicateur.ModeAgregation.SOMME:
                    aggregat_decimal = sum((v for _, v in non_null_rc), Decimal("0"))
                elif indicateur.mode_agregation == Indicateur.ModeAgregation.MOYENNE:
                    tot = sum((v for _, v in non_null_rc), Decimal("0"))
                    aggregat_decimal = tot / Decimal(len(non_null_rc))
                elif indicateur.mode_agregation == Indicateur.ModeAgregation.MINIMUM:
                    aggregat_decimal = min(v for _, v in non_null_rc)
                elif indicateur.mode_agregation == Indicateur.ModeAgregation.MAXIMUM:
                    aggregat_decimal = max(v for _, v in non_null_rc)
                elif indicateur.mode_agregation == Indicateur.ModeAgregation.NOMBRE:
                    aggregat_decimal = Decimal(len(non_null_rc))
                else:
                    aggregat_decimal = non_null_rc[-1][1]

            if aggregat_decimal is not None and isinstance(aggregat_decimal, Decimal):
                aggregat_decimal = aggregat_decimal.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)

            serie_mensuelle = [float(v) if v is not None else None for v in serie_mensuelle_dec]
            mesures_par_mois = {
                m: v for m, v in enumerate(serie_mensuelle_dec, start=1)
                if v is not None
            }
        elif indicateur.mode_calcul == Indicateur.ModeCalcul.FORMULE:
            # Mode FORMULE : Calcul dynamique à partir des composantes et de la formule déclarée
            comps = composantes_par_indicateur.get(indicateur.id, [])
            formule_str = indicateur.formule or ""

            if is_glissant:
                serie_mensuelle_dec = []
                mois_avec_donnees_annee = []

                for m in range(1, 13):
                    vars_m = {}
                    for comp in comps:
                        sum_comp = consolider_fenetre_12_mois_composante(
                            comp.code, indicateur.id, m, annee, valeurs_comp_dict_all, mode_agregation="somme"
                        )
                        vars_m[comp.code] = sum_comp

                    if any((comp.code, indicateur.id, annee, m) in valeurs_comp_dict_all for comp in comps):
                        mois_avec_donnees_annee.append(m)

                    val_m = evaluer_formule_securisee(formule_str, vars_m)
                    if val_m is not None and isinstance(val_m, Decimal):
                        val_m = val_m.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
                    serie_mensuelle_dec.append(val_m)

                if mois_avec_donnees_annee:
                    dernier_mois = max(mois_avec_donnees_annee)
                    aggregat_decimal = serie_mensuelle_dec[dernier_mois - 1]
                else:
                    non_null_glissant = [v for v in serie_mensuelle_dec if v is not None]
                    aggregat_decimal = non_null_glissant[-1] if non_null_glissant else None

                # Chaque point i est la fenêtre glissante de 12 mois terminant en (annee, i+1).
                serie_mensuelle_labels = [
                    f"{MOIS_ABBR[m - 1]} {annee}" for m in range(1, 13)
                ]
                serie_mensuelle = [float(v) if v is not None else None for v in serie_mensuelle_dec]
                mesures_par_mois = {
                    m: v for m, v in enumerate(serie_mensuelle_dec, start=1)
                    if v is not None
                }
            else:
                serie_mensuelle_dec = []
                for m in range(1, 13):
                    vars_m = {
                        comp.code: valeurs_comp_dict_all.get((comp.code, indicateur.id, annee, m))
                        for comp in comps
                    }
                    val_m = evaluer_formule_securisee(formule_str, vars_m)
                    if val_m is not None and isinstance(val_m, Decimal):
                        val_m = val_m.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
                    serie_mensuelle_dec.append(val_m)

                non_null_f = [
                    (m, v) for m, v in enumerate(serie_mensuelle_dec, start=1)
                    if m in mois_filtrer and v is not None
                ]
                if not non_null_f:
                    aggregat_decimal = None
                elif indicateur.mode_agregation == Indicateur.ModeAgregation.DERNIERE_VALEUR:
                    aggregat_decimal = non_null_f[-1][1]
                elif indicateur.mode_agregation == Indicateur.ModeAgregation.SOMME:
                    # Pour un mode FORMULE, consolider les composantes sur la période
                    # afin d'évaluer le ratio global (ex: TF) au lieu d'additionner des ratios mensuels.
                    vars_total = {}
                    for comp in comps:
                        comp_vals = [
                            valeurs_comp_dict_all.get((comp.code, indicateur.id, annee, m))
                            for m in mois_filtrer
                        ]
                        non_null_comp = [v for v in comp_vals if v is not None]
                        if non_null_comp:
                            vars_total[comp.code] = sum(non_null_comp, Decimal("0"))
                        else:
                            vars_total[comp.code] = None

                    val_total = evaluer_formule_securisee(formule_str, vars_total)
                    if val_total is not None:
                        aggregat_decimal = val_total
                    else:
                        aggregat_decimal = sum((v for _, v in non_null_f), Decimal("0"))
                elif indicateur.mode_agregation == Indicateur.ModeAgregation.MOYENNE:
                    tot = sum((v for _, v in non_null_f), Decimal("0"))
                    aggregat_decimal = tot / Decimal(len(non_null_f))
                elif indicateur.mode_agregation == Indicateur.ModeAgregation.MINIMUM:
                    aggregat_decimal = min(v for _, v in non_null_f)
                elif indicateur.mode_agregation == Indicateur.ModeAgregation.MAXIMUM:
                    aggregat_decimal = max(v for _, v in non_null_f)
                elif indicateur.mode_agregation == Indicateur.ModeAgregation.NOMBRE:
                    aggregat_decimal = Decimal(len(non_null_f))
                else:
                    aggregat_decimal = non_null_f[-1][1]

                if aggregat_decimal is not None and isinstance(aggregat_decimal, Decimal):
                    aggregat_decimal = aggregat_decimal.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)

                serie_mensuelle = [float(v) if v is not None else None for v in serie_mensuelle_dec]
                mesures_par_mois = {
                    m: v for m, v in enumerate(serie_mensuelle_dec, start=1)
                    if v is not None
                }
        else:
            # Mode AUTOMATIQUE : Calcul depuis les composantes mensuelles MesureIndicateur
            if is_glissant:
                serie_mensuelle_dec = []
                mois_avec_donnees_annee = []

                for m in range(1, 13):
                    ind_mesures_map = {
                        (y_k, m_k): v
                        for (ind_k, y_k, m_k), v in mesures_dict_all.items()
                        if ind_k == indicateur.id
                    }
                    val_m = calculer_fenetre_12_mois_glissants(indicateur, m, annee, ind_mesures_map)
                    if (indicateur.id, annee, m) in mesures_dict_all:
                        mois_avec_donnees_annee.append(m)

                    serie_mensuelle_dec.append(val_m)

                if mois_avec_donnees_annee:
                    dernier_mois = max(mois_avec_donnees_annee)
                    aggregat_decimal = serie_mensuelle_dec[dernier_mois - 1]
                else:
                    non_null_glissant = [v for v in serie_mensuelle_dec if v is not None]
                    aggregat_decimal = non_null_glissant[-1] if non_null_glissant else None

                # Chaque point i est la fenêtre glissante de 12 mois terminant en (annee, i+1).
                serie_mensuelle_labels = [
                    f"{MOIS_ABBR[m - 1]} {annee}" for m in range(1, 13)
                ]
                serie_mensuelle = [float(v) if v is not None else None for v in serie_mensuelle_dec]
                mesures_par_mois = {
                    m: v for m, v in enumerate(serie_mensuelle_dec, start=1)
                    if v is not None
                }
            else:
                mesures_n = mesures_par_indicateur_annee.get((indicateur.id, annee), [])
                mesures_periode_n = [m for m in mesures_n if m.mois in mois_filtrer or m.mois is None]
                aggregat_decimal = calculer_agregation_indicateur(indicateur, mesures_periode_n, type_periode)

                mesures_par_mois = {m.mois: m.valeur for m in mesures_n if m.mois is not None}
                serie_mensuelle = [
                    float(mesures_par_mois[m]) if m in mesures_par_mois else None
                    for m in range(1, 13)
                ]

        objectif_decimal = objectifs_par_indicateur_annee.get((indicateur.id, annee))
        statut = evaluer_statut_indicateur(indicateur, aggregat_decimal, objectif_decimal)

        if statut_filtre and statut["code"] != statut_filtre:
            continue

        total_indicateurs += 1
        if statut["code"] == "conforme":
            objectifs_atteints_count += 1
        elif statut["code"] == "a_surveiller":
            a_surveiller_count += 1
        elif statut["code"] == "non_conforme":
            non_conformes_count += 1
        elif statut["code"] == "donnee_absente":
            donnees_absentes_count += 1

        mesures_n = mesures_par_indicateur_annee.get((indicateur.id, annee), [])
        mesure_annuelle = next((m.valeur for m in mesures_n if m.mois is None), None)
        if mesure_annuelle is None:
            mesure_annuelle = aggregat_decimal

        serie_mensuelle_json = json.dumps(
            [int(v) if v is not None and type(v) in (int, float) and float(v).is_integer() else v for v in serie_mensuelle]
        )

        formats_raw = indicateur.formats_chart_disponibles
        allowed_types = {"bar", "line", "pie", "doughnut", "radar"}

        if isinstance(formats_raw, list) and formats_raw:
            formats_disponibles = [fmt for fmt in formats_raw if isinstance(fmt, str) and fmt in allowed_types]
        else:
            formats_disponibles = []

        if not formats_disponibles:
            formats_disponibles = ["bar", "line", "pie", "doughnut", "radar"]

        chart_options = generer_options_chart_compatibles(indicateur)

        ecart = (aggregat_decimal - objectif_decimal) if (aggregat_decimal is not None and objectif_decimal is not None) else None

        statut_taux_formate = (
            formater_pourcentage_tdb(statut["taux_atteinte"])
            if statut.get("taux_atteinte") is not None
            else statut.get("label", "—")
        )

        mesures_dict = {
            str(m.mois): formater_nombre_tdb(m.valeur)
            for m in mesures_n
            if m.mois is not None
        }
        mesures_dict_json = json.dumps(mesures_dict)

        is_pct = "%" in indicateur.nom.lower() or "taux" in indicateur.nom.lower() or "pourcentage" in indicateur.nom.lower()
        objectif_affichage = (
            formater_pourcentage_tdb(objectif_decimal)
            if (objectif_decimal is not None and is_pct)
            else formater_nombre_tdb(objectif_decimal)
        )
        agregat_affichage = (
            formater_pourcentage_tdb(aggregat_decimal)
            if (aggregat_decimal is not None and is_pct)
            else formater_nombre_tdb(aggregat_decimal)
        )

        excel_style = resoudre_excel_chart_style(indicateur)
        excel_style_json = json.dumps(excel_style, ensure_ascii=False)

        item_data = {
            "indicateur": indicateur,
            "cumul": aggregat_decimal,
            "cumul_valeur_raw": (f"{aggregat_decimal:g}" if isinstance(aggregat_decimal, Decimal) else str(aggregat_decimal)) if aggregat_decimal is not None else "",
            "agregat": aggregat_decimal,
            "agregat_formate": formater_nombre_tdb(aggregat_decimal),
            "objectif": objectif_decimal,
            "objectif_affichage": objectif_affichage,
            "agregat_affichage": agregat_affichage,
            "objectif_valeur_raw": (f"{objectif_decimal:g}" if isinstance(objectif_decimal, Decimal) else str(objectif_decimal)) if objectif_decimal is not None else "",
            "objectif_formate": formater_nombre_tdb(objectif_decimal),
            "ecart": ecart,
            "ecart_formate": formater_nombre_tdb(ecart),
            "statut": statut,
            "statut_taux_formate": statut_taux_formate,
            "serie_mensuelle": serie_mensuelle,
            "serie_mensuelle_json": serie_mensuelle_json,
            "serie_mensuelle_labels": serie_mensuelle_labels,
            "serie_mensuelle_labels_json": json.dumps(serie_mensuelle_labels, ensure_ascii=False),
            "mesures_par_mois": mesures_par_mois,
            "mesures_dict": mesures_dict,
            "mesures_dict_json": mesures_dict_json,
            "mois_formates": [
                formater_nombre_tdb(mesures_par_mois.get(m)) for m in range(1, 13)
            ],
            "mesure_annuelle": mesure_annuelle,
            "valeur_annuelle_formatee": formater_nombre_tdb(mesure_annuelle),
            "can_edit_measurements": True,
            "formats_chart": [opt["value"] for opt in chart_options],
            "chart_options": chart_options,
            "excel_style": excel_style,
            "excel_style_json": excel_style_json,
        }
        indicateurs_data.append(item_data)

        obj_n = objectif_decimal
        obj_n_1 = objectifs_par_indicateur_annee.get((indicateur.id, annee - 1))
        obj_n_obj = objectifs_obj_par_indicateur_annee.get((indicateur.id, annee))

        diff_obj = (obj_n - obj_n_1) if (obj_n is not None and obj_n_1 is not None) else None
        var_pct = ((diff_obj / obj_n_1) * Decimal("100")) if (diff_obj is not None and obj_n_1 and obj_n_1 != Decimal("0")) else None

        objectifs_data.append({
            "indicateur": indicateur,
            "objectif_n": obj_n,
            "objectif_n_1": obj_n_1,
            "diff_obj": diff_obj,
            "var_pct": float(var_pct) if var_pct is not None else None,
            "realise_n": aggregat_decimal,
            "ecart_realise": ecart,
            "statut": statut,
            "updated_at": obj_n_obj.updated_at if obj_n_obj else None,
        })

    items_by_processus: dict = {}
    for item in indicateurs_data:
        pid = item["indicateur"].processus_id
        items_by_processus.setdefault(pid, []).append(item)

    processus_source = [processus] if processus else tous_processus

    processus_groups = []
    for proc in processus_source:
        proc_items = items_by_processus.get(proc.id, [])
        nb_atteints = sum(1 for i in proc_items if i["statut"]["code"] == "conforme")
        nb_surveiller = sum(1 for i in proc_items if i["statut"]["code"] == "a_surveiller")
        nb_nc = sum(1 for i in proc_items if i["statut"]["code"] == "non_conforme")
        nb_absents = sum(1 for i in proc_items if i["statut"]["code"] == "donnee_absente")
        processus_groups.append({
            "processus": proc,
            "indicateurs_data": proc_items,
            "indicateurs": proc_items,
            "can_add_indicator": True,
            "total_indicateurs": len(proc_items),
            "objectifs_atteints": nb_atteints,
            "a_surveiller": nb_surveiller,
            "non_conformes": nb_nc,
            "donnees_absentes": nb_absents,
        })

    statut_choices = [
        ("conforme", "Objectif atteint"),
        ("a_surveiller", "À surveiller"),
        ("non_conforme", "Objectif non atteint"),
        ("donnee_absente", "Donnée absente"),
        ("non_applicable", "Sans objectif"),
    ]

    AUDIT_STATUS_MAPPING = {
        "conforme": "conforme",
        "objectif_atteint": "conforme",
        "a_surveiller": "defauts_isoles",
        "non_conforme": "non_conforme",
        "objectif_non_atteint": "non_conforme",
        "donnee_absente": "non_audite",
        "sans_objectif": "non_audite",
        "non_applicable": "non_audite",
    }

    audit_counts = {
        "conforme": 0,
        "defauts_isoles": 0,
        "non_conforme": 0,
        "non_audite": 0,
    }

    audits_processus_list = []
    for group in processus_groups:
        proc = group["processus"]
        proc_items = group["indicateurs_data"]

        p_counts = {
            "conforme": 0,
            "defauts_isoles": 0,
            "non_conforme": 0,
            "non_audite": 0,
        }

        for item in proc_items:
            st_code = item["statut"]["code"]
            audit_cat = AUDIT_STATUS_MAPPING.get(st_code, "non_audite")
            p_counts[audit_cat] += 1
            audit_counts[audit_cat] += 1

        total_proc_ind = len(proc_items)
        taux_conf_val = round((p_counts["conforme"] / total_proc_ind * 100), 1) if total_proc_ind > 0 else 0.0

        if p_counts["non_conforme"] > 0:
            statut_global = {"code": "non_conforme", "label": "Non conforme", "badge_class": "bg-danger text-white", "bg_color": "#dc3545"}
        elif p_counts["defauts_isoles"] > 0:
            statut_global = {"code": "defauts_isoles", "label": "Défauts isolés", "badge_class": "bg-warning text-dark", "bg_color": "#ffc107"}
        elif p_counts["conforme"] > 0:
            statut_global = {"code": "conforme", "label": "Conforme", "badge_class": "bg-success text-white", "bg_color": "#28a745"}
        else:
            statut_global = {"code": "non_audite", "label": "Non audité", "badge_class": "bg-secondary text-white", "bg_color": "#6c757d"}

        audits_processus_list.append({
            "processus": proc,
            "total_indicateurs": total_proc_ind,
            "conforme": p_counts["conforme"],
            "defauts_isoles": p_counts["defauts_isoles"],
            "non_conforme": p_counts["non_conforme"],
            "non_audite": p_counts["non_audite"],
            "taux_conformite": formater_nombre_tdb(taux_conf_val) + " %",
            "taux_conformite_valeur": taux_conf_val,
            "statut_global": statut_global,
        })

    total_audits_indicateurs = sum(audit_counts.values())
    taux_conformite_global = round((audit_counts["conforme"] / total_audits_indicateurs * 100), 1) if total_audits_indicateurs > 0 else 0.0

    audits_summary = {
        "counts": audit_counts,
        "total_indicateurs": total_audits_indicateurs,
        "taux_conformite_global": formater_nombre_tdb(taux_conformite_global) + " %",
        "taux_conformite_valeur": taux_conformite_global,
        "chart_labels": ["Conforme", "Défauts isolés", "Non conforme", "Non audité"],
        "chart_data": [audit_counts["conforme"], audit_counts["defauts_isoles"], audit_counts["non_conforme"], audit_counts["non_audite"]],
        "chart_colors": ["#28a745", "#ffc107", "#dc3545", "#6c757d"],
    }

    OBJECTIVE_STATUS_LABELS = {
        "conforme": "Objectifs atteints",
        "a_surveiller": "À surveiller",
        "non_conforme": "Non atteints",
        "donnee_absente": "Sans données",
        "non_applicable": "Sans objectif",
    }
    OBJECTIVE_STATUS_COLORS = {
        "conforme": "#28a745",
        "a_surveiller": "#ffc107",
        "non_conforme": "#dc3545",
        "donnee_absente": "#6c757d",
        "non_applicable": "#17a2b8",
    }

    statuses_counts: Dict[str, int] = {}
    for obj in objectifs_data:
        code = obj["statut"]["code"]
        statuses_counts[code] = statuses_counts.get(code, 0) + 1

    status_labels = []
    status_values = []
    status_colors = []
    for code, label in OBJECTIVE_STATUS_LABELS.items():
        if code in statuses_counts:
            status_labels.append(label)
            status_values.append(statuses_counts[code])
            status_colors.append(OBJECTIVE_STATUS_COLORS[code])
    for code in statuses_counts:
        if code not in OBJECTIVE_STATUS_LABELS:
            status_labels.append(code)
            status_values.append(statuses_counts[code])
            status_colors.append("#343a40")

    objectives_chart_payload = {
        "evolution": {
            "labels": [obj["indicateur"].code for obj in objectifs_data],
            "currentYear": annee,
            "previousYear": annee - 1,
            "currentValues": [
                float(obj["objectif_n"]) if obj["objectif_n"] is not None else None
                for obj in objectifs_data
            ],
            "previousValues": [
                float(obj["objectif_n_1"]) if obj["objectif_n_1"] is not None else None
                for obj in objectifs_data
            ],
        },
        "statuses": {
            "labels": status_labels,
            "values": status_values,
            "colors": status_colors,
        },
    }

    return {
        "AUDIT_STATUS_MAPPING": AUDIT_STATUS_MAPPING,
        "audits_summary": audits_summary,
        "audits_processus_list": audits_processus_list,
        "audits_chart_data_json": json.dumps({
            "labels": audits_summary["chart_labels"],
            "data": audits_summary["chart_data"],
            "colors": audits_summary["chart_colors"],
        }),
        "processus": processus,
        "tous_processus": tous_processus,
        "societes_disponibles": societes_disponibles,
        "societes_list": list(Societe.objects.order_by("nom")),
        "selected_societe": selected_societe,
        "selected_societe_id": selected_societe.id if selected_societe else None,
        "selected_processus_id": processus.id if processus else None,
        "processus_groups": processus_groups,
        "nb_processus_visibles": len(processus_groups),
        "nb_indicateurs_visibles": total_indicateurs,
        "annee": annee,
        "type_periode": type_periode,
        "periode": periode,
        "periodicite_filtre": periodicite_filtre,
        "periodicite_choices": Indicateur.Periodicite.choices,
        "statut_filtre": statut_filtre,
        "statut_choices": statut_choices,
        "recherche": recherche,
        "active_tab": active_tab,
        "annees_disponibles": annees_disponibles,
        "indicateurs_data": indicateurs_data,
        "objectifs_data": objectifs_data,
        "kpi_stats": {
            "total_indicateurs": total_indicateurs,
            "objectifs_atteints": objectifs_atteints_count,
            "a_surveiller": a_surveiller_count,
            "non_conformes": non_conformes_count,
            "donnees_absentes": donnees_absentes_count,
        },
        "objectives_chart_payload": objectives_chart_payload,
        "chart_colors": CHART_COLORS,
        "chart_colors_json": json.dumps(CHART_COLORS),
    }


class TdbDashboardView(TdbAccesRequiredMixin, TemplateView):
    """Vue principale unifiée du Tableau de Bord SMQS comprenant 4 onglets."""

    template_name = "gestion_documentaire/tdb_dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dashboard_data = obtenir_donnees_tableau_de_bord(
            self.request.user,
            self.request.GET,
            kwargs_processus_id=self.kwargs.get("processus_id")
        )
        context.update(dashboard_data)
        # Si une précédente tentative de modification a échoué, on conserve les
        # valeurs saisies pour réafficher la modal pré-remplie.
        form_erreur = self.request.session.pop("tdb_indicateur_form_erreur", None)
        if form_erreur:
            context["indicateur_form_erreur"] = form_erreur
            context["indicateur_form_erreur_json"] = json.dumps(form_erreur, ensure_ascii=False)
        return context


@login_required
@tdb_acces_required
@csrf_exempt
def tdb_export_excel(request):
    """
    Génère un fichier Excel (.xlsx) contenant 4 feuilles reproduisant fidèlement la mise en page
    et la structure visuelle des modèles (Captures de référence 1 à 5).
    Prend en compte dynamiquement les types de graphiques (Courbe, Barres, Aires) sélectionnés dans l'interface web.
    Consomme EXCLUSIVEMENT `obtenir_donnees_tableau_de_bord` sans aucune duplication de logique métier.
    """
    from openpyxl import Workbook
    from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    # Extraire les filtres et les types de graphiques sélectionnés (support GET et POST)
    req_params = request.GET.copy()
    chart_types = {}
    default_chart_type = req_params.get("chart_type", "line")

    if request.method == "POST":
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
            if isinstance(payload, dict):
                chart_types = payload.get("chart_types", {})
                if payload.get("chart_type"):
                    default_chart_type = payload.get("chart_type")
                for k in ["annee", "societe", "processus"]:
                    if k in payload and payload[k]:
                        req_params[k] = str(payload[k])
        except Exception:
            pass
    elif "chart_types" in req_params:
        try:
            chart_types = json.loads(req_params.get("chart_types"))
        except Exception:
            pass

        print("CHART TYPES RECEIVED AT BACKEND:", chart_types)

    def creer_chart_excel(chart_type):
        val = str(chart_type).lower().strip()
        if val in ["bar", "barres", "column"]:
            c = BarChart()
            c.type = "col"
            c.grouping = "clustered"
            c.overlap = 0
            c.style = 10
            c.varyColors = False
            return c
        elif val in ["area", "aires"]:
            c = AreaChart()
            c.varyColors = False
            return c
        else:
            c = LineChart()
            c.varyColors = False
            return c

    data = obtenir_donnees_tableau_de_bord(request.user, req_params)

    import io

    annee = data["annee"]
    processus_filtered = data["processus"]
    filename = (
        f"Tableau_de_bord_{processus_filtered.code}_{annee}.xlsx"
        if processus_filtered
        else f"Tableau_de_bord_SMQS_{annee}.xlsx"
    )

    wb = Workbook()

    # Styles globaux
    title_font = Font(name="Calibri", size=15, bold=True, color="1E293B")
    subtitle_font = Font(name="Calibri", size=11, bold=True, color="475569")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")

    process_header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    process_fill = PatternFill(start_color="E26B00", end_color="E26B00", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    status_fills = {
        "conforme": PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid"),
        "a_surveiller": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "non_conforme": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
        "donnee_absente": PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid"),
        "defauts_isoles": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "non_audite": PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid"),
    }

    status_fonts = {
        "conforme": Font(name="Calibri", size=10, bold=True, color="0F5132"),
        "a_surveiller": Font(name="Calibri", size=10, bold=True, color="664D03"),
        "non_conforme": Font(name="Calibri", size=10, bold=True, color="842029"),
        "donnee_absente": Font(name="Calibri", size=10, bold=True, color="41464B"),
        "defauts_isoles": Font(name="Calibri", size=10, bold=True, color="664D03"),
        "non_audite": Font(name="Calibri", size=10, bold=True, color="41464B"),
    }

    # =========================================================================
    # FEUILLE 1 : Synthèse des indicateurs
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Synthèse des indicateurs"

    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=18)
    t1_cell = ws1.cell(row=1, column=1, value=f"Tableau de Bord SMQS - Synthèse des Indicateurs ({annee})")
    t1_cell.font = title_font
    t1_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 35

    headers_ws1 = [
        "Code & Indicateur", "Fréquence", "Objectif", "Cumul / Moyenne",
        "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
        "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre",
        "Annuel", "Statut"
    ]

    for col_idx, h in enumerate(headers_ws1, start=1):
        cell = ws1.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
        cell.border = thin_border
    ws1.row_dimensions[3].height = 25

    current_row = 4
    for group in data["processus_groups"]:
        proc = group["processus"]
        ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=18)
        p_cell = ws1.cell(row=current_row, column=1, value=f"{proc.code} - {proc.nom}")
        p_cell.font = process_header_font
        p_cell.fill = process_fill
        p_cell.alignment = Alignment(horizontal="left", vertical="center")
        for c in range(1, 19):
            ws1.cell(row=current_row, column=c).border = thin_border
        ws1.row_dimensions[current_row].height = 22
        current_row += 1

        for item in group["indicateurs_data"]:
            ind = item["indicateur"]
            ws1.cell(row=current_row, column=1, value=f"{ind.code} - {ind.nom}").border = thin_border
            ws1.cell(row=current_row, column=2, value=ind.get_periodicite_display()).alignment = Alignment(horizontal="center")
            ws1.cell(row=current_row, column=2).border = thin_border

            obj_val = float(item["objectif"]) if item["objectif"] is not None else None
            obj_cell = ws1.cell(row=current_row, column=3, value=obj_val)
            obj_cell.alignment = Alignment(horizontal="right")
            obj_cell.border = thin_border
            if obj_val is not None:
                obj_cell.number_format = "#,##0.00"

            cumul_val = float(item["cumul"]) if item["cumul"] is not None else None
            cumul_cell = ws1.cell(row=current_row, column=4, value=cumul_val)
            cumul_cell.alignment = Alignment(horizontal="right")
            cumul_cell.border = thin_border
            if cumul_val is not None:
                cumul_cell.number_format = "#,##0.00"

            for m in range(1, 13):
                m_val = item["mesures_par_mois"].get(m)
                m_num = float(m_val) if m_val is not None else None
                m_cell = ws1.cell(row=current_row, column=4 + m, value=m_num)
                m_cell.alignment = Alignment(horizontal="right")
                m_cell.border = thin_border
                if m_num is not None:
                    m_cell.number_format = "#,##0.00"

            ann_val = item["mesure_annuelle"]
            ann_num = float(ann_val) if ann_val is not None else None
            ann_cell = ws1.cell(row=current_row, column=17, value=ann_num)
            ann_cell.alignment = Alignment(horizontal="right")
            ann_cell.border = thin_border
            if ann_num is not None:
                ann_cell.number_format = "#,##0.00"

            st_code = item["statut"]["code"]
            st_label = item["statut"]["label"]
            st_cell = ws1.cell(row=current_row, column=18, value=st_label)
            st_cell.alignment = Alignment(horizontal="center")
            st_cell.border = thin_border
            if st_code in status_fills:
                st_cell.fill = status_fills[st_code]
                st_cell.font = status_fonts[st_code]

            current_row += 1

    ws1.freeze_panes = "A4"

    for col in ws1.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 11)
    ws1.column_dimensions["A"].width = 35

    # =========================================================================
    # FEUILLE 2 : Vue graphique (Structure Grille 2 par ligne avec largeur 13)
    # =========================================================================
    ws2 = wb.create_sheet(title="Vue graphique")

    proc_title_str = processus_filtered.code if processus_filtered else "SMQS"
    title_ws2_text = f"Zoom Indicateurs {proc_title_str} {annee} : Représentation Graphique"

    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
    t2_cell = ws2.cell(row=1, column=1, value=title_ws2_text)
    t2_cell.font = Font(name="Calibri", size=16, bold=True, color="000000")
    t2_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 35

    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=20)
    sub2_cell = ws2.cell(row=2, column=1, value=f"Représentation Graphique par Indicateur ({annee})")
    sub2_cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    sub2_cell.fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
    sub2_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[2].height = 25

    mois_abbr = ["Jan", "Fév", "Mar", "Avr", "Mai", "Juin", "Juil", "Août", "Sept", "Oct", "Nov", "Déc"]

    indicators = data["indicateurs_data"]
    for idx, item in enumerate(indicators):
        ind = item["indicateur"]

        # Grille aérées de 2 blocs par ligne (Cols A=1, J=10)
        # Layout per block: 4 rows KPI + 1 gap + 13 rows table + 2 gap + ~18 rows chart = 38 rows
        col_start = (idx % 2) * 9 + 1
        row_start = (idx // 2) * 38 + 4

        # En-tête de bloc d'indicateur
        ws2.merge_cells(start_row=row_start, start_column=col_start, end_row=row_start, end_column=col_start + 7)
        h_cell = ws2.cell(row=row_start, column=col_start, value=f"{ind.code} - {ind.nom}")
        h_cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        h_cell.fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
        h_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws2.merge_cells(start_row=row_start + 1, start_column=col_start, end_row=row_start + 1, end_column=col_start + 7)
        sub_c = ws2.cell(row=row_start + 1, column=col_start, value=f"Périodicité : {ind.get_periodicite_display()} | Agrégation : {ind.get_mode_agregation_display()}")
        sub_c.font = Font(name="Calibri", size=8, italic=True, color="595959")
        sub_c.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        sub_c.alignment = Alignment(horizontal="center", vertical="center")

        # Cartouche de métriques KPI
        ws2.merge_cells(start_row=row_start + 2, start_column=col_start, end_row=row_start + 2, end_column=col_start + 1)
        ws2.cell(row=row_start + 2, column=col_start, value="Objectif").font = Font(size=8, bold=True, color="595959")
        ws2.cell(row=row_start + 2, column=col_start).alignment = Alignment(horizontal="center")

        ws2.merge_cells(start_row=row_start + 2, start_column=col_start + 3, end_row=row_start + 2, end_column=col_start + 4)
        ws2.cell(row=row_start + 2, column=col_start + 3, value="Valeur Agrégée").font = Font(size=8, bold=True, color="595959")
        ws2.cell(row=row_start + 2, column=col_start + 3).alignment = Alignment(horizontal="center")

        ws2.merge_cells(start_row=row_start + 2, start_column=col_start + 6, end_row=row_start + 2, end_column=col_start + 7)
        ws2.cell(row=row_start + 2, column=col_start + 6, value="Statut").font = Font(size=8, bold=True, color="595959")
        ws2.cell(row=row_start + 2, column=col_start + 6).alignment = Alignment(horizontal="center")

        o_val = float(item["objectif"]) if item["objectif"] is not None else None
        ws2.merge_cells(start_row=row_start + 3, start_column=col_start, end_row=row_start + 3, end_column=col_start + 1)
        oc = ws2.cell(row=row_start + 3, column=col_start, value=o_val)
        oc.font = Font(size=10, bold=True)
        oc.alignment = Alignment(horizontal="center")
        if o_val is not None:
            oc.number_format = "#,##0.00"

        c_val = float(item["cumul"]) if item["cumul"] is not None else None
        ws2.merge_cells(start_row=row_start + 3, start_column=col_start + 3, end_row=row_start + 3, end_column=col_start + 4)
        cc = ws2.cell(row=row_start + 3, column=col_start + 3, value=c_val)
        cc.font = Font(size=10, bold=True, color="0D6EFD")
        cc.alignment = Alignment(horizontal="center")
        if c_val is not None:
            cc.number_format = "#,##0.00"

        st_code = item["statut"]["code"]
        ws2.merge_cells(start_row=row_start + 3, start_column=col_start + 6, end_row=row_start + 3, end_column=col_start + 7)
        sc = ws2.cell(row=row_start + 3, column=col_start + 6, value=item["statut"]["label"])
        sc.alignment = Alignment(horizontal="center")
        if st_code in status_fills:
            sc.fill = status_fills[st_code]
            sc.font = status_fonts[st_code]


        # -----------------------------------------------
        # Table source : Mois | Réalisé | Objectif  (rows: +5..+17)
        # row_start+5  → en-têtes "Mois" / "Réalisé" / "Objectif"
        # row_start+6  → Jan
        # row_start+17 → Déc
        # -----------------------------------------------
        data_hdr_row  = row_start + 5
        data_first_row = row_start + 6
        data_last_row  = row_start + 17   # 12 mois
        month_col  = col_start       # colonne A du bloc
        value_col  = col_start + 1   # colonne B du bloc
        obj_col    = col_start + 2   # colonne C du bloc

        ws2.cell(row=data_hdr_row, column=month_col,  value="Mois").font    = Font(bold=True, size=8)
        ws2.cell(row=data_hdr_row, column=value_col, value="Réalisé").font = Font(bold=True, size=8)
        ws2.cell(row=data_hdr_row, column=obj_col,   value="Objectif").font = Font(bold=True, size=8)

        has_obj = item["objectif"] is not None
        o_val_num = float(item["objectif"]) if has_obj else None

        for m_idx in range(12):
            r_row = data_first_row + m_idx
            labels_mois_excel = item.get("serie_mensuelle_labels") or mois_abbr
            ws2.cell(row=r_row, column=month_col, value=labels_mois_excel[m_idx]).font = Font(size=8)

            raw_v = item["serie_mensuelle"][m_idx]
            if raw_v is not None:
                try:
                    v_m = float(raw_v)
                except (ValueError, TypeError):
                    v_m = None
            else:
                v_m = None

            vm_cell = ws2.cell(row=r_row, column=value_col, value=v_m)
            vm_cell.font = Font(size=8)
            if v_m is not None:
                vm_cell.number_format = "#,##0.00"

            om_cell = ws2.cell(row=r_row, column=obj_col, value=o_val_num)
            om_cell.font = Font(size=8)
            if o_val_num is not None:
                om_cell.number_format = "#,##0.00"

        # -----------------------------------------------
        # Graphique — ancré SOUS la table
        # -----------------------------------------------
        c_type = (
            chart_types.get(ind.code)
            or chart_types.get(str(ind.code))
            or chart_types.get(str(ind.pk))
            or default_chart_type
        )
        c_type_clean = str(c_type).lower().strip()
        is_bar_chart = c_type_clean in ["bar", "barres", "column"]

        chart = creer_chart_excel(c_type)

        raw_title = f"{ind.code} - {ind.nom}"
        chart.title = raw_title[:52] + "..." if len(raw_title) > 55 else raw_title

        chart.width  = 13
        chart.height = 8   # hauteur suffisante pour que les labels soient visibles

        # categories Reference = uniquement les 12 mois (sans en-tête)
        chart_cats_ref = Reference(
            ws2,
            min_col=month_col,
            max_col=month_col,
            min_row=data_first_row,   # Jan
            max_row=data_last_row,    # Déc
        )

        if is_bar_chart and has_obj:
            # Graphique combiné : BarChart pour Réalisé + LineChart (ligne rouge) pour Objectif
            chart_data_ref = Reference(
                ws2,
                min_col=value_col,
                max_col=value_col,
                min_row=data_hdr_row,
                max_row=data_last_row,
            )
            chart.add_data(chart_data_ref, titles_from_data=True)
            chart.set_categories(chart_cats_ref)

            realized_hex = CHART_COLORS["realized"].lstrip("#")
            target_hex = CHART_COLORS["target"].lstrip("#")
            if chart.series:
                s_real = chart.series[0]
                if hasattr(s_real, "graphicalProperties") and hasattr(s_real.graphicalProperties, "solidFill"):
                    s_real.graphicalProperties.solidFill = realized_hex

            line_chart = LineChart()
            obj_data_ref = Reference(
                ws2,
                min_col=obj_col,
                max_col=obj_col,
                min_row=data_hdr_row,
                max_row=data_last_row,
            )
            line_chart.add_data(obj_data_ref, titles_from_data=True)
            line_chart.set_categories(chart_cats_ref)
            if line_chart.series:
                s_obj = line_chart.series[0]
                if hasattr(s_obj, "graphicalProperties") and hasattr(s_obj.graphicalProperties, "line"):
                    s_obj.graphicalProperties.line.solidFill = target_hex

            chart += line_chart
        else:
            # Graphique simple ou LineChart avec 2 séries
            max_c = obj_col if has_obj else value_col
            chart_data_ref = Reference(
                ws2,
                min_col=value_col,
                max_col=max_c,
                min_row=data_hdr_row,
                max_row=data_last_row,
            )
            chart.add_data(chart_data_ref, titles_from_data=True)
            chart.set_categories(chart_cats_ref)

            realized_hex = CHART_COLORS["realized"].lstrip("#")
            target_hex = CHART_COLORS["target"].lstrip("#")
            if chart.series:
                s_real = chart.series[0]
                if hasattr(s_real, "graphicalProperties") and hasattr(s_real.graphicalProperties, "line"):
                    s_real.graphicalProperties.line.solidFill = realized_hex

            if has_obj and len(chart.series) > 1:
                s_obj = chart.series[1]
                if hasattr(s_obj, "graphicalProperties") and hasattr(s_obj.graphicalProperties, "line"):
                    s_obj.graphicalProperties.line.solidFill = target_hex

        # ── Axe X (catégories = mois) ──────────────────────────────────────
        chart.x_axis.title     = None
        chart.x_axis.delete    = False      # ne pas cacher l'axe
        chart.x_axis.tickLblPos = "low"    # labels sous les barres
        chart.x_axis.numFmt    = "General"

        # ── Axe Y (valeurs) ────────────────────────────────────────────────
        chart.y_axis.title      = None
        chart.y_axis.delete     = False     # ne pas cacher l'axe
        chart.y_axis.tickLblPos = "nextTo"  # labels à gauche
        chart.y_axis.scaling.min = 0
        chart.y_axis.numFmt    = "General"

        chart.varyColors = False

        # Ancrer le graphique 2 lignes sous la table
        chart_anchor = f"{get_column_letter(col_start)}{data_last_row + 2}"
        ws2.add_chart(chart, chart_anchor)



    # Configuration des largeurs de colonnes
    for c in range(1, 20):
        col_let = get_column_letter(c)
        if c in [9, 18]:
            ws2.column_dimensions[col_let].width = 3
        else:
            ws2.column_dimensions[col_let].width = 10

    # =========================================================================
    # FEUILLE 3 : Objectifs annuels (Structure Capture 4 - Groupement Processus/Société)
    # =========================================================================
    ws3 = wb.create_sheet(title="Objectifs annuels")

    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    t3_cell = ws3.cell(row=1, column=1, value=f"Tableau de Bord SMQS - Objectifs annuels ({annee - 1} / {annee})")
    t3_cell.font = title_font
    t3_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 35

    headers_ws3 = ["Indicateurs", f"Objectifs {annee - 1}", f"Objectifs {annee}"]

    for col_idx, h in enumerate(headers_ws3, start=1):
        cell = ws3.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
        cell.border = thin_border
    ws3.row_dimensions[3].height = 25

    # Groupement des objectifs par Processus & Société
    soc_colors = {
        "AB Serve Group": PatternFill(start_color="E26B00", end_color="E26B00", fill_type="solid"),
        "AB Serve": PatternFill(start_color="005580", end_color="005580", fill_type="solid"),
        "Valoris Emploi": PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),
        "AB Serve GERMANY": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
    }

    soc_fonts = {
        "AB Serve GERMANY": Font(name="Calibri", size=10, bold=True, color="000000"),
    }

    target_n_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    target_n_font = Font(name="Calibri", size=10, bold=True, color="006100")

    current_row_ws3 = 4
    for group in data["processus_groups"]:
        proc = group["processus"]

        # Bandeau Processus (Orange Foncé, texte blanc)
        ws3.merge_cells(start_row=current_row_ws3, start_column=1, end_row=current_row_ws3, end_column=3)
        p_cell = ws3.cell(row=current_row_ws3, column=1, value=f"{proc.code} : {proc.nom}")
        p_cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        p_cell.fill = PatternFill(start_color="E26B00", end_color="E26B00", fill_type="solid")
        p_cell.alignment = Alignment(horizontal="left", vertical="center")
        for c in range(1, 4):
            ws3.cell(row=current_row_ws3, column=c).border = thin_border
        ws3.row_dimensions[current_row_ws3].height = 24
        current_row_ws3 += 1

        proc_objs = [obj for obj in data["objectifs_data"] if obj["indicateur"].processus_id == proc.pk]

        objs_by_soc = {}
        for obj in proc_objs:
            soc_nom = proc.societe.nom if proc.societe else "Général"
            if soc_nom not in objs_by_soc:
                objs_by_soc[soc_nom] = []
            objs_by_soc[soc_nom].append(obj)

        for soc_nom, obj_list in objs_by_soc.items():
            if len(objs_by_soc) > 1 or (soc_nom and soc_nom != "Général" and soc_nom != proc.nom):
                ws3.merge_cells(start_row=current_row_ws3, start_column=1, end_row=current_row_ws3, end_column=3)
                s_cell = ws3.cell(row=current_row_ws3, column=1, value=soc_nom)
                s_fill = soc_colors.get(soc_nom, PatternFill(start_color="005580", end_color="005580", fill_type="solid"))
                s_font = soc_fonts.get(soc_nom, Font(name="Calibri", size=10, bold=True, color="FFFFFF"))
                s_cell.fill = s_fill
                s_cell.font = s_font
                s_cell.alignment = Alignment(horizontal="left", vertical="center")
                for c in range(1, 4):
                    ws3.cell(row=current_row_ws3, column=c).border = thin_border
                ws3.row_dimensions[current_row_ws3].height = 20
                current_row_ws3 += 1

            for obj in obj_list:
                ind = obj["indicateur"]

                ind_cell = ws3.cell(row=current_row_ws3, column=1, value=ind.nom)
                ind_cell.font = Font(name="Calibri", size=10)
                ind_cell.border = thin_border

                is_pct = "%" in ind.nom.lower() or "taux" in ind.nom.lower() or "pourcentage" in ind.nom.lower()

                o_prev = float(obj["objectif_n_1"]) if obj["objectif_n_1"] is not None else None
                o_prev_cell = ws3.cell(row=current_row_ws3, column=2)
                o_prev_cell.border = thin_border
                o_prev_cell.alignment = Alignment(horizontal="center", vertical="center")
                if o_prev is not None:
                    if is_pct:
                        o_prev_cell.value = (o_prev / 100.0) if o_prev > 1.0 else o_prev
                        o_prev_cell.number_format = "0%" if (o_prev % 1 == 0) else "0.0%"
                    else:
                        if o_prev % 1 == 0:
                            o_prev_cell.value = int(o_prev)
                            o_prev_cell.number_format = "#,##0"
                        else:
                            o_prev_cell.value = o_prev
                            o_prev_cell.number_format = "#,##0.00"

                o_curr = float(obj["objectif_n"]) if obj["objectif_n"] is not None else None
                o_curr_cell = ws3.cell(row=current_row_ws3, column=3)
                o_curr_cell.border = thin_border
                o_curr_cell.fill = target_n_fill
                o_curr_cell.font = target_n_font
                o_curr_cell.alignment = Alignment(horizontal="center", vertical="center")
                if o_curr is not None:
                    if is_pct:
                        o_curr_cell.value = (o_curr / 100.0) if o_curr > 1.0 else o_curr
                        o_curr_cell.number_format = "0%" if (o_curr % 1 == 0) else "0.0%"
                    else:
                        if o_curr % 1 == 0:
                            o_curr_cell.value = int(o_curr)
                            o_curr_cell.number_format = "#,##0"
                        else:
                            o_curr_cell.value = o_curr
                            o_curr_cell.number_format = "#,##0.00"

                current_row_ws3 += 1

    ws3.freeze_panes = "A4"
    ws3.column_dimensions["A"].width = 50
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 20

    # =========================================================================
    # FEUILLE 4 : Audits et conformité (Structure Capture 5)
    # =========================================================================
    ws4 = wb.create_sheet(title="Audits et conformité")

    # ── 1. Titre général ─────────────────────────────────────────────────────
    ws4.merge_cells("A1:L2")
    t4_cell = ws4["A1"]
    t4_cell.value = f"Tableau de Bord SMQS - Audits et conformité ({annee})"
    t4_cell.font = Font(name="Calibri", size=15, bold=True, color="1E293B")
    t4_cell.alignment = Alignment(horizontal="center", vertical="center")
    t4_cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    ws4.row_dimensions[1].height = 28
    ws4.row_dimensions[2].height = 10

    # ── 2. KPI en vraies cartes côte à côte ──────────────────────────────────
    kpi_labels = [
        "Conformes / Atteints",
        "Défauts isolés / À surveiller",
        "Non conformes",
        "Non audités / Absents",
    ]
    kpi_keys   = ["conforme", "defauts_isoles", "non_conforme", "non_audite"]
    counts     = data["audits_summary"]["counts"]

    kpi_colors = {
        "conforme":      ("D1FAE5", "065F46"),   # fond vert clair / texte vert foncé
        "defauts_isoles":("FEF9C3", "92400E"),   # fond jaune / texte orange
        "non_conforme":  ("FEE2E2", "991B1B"),   # fond rouge / texte rouge
        "non_audite":    ("F1F5F9", "475569"),   # fond gris / texte bleu-gris
    }

    # 4 cartes : colonnes A-C, D-F, G-I, J-L  (rows 3-5)
    card_cols = [(1, 3), (4, 6), (7, 9), (10, 12)]
    ws4.row_dimensions[3].height = 18
    ws4.row_dimensions[4].height = 28
    ws4.row_dimensions[5].height = 6

    for (c_start, c_end), label, key in zip(card_cols, kpi_labels, kpi_keys):
        bg_hex, fg_hex = kpi_colors[key]
        card_fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")

        # Ligne de titre de la carte
        ws4.merge_cells(start_row=3, start_column=c_start, end_row=3, end_column=c_end)
        lbl_cell = ws4.cell(row=3, column=c_start, value=label)
        lbl_cell.font = Font(name="Calibri", size=9, bold=True, color=fg_hex)
        lbl_cell.alignment = Alignment(horizontal="center", vertical="center")
        lbl_cell.fill = card_fill

        # Ligne de valeur
        ws4.merge_cells(start_row=4, start_column=c_start, end_row=4, end_column=c_end)
        val_cell = ws4.cell(row=4, column=c_start, value=counts[key])
        val_cell.font = Font(name="Calibri", size=18, bold=True, color=fg_hex)
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        val_cell.fill = card_fill

    # ── 3. Sous-titre tableau ─────────────────────────────────────────────────
    ws4.merge_cells("A6:L6")
    sub_cell = ws4["A6"]
    sub_cell.value = "Synthèse de conformité d'audit par processus"
    sub_cell.font = subtitle_font
    sub_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws4.row_dimensions[6].height = 20

    # ── 4. En-têtes tableau principal ─────────────────────────────────────────
    headers_ws4 = [
        "Code & Processus", "Total Indicateurs", "Conformes",
        "Défauts isolés", "Non conformes", "Non audités",
        "Taux de conformité", "Statut global",
    ]

    for col_idx, h in enumerate(headers_ws4, start=1):
        cell = ws4.cell(row=7, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center" if col_idx > 1 else "left",
            vertical="center",
            wrap_text=True,
        )
        cell.border = thin_border
    ws4.row_dimensions[7].height = 28

    # ── 5. Lignes de données ──────────────────────────────────────────────────
    current_row_ws4 = 8
    for audit in data["audits_processus_list"]:
        proc = audit["processus"]
        ws4.row_dimensions[current_row_ws4].height = 18

        # Colonne A : Code & Processus (gauche)
        c_proc = ws4.cell(row=current_row_ws4, column=1, value=f"{proc.code} - {proc.nom}")
        c_proc.alignment = Alignment(horizontal="left", vertical="center")
        c_proc.border = thin_border

        # Colonnes B-F : valeurs numériques (centré)
        for col_offset, col_key in enumerate(
            ["total_indicateurs", "conforme", "defauts_isoles", "non_conforme", "non_audite"], start=2
        ):
            c = ws4.cell(row=current_row_ws4, column=col_offset, value=audit[col_key])
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border

        # Colonne G : taux de conformité (%)
        taux_valeur_pct = audit["taux_conformite_valeur"] / 100.0
        taux_cell = ws4.cell(row=current_row_ws4, column=7, value=taux_valeur_pct)
        taux_cell.number_format = "0.0%"
        taux_cell.alignment = Alignment(horizontal="center", vertical="center")
        taux_cell.border = thin_border

        # Colonne H : statut global coloré
        st_g = audit["statut_global"]
        st_cell = ws4.cell(row=current_row_ws4, column=8, value=st_g["label"])
        st_cell.alignment = Alignment(horizontal="center", vertical="center")
        st_cell.border = thin_border
        st_code = st_g["code"]
        if st_code in status_fills:
            st_cell.fill = status_fills[st_code]
            st_cell.font = Font(
                name="Calibri",
                bold=True,
                size=9,
                color=status_fonts[st_code].color if st_code in status_fonts else "000000",
            )

        current_row_ws4 += 1

    table_end_row = current_row_ws4 - 1

    # ── 6. Petit tableau Statut / Nombre (col J-K, aligné avec en-tête table) ─
    stat_tbl_start = 7
    ws4.cell(row=stat_tbl_start, column=10, value="Statut").font = Font(
        name="Calibri", bold=True, size=9, color="FFFFFF"
    )
    ws4.cell(row=stat_tbl_start, column=10).fill = header_fill
    ws4.cell(row=stat_tbl_start, column=10).alignment = Alignment(horizontal="left", vertical="center")
    ws4.cell(row=stat_tbl_start, column=10).border = thin_border

    ws4.cell(row=stat_tbl_start, column=11, value="Nombre").font = Font(
        name="Calibri", bold=True, size=9, color="FFFFFF"
    )
    ws4.cell(row=stat_tbl_start, column=11).fill = header_fill
    ws4.cell(row=stat_tbl_start, column=11).alignment = Alignment(horizontal="center", vertical="center")
    ws4.cell(row=stat_tbl_start, column=11).border = thin_border

    chart_data_start_row = stat_tbl_start + 1
    for i, (label, key) in enumerate(zip(kpi_labels, kpi_keys)):
        r = stat_tbl_start + 1 + i
        bg_hex, fg_hex = kpi_colors[key]
        row_fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")

        lc = ws4.cell(row=r, column=10, value=label)
        lc.font = Font(name="Calibri", size=9, color=fg_hex)
        lc.fill = row_fill
        lc.alignment = Alignment(horizontal="left", vertical="center")
        lc.border = thin_border

        nc = ws4.cell(row=r, column=11, value=counts[key])
        nc.font = Font(name="Calibri", bold=True, size=9, color=fg_hex)
        nc.fill = row_fill
        nc.alignment = Alignment(horizontal="center", vertical="center")
        nc.border = thin_border

    # ── 7. DoughnutChart (donut) ──────────────────────────────────────────────
    try:
        from openpyxl.chart import DoughnutChart
        donut = DoughnutChart()
        donut.holeSize = 60   # trou plus large = donut moins écrasé
    except (ImportError, AttributeError):
        donut = PieChart()

    donut.title  = "Répartition globale de conformité d'audit"
    donut.height = 8.5   # hauteur suffisante pour le titre + donut + légende
    donut.width  = 15    # largeur large pour que la légende ne chevauche pas le donut

    chart_labels_ref = Reference(
        ws4, min_col=10,
        min_row=chart_data_start_row,
        max_row=chart_data_start_row + 3,
    )
    chart_data_ref = Reference(
        ws4, min_col=11,
        min_row=stat_tbl_start,        # inclut l'en-tête "Nombre" → légende
        max_row=chart_data_start_row + 3,
    )
    donut.add_data(chart_data_ref, titles_from_data=True)
    donut.set_categories(chart_labels_ref)

    # Légende à droite, toujours créée
    from openpyxl.chart.legend import Legend
    donut.legend = Legend()
    donut.legend.position = "r"   # droite
    donut.legend.overlay  = False  # ne pas superposer au plot

    chart_anchor_row = table_end_row + 2
    ws4.add_chart(donut, f"A{chart_anchor_row}")

    # ── 8. Largeurs de colonnes ───────────────────────────────────────────────
    col_widths = {
        "A": 42,   # Code & Processus
        "B": 15,   # Total Indicateurs
        "C": 13,   # Conformes
        "D": 16,   # Défauts isolés
        "E": 15,   # Non conformes
        "F": 14,   # Non audités
        "G": 18,   # Taux de conformité
        "H": 18,   # Statut global
        "I":  3,   # séparateur
        "J": 30,   # Statut (libellé long)
        "K": 10,   # Nombre
        "L":  3,   # marge droite
    }
    for col_letter, width in col_widths.items():
        ws4.column_dimensions[col_letter].width = width



    try:
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    except Exception as exc:
        import traceback
        from django.http import JsonResponse as _JsonResponse
        print("EXCEL EXPORT ERROR:\n", traceback.format_exc())
        return _JsonResponse(
            {"error": str(exc), "traceback": traceback.format_exc()},
            status=500,
        )


@login_required
@tdb_acces_required
def tdb_saisie_mesures(request, processus_id: int, annee: int, mois: int):
    """
    Vue de saisie/modification mensuelle des mesures d'indicateurs d'un processus.
    CONSERVATION OBLIGATOIRE : Un champ soumis vide N'EST PAS supprimé automatiquement.
    """
    processus = get_object_or_404(Processus, pk=processus_id)
    if not (1 <= mois <= 12):
        messages.error(request, "Numéro de mois invalide.")
        return redirect("gestion_documentaire:tdb_dashboard_processus", processus_id=processus.pk)

    indicateurs = processus.indicateurs.filter(is_active=True).order_by("code")
    if not indicateurs.exists():
        messages.warning(request, "Aucun indicateur actif trouvé pour ce processus.")
        return redirect("gestion_documentaire:tdb_dashboard_processus", processus_id=processus.pk)

    mesures_existantes = {
        m.indicateur_id: m
        for m in MesureIndicateur.objects.filter(
            indicateur__in=indicateurs, annee=annee, mois=mois
        )
    }

    if request.method == "POST":
        erreurs = []
        modifications = 0
        try:
            with transaction.atomic():
                for ind in indicateurs:
                    field_name = f"valeur_{ind.id}"
                    raw_val = request.POST.get(field_name, "").strip()

                    if raw_val != "":
                        try:
                            val_decimal = Decimal(raw_val.replace(",", "."))
                            MesureIndicateur.objects.update_or_create(
                                indicateur=ind,
                                annee=annee,
                                mois=mois,
                                defaults={
                                    "valeur": val_decimal,
                                    "saisie_par": request.user,
                                },
                            )
                            modifications += 1
                        except (InvalidOperation, ValueError):
                            erreurs.append(f"Valeur numérique invalide pour l'indicateur {ind.code} ({ind.nom}).")

            if not erreurs:
                messages.success(
                    request,
                    f"Les mesures du mois {mois:02d}/{annee} ont été enregistrées avec succès.",
                )
                return redirect(
                    f"{reverse('gestion_documentaire:tdb_dashboard_processus', kwargs={'processus_id': processus.pk})}?annee={annee}"
                )
            else:
                for err in erreurs:
                    messages.error(request, err)
        except DatabaseError:
            messages.error(request, "Erreur lors de l'enregistrement en base de données.")

    lignes_saisie = []
    for ind in indicateurs:
        mesure_obj = mesures_existantes.get(ind.id)
        valeur_str = ""
        if mesure_obj is not None:
            valeur_str = f"{mesure_obj.valeur:g}".replace(".", ",")

        lignes_saisie.append({
            "indicateur": ind,
            "valeur_actuelle": valeur_str,
            "mesure": mesure_obj,
        })

    return redirect(
        f"{reverse('gestion_documentaire:tdb_dashboard_processus', kwargs={'processus_id': processus.pk})}?annee={annee}"
    )


@login_required
@tdb_acces_required
def tdb_saisie_mesures_indicateur(request, indicateur_id: int, annee: int):
    """Saisie / chargement des mesures mensuelles, des réalisés consolidés et de l'objectif pour un indicateur."""
    indicateur = get_object_or_404(Indicateur, pk=indicateur_id)

    if request.method == "GET":
        mesures_qs = MesureIndicateur.objects.filter(
            indicateur=indicateur,
            annee=annee,
            mois__isnull=False,
        )
        mesures_dict = {
            str(m.mois): formater_nombre_tdb(m.valeur)
            for m in mesures_qs
        }
        realises_qs = RealiseConsolideIndicateur.objects.filter(
            indicateur=indicateur,
            annee_reference=annee,
        )
        realises_dict = {
            str(r.mois_reference): formater_nombre_tdb(r.valeur)
            for r in realises_qs
        }

        composantes_qs = ComposanteIndicateur.objects.filter(indicateur=indicateur)
        valeurs_comp_qs = ValeurComposanteIndicateur.objects.filter(
            composante__indicateur=indicateur,
            annee=annee,
        ).select_related("composante")

        valeurs_comp_dict = {}
        for vc in valeurs_comp_qs:
            valeurs_comp_dict.setdefault(vc.composante.code, {})[str(vc.mois)] = formater_nombre_tdb(vc.valeur)

        data = {
            "mesures": mesures_dict,
            "realises_consolides": realises_dict,
            "mode_calcul": indicateur.mode_calcul,
            "periodicite": indicateur.periodicite,
            "formule": indicateur.formule or "",
            "composantes": [
                {
                    "id": c.pk,
                    "code": c.code,
                    "libelle": c.libelle,
                    "ordre": c.ordre,
                }
                for c in composantes_qs
            ],
            "valeurs_composantes": valeurs_comp_dict,
        }
        try:
            obj = ObjectifIndicateur.objects.get(indicateur=indicateur, annee=annee)
            data["objectif"] = formater_nombre_tdb(obj.valeur_objectif)
        except ObjectifIndicateur.DoesNotExist:
            data["objectif"] = ""

        return JsonResponse(data)

    next_url = request.POST.get("next") or reverse("gestion_documentaire:tdb_dashboard")

    modifications = 0
    erreurs = []

    try:
        with transaction.atomic():
            # 1. Traitement des composantes (Mode Formule)
            if indicateur.mode_calcul == Indicateur.ModeCalcul.FORMULE:
                for comp in indicateur.composantes.all():
                    for m in range(1, 13):
                        field_name = f"comp_{comp.code}_mois_{m}"
                        if field_name in request.POST:
                            raw_val = request.POST.get(field_name, "").strip()
                            if raw_val != "":
                                try:
                                    val_decimal = Decimal(raw_val.replace(",", "."))
                                    ValeurComposanteIndicateur.objects.update_or_create(
                                        composante=comp,
                                        annee=annee,
                                        mois=m,
                                        defaults={
                                            "valeur": val_decimal,
                                            "saisie_par": request.user,
                                        },
                                    )
                                    modifications += 1
                                except (InvalidOperation, ValueError):
                                    erreurs.append(f"Valeur invalide pour {comp.libelle} (mois {m}).")
                            else:
                                ValeurComposanteIndicateur.objects.filter(
                                    composante=comp, annee=annee, mois=m
                                ).delete()

            # 2. Traitement des 12 mois de mesures simples
            for m in range(1, 13):
                field_name = f"mois_{m}"
                if field_name in request.POST:
                    raw_val = request.POST.get(field_name, "").strip()
                    if raw_val != "":
                        try:
                            val_decimal = Decimal(raw_val.replace(",", "."))
                            MesureIndicateur.objects.update_or_create(
                                indicateur=indicateur,
                                annee=annee,
                                mois=m,
                                defaults={
                                    "valeur": val_decimal,
                                    "saisie_par": request.user,
                                },
                            )
                            modifications += 1
                        except (InvalidOperation, ValueError):
                            erreurs.append(f"Valeur invalide pour le mois {m}.")
                    else:
                        MesureIndicateur.objects.filter(
                            indicateur=indicateur, annee=annee, mois=m
                        ).delete()

                # Traitement des réalisés consolidés (mode manuel 12 mois glissants)
                rc_field = f"realise_consolide_{m}"
                if rc_field in request.POST:
                    raw_rc = request.POST.get(rc_field, "").strip()
                    if raw_rc != "":
                        try:
                            rc_decimal = Decimal(raw_rc.replace(",", "."))
                            RealiseConsolideIndicateur.objects.update_or_create(
                                indicateur=indicateur,
                                annee_reference=annee,
                                mois_reference=m,
                                defaults={
                                    "valeur": rc_decimal,
                                    "saisie_par": request.user,
                                },
                            )
                            modifications += 1
                        except (InvalidOperation, ValueError):
                            erreurs.append(f"Valeur de réalisé consolidé invalide pour le mois {m}.")
                    else:
                        RealiseConsolideIndicateur.objects.filter(
                            indicateur=indicateur, annee_reference=annee, mois_reference=m
                        ).delete()

            # 3. Traitement de l'objectif annuel
            if "valeur_objectif" in request.POST:
                raw_obj = request.POST.get("valeur_objectif", "").strip()
                if raw_obj != "":
                    try:
                        obj_decimal = Decimal(raw_obj.replace(",", "."))
                        ObjectifIndicateur.objects.update_or_create(
                            indicateur=indicateur,
                            annee=annee,
                            defaults={"valeur_objectif": obj_decimal},
                        )
                        modifications += 1
                    except (InvalidOperation, ValueError):
                        erreurs.append("Valeur d'objectif annuel invalide.")
                else:
                    ObjectifIndicateur.objects.filter(
                        indicateur=indicateur, annee=annee
                    ).delete()

        if erreurs:
            for err in erreurs:
                messages.error(request, err)
        else:
            messages.success(
                request,
                f"L'indicateur {indicateur.code} ({indicateur.nom}) a été mis à jour pour l'année {annee}.",
            )
    except DatabaseError:
        messages.error(request, "Erreur lors de l'enregistrement en base de données.")

    return redirect(next_url)


@login_required
@tdb_acces_required
@require_POST
def tdb_supprimer_mesure(request, processus_id: int, mesure_id: int):
    """Action explicite de suppression d'une mesure enregistrée avec confirmation."""
    mesure = get_object_or_404(MesureIndicateur, pk=mesure_id, indicateur__processus_id=processus_id)
    annee = mesure.annee
    mesure.delete()
    messages.success(request, "La mesure a été supprimée avec succès.")
    return redirect(
        f"{reverse('gestion_documentaire:tdb_dashboard_processus', kwargs={'processus_id': processus_id})}?annee={annee}"
    )


class ProcessusCreateView(LoginRequiredMixin, CreateView):
    """Vue de création d'un nouveau processus SMQS — réservée aux superusers."""

    model = Processus
    form_class = ProcessusForm

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            from django.contrib.auth.views import redirect_to_login
            return redirect_to_login(request.get_full_path())
        if not request.user.is_superuser:
            raise PermissionDenied("Seuls les administrateurs peuvent créer un processus.")
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        return redirect(reverse("gestion_documentaire:tdb_dashboard") + "?open_modal=processus")

    def get_success_url(self):
        next_url = self.request.GET.get("next") or self.request.POST.get("next")
        if next_url:
            return next_url
        return reverse("gestion_documentaire:tdb_dashboard")

    def form_valid(self, form):
        processus = form.save()

        # Créer les indicateurs en attente — code généré automatiquement
        ind_index = 1
        indicateurs_crees = 0
        while True:
            nom_key = f"ind_nom_{ind_index}"
            # Arrêter si la clé n'est pas soumise
            if nom_key not in self.request.POST:
                break
            nom = self.request.POST.get(nom_key, "").strip()
            periodicite = self.request.POST.get(f"ind_periodicite_{ind_index}", "mensuel").strip()
            agregation = self.request.POST.get(f"ind_agregation_{ind_index}", "somme").strip()
            sens_obj = self.request.POST.get(f"ind_sens_{ind_index}", "atteindre").strip()
            if nom:
                try:
                    code = generer_code_indicateur(processus, extra_offset=indicateurs_crees)
                    Indicateur.objects.create(
                        processus=processus,
                        code=code,
                        nom=nom,
                        periodicite=periodicite,
                        mode_agregation=agregation,
                        sens_objectif=sens_obj,
                        is_active=True,
                    )
                    indicateurs_crees += 1
                except Exception:
                    pass
            ind_index += 1

        msg = f"Le processus « {processus.code} — {processus.nom} » a été créé avec succès."
        if indicateurs_crees:
            msg += f" {indicateurs_crees} indicateur(s) rattaché(s)."
        messages.success(self.request, msg)
        return redirect(self.get_success_url())

    def form_invalid(self, form):
        messages.error(self.request, "Veuillez corriger les erreurs dans le formulaire de création du processus.")
        return redirect(self.get_success_url())

from .forms import (
    DocumentFilterForm,
    DocumentForm,
    DossierDocumentaireForm,
    DossierParametresForm,
    FichierBibliothequeForm,
    IndicateurForm,
    NouvelleRegleAccesFormSet,
    ProcessusForm,
    RegleAccesDossierFormSet,
    VersionDocumentForm,
)


class IndicateurCreateView(LoginRequiredMixin, CreateView):
    """Vue de création d'un nouvel indicateur SMQS."""

    model = Indicateur
    form_class = IndicateurForm

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            from django.contrib.auth.views import redirect_to_login
            return redirect_to_login(request.get_full_path())
        if not user_peut_utiliser_indicateurs_smqs(request.user):
            raise PermissionDenied("Accès réservé aux RS, RO et CE.")
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        proc_id = request.GET.get("processus", "")
        url = reverse("gestion_documentaire:tdb_dashboard") + "?open_modal=indicateur"
        if proc_id:
            url += f"&processus={proc_id}"
        return redirect(url)

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields["processus"].queryset = processus_perimetre_smqs_qs(self.request.user).order_by("code")
        return form

    def get_initial(self):
        initial = super().get_initial()
        processus_id = self.request.GET.get("processus")
        if processus_id:
            initial["processus"] = processus_id
        return initial

    def get_success_url(self):
        next_url = self.request.GET.get("next") or self.request.POST.get("next")
        if next_url:
            return next_url
        return reverse("gestion_documentaire:tdb_dashboard")

    def form_valid(self, form):
        indicateur = form.save(commit=False)
        # Générer le code automatiquement si non renseigné
        if not indicateur.code:
            indicateur.code = generer_code_indicateur(indicateur.processus)

        # Mode formule : valider la formule + les composantes avant la sauvegarde
        comps_to_create = []
        if indicateur.mode_calcul == Indicateur.ModeCalcul.FORMULE:
            formule_val = self.request.POST.get("formule", "").strip()
            comp_codes, comp_libelles = _post_composantes_formule(self.request)
            if not formule_val:
                messages.error(self.request, "La formule est obligatoire pour le mode de calcul automatique par formule.")
                return redirect(self.get_success_url())
            try:
                comps_to_create, codes_clean = _normaliser_composantes_formule(comp_codes, comp_libelles)
            except ValueError as err:
                messages.error(self.request, str(err))
                return redirect(self.get_success_url())
            # Normaliser la casse de la formule pour correspondre aux codes de
            # composantes (stockés en minuscules) — ex : Accidents -> accidents.
            formule_val = formule_val.lower()
            valide, msg_err = valider_formule_securisee(formule_val, codes_clean)
            if not valide:
                messages.error(self.request, f"Formule invalide : {msg_err}")
                return redirect(self.get_success_url())
            indicateur.formule = formule_val

        indicateur.save()
        form.save_m2m()

        # Créer les composantes en mode formule
        if indicateur.mode_calcul == Indicateur.ModeCalcul.FORMULE and comps_to_create:
            with transaction.atomic():
                for ordre_idx, (c_code, c_lib) in enumerate(comps_to_create, start=1):
                    ComposanteIndicateur.objects.create(
                        indicateur=indicateur,
                        code=c_code,
                        libelle=c_lib,
                        ordre=ordre_idx,
                    )

        # Créer l'objectif initial si fourni
        raw_val = self.request.POST.get("valeur_objectif", "").strip()
        raw_annee = self.request.POST.get("objectif_annee", "").strip()
        if raw_val and raw_annee:
            try:
                val_obj = Decimal(raw_val.replace(",", "."))
                annee_obj = int(raw_annee)
                ObjectifIndicateur.objects.update_or_create(
                    indicateur=indicateur,
                    annee=annee_obj,
                    defaults={"valeur_objectif": val_obj},
                )
            except (InvalidOperation, ValueError):
                messages.warning(
                    self.request,
                    f"L'indicateur « {indicateur.code} » a été créé, mais la valeur d'objectif fournie est invalide et n'a pas été enregistrée.",
                )

        messages.success(
            self.request,
            f"L'indicateur « {indicateur.code} — {indicateur.nom} » a été créé avec succès.",
        )

        return redirect(self.get_success_url())

    def form_invalid(self, form):
        for field, errors in form.errors.items():
            field_label = form.fields[field].label if field in form.fields else field
            for error in errors:
                messages.error(self.request, f"{field_label} : {error}")
        return redirect(self.get_success_url())


@login_required
@tdb_acces_required
@require_POST
def tdb_processus_update(request, processus_id: int):
    """Modification d'un processus depuis la modal du tableau de bord."""
    processus = get_object_or_404(Processus, pk=processus_id)
    if not (request.user.is_superuser or request.user.is_staff or getattr(request.user, "is_auditeur", False) or getattr(request.user, "is_RO", False) or getattr(request.user, "is_RS", False) or getattr(request.user, "is_CE", False)):
        raise PermissionDenied("Seuls les pilotes et administrateurs peuvent modifier un processus.")

    nom = request.POST.get("nom", "").strip()
    code = request.POST.get("code", "").strip()
    description = request.POST.get("description", "").strip()
    societe_id = request.POST.get("societe", "").strip()
    is_active = request.POST.get("is_active") == "true" or "is_active" in request.POST

    if nom:
        processus.nom = nom
        if code:
            processus.code = code
        processus.description = description
        if societe_id:
            try:
                processus.societe_id = int(societe_id)
            except ValueError:
                pass
        processus.is_active = is_active
        processus.save()
        messages.success(request, f"Le processus « {processus.code} — {processus.nom} » a été modifié avec succès.")
    else:
        messages.error(request, "Veuillez saisir un nom valide pour le processus.")

    next_url = request.POST.get("next") or reverse("gestion_documentaire:tdb_dashboard")
    return redirect(next_url)


@login_required
@tdb_acces_required
def tdb_processus_detail_json(request, processus_id: int):
    """Retourne les détails d'un processus au format JSON pour pré-remplir la modal de modification."""
    processus = get_object_or_404(Processus, pk=processus_id)
    data = {
        "id": processus.pk,
        "code": processus.code,
        "nom": processus.nom,
        "description": processus.description or "",
        "societe_id": processus.societe_id or "",
        "is_active": processus.is_active,
    }
    return JsonResponse(data)


@login_required
@tdb_acces_required
@require_POST
def tdb_processus_delete(request, processus_id: int):
    """Suppression d'un processus depuis la modal du tableau de bord."""
    processus = get_object_or_404(Processus, pk=processus_id)
    if not (request.user.is_superuser or request.user.is_staff or getattr(request.user, "is_auditeur", False) or getattr(request.user, "is_RO", False) or getattr(request.user, "is_RS", False) or getattr(request.user, "is_CE", False)):
        raise PermissionDenied("Seuls les pilotes et administrateurs peuvent supprimer un processus.")

    nom_format = f"{processus.code} — {processus.nom}"
    processus.delete()
    messages.success(request, f"Le processus « {nom_format} » et tous ses indicateurs rattachés ont été supprimés.")

    next_url = request.POST.get("next") or reverse("gestion_documentaire:tdb_dashboard")
    return redirect(next_url)


@login_required
@tdb_acces_required
def tdb_indicateur_detail_json(request, indicateur_id: int):
    """Retourne les détails d'un indicateur au format JSON (formule et composantes)."""
    indicateur = get_object_or_404(Indicateur, pk=indicateur_id)
    comps = list(indicateur.composantes.all().values("id", "code", "libelle", "ordre"))
    data = {
        "id": indicateur.pk,
        "code": indicateur.code,
        "nom": indicateur.nom,
        "periodicite": indicateur.periodicite,
        "mode_agregation": indicateur.mode_agregation,
        "sens_objectif": indicateur.sens_objectif,
        "mode_calcul": indicateur.mode_calcul,
        "formule": indicateur.formule or "",
        "composantes": comps,
        "is_active": indicateur.is_active,
    }
    return JsonResponse(data)


def _post_composantes_formule(request):
    """Lit les composantes (code/libellé) depuis le POST, avec repli sur les champs numérotés."""
    comp_codes = request.POST.getlist("comp_code")
    comp_libelles = request.POST.getlist("comp_libelle")
    if not comp_codes:
        idx = 1
        while f"comp_code_{idx}" in request.POST:
            c_code = request.POST.get(f"comp_code_{idx}", "").strip()
            c_lib = request.POST.get(f"comp_libelle_{idx}", "").strip()
            if c_code:
                comp_codes.append(c_code)
                comp_libelles.append(c_lib or c_code)
            idx += 1
    return comp_codes, comp_libelles


def _normaliser_composantes_formule(comp_codes, comp_libelles):
    """Normalise les composantes d'une formule et renvoie (comps_to_create, codes_clean)."""
    import re
    comps_to_create = []
    codes_clean = set()
    for i, c_code in enumerate(comp_codes):
        c_clean = c_code.strip().lower()
        c_lib = comp_libelles[i].strip() if i < len(comp_libelles) else c_clean
        if c_clean:
            if not re.match(r"^[a-z0-9_]+$", c_clean):
                raise ValueError(f"Code de composante invalide « {c_clean} ». Utilisez uniquement des lettres minuscules, chiffres et underscores.")
            codes_clean.add(c_clean)
            comps_to_create.append((c_clean, c_lib or c_clean))
    return comps_to_create, codes_clean


def _synchroniser_composantes_indicateur(indicateur, comps_to_create):
    """Crée/met à jour les composantes d'un indicateur et supprime celles qui ne sont plus listées."""
    with transaction.atomic():
        existing_comps = {c.code: c for c in indicateur.composantes.all()}
        kept_codes = set()
        for ordre_idx, (c_code, c_lib) in enumerate(comps_to_create, start=1):
            kept_codes.add(c_code)
            if c_code in existing_comps:
                comp_obj = existing_comps[c_code]
                comp_obj.libelle = c_lib
                comp_obj.ordre = ordre_idx
                comp_obj.save()
            else:
                ComposanteIndicateur.objects.create(
                    indicateur=indicateur,
                    code=c_code,
                    libelle=c_lib,
                    ordre=ordre_idx,
                )
        indicateur.composantes.exclude(code__in=kept_codes).delete()


def _stocker_formulaire_indicateur_erreur(request, indicateur, nom):
    """Conserve les valeurs saisies afin de réafficher la modal avec les données en cas d'erreur."""
    request.session["tdb_indicateur_form_erreur"] = {
        "indicateur_id": indicateur.pk,
        "code": indicateur.code,
        "nom": nom,
        "periodicite": request.POST.get("periodicite", "").strip(),
        "mode_agregation": request.POST.get("mode_agregation", "").strip(),
        "sens_objectif": request.POST.get("sens_objectif", "").strip(),
        "mode_calcul": request.POST.get("mode_calcul", "").strip(),
        "is_active": request.POST.get("is_active") == "true" or "is_active" in request.POST,
        "formule": request.POST.get("formule", "").strip(),
        "type_formule": request.POST.get("type_formule", "").strip(),
        "comp_code": request.POST.getlist("comp_code"),
        "comp_libelle": request.POST.getlist("comp_libelle"),
        "valeur_objectif": request.POST.get("valeur_objectif", "").strip(),
        "objectif_annee": request.POST.get("objectif_annee", "").strip(),
    }


@login_required
@tdb_acces_required
@require_POST
def tdb_indicateur_update(request, indicateur_id: int):
    """Modification d'un indicateur depuis la modal du tableau de bord."""
    indicateur = get_object_or_404(Indicateur, pk=indicateur_id)
    nom = request.POST.get("nom", "").strip()
    periodicite = request.POST.get("periodicite", "").strip()
    agregation = request.POST.get("mode_agregation", "").strip()
    sens_objectif = request.POST.get("sens_objectif", "").strip()
    mode_calcul = request.POST.get("mode_calcul", "").strip()
    is_active = request.POST.get("is_active") == "true" or "is_active" in request.POST
    next_url = request.POST.get("next") or reverse("gestion_documentaire:tdb_dashboard")

    if nom:
        indicateur.nom = nom
        if periodicite:
            indicateur.periodicite = periodicite
        if agregation:
            indicateur.mode_agregation = agregation
        if sens_objectif:
            indicateur.sens_objectif = sens_objectif
        if mode_calcul:
            indicateur.mode_calcul = mode_calcul
        indicateur.is_active = is_active

        # Traitement du mode formule et des composantes
        if mode_calcul == Indicateur.ModeCalcul.FORMULE:
            formule_val = request.POST.get("formule", "").strip()
            comp_codes, comp_libelles = _post_composantes_formule(request)
            if not formule_val:
                _stocker_formulaire_indicateur_erreur(request, indicateur, nom)
                messages.error(request, "La formule est obligatoire pour le mode de calcul automatique par formule.")
                return redirect(next_url)
            try:
                comps_to_create, codes_clean = _normaliser_composantes_formule(comp_codes, comp_libelles)
            except ValueError as err:
                _stocker_formulaire_indicateur_erreur(request, indicateur, nom)
                messages.error(request, str(err))
                return redirect(next_url)
            # Normaliser la casse de la formule pour correspondre aux codes de
            # composantes (stockés en minuscules) — ex : Accidents -> accidents.
            formule_val = formule_val.lower()
            valide, msg_err = valider_formule_securisee(formule_val, codes_clean)
            if not valide:
                _stocker_formulaire_indicateur_erreur(request, indicateur, nom)
                messages.error(request, f"Formule invalide : {msg_err}")
                return redirect(next_url)

            indicateur.formule = formule_val

        indicateur.save()

        if mode_calcul == Indicateur.ModeCalcul.FORMULE:
            _synchroniser_composantes_indicateur(indicateur, comps_to_create)

        # Traitement de l'objectif initial / cible si renseigné
        raw_val = request.POST.get("valeur_objectif", "").strip()
        raw_annee = request.POST.get("objectif_annee", "").strip()
        skip_obj = request.POST.get("skip_objectif") == "true" or "skip_objectif" in request.POST

        if raw_val and raw_annee and not skip_obj:
            try:
                val_obj = Decimal(raw_val.replace(",", "."))
                annee_obj = int(raw_annee)
                ObjectifIndicateur.objects.update_or_create(
                    indicateur=indicateur,
                    annee=annee_obj,
                    defaults={"valeur_objectif": val_obj},
                )
            except (InvalidOperation, ValueError):
                messages.warning(
                    request,
                    f"L'indicateur « {indicateur.code} » a été modifié, mais la valeur d'objectif fournie est invalide et n'a pas été enregistrée.",
                )

        messages.success(request, f"L'indicateur « {indicateur.code} — {indicateur.nom} » a été modifié avec succès.")
    else:
        messages.error(request, "Veuillez saisir un nom valide pour l'indicateur.")

    return redirect(next_url)


@login_required
@tdb_acces_required
@require_POST
def tdb_indicateur_delete(request, indicateur_id: int):
    """Suppression d'un indicateur depuis la modal du tableau de bord."""
    indicateur = get_object_or_404(Indicateur, pk=indicateur_id)
    nom_format = f"{indicateur.code} — {indicateur.nom}"
    indicateur.delete()
    messages.success(request, f"L'indicateur « {nom_format} » a été supprimé avec succès.")

    next_url = request.POST.get("next") or reverse("gestion_documentaire:tdb_dashboard")
    return redirect(next_url)


