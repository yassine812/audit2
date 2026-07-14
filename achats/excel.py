"""Export Excel des demandes d'achat."""

from copy import copy
from io import BytesIO
from pathlib import Path
import unicodedata

from openpyxl import load_workbook


TEMPLATE_PATH = Path(__file__).resolve().parent / "excel_templates" / "Doc07_demande_achat.xlsx"
LINE_START_ROW = 10
LINE_TEMPLATE_ROWS = 8
APPROVAL_START_ROW = 18


def _normalize_text(value):
    normalized = unicodedata.normalize("NFKD", value or "")
    return "".join(character for character in normalized if not unicodedata.combining(character)).lower().strip()


def _select_sheet(workbook, societe_name):
    if not societe_name:
        return workbook[workbook.sheetnames[0]]

    normalized_target = _normalize_text(societe_name)
    exact_match = None
    fallback_match = None
    for sheet_name in workbook.sheetnames:
        normalized_sheet = _normalize_text(sheet_name)
        if normalized_sheet == normalized_target:
            exact_match = sheet_name
            break
        if normalized_target in normalized_sheet or normalized_sheet in normalized_target:
            fallback_match = fallback_match or sheet_name

    return workbook[exact_match or fallback_match or workbook.sheetnames[0]]


def _copy_row_style(worksheet, source_row, target_row):
    for column in range(1, 13):
        source_cell = worksheet.cell(source_row, column)
        target_cell = worksheet.cell(target_row, column)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)

    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
    worksheet.merge_cells(start_row=target_row, start_column=2, end_row=target_row, end_column=3)
    worksheet.merge_cells(start_row=target_row, start_column=10, end_row=target_row, end_column=11)


def _ensure_line_capacity(worksheet, line_count):
    extra_rows = max(0, line_count - LINE_TEMPLATE_ROWS)
    if extra_rows == 0:
        return

    worksheet.insert_rows(APPROVAL_START_ROW, amount=extra_rows)
    source_row = APPROVAL_START_ROW - 1
    for offset in range(extra_rows):
        _copy_row_style(worksheet, source_row, source_row + 1 + offset)


def _format_demandeur_name(demande):
    full_name = demande.demandeur.get_full_name().strip()
    return full_name or demande.demandeur.username


def build_demande_excel(demande):
    workbook = load_workbook(TEMPLATE_PATH)
    societe_name = ""
    if getattr(demande.demandeur, "societe", None):
        societe_name = demande.demandeur.societe.nom or ""
    elif getattr(demande.section_analytique, "societe", None):
        societe_name = demande.section_analytique.societe.nom or ""

    worksheet = _select_sheet(workbook, societe_name)
    for sheet_name in list(workbook.sheetnames):
        if sheet_name != worksheet.title:
            del workbook[sheet_name]
    workbook.active = 0
    worksheet = workbook.active

    lignes = list(demande.lignes.select_related("article_catalogue", "fournisseur_retenu").all())
    _ensure_line_capacity(worksheet, len(lignes))

    demandeur = _format_demandeur_name(demande)
    telephone = getattr(demande.demandeur, "telephone", "") or "-"
    email = demande.demandeur.email or "-"
    delai = demande.delai_souhaite.strftime("%d/%m/%Y") if demande.delai_souhaite else "-"
    date_emission = demande.date_emission.strftime("%d/%m/%Y %H:%M")

    worksheet["A5"] = f"Demandeur : {demandeur}"
    worksheet["D5"] = f"N° de téléphone : {telephone}"
    worksheet["I5"] = f"Mail : {email}"
    worksheet["A6"] = f"Date d'émission : {date_emission}"
    worksheet["D6"] = f"Délai de réception souhaité : {delai}"
    worksheet["I6"] = f"Adresse de livraison :\n{demande.adresse_livraison}"
    worksheet["C3"] = f"{worksheet['C3'].value} - {demande.get_categorie_display()}"

    for index, ligne in enumerate(lignes, start=LINE_START_ROW):
        designation = (
            ligne.article_catalogue.designation if ligne.article_catalogue else ligne.nouvelle_designation
        )
        fournisseur = ligne.fournisseur_retenu.nom if ligne.fournisseur_retenu else ""
        worksheet[f"A{index}"] = demande.section_analytique.Nom
        worksheet[f"B{index}"] = designation or ""
        worksheet[f"D{index}"] = ligne.quantite
        worksheet[f"E{index}"] = ligne.qte_stock
        worksheet[f"F{index}"] = float(ligne.prix_unitaire) if ligne.prix_unitaire is not None else None
        worksheet[f"G{index}"] = float(ligne.prix_total) if ligne.prix_total is not None else None
        worksheet[f"H{index}"] = ligne.reference_fournisseur or ""
        worksheet[f"I{index}"] = fournisseur
        worksheet[f"J{index}"] = ligne.commentaire or ""
        worksheet[f"L{index}"] = ligne.engin_concerne or ""

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()