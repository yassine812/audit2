from decimal import Decimal

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.exceptions import ValidationError
from django.db import transaction
from django.test import TestCase
from django.urls import reverse

from .forms import VersionDocumentForm
from .models import (
    Document,
    DossierDocumentaire,
    FichierBibliotheque,
    Indicateur,
    MesureIndicateur,
    ObjectifIndicateur,
    Processus,
    ProcessusService,
    RegleAccesDossier,
)
from .permissions import ROLE_DIRECTION, ROLE_PILOTE_PROCESSUS, ROLE_QSE, ROLE_UTILISATEUR


User = get_user_model()


class DocumentTransitionTests(TestCase):
    def setUp(self):
        self.processus = ProcessusService.objects.create(code="PM02", libelle="Processus documentaire")
        self.user = User.objects.create_user(username="redacteur", password="testpass123")
        self.document = Document.objects.create(
            type_document=Document.TypeDocument.PROCEDURE,
            processus_service=self.processus,
            numero_ordre=1,
            titre="Gestion documentaire",
            statut=Document.Statut.BROUILLON,
            cree_par=self.user,
        )

    def test_transition_directe_brouillon_vers_applicable_interdite(self):
        with self.assertRaises(ValidationError):
            self.document.transitionner_statut(Document.Statut.APPLICABLE, utilisateur=self.user)

    def test_transition_sequence_complete_autorisee(self):
        self.document.transitionner_statut(Document.Statut.EN_VERIFICATION, utilisateur=self.user)
        self.document.refresh_from_db()
        self.assertEqual(self.document.statut, Document.Statut.EN_VERIFICATION)

        self.document.transitionner_statut(Document.Statut.EN_APPROBATION, utilisateur=self.user)
        self.document.refresh_from_db()
        self.assertEqual(self.document.statut, Document.Statut.EN_APPROBATION)

        self.document.transitionner_statut(Document.Statut.APPLICABLE, utilisateur=self.user)
        self.document.refresh_from_db()
        self.assertEqual(self.document.statut, Document.Statut.APPLICABLE)
        self.assertIsNotNone(self.document.date_application)


class UniqueApplicableConstraintTests(TestCase):
    def setUp(self):
        self.processus = ProcessusService.objects.create(code="PM02", libelle="Processus documentaire")
        self.user = User.objects.create_user(username="qse_user", password="testpass123")

    def test_un_seul_document_applicable_par_code_documentaire(self):
        doc1 = Document.objects.create(
            type_document=Document.TypeDocument.PROCEDURE,
            processus_service=self.processus,
            numero_ordre=1,
            titre="Doc v1",
            statut=Document.Statut.APPLICABLE,
            cree_par=self.user,
        )

        doc2 = Document.objects.create(
            type_document=Document.TypeDocument.PROCEDURE,
            processus_service=self.processus,
            numero_ordre=1,
            titre="Doc v2",
            statut=Document.Statut.BROUILLON,
            cree_par=self.user,
        )

        doc2.statut = Document.Statut.APPLICABLE
        with self.assertRaises(ValidationError):
            with transaction.atomic():
                doc2.save()

        doc1.refresh_from_db()
        doc2.refresh_from_db()
        self.assertEqual(doc1.statut, Document.Statut.APPLICABLE)
        self.assertEqual(doc2.statut, Document.Statut.BROUILLON)


class PermissionsByRoleTests(TestCase):
    def setUp(self):
        self.processus = ProcessusService.objects.create(code="PM02", libelle="Processus documentaire")

        for role in [ROLE_QSE, ROLE_PILOTE_PROCESSUS, ROLE_DIRECTION, ROLE_UTILISATEUR]:
            Group.objects.get_or_create(name=role)

        self.qse = User.objects.create_user(
            username="qse", password="testpass123", is_auditeur=True
        )
        self.pilote = User.objects.create_user(
            username="pilote", password="testpass123", is_RO=True, is_auditeur=True
        )
        self.direction = User.objects.create_user(username="direction", password="testpass123")
        self.standard = User.objects.create_user(username="standard", password="testpass123")

        self.qse.groups.add(Group.objects.get(name=ROLE_QSE))
        self.pilote.groups.add(Group.objects.get(name=ROLE_PILOTE_PROCESSUS))
        self.direction.groups.add(Group.objects.get(name=ROLE_DIRECTION))
        self.standard.groups.add(Group.objects.get(name=ROLE_UTILISATEUR))

        self.doc_applicable = Document.objects.create(
            type_document=Document.TypeDocument.PROCEDURE,
            processus_service=self.processus,
            numero_ordre=10,
            titre="Applicable",
            statut=Document.Statut.APPLICABLE,
            cree_par=self.qse,
        )
        self.doc_brouillon = Document.objects.create(
            type_document=Document.TypeDocument.MODE_OPERATOIRE,
            processus_service=self.processus,
            numero_ordre=11,
            titre="Brouillon",
            statut=Document.Statut.BROUILLON,
            cree_par=self.qse,
        )

    def test_utilisateur_sans_profil_autorise_est_refuse(self):
        self.client.force_login(self.standard)
        response = self.client.get(reverse("gestion_documentaire:document_list"))
        self.assertEqual(response.status_code, 403)

    def test_qse_peut_acceder_dashboard_qse(self):
        self.client.force_login(self.qse)
        response = self.client.get(reverse("gestion_documentaire:dashboard_qse"))
        self.assertEqual(response.status_code, 200)

    def test_utilisateur_standard_refuse_dashboard_qse(self):
        self.client.force_login(self.standard)
        response = self.client.get(reverse("gestion_documentaire:dashboard_qse"))
        self.assertEqual(response.status_code, 403)

    def test_pilote_peut_soumettre_en_verification(self):
        self.client.force_login(self.pilote)
        response = self.client.post(
            reverse("gestion_documentaire:soumettre_verification", kwargs={"pk": self.doc_brouillon.pk})
        )
        self.assertEqual(response.status_code, 302)
        self.doc_brouillon.refresh_from_db()
        self.assertEqual(self.doc_brouillon.statut, Document.Statut.EN_VERIFICATION)


class BibliothequeDocumentaireTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="bibliothecaire", password="testpass123", is_auditeur=True
        )
        self.user.groups.add(Group.objects.get_or_create(name=ROLE_QSE)[0])

    def test_navigation_dans_un_dossier(self):
        dossier = DossierDocumentaire.objects.create(nom="Procédures", cree_par=self.user)
        fichier = FichierBibliotheque.objects.create(
            dossier=dossier,
            fichier="gestion_documentaire/bibliotheque/Procedures/procedure.pdf",
            nom="procedure.pdf",
            taille=9,
            ajoute_par=self.user,
        )
        self.client.force_login(self.user)
        response = self.client.get(
            reverse("gestion_documentaire:dossier_detail", kwargs={"dossier_id": dossier.pk})
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, fichier.nom)

    def test_une_version_vide_est_refusee(self):
        form = VersionDocumentForm(data={"type_increment": "mineur", "resume_changements": ""})
        self.assertFalse(form.is_valid())
        self.assertIn("Ajoutez au moins un fichier", form.non_field_errors()[0])

    def test_les_acces_sont_distincts_pour_chaque_action(self):
        ajout_user = User.objects.create_user(
            username="ajout_seulement", password="testpass123",
            is_RO=True, is_auditeur=True,
        )
        modification_user = User.objects.create_user(
            username="modification_seulement", password="testpass123",
            is_RO=True, is_auditeur=True,
        )
        dossier = DossierDocumentaire.objects.create(
            nom="Dossier strict", acces_restreint=True
        )
        consultation = RegleAccesDossier.objects.create(
            dossier=dossier, actions_autorisees=["lire"]
        )
        consultation.utilisateurs_autorises.add(ajout_user, modification_user)
        ajout = RegleAccesDossier.objects.create(
            dossier=dossier, actions_autorisees=["modifier"]
        )
        ajout.utilisateurs_autorises.add(ajout_user)
        modification = RegleAccesDossier.objects.create(
            dossier=dossier, actions_autorisees=["telecharger"]
        )
        modification.utilisateurs_autorises.add(modification_user)

        self.assertTrue(dossier.utilisateur_autorise(ajout_user, "modifier"))
        self.assertFalse(dossier.utilisateur_autorise(ajout_user, "telecharger"))
        self.assertTrue(dossier.utilisateur_autorise(modification_user, "telecharger"))
        self.assertFalse(dossier.utilisateur_autorise(modification_user, "modifier"))


# ============================================================
# MODULE TABLEAU DE BORD SMQS
# Tests automatisés pour le Tableau de Bord des Indicateurs
# ============================================================

from decimal import Decimal
from accounts.models import Societe
from .models import (
    Processus,
    Indicateur,
    ObjectifIndicateur,
    MesureIndicateur,
    RealiseConsolideIndicateur,
    ComposanteIndicateur,
    ValeurComposanteIndicateur,
)
from .views import (
    obtenir_mois_periode,
    calculer_agregation_indicateur,
    evaluer_statut_indicateur,
    obtenir_donnees_tableau_de_bord,
)


class TdbModuleTests(TestCase):
    def setUp(self):
        self.societe = Societe.objects.create(nom="Société Test TdB")
        self.superuser = User.objects.create_superuser(
            username="admin_tdb", email="admin@test.com", password="password123"
        )
        self.ro_user = User.objects.create_user(
            username="ro_user", password="password123", is_RO=True, societe=self.societe
        )
        self.rs_user = User.objects.create_user(
            username="rs_user", password="password123", is_RS=True, societe=self.societe
        )
        self.ce_user = User.objects.create_user(
            username="ce_user", password="password123", is_CE=True, societe=self.societe
        )
        self.autre_user = User.objects.create_user(
            username="autre_user", password="password123"
        )

        self.processus = Processus.objects.create(
            societe=self.societe,
            code="PM02",
            nom="Management SMQS",
            is_active=True,
        )
        self.processus.RO.add(self.ro_user)
        self.processus.RS.add(self.rs_user)
        self.processus.CE.add(self.ce_user)

        self.ind_somme = Indicateur.objects.create(
            processus=self.processus,
            code="IND01",
            nom="Nombre d'audits",
            periodicite=Indicateur.Periodicite.MENSUEL,
            mode_agregation=Indicateur.ModeAgregation.SOMME,
            formats_chart_disponibles=["bar", "line"],
            is_active=True,
        )
        self.ind_moyenne = Indicateur.objects.create(
            processus=self.processus,
            code="IND02",
            nom="Taux de conformité",
            periodicite=Indicateur.Periodicite.MENSUEL,
            mode_agregation=Indicateur.ModeAgregation.MOYENNE,
            formats_chart_disponibles=["line"],
            is_active=True,
        )

        ObjectifIndicateur.objects.create(
            indicateur=self.ind_somme,
            annee=2026,
            valeur_objectif=Decimal("10.0000"),
        )

    def test_permissions_access_control(self):
        url_direct = reverse("gestion_documentaire:tdb_dashboard")
        url_proc = reverse("gestion_documentaire:tdb_dashboard_processus", kwargs={"processus_id": self.processus.pk})

        # Superuser autorisé sur l'accès direct
        self.client.force_login(self.superuser)
        res = self.client.get(url_direct)
        self.assertEqual(res.status_code, 200)

        # RO affecté autorisé sur l'accès direct
        self.client.force_login(self.ro_user)
        res = self.client.get(url_direct)
        self.assertEqual(res.status_code, 200)

        # RS affecté autorisé sur l'accès direct et sur le processus
        self.client.force_login(self.rs_user)
        res = self.client.get(url_direct)
        self.assertEqual(res.status_code, 200)
        res = self.client.get(url_proc)
        self.assertEqual(res.status_code, 200)

        # CE affecté autorisé sur l'accès direct et sur le processus
        self.client.force_login(self.ce_user)
        res = self.client.get(url_direct)
        self.assertEqual(res.status_code, 200)
        res = self.client.get(url_proc)
        self.assertEqual(res.status_code, 200)

        # Utilisateur non affecté et sans rôle REFUSÉ (403)
        self.client.force_login(self.autre_user)
        res = self.client.get(url_direct)
        self.assertEqual(res.status_code, 403)

    def test_permissions_role_flag_sans_affectation(self):
        # Un utilisateur portant le rôle RS/RO/CE doit pouvoir ouvrir l'accès direct
        # même s'il n'est pas affecté à un processus (le périmètre visible restera vide).
        url_direct = reverse("gestion_documentaire:tdb_dashboard")

        for user in (self.ro_user, self.rs_user, self.ce_user):
            self.processus.RO.remove(user)
            self.processus.RS.remove(user)
            self.processus.CE.remove(user)
            self.client.force_login(user)
            res = self.client.get(url_direct)
            self.assertEqual(res.status_code, 200)

        # Sans rôle ni affectation, accès direct toujours refusé
        self.client.force_login(self.autre_user)
        res = self.client.get(url_direct)
        self.assertEqual(res.status_code, 403)

    def test_perimetre_processus_par_societe(self):
        from gestion_documentaire.views import obtenir_donnees_tableau_de_bord

        societe_b = Societe.objects.create(nom="Société B TdB")

        proc_a2 = Processus.objects.create(
            societe=self.societe, code="PA01", nom="Processus A2", is_active=True
        )
        proc_a_inactif = Processus.objects.create(
            societe=self.societe, code="PAIN", nom="Processus A inactif", is_active=False
        )
        proc_b = Processus.objects.create(
            societe=societe_b, code="PB01", nom="Processus B", is_active=True
        )

        # RS/RO/CE de la société A : voient uniquement les processus actifs de A,
        # indépendamment de toute affectation nominative M2M.
        for user in (self.ro_user, self.rs_user, self.ce_user):
            user.societe = self.societe
            user.save()
            user.processus_RO.clear()
            user.processus_RS.clear()
            user.processus_CE.clear()
            data = obtenir_donnees_tableau_de_bord(user, {})
            codes = {p.code for p in data["tous_processus"]}
            self.assertEqual(codes, {"PM02", "PA01"})

        # RS/RO/CE sans société principale : zéro processus
        user_sans_societe = User.objects.create_user(
            username="ro_sans_societe", password="password123", is_RO=True
        )
        data = obtenir_donnees_tableau_de_bord(user_sans_societe, {})
        self.assertEqual(data["tous_processus"], [])

        # Superuser : voit tous les processus actifs, inactifs exclus
        data = obtenir_donnees_tableau_de_bord(self.superuser, {})
        codes = {p.code for p in data["tous_processus"]}
        self.assertIn("PM02", codes)
        self.assertIn("PA01", codes)
        self.assertIn("PB01", codes)
        self.assertNotIn("PAIN", codes)

        # Utilisateur non RS/RO/CE : périmètre M2M inchangé
        self.autre_user.societe = self.societe
        self.autre_user.save()
        proc_b.RO.add(self.autre_user)
        data = obtenir_donnees_tableau_de_bord(self.autre_user, {})
        codes = {p.code for p in data["tous_processus"]}
        self.assertEqual(codes, {"PB01"})

    def test_perimetre_societe_via_vue_http(self):
        self.ro_user.societe = self.societe
        self.ro_user.save()
        self.ro_user.processus_RO.clear()
        self.ro_user.processus_RS.clear()
        self.ro_user.processus_CE.clear()

        Processus.objects.create(societe=self.societe, code="PA01", nom="Processus A2", is_active=True)
        societe_b = Societe.objects.create(nom="Société B TdB")
        Processus.objects.create(societe=societe_b, code="PB01", nom="Processus B", is_active=True)

        url = reverse("gestion_documentaire:tdb_dashboard")
        self.client.force_login(self.ro_user)
        res = self.client.get(url)
        self.assertEqual(res.status_code, 200)
        codes = {p.code for p in res.context["tous_processus"]}
        self.assertEqual(codes, {"PM02", "PA01"})

    def test_actions_completes_rs_ro_ce(self):
        societe_b = Societe.objects.create(nom="Société B TdB")
        proc_b = Processus.objects.create(
            societe=societe_b, code="PB01", nom="Processus B", is_active=True
        )

        url_dashboard = reverse("gestion_documentaire:tdb_dashboard")
        url_dashboard_proc = reverse(
            "gestion_documentaire:tdb_dashboard_processus",
            kwargs={"processus_id": self.processus.pk},
        )
        url_dashboard_proc_b = reverse(
            "gestion_documentaire:tdb_dashboard_processus",
            kwargs={"processus_id": proc_b.pk},
        )
        url_export = reverse("gestion_documentaire:tdb_export_excel")
        url_json_proc = reverse(
            "gestion_documentaire:tdb_processus_detail_json",
            kwargs={"processus_id": self.processus.pk},
        )
        url_saisie_ind = reverse(
            "gestion_documentaire:tdb_saisie_mesures_indicateur",
            kwargs={"indicateur_id": self.ind_somme.pk, "annee": 2026},
        )
        url_saisie_mois = reverse(
            "gestion_documentaire:tdb_saisie_mesures",
            kwargs={"processus_id": self.processus.pk, "annee": 2026, "mois": 1},
        )
        url_update_proc_b = reverse(
            "gestion_documentaire:tdb_processus_update",
            kwargs={"processus_id": proc_b.pk},
        )

        for user in (self.ro_user, self.rs_user, self.ce_user):
            user.societe = self.societe
            user.save()
            user.processus_RO.clear()
            user.processus_RS.clear()
            user.processus_CE.clear()

            self.client.force_login(user)

            # Accès direct au module
            self.assertEqual(self.client.get(url_dashboard).status_code, 200)
            # Processus de sa société (même sans affectation M2M)
            self.assertEqual(self.client.get(url_dashboard_proc).status_code, 200)
            # Processus d'une autre société -> 403 (périmètre)
            self.assertEqual(self.client.get(url_dashboard_proc_b).status_code, 403)

            # Export Excel (POST) -> 200, pas de 403
            self.assertEqual(self.client.post(url_export).status_code, 200)

            # Endpoints AJAX / JSON
            self.assertEqual(self.client.get(url_json_proc).status_code, 200)
            self.assertEqual(self.client.get(url_saisie_ind).status_code, 200)
            self.assertIn(self.client.get(url_saisie_mois).status_code, (200, 302))

            # Action d'écriture sur un processus hors périmètre -> 403
            self.assertEqual(self.client.post(url_update_proc_b, {"nom": "X"}).status_code, 403)

    def test_export_excel_sans_societe(self):
        url_export = reverse("gestion_documentaire:tdb_export_excel")
        user_sans_societe = User.objects.create_user(
            username="ro_sans_societe", password="password123", is_RO=True
        )
        self.client.force_login(user_sans_societe)
        res = self.client.post(url_export)
        self.assertEqual(res.status_code, 200)

        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(res.content))
        self.assertEqual(len(wb.sheetnames), 4)

    def test_export_excel_non_autorise(self):
        url_export = reverse("gestion_documentaire:tdb_export_excel")
        self.client.force_login(self.autre_user)
        res = self.client.post(url_export)
        self.assertEqual(res.status_code, 403)

    def test_calcul_agregation_avec_mesure_annuelle_mois_none(self):
        # M1=2.0, M2=0.0 (vraie mesure zero!), et mesure annuelle mois=None=5.0
        m1 = MesureIndicateur.objects.create(
            indicateur=self.ind_somme, annee=2026, mois=1, valeur=Decimal("2.0000"), saisie_par=self.ro_user
        )
        m2 = MesureIndicateur.objects.create(
            indicateur=self.ind_somme, annee=2026, mois=2, valeur=Decimal("0.0000"), saisie_par=self.ro_user
        )
        m_annuelle = MesureIndicateur.objects.create(
            indicateur=self.ind_somme, annee=2026, mois=None, valeur=Decimal("5.0000"), saisie_par=self.ro_user
        )

        mesures = [m1, m2, m_annuelle]

        # En Année complète : les 3 sont sommées (2 + 0 + 5 = 7)
        somme_annee = calculer_agregation_indicateur(self.ind_somme, mesures, type_periode="annee_complete")
        self.assertEqual(somme_annee, Decimal("7.0000"))

        # En Trimestre T1 : la mesure annuelle mois=None est EXCLUE (2 + 0 = 2)
        somme_t1 = calculer_agregation_indicateur(self.ind_somme, mesures, type_periode="trimestre")
        self.assertEqual(somme_t1, Decimal("2.0000"))

    def test_valeur_annuelle_repli_sur_agregat_sans_mesure_stockee(self):
        from gestion_documentaire.views import obtenir_donnees_tableau_de_bord
        MesureIndicateur.objects.create(
            indicateur=self.ind_somme, annee=2026, mois=1, valeur=Decimal("2.0000"), saisie_par=self.ro_user
        )
        MesureIndicateur.objects.create(
            indicateur=self.ind_somme, annee=2026, mois=2, valeur=Decimal("3.0000"), saisie_par=self.ro_user
        )

        data = obtenir_donnees_tableau_de_bord(
            self.ro_user, {"annee": "2026", "processus": str(self.processus.id)}
        )
        ind_item = next(
            (item for item in data["indicateurs_data"] if item["indicateur"].id == self.ind_somme.id), None
        )
        self.assertIsNotNone(ind_item)
        # Pas de mesure annuelle stockée (mois=None) : repli sur l'agrégat calculé
        self.assertEqual(ind_item["cumul"], Decimal("5.0000"))
        self.assertEqual(ind_item["mesure_annuelle"], Decimal("5.0000"))
        self.assertEqual(ind_item["valeur_annuelle_formatee"], "5")

    def test_valeur_annuelle_priorite_mesure_stockee(self):
        from gestion_documentaire.views import obtenir_donnees_tableau_de_bord
        MesureIndicateur.objects.create(
            indicateur=self.ind_somme, annee=2026, mois=1, valeur=Decimal("2.0000"), saisie_par=self.ro_user
        )
        MesureIndicateur.objects.create(
            indicateur=self.ind_somme, annee=2026, mois=2, valeur=Decimal("3.0000"), saisie_par=self.ro_user
        )
        MesureIndicateur.objects.create(
            indicateur=self.ind_somme, annee=2026, mois=None, valeur=Decimal("7.0000"), saisie_par=self.ro_user
        )

        data = obtenir_donnees_tableau_de_bord(
            self.ro_user, {"annee": "2026", "processus": str(self.processus.id)}
        )
        ind_item = next(
            (item for item in data["indicateurs_data"] if item["indicateur"].id == self.ind_somme.id), None
        )
        self.assertIsNotNone(ind_item)
        # L'agrégat Cumul/Moy. en année complète inclut la mesure annuelle (2+3+7=12),
        # règle métier inchangée ; Ann. doit utiliser la mesure stockée (pas de repli)
        self.assertEqual(ind_item["cumul"], Decimal("12.0000"))
        self.assertEqual(ind_item["mesure_annuelle"], Decimal("7.0000"))
        self.assertEqual(ind_item["valeur_annuelle_formatee"], "7")

    def test_valeur_annuelle_zero_stockee_preservee(self):
        from gestion_documentaire.views import obtenir_donnees_tableau_de_bord
        MesureIndicateur.objects.create(
            indicateur=self.ind_somme, annee=2026, mois=1, valeur=Decimal("2.0000"), saisie_par=self.ro_user
        )
        MesureIndicateur.objects.create(
            indicateur=self.ind_somme, annee=2026, mois=None, valeur=Decimal("0.0000"), saisie_par=self.ro_user
        )

        data = obtenir_donnees_tableau_de_bord(
            self.ro_user, {"annee": "2026", "processus": str(self.processus.id)}
        )
        ind_item = next(
            (item for item in data["indicateurs_data"] if item["indicateur"].id == self.ind_somme.id), None
        )
        self.assertIsNotNone(ind_item)
        # La valeur 0 est valide : pas de repli sur l'agrégat (2), pas de "—"
        self.assertEqual(ind_item["mesure_annuelle"], Decimal("0.0000"))
        self.assertEqual(ind_item["valeur_annuelle_formatee"], "0")

    def test_valeur_annuelle_absente_sans_agregat(self):
        from gestion_documentaire.views import obtenir_donnees_tableau_de_bord
        ind_sans = Indicateur.objects.create(
            processus=self.processus,
            code="IND99",
            nom="Indicateur sans données",
            periodicite=Indicateur.Periodicite.MENSUEL,
            mode_agregation=Indicateur.ModeAgregation.SOMME,
            is_active=True,
        )

        data = obtenir_donnees_tableau_de_bord(
            self.ro_user, {"annee": "2026", "processus": str(self.processus.id)}
        )
        ind_item = next(
            (item for item in data["indicateurs_data"] if item["indicateur"].id == ind_sans.id), None
        )
        self.assertIsNotNone(ind_item)
        # Ni mesure annuelle ni agrégat calculable : Ann. doit rester "—"
        self.assertIsNone(ind_item["mesure_annuelle"])
        self.assertEqual(ind_item["valeur_annuelle_formatee"], "—")

    def test_saisie_mesures_conservation_champ_vide(self):
        url_saisie = reverse(
            "gestion_documentaire:tdb_saisie_mesures",
            kwargs={"processus_id": self.processus.pk, "annee": 2026, "mois": 1},
        )
        self.client.force_login(self.ro_user)

        # 1. Création initiale d'une mesure M1 = 15.5 pour IND01
        MesureIndicateur.objects.create(
            indicateur=self.ind_somme, annee=2026, mois=1, valeur=Decimal("15.5000"), saisie_par=self.ro_user
        )

        # 2. Soumission d'une chaîne vide sur IND01 => RÈGLE : CONSERVER la mesure existante (pas de delete)
        res = self.client.post(url_saisie, {
            f"valeur_{self.ind_somme.id}": "",
            f"valeur_{self.ind_moyenne.id}": "8",
        })
        self.assertEqual(res.status_code, 302)

        # Vérification que la mesure IND01 existe TOUJOURS
        m_somme = MesureIndicateur.objects.get(indicateur=self.ind_somme, annee=2026, mois=1)
        self.assertEqual(m_somme.valeur, Decimal("15.5000"))

    def test_suppression_explicite_mesure(self):
        m = MesureIndicateur.objects.create(
            indicateur=self.ind_somme, annee=2026, mois=1, valeur=Decimal("10.0000"), saisie_par=self.ro_user
        )
        url_supprimer = reverse(
            "gestion_documentaire:tdb_supprimer_mesure",
            kwargs={"processus_id": self.processus.pk, "mesure_id": m.pk},
        )
        self.client.force_login(self.ro_user)
        res = self.client.post(url_supprimer)
        self.assertEqual(res.status_code, 302)
        self.assertFalse(MesureIndicateur.objects.filter(pk=m.pk).exists())

    def test_onglet_audits_etat_vide_sans_mapping(self):
        url = reverse("gestion_documentaire:tdb_dashboard") + "?tab=audits"
        self.client.force_login(self.ro_user)
        res = self.client.get(url)
        self.assertEqual(res.status_code, 200)
        self.assertContains(res, "Répartition Globale des Audits")

    def test_evaluer_statut_indicateur_sens_ne_pas_depasser_tf(self):
        ind_tf = Indicateur.objects.create(
            processus=self.processus,
            code="TF01",
            nom="Taux de fréquence",
            periodicite=Indicateur.Periodicite.MENSUEL,
            mode_agregation=Indicateur.ModeAgregation.MOYENNE,
            sens_objectif=Indicateur.SensObjectif.NE_PAS_DEPASSER,
            is_active=True,
        )
        obj = Decimal("15.0000")

        # Exemple TF : Objectif = 15, Réalisé = 19.26 => Objectif non atteint (non_conforme)
        res_non_conforme = evaluer_statut_indicateur(ind_tf, Decimal("19.2600"), obj)
        self.assertEqual(res_non_conforme["code"], "non_conforme")
        self.assertEqual(res_non_conforme["label"], "Objectif non atteint")
        self.assertEqual(res_non_conforme["badge_class"], "badge-danger")
        self.assertAlmostEqual(res_non_conforme["taux_atteinte"], 77.88, places=2)

        # Exemple TF : Objectif = 15, Réalisé = 12 => Objectif atteint (conforme)
        res_conforme = evaluer_statut_indicateur(ind_tf, Decimal("12.0000"), obj)
        self.assertEqual(res_conforme["code"], "conforme")
        self.assertEqual(res_conforme["label"], "Objectif atteint")
        self.assertEqual(res_conforme["badge_class"], "badge-success")
        self.assertAlmostEqual(res_conforme["taux_atteinte"], 125.00, places=2)

        # Cas Réalisé = 0 => Objectif atteint, pas de division par zéro
        res_zero = evaluer_statut_indicateur(ind_tf, Decimal("0.0000"), obj)
        self.assertEqual(res_zero["code"], "conforme")
        self.assertEqual(res_zero["taux_atteinte"], 100.00)

    def test_creation_et_modification_sens_objectif(self):
        url_update = reverse(
            "gestion_documentaire:tdb_indicateur_update",
            kwargs={"indicateur_id": self.ind_somme.pk},
        )
        self.client.force_login(self.ro_user)

        res = self.client.post(url_update, {
            "nom": "Nombre d'audits modifié",
            "periodicite": "mensuel",
            "mode_agregation": "somme",
            "sens_objectif": "ne_pas_depasser",
            "is_active": "true",
        })
        self.assertEqual(res.status_code, 302)

        self.ind_somme.refresh_from_db()
        self.assertEqual(self.ind_somme.sens_objectif, Indicateur.SensObjectif.NE_PAS_DEPASSER)

    def test_export_excel_graphique_combine_et_ligne_objectif(self):
        url_export = reverse("gestion_documentaire:tdb_export_excel")
        self.client.force_login(self.ro_user)
        res = self.client.post(url_export, data='{"chart_type": "bar"}', content_type="application/json")
        self.assertEqual(res.status_code, 200)

        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(res.content))
        ws2 = wb["Vue graphique"]
        # Vérification des en-têtes de la table source (Mois, Réalisé, Objectif)
        self.assertEqual(ws2.cell(row=9, column=1).value, "Mois")
        self.assertEqual(ws2.cell(row=9, column=2).value, "Réalisé")
        self.assertEqual(ws2.cell(row=9, column=3).value, "Objectif")
        self.assertEqual(len(ws2._charts), 2)  # au moins 2 graphiques présents

    def test_indicateur_mode_calcul_et_realise_consolide(self):
        from gestion_documentaire.models import RealiseConsolideIndicateur
        ind_glissant = Indicateur.objects.create(
            processus=self.processus,
            code="PM02-TF",
            nom="Taux de fréquence 12M",
            periodicite=Indicateur.Periodicite.GLISSANT_12_MOIS,
            mode_agregation=Indicateur.ModeAgregation.MOYENNE,
            sens_objectif=Indicateur.SensObjectif.NE_PAS_DEPASSER,
            mode_calcul=Indicateur.ModeCalcul.MANUEL,
        )
        ObjectifIndicateur.objects.create(indicateur=ind_glissant, annee=2026, valeur_objectif=Decimal("15"))

        # Coexistence d'une composante mensuelle (3.4) et d'un réalisé consolidé manuel (19.26)
        MesureIndicateur.objects.create(indicateur=ind_glissant, annee=2026, mois=3, valeur=Decimal("3.40"))
        RealiseConsolideIndicateur.objects.create(
            indicateur=ind_glissant, annee_reference=2026, mois_reference=3, valeur=Decimal("19.26")
        )

        data = obtenir_donnees_tableau_de_bord(self.ro_user, {"annee": "2026", "processus": str(self.processus.id)})
        ind_item = next((item for item in data["indicateurs_data"] if item["indicateur"].id == ind_glissant.id), None)

        self.assertIsNotNone(ind_item)
        # La valeur agrégée KPI doit être le dernier réalisé glissant disponible (19.26)
        self.assertEqual(ind_item["cumul"], Decimal("19.26"))
        self.assertEqual(ind_item["statut"]["code"], "non_conforme")  # 19.26 > 15 pour ne_pas_depasser

    def test_mode_manuel_indicateur_mensuel_exclusivite_realise_consolide(self):
        from gestion_documentaire.models import RealiseConsolideIndicateur
        ind_mensuel_manuel = Indicateur.objects.create(
            processus=self.processus,
            code="PM02-MAN",
            nom="Indicateur Mensuel Manuel Exclusif",
            periodicite=Indicateur.Periodicite.MENSUEL,
            mode_agregation=Indicateur.ModeAgregation.SOMME,
            mode_calcul=Indicateur.ModeCalcul.MANUEL,
        )
        # Mesures mensuelles qui donneraient 9 en automatique (2+3+4)
        MesureIndicateur.objects.create(indicateur=ind_mensuel_manuel, annee=2026, mois=1, valeur=Decimal("2"))
        MesureIndicateur.objects.create(indicateur=ind_mensuel_manuel, annee=2026, mois=2, valeur=Decimal("3"))
        MesureIndicateur.objects.create(indicateur=ind_mensuel_manuel, annee=2026, mois=3, valeur=Decimal("4"))

        # Réalisé consolidé manuel saisi = 19.26
        RealiseConsolideIndicateur.objects.create(
            indicateur=ind_mensuel_manuel, annee_reference=2026, mois_reference=3, valeur=Decimal("19.26")
        )

        data = obtenir_donnees_tableau_de_bord(self.ro_user, {"annee": "2026", "processus": str(self.processus.id)})
        ind_item = next((item for item in data["indicateurs_data"] if item["indicateur"].id == ind_mensuel_manuel.id), None)

        self.assertIsNotNone(ind_item)
        # La valeur agrégée doit être 19.26 et surtout PAS 9.00
        self.assertEqual(ind_item["cumul"], Decimal("19.26"))
        self.assertNotEqual(ind_item["cumul"], Decimal("9.00"))

    def test_calcul_fenetre_12_mois_glissants_automatique(self):
        ind_auto = Indicateur.objects.create(
            processus=self.processus,
            code="PM02-AUTO",
            nom="Indicateur Glissant Auto",
            periodicite=Indicateur.Periodicite.GLISSANT_12_MOIS,
            mode_agregation=Indicateur.ModeAgregation.SOMME,
            mode_calcul=Indicateur.ModeCalcul.AUTOMATIQUE,
        )

        # 9 mois en 2025 (Avril à Décembre, 10 par mois = 90)
        for m in range(4, 13):
            MesureIndicateur.objects.create(indicateur=ind_auto, annee=2025, mois=m, valeur=Decimal("10"))
        # 3 mois en 2026 (Janvier à Mars, 10 par mois = 30)
        for m in range(1, 4):
            MesureIndicateur.objects.create(indicateur=ind_auto, annee=2026, mois=m, valeur=Decimal("10"))

        data = obtenir_donnees_tableau_de_bord(self.ro_user, {"annee": "2026", "processus": str(self.processus.id)})
        ind_item = next((item for item in data["indicateurs_data"] if item["indicateur"].id == ind_auto.id), None)

        self.assertIsNotNone(ind_item)
        # Fenêtre de 12 mois (Avril 2025 à Mars 2026) : 12 * 10 = 120.0000
        self.assertEqual(ind_item["cumul"], Decimal("120.0000"))
        # L'axe X affiche les mois de fin de chaque fenêtre glissante, avec l'année
        self.assertEqual(ind_item["serie_mensuelle_labels"], [
            "Jan 2026", "Fév 2026", "Mar 2026", "Avr 2026", "Mai 2026", "Juin 2026",
            "Juil 2026", "Août 2026", "Sept 2026", "Oct 2026", "Nov 2026", "Déc 2026",
        ])

    def test_preservation_donnees_basculement_mode(self):
        from gestion_documentaire.models import RealiseConsolideIndicateur
        ind = Indicateur.objects.create(
            processus=self.processus,
            code="PM02-SW",
            nom="Indicateur Switch Mode",
            periodicite=Indicateur.Periodicite.GLISSANT_12_MOIS,
            mode_calcul=Indicateur.ModeCalcul.MANUEL,
        )
        rc = RealiseConsolideIndicateur.objects.create(
            indicateur=ind, annee_reference=2026, mois_reference=1, valeur=Decimal("50")
        )

        # Basculement en mode AUTOMATIQUE
        ind.mode_calcul = Indicateur.ModeCalcul.AUTOMATIQUE
        ind.save()

        # Vérification : l'enregistrement RealiseConsolideIndicateur n'est PAS supprimé
        self.assertTrue(RealiseConsolideIndicateur.objects.filter(pk=rc.pk).exists())


class TdbCrudAndCodeGenerationTests(TestCase):
    def setUp(self):
        from accounts.models import Societe
        self.admin_user = User.objects.create_superuser(
            username="admin_test", email="admin@test.com", password="adminpassword123"
        )
        self.societe = Societe.objects.create(nom="Société AB Serve")
        self.processus = Processus.objects.create(
            code="CE", nom="Contrôle outil", description="Processus de test", societe=self.societe
        )
        self.indicateur = Indicateur.objects.create(
            processus=self.processus,
            code="CE01",
            nom="Contrôle outil ind",
            periodicite="mensuel",
            mode_agregation="somme",
        )

    def test_generer_code_processus_depuis_nom(self):
        from .views import generer_code_processus_depuis_nom
        code_ce = generer_code_processus_depuis_nom("Contrôle outil")
        self.assertTrue(code_ce.startswith("CE"))

        code_pa = generer_code_processus_depuis_nom("Pilotage et amélioration")
        self.assertEqual(code_pa, "PA")

        code_mq = generer_code_processus_depuis_nom("Management de la qualité")
        self.assertEqual(code_mq, "MQ")

    def test_generer_code_indicateur_format(self):
        from .views import generer_code_indicateur
        proc = Processus.objects.create(code="TEST", nom="Processus test code", societe=self.societe)
        code1 = generer_code_indicateur(proc)
        self.assertEqual(code1, "TEST01")

        Indicateur.objects.create(processus=proc, code=code1, nom="Ind 1")
        code2 = generer_code_indicateur(proc)
        self.assertEqual(code2, "TEST02")

        code3 = generer_code_indicateur(proc, extra_offset=1)
        self.assertEqual(code3, "TEST03")

    def test_saisie_mesures_indicateur_get_json_strict(self):
        # Mois 1 = 1, Mois 2 = 0, Mois 3..12 non mesurés
        MesureIndicateur.objects.create(
            indicateur=self.indicateur, annee=2026, mois=1, valeur=Decimal("1.0000"), saisie_par=self.admin_user
        )
        MesureIndicateur.objects.create(
            indicateur=self.indicateur, annee=2026, mois=2, valeur=Decimal("0.0000"), saisie_par=self.admin_user
        )

        url = reverse(
            "gestion_documentaire:tdb_saisie_mesures_indicateur",
            kwargs={"indicateur_id": self.indicateur.pk, "annee": 2026},
        )
        self.client.force_login(self.admin_user)
        res = self.client.get(url)
        self.assertEqual(res.status_code, 200)
        data = res.json()

        # Seuls les mois 1 et 2 doivent figurer
        self.assertIn("mesures", data)
        self.assertEqual(data["mesures"], {"1": "1", "2": "0"})
        self.assertNotIn("3", data["mesures"])

    def test_processus_update_view(self):
        url = reverse(
            "gestion_documentaire:tdb_processus_update",
            kwargs={"processus_id": self.processus.pk},
        )
        self.client.force_login(self.admin_user)
        res = self.client.post(url, {
            "code": "CE",
            "nom": "Contrôle outil modifié",
            "description": "Nouvelle description",
            "societe": str(self.societe.pk),
            "is_active": "true",
        })
        self.assertEqual(res.status_code, 302)
        self.processus.refresh_from_db()
        self.assertEqual(self.processus.nom, "Contrôle outil modifié")
        self.assertEqual(self.processus.description, "Nouvelle description")

    def test_processus_delete_view(self):
        url = reverse(
            "gestion_documentaire:tdb_processus_delete",
            kwargs={"processus_id": self.processus.pk},
        )
        self.client.force_login(self.admin_user)
        res = self.client.post(url)
        self.assertEqual(res.status_code, 302)
        self.assertFalse(Processus.objects.filter(pk=self.processus.pk).exists())

    def test_indicateur_update_view(self):
        url = reverse(
            "gestion_documentaire:tdb_indicateur_update",
            kwargs={"indicateur_id": self.indicateur.pk},
        )
        self.client.force_login(self.admin_user)
        res = self.client.post(url, {
            "nom": "Contrôle outil ind maj",
            "periodicite": "trimestriel",
            "mode_agregation": "moyenne",
            "is_active": "true",
            "valeur_objectif": "95.5",
            "objectif_annee": "2026",
        })
        self.assertEqual(res.status_code, 302)
        self.indicateur.refresh_from_db()
        self.assertEqual(self.indicateur.nom, "Contrôle outil ind maj")
        self.assertEqual(self.indicateur.periodicite, "trimestriel")
        self.assertEqual(self.indicateur.mode_agregation, "moyenne")
        obj = ObjectifIndicateur.objects.get(indicateur=self.indicateur, annee=2026)
        self.assertEqual(obj.valeur_objectif, Decimal("95.5000"))

    def test_indicateur_agregation_modes(self):
        ind_min = Indicateur.objects.create(
            processus=self.processus,
            code="IND_MIN",
            nom="Min test",
            periodicite=Indicateur.Periodicite.MENSUEL,
            mode_agregation=Indicateur.ModeAgregation.MINIMUM,
        )
        ind_max = Indicateur.objects.create(
            processus=self.processus,
            code="IND_MAX",
            nom="Max test",
            periodicite=Indicateur.Periodicite.MENSUEL,
            mode_agregation=Indicateur.ModeAgregation.MAXIMUM,
        )
        ind_num = Indicateur.objects.create(
            processus=self.processus,
            code="IND_NUM",
            nom="Nombre test",
            periodicite=Indicateur.Periodicite.MENSUEL,
            mode_agregation=Indicateur.ModeAgregation.NOMBRE,
        )

        MesureIndicateur.objects.create(indicateur=ind_min, annee=2026, mois=1, valeur=Decimal("15.5"))
        MesureIndicateur.objects.create(indicateur=ind_min, annee=2026, mois=2, valeur=Decimal("5.0"))
        MesureIndicateur.objects.create(indicateur=ind_min, annee=2026, mois=3, valeur=Decimal("20.0"))

        MesureIndicateur.objects.create(indicateur=ind_max, annee=2026, mois=1, valeur=Decimal("15.5"))
        MesureIndicateur.objects.create(indicateur=ind_max, annee=2026, mois=2, valeur=Decimal("5.0"))
        MesureIndicateur.objects.create(indicateur=ind_max, annee=2026, mois=3, valeur=Decimal("20.0"))

        MesureIndicateur.objects.create(indicateur=ind_num, annee=2026, mois=1, valeur=Decimal("15.5"))
        MesureIndicateur.objects.create(indicateur=ind_num, annee=2026, mois=2, valeur=Decimal("5.0"))

        mesures_min = list(MesureIndicateur.objects.filter(indicateur=ind_min))
        mesures_max = list(MesureIndicateur.objects.filter(indicateur=ind_max))
        mesures_num = list(MesureIndicateur.objects.filter(indicateur=ind_num))

        self.assertEqual(calculer_agregation_indicateur(ind_min, mesures_min), Decimal("5.0000"))
        self.assertEqual(calculer_agregation_indicateur(ind_max, mesures_max), Decimal("20.0000"))
        self.assertEqual(calculer_agregation_indicateur(ind_num, mesures_num), Decimal("2.0000"))

    def test_processus_create_with_indicator_agregation(self):
        url = reverse("gestion_documentaire:tdb_processus_create")
        self.client.force_login(self.admin_user)
        res = self.client.post(url, {
            "code": "PR_TEST",
            "nom": "Processus avec Agrégation",
            "description": "Test agregation",
            "societe": self.processus.societe_id,
            "ind_nom_1": "Indicateur Minimum",
            "ind_periodicite_1": "mensuel",
            "ind_agregation_1": "minimum",
        })
        self.assertEqual(res.status_code, 302)
        proc = Processus.objects.get(code="PR_TEST")
        ind = proc.indicateurs.first()
        self.assertIsNotNone(ind)
    def test_indicateur_create_modal_without_code(self):
        url = reverse("gestion_documentaire:indicateur_create")
        self.client.force_login(self.admin_user)
        res = self.client.post(url, {
            "processus": self.processus.id,
            "nom": "Taux de satisfaction client",
            "periodicite": "mensuel",
            "mode_agregation": "somme",
            "is_active": "true",
            "next": "/gestion-documentaire/tableau-de-bord/",
        })
        self.assertEqual(res.status_code, 302)
        ind = Indicateur.objects.get(nom="Taux de satisfaction client")
        self.assertTrue(ind.code.startswith(self.processus.code))


class TdbExportExcelTest(TestCase):
    def setUp(self):
        from accounts.models import Societe
        from django.contrib.auth import get_user_model
        User = get_user_model()
        self.user = User.objects.create_superuser("excel_user", "excel@test.com", "pass123")
        self.societe = Societe.objects.create(nom="Société Test Excel")
        self.processus = Processus.objects.create(code="PM01", nom="Pilotage", societe=self.societe)
        self.processus.RO.add(self.user)

    def test_export_excel_endpoint(self):
        import io
        from openpyxl import load_workbook
        url = reverse("gestion_documentaire:tdb_export_excel")
        self.client.force_login(self.user)
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        self.assertIn("attachment; filename=", response["Content-Disposition"])

        wb = load_workbook(io.BytesIO(response.content))
        sheet_names = wb.sheetnames

        self.assertEqual(len(sheet_names), 4)
        self.assertIn("Synthèse des indicateurs", sheet_names)
        self.assertIn("Vue graphique", sheet_names)
        self.assertIn("Objectifs annuels", sheet_names)
        self.assertIn("Audits et conformité", sheet_names)

    def test_export_excel_post_chart_types(self):
        import io
        from openpyxl import load_workbook
        url = reverse("gestion_documentaire:tdb_export_excel")
        self.client.force_login(self.user)
        payload = {
            "chart_types": {"PM01-01": "bar", "PM01-02": "area"}
        }
        response = self.client.post(url, data=payload, content_type="application/json")
        self.assertEqual(response.status_code, 200)
        wb = load_workbook(io.BytesIO(response.content))
        self.assertIn("Vue graphique", wb.sheetnames)

    def test_export_excel_labels_glissants_avec_annee(self):
        """Un indicateur GLISSANT_12_MOIS doit inscrire ses labels roulants (mois + année) dans le classeur."""
        import io
        from openpyxl import load_workbook

        ind_glissant = Indicateur.objects.create(
            processus=self.processus,
            code="PM01-GL",
            nom="Glissant Excel",
            periodicite=Indicateur.Periodicite.GLISSANT_12_MOIS,
            mode_agregation=Indicateur.ModeAgregation.SOMME,
            mode_calcul=Indicateur.ModeCalcul.AUTOMATIQUE,
        )
        # 12 mesures traversant 2025/2026 (fenêtres roulantes terminant chaque mois de 2026)
        for m in range(4, 13):
            MesureIndicateur.objects.create(indicateur=ind_glissant, annee=2025, mois=m, valeur=Decimal("10"))
        for m in range(1, 4):
            MesureIndicateur.objects.create(indicateur=ind_glissant, annee=2026, mois=m, valeur=Decimal("10"))

        url = reverse("gestion_documentaire:tdb_export_excel")
        self.client.force_login(self.user)
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        wb = load_workbook(io.BytesIO(response.content))
        ws2 = wb["Vue graphique"]

        valeurs_mois_col = []
        for row in range(1, ws2.max_row + 1):
            for col in range(1, ws2.max_column + 1):
                val = ws2.cell(row=row, column=col).value
                if val == "Mois":
                    valeurs_mois_col = [
                        ws2.cell(row=row + i, column=col).value for i in range(1, 13)
                    ]
                    break
        self.assertEqual(valeurs_mois_col[0], "Jan 2026")
        self.assertEqual(valeurs_mois_col[8], "Sept 2026")
        self.assertEqual(valeurs_mois_col[-1], "Déc 2026")


class NouvelIndicateurSansMesuresTests(TestCase):
    def setUp(self):
        from accounts.models import Societe
        self.user = User.objects.create_superuser("admin_no_m", "admin_no_m@test.com", "pass123")
        self.societe = Societe.objects.create(nom="Société Test")
        self.processus = Processus.objects.create(code="PM01", nom="Pilotage", societe=self.societe)
        self.processus.RO.add(self.user)

        # Indicateur A avec des mesures
        self.ind_a = Indicateur.objects.create(
            processus=self.processus,
            code="PM01-01",
            nom="Indicateur A avec mesures",
            periodicite=Indicateur.Periodicite.MENSUEL,
            mode_calcul=Indicateur.ModeCalcul.AUTOMATIQUE,
        )
        MesureIndicateur.objects.create(indicateur=self.ind_a, annee=2026, mois=1, valeur=Decimal("5"))
        MesureIndicateur.objects.create(indicateur=self.ind_a, annee=2026, mois=2, valeur=Decimal("8"))

        # Indicateur B nouvellement créé SANS aucune mesure (mode MANUEL)
        self.ind_b = Indicateur.objects.create(
            processus=self.processus,
            code="PM01-02",
            nom="Indicateur B sans mesure (Manuel)",
            periodicite=Indicateur.Periodicite.MENSUEL,
            mode_calcul=Indicateur.ModeCalcul.MANUEL,
        )

    def test_nouvel_indicateur_sans_mesures_affiche_12_mois_vides(self):
        from gestion_documentaire.views import obtenir_donnees_tableau_de_bord
        donnees = obtenir_donnees_tableau_de_bord(self.user, {"annee": "2026"})

        items_b = [
            item for group in donnees["processus_groups"]
            for item in group["indicateurs_data"]
            if item["indicateur"].pk == self.ind_b.pk
        ]
        self.assertEqual(len(items_b), 1)
        item_b = items_b[0]

        # Agrégat et 12 mois formates doivent être "—" (vides)
        self.assertEqual(item_b["agregat_formate"], "—")
        self.assertEqual(item_b["mois_formates"], ["—"] * 12)
        self.assertEqual(item_b["serie_mensuelle"], [None] * 12)

    def test_calculer_fenetre_12_mois_glissants_multi_annees(self):
        from gestion_documentaire.views import calculer_fenetre_12_mois_glissants
        ind = Indicateur.objects.create(
            processus=self.processus,
            code="PM01-03",
            nom="Indicateur Glissant 12M",
            periodicite=Indicateur.Periodicite.GLISSANT_12_MOIS,
            mode_agregation=Indicateur.ModeAgregation.SOMME,
        )
        # Mesure en Avril 2025 = 10, Mesure en Mars 2026 = 5
        mesures_dict = {
            (2025, 4): Decimal("10"),
            (2026, 3): Decimal("5"),
        }
        # Fenêtre pour Mars 2026 (mois_ref=3, annee_ref=2026) -> Avril 2025 à Mars 2026
        valeur = calculer_fenetre_12_mois_glissants(ind, 3, 2026, mesures_dict)
        self.assertEqual(valeur, Decimal("15.0000"))


class IndicateurFormuleTests(TestCase):
    def setUp(self):
        from accounts.models import Societe
        self.user = User.objects.create_superuser("admin_formule", "formula@test.com", "pass123")
        self.societe = Societe.objects.create(nom="Société Formule")
        self.processus, _ = Processus.objects.get_or_create(code="PM02_TEST_FORMULA", defaults={"nom": "Sécurité & Qualité", "societe": self.societe})
        self.processus.RO.add(self.user)

        # Indicateur TF (Taux de fréquence) en Mode Formule
        self.ind_tf = Indicateur.objects.create(
            processus=self.processus,
            code="PM02-TF",
            nom="Taux de fréquence (TF)",
            periodicite=Indicateur.Periodicite.GLISSANT_12_MOIS,
            mode_calcul=Indicateur.ModeCalcul.FORMULE,
            formule="(accidents * 1000000) / heures",
        )
        self.comp_accidents = ComposanteIndicateur.objects.create(
            indicateur=self.ind_tf, code="accidents", libelle="Nombre d'accidents", ordre=1
        )
        self.comp_heures = ComposanteIndicateur.objects.create(
            indicateur=self.ind_tf, code="heures", libelle="Heures travaillées", ordre=2
        )

    def test_formula_simple_tf_evaluation(self):
        from gestion_documentaire.utils_formule import evaluer_formule_securisee
        res = evaluer_formule_securisee("(accidents * 1000000) / heures", {
            "accidents": Decimal("3"),
            "heures": Decimal("200000"),
        })
        self.assertEqual(res, Decimal("15"))

    def test_formula_division_by_zero_returns_none(self):
        from gestion_documentaire.utils_formule import evaluer_formule_securisee
        res = evaluer_formule_securisee("(accidents * 1000000) / heures", {
            "accidents": Decimal("3"),
            "heures": Decimal("0"),
        })
        self.assertIsNone(res)

    def test_formula_glissant_12_mois_consolidation(self):
        from gestion_documentaire.views import obtenir_donnees_tableau_de_bord
        # Saisies sur 12 mois (ex: Avril 2025 à Mars 2026)
        # 3 accidents au total (1 en Mai 2025, 1 en Nov 2025, 1 en Fév 2026)
        ValeurComposanteIndicateur.objects.create(
            composante=self.comp_accidents, annee=2025, mois=5, valeur=Decimal("1")
        )
        ValeurComposanteIndicateur.objects.create(
            composante=self.comp_accidents, annee=2025, mois=11, valeur=Decimal("1")
        )
        ValeurComposanteIndicateur.objects.create(
            composante=self.comp_accidents, annee=2026, mois=2, valeur=Decimal("1")
        )

        # 200 000 heures réparties (20 000 h par mois sur 10 mois)
        for m in range(4, 13):
            ValeurComposanteIndicateur.objects.create(
                composante=self.comp_heures, annee=2025, mois=m, valeur=Decimal("20000")
            )
        ValeurComposanteIndicateur.objects.create(
            composante=self.comp_heures, annee=2026, mois=1, valeur=Decimal("10000")
        )
        ValeurComposanteIndicateur.objects.create(
            composante=self.comp_heures, annee=2026, mois=2, valeur=Decimal("10000")
        )

        donnees = obtenir_donnees_tableau_de_bord(self.user, {"annee": "2026"})
        items = [
            item for group in donnees["processus_groups"]
            for item in group["indicateurs_data"]
            if item["indicateur"].pk == self.ind_tf.pk
        ]
        self.assertEqual(len(items), 1)
        tf_item = items[0]

        # TF consolidé en Mars 2026 (mois 3):
        # Sum accidents = 3, Sum heures = 200,000 -> TF = (3 * 1,000,000) / 200,000 = 15.00
        self.assertIsNotNone(tf_item["serie_mensuelle"][2]) # Mars (index 2)
        self.assertEqual(Decimal(str(tf_item["serie_mensuelle"][2])), Decimal("15"))
        # L'axe X affiche le mois de fin de chaque fenêtre glissante, avec l'année
        self.assertEqual(tf_item["serie_mensuelle_labels"], [
            "Jan 2026", "Fév 2026", "Mar 2026", "Avr 2026", "Mai 2026", "Juin 2026",
            "Juil 2026", "Août 2026", "Sept 2026", "Oct 2026", "Nov 2026", "Déc 2026",
        ])

    def test_tf_monthly_and_12m_consolidation_exact_user_spec(self):
        from gestion_documentaire.views import obtenir_donnees_tableau_de_bord
        # Nettoyer les valeurs de composantes pour ce test
        ValeurComposanteIndicateur.objects.filter(composante__in=[self.comp_accidents, self.comp_heures]).delete()

        # 1. Test des calculs mensuels direct :
        # accidents=2, heures=100000 -> TF = 20.0
        # accidents=1, heures=80000  -> TF = 12.5
        # accidents=0, heures=120000 -> TF = 0.0
        self.ind_tf.periodicite = Indicateur.Periodicite.MENSUEL
        self.ind_tf.mode_agregation = Indicateur.ModeAgregation.SOMME
        self.ind_tf.save()

        # Mois 1: 2 accidents, 100 000 h
        ValeurComposanteIndicateur.objects.create(composante=self.comp_accidents, annee=2026, mois=1, valeur=Decimal("2"))
        ValeurComposanteIndicateur.objects.create(composante=self.comp_heures, annee=2026, mois=1, valeur=Decimal("100000"))

        # Mois 2: 0 accidents, 120 000 h
        ValeurComposanteIndicateur.objects.create(composante=self.comp_accidents, annee=2026, mois=2, valeur=Decimal("0"))
        ValeurComposanteIndicateur.objects.create(composante=self.comp_heures, annee=2026, mois=2, valeur=Decimal("120000"))

        # Mois 3: 1 accident, 80 000 h
        ValeurComposanteIndicateur.objects.create(composante=self.comp_accidents, annee=2026, mois=3, valeur=Decimal("1"))
        ValeurComposanteIndicateur.objects.create(composante=self.comp_heures, annee=2026, mois=3, valeur=Decimal("80000"))

        # Mois 4: 3 accidents, 150 000 h
        ValeurComposanteIndicateur.objects.create(composante=self.comp_accidents, annee=2026, mois=4, valeur=Decimal("3"))
        ValeurComposanteIndicateur.objects.create(composante=self.comp_heures, annee=2026, mois=4, valeur=Decimal("150000"))

        # Mois 5: 0 accidents, 100 000 h
        ValeurComposanteIndicateur.objects.create(composante=self.comp_accidents, annee=2026, mois=5, valeur=Decimal("0"))
        ValeurComposanteIndicateur.objects.create(composante=self.comp_heures, annee=2026, mois=5, valeur=Decimal("100000"))

        # Mois 6: 4 accidents, 200 000 h
        ValeurComposanteIndicateur.objects.create(composante=self.comp_accidents, annee=2026, mois=6, valeur=Decimal("4"))
        ValeurComposanteIndicateur.objects.create(composante=self.comp_heures, annee=2026, mois=6, valeur=Decimal("200000"))

        donnees = obtenir_donnees_tableau_de_bord(self.user, {"annee": "2026"})
        tf_item = [
            item for group in donnees["processus_groups"]
            for item in group["indicateurs_data"]
            if item["indicateur"].pk == self.ind_tf.pk
        ][0]

        # Vérification des TF mensuels
        self.assertEqual(Decimal(str(tf_item["serie_mensuelle"][0])), Decimal("20.0"))   # Janvier
        self.assertEqual(Decimal(str(tf_item["serie_mensuelle"][1])), Decimal("0.0"))    # Février
        self.assertEqual(Decimal(str(tf_item["serie_mensuelle"][2])), Decimal("12.5"))   # Mars
        self.assertEqual(Decimal(str(tf_item["serie_mensuelle"][3])), Decimal("20.0"))   # Avril
        self.assertEqual(Decimal(str(tf_item["serie_mensuelle"][4])), Decimal("0.0"))    # Mai
        self.assertEqual(Decimal(str(tf_item["serie_mensuelle"][5])), Decimal("20.0"))   # Juin

        # 2. Consolidation de la formule :
        # Total accidents = 2+0+1+3+0+4 = 10, Total heures = 100k+120k+80k+150k+100k+200k = 750 000
        # TF Consolidé = (10 * 1 000 000) / 750 000 = 13.3333
        self.assertEqual(Decimal(str(tf_item["cumul_valeur_raw"])), Decimal("13.3333"))

        # 3. Test du mode glissant 12 mois
        self.ind_tf.periodicite = Indicateur.Periodicite.GLISSANT_12_MOIS
        self.ind_tf.save()

        donnees_glissant = obtenir_donnees_tableau_de_bord(self.user, {"annee": "2026"})
        tf_glissant_item = [
            item for group in donnees_glissant["processus_groups"]
            for item in group["indicateurs_data"]
            if item["indicateur"].pk == self.ind_tf.pk
        ][0]

        # La 6ème fenêtre (terminant en Juin) cumule 10 accidents et 750 000 h -> 13.3333
        self.assertEqual(Decimal(str(tf_glissant_item["serie_mensuelle"][5])), Decimal("13.3333"))
        self.assertEqual(Decimal(str(tf_glissant_item["cumul_valeur_raw"])), Decimal("13.3333"))

    def test_evaluer_statut_indicateur_ne_pas_depasser_5_cases(self):
        from gestion_documentaire.views import evaluer_statut_indicateur
        self.ind_tf.sens_objectif = Indicateur.SensObjectif.NE_PAS_DEPASSER

        # Cas 1: objectif=90, réalisé=25.1852 => taux=100.00 %, conforme
        res1 = evaluer_statut_indicateur(self.ind_tf, Decimal("25.1852"), Decimal("90"))
        self.assertEqual(res1["taux_atteinte"], 100.0)
        self.assertEqual(res1["code"], "conforme")

        # Cas 2: objectif=90, réalisé=90 => taux=100.00 %, conforme
        res2 = evaluer_statut_indicateur(self.ind_tf, Decimal("90"), Decimal("90"))
        self.assertEqual(res2["taux_atteinte"], 100.0)
        self.assertEqual(res2["code"], "conforme")

        # Cas 3: objectif=90, réalisé=100 => taux=90.00 %, a_surveiller (dépassement faible)
        res3 = evaluer_statut_indicateur(self.ind_tf, Decimal("100"), Decimal("90"))
        self.assertEqual(res3["taux_atteinte"], 90.0)
        self.assertIn(res3["code"], ("a_surveiller", "non_conforme"))

        # Cas 4: objectif=90, réalisé=120 => taux=75.00 %, non conforme
        res4 = evaluer_statut_indicateur(self.ind_tf, Decimal("120"), Decimal("90"))
        self.assertEqual(res4["taux_atteinte"], 75.0)
        self.assertEqual(res4["code"], "non_conforme")

        # Cas 5: objectif=90, réalisé=0 => taux=100.00 %, conforme
        res5 = evaluer_statut_indicateur(self.ind_tf, Decimal("0"), Decimal("90"))
        self.assertEqual(res5["taux_atteinte"], 100.0)
        self.assertEqual(res5["code"], "conforme")



    def test_mode_manuel_vs_formule_independence(self):
        # Créer une valeur de composante
        ValeurComposanteIndicateur.objects.create(
            composante=self.comp_accidents, annee=2026, mois=3, valeur=Decimal("5")
        )

        # Passer temporairement en mode MANUEL et ajouter un réalisé consolidé
        self.ind_tf.mode_calcul = Indicateur.ModeCalcul.MANUEL
        self.ind_tf.save()
        RealiseConsolideIndicateur.objects.create(
            indicateur=self.ind_tf, annee_reference=2026, mois_reference=3, valeur=Decimal("42.50")
        )

        # Revenir en mode FORMULE
        self.ind_tf.mode_calcul = Indicateur.ModeCalcul.FORMULE
        self.ind_tf.save()

        # Vérifier que les deux enregistrements subsistent de façon indépendante sans écrasement
        self.assertTrue(ValeurComposanteIndicateur.objects.filter(composante=self.comp_accidents, annee=2026, mois=3).exists())
        self.assertTrue(RealiseConsolideIndicateur.objects.filter(indicateur=self.ind_tf, annee_reference=2026, mois_reference=3).exists())

    def test_creation_formule_personnalisee_est_sauvegardee(self):
        """La modal de création doit enregistrer la formule personnalisée (ex: (5*5)/100)."""
        url = reverse("gestion_documentaire:indicateur_create")
        self.client.force_login(self.user)
        res = self.client.post(url, {
            "processus": self.processus.id,
            "nom": "Perso constante",
            "periodicite": "mensuel",
            "mode_agregation": "somme",
            "mode_calcul": Indicateur.ModeCalcul.FORMULE,
            "formule": "(5*5)/100",
            "is_active": "true",
        })
        self.assertEqual(res.status_code, 302)
        ind = Indicateur.objects.get(nom="Perso constante")
        self.assertEqual(ind.formule, "(5*5)/100")
        self.assertEqual(ind.composantes.count(), 0)

    def test_creation_formule_avec_composantes_est_sauvegardee(self):
        """La création doit enregistrer formule et composantes (workflow Fabien / TF)."""
        url = reverse("gestion_documentaire:indicateur_create")
        self.client.force_login(self.user)
        res = self.client.post(url, {
            "processus": self.processus.id,
            "nom": "TF creation",
            "periodicite": "mensuel",
            "mode_agregation": "somme",
            "mode_calcul": Indicateur.ModeCalcul.FORMULE,
            "formule": "(accidents * 1000000) / heures",
            "comp_code": ["accidents", "heures"],
            "comp_libelle": ["Nombre d'accidents", "Nombre d'heures travaillées"],
            "is_active": "true",
        })
        self.assertEqual(res.status_code, 302)
        ind = Indicateur.objects.get(nom="TF creation")
        self.assertEqual(ind.formule, "(accidents * 1000000) / heures")
        self.assertEqual(ind.composantes.count(), 2)
        self.assertEqual(set(ind.composantes.values_list("code", flat=True)), {"accidents", "heures"})

    def test_creation_formule_invalide_rejetee(self):
        """Une formule référençant une variable non déclarée ne doit pas créer l'indicateur."""
        url = reverse("gestion_documentaire:indicateur_create")
        self.client.force_login(self.user)
        res = self.client.post(url, {
            "processus": self.processus.id,
            "nom": "TF invalide",
            "periodicite": "mensuel",
            "mode_agregation": "somme",
            "mode_calcul": Indicateur.ModeCalcul.FORMULE,
            "formule": "(inconnue * 1000000) / heures",
            "comp_code": ["accidents", "heures"],
            "comp_libelle": ["Accidents", "Heures"],
            "is_active": "true",
        })
        self.assertFalse(Indicateur.objects.filter(nom="TF invalide").exists())

    def test_update_formule_personnalisee_constante_sans_composante(self):
        """L'update doit accepter une formule constante sans composante et nettoyer les anciennes."""
        url = reverse("gestion_documentaire:tdb_indicateur_update", kwargs={"indicateur_id": self.ind_tf.pk})
        self.client.force_login(self.user)
        res = self.client.post(url, {
            "nom": self.ind_tf.nom,
            "periodicite": "mensuel",
            "mode_agregation": "somme",
            "mode_calcul": Indicateur.ModeCalcul.FORMULE,
            "formule": "(7*8)",
            "is_active": "true",
        })
        self.assertEqual(res.status_code, 302)
        self.ind_tf.refresh_from_db()
        self.assertEqual(self.ind_tf.formule, "(7*8)")
        self.assertEqual(self.ind_tf.composantes.count(), 0)

    def test_detail_json_renvoie_formule_sauvegardee(self):
        """Le JSON de détail doit renvoyer la formule sauvegardée pour le préremplissage."""
        self.client.force_login(self.user)
        url = reverse("gestion_documentaire:tdb_indicateur_detail_json", kwargs={"indicateur_id": self.ind_tf.pk})
        res = self.client.get(url, HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertEqual(res.status_code, 200)
        data = res.json()
        self.assertEqual(data["formule"], "(accidents * 1000000) / heures")
        self.assertEqual(len(data["composantes"]), 2)

    def _passer_en_manuel(self):
        """Bascule l'indicateur en mode MANUEL sans formule ni composantes (état de départ du scénario)."""
        self.ind_tf.mode_calcul = Indicateur.ModeCalcul.MANUEL
        self.ind_tf.formule = ""
        self.ind_tf.composantes.all().delete()
        self.ind_tf.save()

    def test_update_manuel_vers_formule_sauvegarde_ok(self):
        """MANUEL → FORMULE avec composantes accidents/heures (type taux de fréquence) doit sauvegarder."""
        self._passer_en_manuel()
        url = reverse("gestion_documentaire:tdb_indicateur_update", kwargs={"indicateur_id": self.ind_tf.pk})
        self.client.force_login(self.user)
        res = self.client.post(url, {
            "nom": self.ind_tf.nom,
            "periodicite": "glissant_12_mois",
            "mode_agregation": "moyenne",
            "sens_objectif": "ne_pas_depasser",
            "mode_calcul": Indicateur.ModeCalcul.FORMULE,
            "type_formule": "taux_frequence",
            "formule": "(accidents * 1000000) / heures",
            "comp_code": ["accidents", "heures"],
            "comp_libelle": ["Nombre d'accidents", "Heures travaillées"],
            "is_active": "true",
        })
        self.assertEqual(res.status_code, 302)
        self.ind_tf.refresh_from_db()
        self.assertEqual(self.ind_tf.mode_calcul, Indicateur.ModeCalcul.FORMULE)
        self.assertEqual(self.ind_tf.formule, "(accidents * 1000000) / heures")
        self.assertEqual(set(self.ind_tf.composantes.values_list("code", flat=True)), {"accidents", "heures"})

    def test_update_formule_normalise_casse_des_identifiants(self):
        """Saisie Accidents/Heures (majuscules) → formule normalisée en minuscules et composantes en minuscules."""
        self._passer_en_manuel()
        url = reverse("gestion_documentaire:tdb_indicateur_update", kwargs={"indicateur_id": self.ind_tf.pk})
        self.client.force_login(self.user)
        res = self.client.post(url, {
            "nom": self.ind_tf.nom,
            "periodicite": "glissant_12_mois",
            "mode_agregation": "moyenne",
            "mode_calcul": Indicateur.ModeCalcul.FORMULE,
            "type_formule": "taux_frequence",
            "formule": "(Accidents * 1000000) / Heures",
            "comp_code": ["Accidents", "Heures"],
            "comp_libelle": ["Nombre d'accidents", "Heures travaillées"],
            "is_active": "true",
        })
        self.assertEqual(res.status_code, 302)
        self.ind_tf.refresh_from_db()
        self.assertEqual(self.ind_tf.formule, "(accidents * 1000000) / heures")
        self.assertEqual(set(self.ind_tf.composantes.values_list("code", flat=True)), {"accidents", "heures"})

    def test_update_formule_vide_bloquee_avec_erreur_visible_et_valeurs_conservees(self):
        """MANUEL → FORMULE avec formule vide : sauvegarde bloquée, erreur visible, valeurs conservées."""
        self._passer_en_manuel()
        url = reverse("gestion_documentaire:tdb_indicateur_update", kwargs={"indicateur_id": self.ind_tf.pk})
        self.client.force_login(self.user)
        res = self.client.post(url, {
            "nom": self.ind_tf.nom,
            "periodicite": "glissant_12_mois",
            "mode_agregation": "moyenne",
            "mode_calcul": Indicateur.ModeCalcul.FORMULE,
            "type_formule": "taux_frequence",
            "formule": "",
            "comp_code": ["accidents", "heures"],
            "comp_libelle": ["Nombre d'accidents", "Heures travaillées"],
            "is_active": "true",
        })
        self.assertEqual(res.status_code, 302)

        # BD inchangée : l'indicateur reste en MANUEL, formule vide
        self.ind_tf.refresh_from_db()
        self.assertEqual(self.ind_tf.mode_calcul, Indicateur.ModeCalcul.MANUEL)
        self.assertEqual(self.ind_tf.formule, "")

        # Valeurs saisies conservées dans la session pour réafficher la modal
        conserve = self.client.session.get("tdb_indicateur_form_erreur")
        self.assertIsNotNone(conserve)
        self.assertEqual(conserve["mode_calcul"], Indicateur.ModeCalcul.FORMULE)
        self.assertEqual(conserve["formule"], "")
        self.assertEqual(conserve["type_formule"], "taux_frequence")
        self.assertEqual(conserve["comp_code"], ["accidents", "heures"])

        # Erreur visible après redirection
        page = self.client.get(res.url)
        self.assertContains(page, "La formule est obligatoire pour le mode de calcul automatique par formule.")

    def test_update_formule_reexamen_detail_json(self):
        """Après un passage réussi MANUEL → FORMULE, le JSON de détail expose mode_calcul, formule et composantes."""
        self._passer_en_manuel()
        url = reverse("gestion_documentaire:tdb_indicateur_update", kwargs={"indicateur_id": self.ind_tf.pk})
        self.client.force_login(self.user)
        res = self.client.post(url, {
            "nom": self.ind_tf.nom,
            "periodicite": "glissant_12_mois",
            "mode_agregation": "moyenne",
            "mode_calcul": Indicateur.ModeCalcul.FORMULE,
            "type_formule": "taux_frequence",
            "formule": "(accidents * 1000000) / heures",
            "comp_code": ["accidents", "heures"],
            "comp_libelle": ["Nombre d'accidents", "Heures travaillées"],
            "is_active": "true",
        })
        self.assertEqual(res.status_code, 302)

        detail_url = reverse("gestion_documentaire:tdb_indicateur_detail_json", kwargs={"indicateur_id": self.ind_tf.pk})
        res = self.client.get(detail_url, HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertEqual(res.status_code, 200)
        data = res.json()
        self.assertEqual(data["mode_calcul"], Indicateur.ModeCalcul.FORMULE)
        self.assertEqual(data["formule"], "(accidents * 1000000) / heures")
        self.assertEqual(len(data["composantes"]), 2)



class ModeManuelGlissantTests(TestCase):
    """
    Le mode MANUEL d'un indicateur GLISSANT_12_MOIS doit lire la série sur les
    12 derniers mois terminant au mois de référence (fin de période filtrée),
    en traversant l'année précédente — et non seulement Janvier → Décembre de
    l'année sélectionnée.
    """

    def setUp(self):
        from accounts.models import Societe
        from .models import RealiseConsolideIndicateur

        self.user = User.objects.create_superuser("admin_glissant", "glissant@test.com", "pass123")
        self.societe = Societe.objects.create(nom="Société Glissant")
        self.processus = Processus.objects.create(code="PM02", nom="Sécurité & Qualité", societe=self.societe)
        self.processus.RO.add(self.user)

        self.ind_tf = Indicateur.objects.create(
            processus=self.processus,
            code="PM02-TFG",
            nom="Taux de fréquence (TF) 12M glissants — Manuel",
            periodicite=Indicateur.Periodicite.GLISSANT_12_MOIS,
            mode_calcul=Indicateur.ModeCalcul.MANUEL,
            mode_agregation=Indicateur.ModeAgregation.MOYENNE,
            sens_objectif=Indicateur.SensObjectif.NE_PAS_DEPASSER,
        )
        valeurs = {
            (2025, 9): "10.0", (2025, 10): "10.5", (2025, 11): "11.0", (2025, 12): "11.5",
            (2026, 1): "12.0", (2026, 2): "12.5", (2026, 3): "13.0", (2026, 4): "13.5",
            (2026, 5): "14.0", (2026, 6): "14.5", (2026, 7): "15.0", (2026, 8): "15.5",
        }
        for (annee, mois), val in valeurs.items():
            RealiseConsolideIndicateur.objects.create(
                indicateur=self.ind_tf,
                annee_reference=annee,
                mois_reference=mois,
                valeur=Decimal(val),
                saisie_par=self.user,
            )

    def test_fenetre_12_mois_traverse_les_annees_manuelle(self):
        """Référence Août 2026 : la série couvre Sept 2025 → Août 2026."""
        from gestion_documentaire.views import obtenir_donnees_tableau_de_bord

        donnees = obtenir_donnees_tableau_de_bord(
            self.user,
            {"annee": "2026", "type_periode": "mois", "periode": "8"},
        )
        item = next(
            it for group in donnees["processus_groups"]
            for it in group["indicateurs_data"]
            if it["indicateur"].pk == self.ind_tf.pk
        )
        expected = [
            Decimal("10.0"), Decimal("10.5"), Decimal("11.0"), Decimal("11.5"),
            Decimal("12.0"), Decimal("12.5"), Decimal("13.0"), Decimal("13.5"),
            Decimal("14.0"), Decimal("14.5"), Decimal("15.0"), Decimal("15.5"),
        ]
        self.assertEqual(len(item["serie_mensuelle"]), 12)
        self.assertEqual(
            [Decimal(str(v)) if v is not None else None for v in item["serie_mensuelle"]],
            expected,
        )
        self.assertEqual(item["serie_mensuelle_labels"][0], "Sept 2025")
        self.assertEqual(item["serie_mensuelle_labels"][-1], "Août 2026")
        # KPI = valeur au mois de référence (dernier point de la fenêtre)
        self.assertEqual(Decimal(str(item["agregat"])), Decimal("15.5"))

    def test_annee_complete_garde_janvier_decembre(self):
        """Année complète : la fenêtre reste Janvier → Décembre (comportement inchangé)."""
        from gestion_documentaire.views import obtenir_donnees_tableau_de_bord

        donnees = obtenir_donnees_tableau_de_bord(self.user, {"annee": "2026"})
        item = next(
            it for group in donnees["processus_groups"]
            for it in group["indicateurs_data"]
            if it["indicateur"].pk == self.ind_tf.pk
        )
        self.assertEqual(item["serie_mensuelle"][0], 12.0)  # Janvier 2026
        self.assertIsNone(item["serie_mensuelle"][8])       # Septembre 2026 : aucune saisie
        self.assertEqual(item["serie_mensuelle_labels"], [
            "Jan", "Fév", "Mar", "Avr", "Mai", "Juin", "Juil", "Août", "Sept", "Oct", "Nov", "Déc",
        ])








