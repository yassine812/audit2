"""Vues du module Accidents de Travail."""

import io
import json
import logging
from datetime import date

from django.contrib import messages
from django.contrib.admin.models import LogEntry, ADDITION, CHANGE
from django.contrib.auth.decorators import login_required
from django.contrib.contenttypes.models import ContentType
from django.core.paginator import Paginator
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.generic import DetailView, ListView, UpdateView, CreateView

from .signals import log_accident_history, log_analyse_history, log_lap_history
from .forms import (
    AccidentForm24h,
    AccidentQuickCreateForm,
    ActionCorrectiveFormSet,
    ActionCorrectiveImmediateFormSet,
    Analyse48hForm,
    CauseRacineFormSet,
    LAP8JForm,
    QuestionnaireItemForm,
    QuestionnaireTemplateForm,
)
from .mixins import AccidentPermissionMixin, at_roles_required, user_has_at_access
from .models import (
    AccidentTravail,
    ActionCorrectiveImmédiate,
    Analyse48h,
    CauseRacine,
    LAP8Jours,
    QuestionnaireItem,
    QuestionnaireTemplate,
)

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Dashboard
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@at_roles_required
def dashboard(request):
    """Tableau de bord + liste complète paginée des accidents."""
    user = request.user
    qs = AccidentTravail.objects.all()
    if user.is_superuser:
        pass  # voit tous les accidents
    else:
        if user.section:
            qs = qs.filter(section=user.section)
        else:
            qs = qs.none()

    today = date.today()
    mois_en_cours = qs.filter(date_accident__year=today.year, date_accident__month=today.month)

    kpis = {
        "total_mois": mois_en_cours.count(),
        "avec_arret": mois_en_cours.filter(consequence="avec_arret").count(),
        "sans_arret": mois_en_cours.filter(consequence="sans_arret").count(),
        "en_cours": qs.exclude(statut="cloture").count(),
    }

    # Accidents en retard
    now = timezone.now()
    en_retard = qs.filter(
        Q(statut="24h", echeance_48h__lt=now) | Q(statut="48h", echeance_8j__lt=now)
    ).order_by("date_accident")[:10]

    # Graphique : 12 derniers mois
    from datetime import timedelta
    douze_mois = []
    for i in range(11, -1, -1):
        m = today.month - i
        y = today.year + (m - 1) // 12
        m = ((m - 1) % 12) + 1
        douze_mois.append({"annee": y, "mois": m})

    chart_labels = []
    chart_data = []
    for d in douze_mois:
        count = qs.filter(date_accident__year=d["annee"], date_accident__month=d["mois"]).count()
        chart_labels.append(f"{d['mois']:02d}/{d['annee']}")
        chart_data.append(count)

    type_data = {
        "travail": qs.filter(type_accident="travail").count(),
        "trajet": qs.filter(type_accident="trajet").count(),
    }

    # Filtrage + pagination pour la liste
    q = request.GET.get("q", "").strip()
    statut_filtre = request.GET.get("statut", "")
    type_filtre = request.GET.get("type_accident", "")
    qs_liste = qs.select_related("societe", "section")
    if q:
        qs_liste = qs_liste.filter(
            Q(reference__icontains=q)
            | Q(victime_nom_prenom__icontains=q)
            | Q(circonstances_detaillees__icontains=q)
        )
    if statut_filtre:
        qs_liste = qs_liste.filter(statut=statut_filtre)
    if type_filtre:
        qs_liste = qs_liste.filter(type_accident=type_filtre)

    paginator = Paginator(qs_liste, 20)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    return render(request, "accident_travail/dashboard.html", {
        "kpis": kpis,
        "en_retard": en_retard,
        "chart_labels": json.dumps(chart_labels),
        "chart_data": json.dumps(chart_data),
        "type_data": json.dumps(type_data),
        # liste
        "accidents": page_obj,
        "page_obj": page_obj,
        "paginator": paginator,
        "is_paginated": paginator.num_pages > 1,
        "statuts": AccidentTravail.STATUT_CHOICES,
        "types": AccidentTravail.TYPE_CHOICES,
        "q": q,
        "statut_filtre": statut_filtre,
        "type_filtre": type_filtre,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Liste et détail
# ─────────────────────────────────────────────────────────────────────────────

class AccidentListView(AccidentPermissionMixin, ListView):
    """Redirige vers le tableau de bord fusionné."""
    template_name = "accident_travail/liste.html"
    context_object_name = "accidents"
    paginate_by = 20

    def get(self, request, *args, **kwargs):
        # Preserve any query params when redirecting
        from django.urls import reverse
        params = request.GET.urlencode()
        url = reverse("at:at_dashboard")
        if params:
            url = f"{url}?{params}"
        return redirect(url)

    def get_queryset(self):
        qs = super().get_queryset()
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["statuts"] = AccidentTravail.STATUT_CHOICES
        ctx["types"] = AccidentTravail.TYPE_CHOICES
        ctx["q"] = self.request.GET.get("q", "")
        ctx["statut_filtre"] = self.request.GET.get("statut", "")
        ctx["type_filtre"] = self.request.GET.get("type_accident", "")
        return ctx


class AccidentDetailView(AccidentPermissionMixin, DetailView):
    template_name = "accident_travail/detail.html"
    context_object_name = "accident"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        accident = self.object
        ctx["has_analyse"] = hasattr(accident, "analyse_48h")
        ctx["has_lap"] = hasattr(accident, "lap_8j")
        ctx["actions_immediates"] = accident.actions_correctives_immediates.all()
        ctx["causes"] = (
            accident.analyse_48h.causes_racines.all()
            if hasattr(accident, "analyse_48h")
            else []
        )
        ctx["pourquoi_data_json"] = json.dumps(
            accident.analyse_48h.pourquoi_data
            if hasattr(accident, "analyse_48h")
            else {}
        )

        # Informations de retard par étape
        now = timezone.now()
        ctx["analyse_soumise_en_retard"] = (
            hasattr(accident, "analyse_48h")
            and accident.analyse_48h.soumis_en_retard
        )
        ctx["lap_soumis_en_retard"] = (
            hasattr(accident, "lap_8j")
            and accident.lap_8j.soumis_en_retard
        )
        ctx["declaration_soumise_en_retard"] = (
            accident.statut not in (
                AccidentTravail.STATUT_BROUILLON, AccidentTravail.STATUT_24H
            )
            and accident.echeance_48h
            and hasattr(accident, "analyse_48h")
            and accident.analyse_48h.soumis_le
            # La déclaration 24h était en retard si l'analyse a été soumise après écheance_48h
        )

        # Historique (LogEntry) pour cet accident — paginé (10 par page, superadmin seulement)
        ct = ContentType.objects.get_for_model(accident)
        hist_qs = (
            LogEntry.objects.filter(content_type=ct, object_id=accident.pk)
            .select_related("user")
            .order_by("-action_time")
        )
        hist_paginator = Paginator(hist_qs, 10)
        hist_page_obj = hist_paginator.get_page(self.request.GET.get("hist_page", 1))
        ctx["historique"] = hist_page_obj
        ctx["hist_paginator"] = hist_paginator
        ctx["hist_page_obj"] = hist_page_obj
        ctx["hist_is_paginated"] = hist_paginator.num_pages > 1
        return ctx


# ─────────────────────────────────────────────────────────────────────────────
# Création / Modification — Étape 24h
# ─────────────────────────────────────────────────────────────────────────────

class AccidentCreate24hView(AccidentPermissionMixin, CreateView):
    model = AccidentTravail
    form_class = AccidentForm24h
    template_name = "accident_travail/form_24h.html"

    def get_form_kwargs(self):
        kw = super().get_form_kwargs()
        kw["user"] = self.request.user
        return kw

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["title"] = "Nouvelle déclaration d'accident"
        ctx["action_formset"] = ActionCorrectiveImmediateFormSet(
            self.request.POST or None, prefix="actions"
        )
        return ctx

    def form_valid(self, form):
        ctx = self.get_context_data()
        action_formset = ctx["action_formset"]
        if not action_formset.is_valid():
            return self.form_invalid(form)

        accident = form.save(commit=False)
        accident.created_by = self.request.user
        accident.statut = AccidentTravail.STATUT_24H
        accident.save()
        form.save_m2m()

        action_formset.instance = accident
        action_formset.save()

        log_accident_history(self.request.user, accident, "Accident créé (déclaration 24h)", action_flag=ADDITION)
        logger.info("Accident créé : %s par %s", accident.reference, self.request.user)
        from .tasks import notifier_accident_alerte
        notifier_accident_alerte(accident, "creation", acteur=self.request.user)
        messages.success(self.request, f"Accident {accident.reference} créé avec succès.")
        return redirect("at:at_detail", pk=accident.pk)

    def form_invalid(self, form):
        messages.error(self.request, "Veuillez corriger les erreurs ci-dessous.")
        return super().form_invalid(form)


@login_required
@at_roles_required
def at_create_quick(request):
    """Création rapide d'un brouillon d'accident (champs minimum)."""
    if request.method == "POST":
        form = AccidentQuickCreateForm(request.POST, user=request.user)
        if form.is_valid():
            accident = form.save(commit=False)
            accident.created_by = request.user
            accident.statut = AccidentTravail.STATUT_BROUILLON
            accident.save()
            log_accident_history(request.user, accident, "Brouillon rapide créé", action_flag=ADDITION)
            from django.urls import reverse
            url = reverse("at:at_update_24h", args=[accident.pk]) + "?from_quick=1"
            return redirect(url)
    else:
        form = AccidentQuickCreateForm(user=request.user)
    return render(request, "accident_travail/form_quick.html", {"form": form})


class AccidentUpdate24hView(AccidentPermissionMixin, UpdateView):
    model = AccidentTravail
    form_class = AccidentForm24h
    template_name = "accident_travail/form_24h.html"

    def get_form_kwargs(self):
        kw = super().get_form_kwargs()
        kw["user"] = self.request.user
        return kw

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["title"] = f"Modifier la déclaration {self.object.reference}"
        ctx["from_quick"] = self.request.GET.get("from_quick") == "1" or self.request.POST.get("from_quick") == "1"
        ctx["action_formset"] = ActionCorrectiveImmediateFormSet(
            self.request.POST or None,
            instance=self.object,
            prefix="actions",
        )
        return ctx

    def form_valid(self, form):
        ctx = self.get_context_data()
        action_formset = ctx["action_formset"]
        if not action_formset.is_valid():
            return self.form_invalid(form)

        accident = form.save(commit=False)
        accident.updated_by = self.request.user
        accident.save()
        form.save_m2m()
        action_formset.save()

        from .tasks import notifier_accident_alerte

        # Transition brouillon → 24h si l'accident était un brouillon
        if accident.statut == AccidentTravail.STATUT_BROUILLON:
            accident.statut = AccidentTravail.STATUT_24H
            accident.save(update_fields=["statut"])
            log_accident_history(self.request.user, accident, "Brouillon complété → déclaration 24h activée")
            notifier_accident_alerte(accident, "creation", acteur=self.request.user)
            messages.success(self.request, "Déclaration complétée. Étape 24h démarrée.")
            return redirect("at:at_detail", pk=accident.pk)

        # Transition 24h → 48h
        if "soumettre_48h" in self.request.POST and accident.statut == AccidentTravail.STATUT_24H:
            now = timezone.now()
            retard = accident.echeance_48h and now > accident.echeance_48h
            accident.statut = AccidentTravail.STATUT_48H
            accident.soumis_le = now
            accident.save(update_fields=["statut", "soumis_le"])
            msg = "Déclaration 24h soumise en RETARD" if retard else "Déclaration 24h soumise"
            log_accident_history(self.request.user, accident, f"{msg} → 48h démarrée")
            notifier_accident_alerte(accident, "soumission_24h", acteur=self.request.user)
            messages.success(self.request, "Déclaration soumise. Étape 48h démarrée.")
            if retard:
                messages.warning(self.request, "⚠️ Cette déclaration a été soumise après l'échéance 48h.")
            return redirect("at:at_analyse_48h", pk=accident.pk)

        log_accident_history(self.request.user, accident, "Déclaration 24h mise à jour")
        notifier_accident_alerte(accident, "modification_24h", acteur=self.request.user)
        messages.success(self.request, "Déclaration mise à jour.")
        return redirect("at:at_detail", pk=accident.pk)


# ─────────────────────────────────────────────────────────────────────────────
# Analyse 48h
# ─────────────────────────────────────────────────────────────────────────────

class Analyse48hView(AccidentPermissionMixin, UpdateView):
    model = Analyse48h
    form_class = Analyse48hForm
    template_name = "accident_travail/form_48h.html"

    def get_object(self, queryset=None):
        accident = get_object_or_404(
            self.get_queryset().model if hasattr(self, "_accident") else AccidentTravail,
            pk=self.kwargs["pk"],
        )
        self._accident = accident
        analyse, _ = Analyse48h.objects.get_or_create(accident=accident)
        # Générer les items du questionnaire s'ils n'existent pas encore
        if not analyse.questionnaire_items.exists():
            self._generate_questionnaire_items(analyse)
        return analyse

    def _generate_questionnaire_items(self, analyse):
        templates = QuestionnaireTemplate.objects.all()
        items = [
            QuestionnaireItem(
                analyse=analyse,
                code=t.code,
                section=t.section,
                question=t.question,
                est_sous_question=t.est_sous_question,
                parent_code=t.parent_code,
                condition_affichage=t.condition_affichage,
            )
            for t in templates
        ]
        QuestionnaireItem.objects.bulk_create(items, ignore_conflicts=True)

    # Formation types matching the Excel template
    FORMATION_TYPES = [
        "Formation au poste de travail",
        "Formation renforcée intérimaires",
        "Habilitation électrique",
        "CACES",
        "PRAP",
        "SST",
        "Autres formations",
    ]

    # Nuisance types matching the Excel template
    NUISANCE_KEYS = [
        ("bruit", "Bruit"),
        ("vibrations", "Vibrations"),
        ("vapeurs", "Vapeurs / aérosols / gaz"),
        ("climatiques", "Conditions climatiques"),
        ("eclairage", "Éclairage"),
        ("poussieres", "Poussières"),
        ("thermiques", "Ambiances thermiques"),
        ("autres", "Autres"),
    ]

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["accident"] = self._accident
        # Grouper les items par section
        items = self._accident.analyse_48h.questionnaire_items.all()
        sections = {}
        for item in items:
            sections.setdefault(item.section, []).append(item)
        ctx["questionnaire_sections"] = sections
        # Deadline info
        now = timezone.now()
        echeance = self._accident.echeance_48h
        ctx["echeance_48h"] = echeance
        ctx["en_retard_48h"] = echeance and now > echeance
        ctx["heures_restantes_48h"] = round((echeance - now).total_seconds() / 3600, 1) if echeance else None
        # Formations data for template rendering
        formations_raw = self._accident.analyse_48h.formations or []
        formations_data = []
        for i, nom in enumerate(self.FORMATION_TYPES):
            existing = next((f for f in formations_raw if f.get("nom") == nom), {})
            formations_data.append({
                "index": i,
                "nom": nom,
                "realise": existing.get("realise", "NA"),
                "date": existing.get("date", ""),
                "commentaire": existing.get("commentaire", ""),
            })
        ctx["formations_data"] = formations_data
        # Nuisances data
        nuisances_raw = self._accident.analyse_48h.nuisances or {}
        ctx["nuisances_data"] = [
            {"key": k, "label": lbl, "checked": nuisances_raw.get(k, False)}
            for k, lbl in self.NUISANCE_KEYS
        ]
        # Causes racines formset (conservé pour rétrocompat. — non utilisé dans le nouveau UI)
        # ctx["causes_formset"] intentionally removed; pourquoi_data gère désormais les causes
        # Pass all questions as JSON for JS 5-pourquoi widget
        questions_list = []
        for sec_name, sec_items in sections.items():
            for item in sec_items:
                questions_list.append({
                    "code": item.code,
                    "section": sec_name,
                    "question": item.question,
                    "reponse": item.reponse or "NA",
                })
        ctx["questions_json"] = json.dumps(questions_list)
        ctx["pourquoi_data_json"] = json.dumps(self.object.pourquoi_data or {})
        # Afficher l'onglet Causes uniquement après une première sauvegarde
        analyse = self._accident.analyse_48h
        ctx["show_causes_tab"] = (
            analyse.questionnaire_items.filter(reponse__in=["O", "N"]).exists()
            or bool(analyse.pourquoi_data)
            or self.request.GET.get("tab") == "arbre"
        )
        return ctx

    def form_valid(self, form):
        analyse = form.save(commit=False)
        # Sauvegarder les réponses au questionnaire
        items = analyse.questionnaire_items.all()
        for item in items:
            # Le template utilise |cut:'.' → les points sont supprimés (pas remplacés)
            key_reponse    = f"q_{item.code.replace('.', '')}_reponse"
            key_precisions = f"q_{item.code.replace('.', '')}_precisions"
            reponse = self.request.POST.get(key_reponse, "")
            precisions = self.request.POST.get(key_precisions, "")
            if reponse in ("NA", "O", "N"):
                item.reponse = reponse
                item.precisions = precisions
                item.save(update_fields=["reponse", "precisions"])
        # Sauvegarder les formations
        formations = []
        for i, nom in enumerate(self.FORMATION_TYPES):
            realise = self.request.POST.get(f"formation_{i}_realise", "NA")
            date_str = self.request.POST.get(f"formation_{i}_date", "")
            commentaire = self.request.POST.get(f"formation_{i}_commentaire", "")
            formations.append({
                "nom": nom,
                "realise": realise if realise in ("NA", "O", "N") else "NA",
                "date": date_str,
                "commentaire": commentaire,
            })
        analyse.formations = formations
        # Sauvegarder les nuisances
        analyse.nuisances = {
            k: (f"nuisance_{k}" in self.request.POST)
            for k, _ in self.NUISANCE_KEYS
        }
        analyse.save()

        # Sauvegarder les données 5 Pourquoi
        pourquoi_json = self.request.POST.get("pourquoi_data_json", "{}")
        try:
            pourquoi_data = json.loads(pourquoi_json)
            if not isinstance(pourquoi_data, dict):
                pourquoi_data = {}
        except (json.JSONDecodeError, ValueError):
            pourquoi_data = {}
        analyse.pourquoi_data = pourquoi_data
        analyse.save(update_fields=["pourquoi_data"])

        # Synchroniser CauseRacine depuis les P5 (cause racine = dernier pourquoi)
        # Cela maintient la compatibilité avec les statistiques globales des causes.
        analyse.causes_racines.all().delete()
        ordre = 0
        for code, pdata in pourquoi_data.items():
            if not isinstance(pdata, dict):
                continue
            p5 = ((pdata.get("pourquoi") or ["", "", "", "", ""])[4] or "").strip()
            if p5:
                CauseRacine.objects.create(analyse=analyse, texte=p5, ordre=ordre)
                ordre += 1

        log_analyse_history(self.request.user, analyse, "Analyse 48h mise à jour")

        from .tasks import notifier_accident_alerte
        # Transition de statut
        accident = analyse.accident
        if "soumettre_8j" in self.request.POST and accident.statut in (AccidentTravail.STATUT_24H, AccidentTravail.STATUT_48H):
            now = timezone.now()
            retard = accident.echeance_48h and now > accident.echeance_48h
            # Enregistrer la date de soumission
            analyse.soumis_le = now
            analyse.save(update_fields=["soumis_le"])
            accident.statut = AccidentTravail.STATUT_8J
            accident.updated_by = self.request.user
            accident.save(update_fields=["statut", "updated_by"])
            msg = "Analyse 48h soumise en RETARD" if retard else "Analyse 48h soumise"
            log_analyse_history(self.request.user, analyse, f"{msg} → LAP 8j démarré")
            notifier_accident_alerte(accident, "soumission_48h", acteur=self.request.user)
            messages.success(self.request, "Analyse 48h soumise. Étape LAP 8 jours démarrée.")
            if retard:
                messages.warning(self.request, "⚠️ Cette analyse a été soumise après l'échéance 48h.")
            return redirect("at:at_lap_8j", pk=accident.pk)

        notifier_accident_alerte(accident, "modification_48h", acteur=self.request.user)
        messages.success(self.request, "Analyse 48h sauvegardée.")
        from django.http import HttpResponseRedirect
        from django.urls import reverse
        url = reverse("at:at_analyse_48h", kwargs={"pk": accident.pk}) + "?tab=arbre"
        return HttpResponseRedirect(url)

    def get_success_url(self):
        return redirect("at:at_detail", pk=self._accident.pk)


# ─────────────────────────────────────────────────────────────────────────────
# LAP 8 jours
# ─────────────────────────────────────────────────────────────────────────────

class LAP8JView(AccidentPermissionMixin, UpdateView):
    model = LAP8Jours
    form_class = LAP8JForm
    template_name = "accident_travail/form_lap_8j.html"

    def get_object(self, queryset=None):
        accident = get_object_or_404(AccidentTravail, pk=self.kwargs["pk"])
        self._accident = accident
        lap, _ = LAP8Jours.objects.get_or_create(accident=accident)
        return lap

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["accident"] = self._accident
        ctx["action_formset"] = ActionCorrectiveFormSet(
            self.request.POST or None,
            instance=self.object,
            prefix="lap_actions",
        )
        # Deadline info
        now = timezone.now()
        echeance = self._accident.echeance_8j
        ctx["echeance_8j"] = echeance
        ctx["en_retard_8j"] = echeance and now > echeance
        ctx["heures_restantes_8j"] = round((echeance - now).total_seconds() / 3600, 1) if echeance else None
        # Pré-remplissage participants depuis l'analyse 48h
        p48h = {}
        if hasattr(self._accident, 'analyse_48h'):
            a = self._accident.analyse_48h
            p48h = {
                'exploitation': a.participants_exploitation or '',
                'cssct': a.participants_cssct or '',
                'rh': a.participants_rh or '',
                'qse': a.participants_qse or '',
                'direction': a.participants_direction or '',
                'finance': a.participants_finance or '',
                'achat': a.participants_achat or '',
                'commercial': a.participants_commercial or '',
                'autre': a.participants_autre or '',
                'client': a.participants_client or '',
            }
        ctx["participants_48h_json"] = json.dumps(p48h)
        return ctx

    def form_valid(self, form):
        ctx = self.get_context_data()
        action_formset = ctx["action_formset"]
        if not action_formset.is_valid():
            return self.form_invalid(form)

        lap = form.save()
        action_formset.instance = lap
        action_formset.save()
        log_lap_history(self.request.user, lap, "LAP 8j mis à jour")

        from .tasks import notifier_accident_alerte
        # Clôture
        accident = lap.accident
        if "cloturer" in self.request.POST and accident.statut == AccidentTravail.STATUT_8J:
            now = timezone.now()
            retard = accident.echeance_8j and now > accident.echeance_8j
            lap.soumis_le = now
            lap.save(update_fields=["soumis_le"])
            accident.statut = AccidentTravail.STATUT_CLOTURE
            accident.updated_by = self.request.user
            accident.save(update_fields=["statut", "updated_by"])
            msg = "LAP 8j soumis en RETARD → clôture" if retard else "LAP 8j soumis → clôture"
            log_lap_history(self.request.user, lap, msg)
            notifier_accident_alerte(accident, "cloture", acteur=self.request.user)
            messages.success(self.request, "Accident clôturé avec succès.")
            if retard:
                messages.warning(self.request, "⚠️ Ce LAP 8 jours a été soumis après l'échéance.")
            return redirect("at:at_detail", pk=accident.pk)

        messages.success(self.request, "LAP 8 jours sauvegardée.")
        return redirect("at:at_detail", pk=accident.pk)


# ─────────────────────────────────────────────────────────────────────────────
# Autocomplete causes racines
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@at_roles_required
def causes_autocomplete(request):
    """Retourne les libellés de causes correspondant à la recherche (JSON)."""
    q = (request.GET.get("q") or "").strip()
    if len(q) < 2:
        return JsonResponse({"results": []})
    results = (
        CauseRacine.objects
        .filter(texte__icontains=q)
        .values_list("texte", flat=True)
        .distinct()
        .order_by("texte")[:15]
    )
    return JsonResponse({"results": list(results)})


# ─────────────────────────────────────────────────────────────────────────────
# Global causes statistics
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@at_roles_required
def global_causes_view(request):
    """Statistiques globales des causes d'accidents (questionnaire 48h + arbre des causes)."""
    from django.db.models import Count
    from .models import QuestionnaireItem

    user = request.user
    qs = AccidentTravail.objects.all()
    if user.is_superuser:
        pass
    else:
        if user.section:
            qs = qs.filter(section=user.section)
        else:
            qs = qs.none()

    accident_ids = qs.values_list("pk", flat=True)
    total_accidents = qs.count()
    total_avec_analyse = qs.filter(analyse_48h__isnull=False).count()

    # Questionnaire 48h — questions répondues "Oui" (O) = causes identifiées
    causes_oui = list(
        QuestionnaireItem.objects
        .filter(analyse__accident__in=accident_ids, reponse="O")
        .values("code", "question", "section", "est_sous_question", "parent_code")
        .annotate(count=Count("id"))
        .order_by("-count")[:20]
    )
    max_count = causes_oui[0]["count"] if causes_oui else 1
    most_popular = causes_oui[0] if causes_oui else None

    # Chart data (top 10)
    top10 = causes_oui[:10]
    chart_labels = [
        (item["question"][:50] + "…") if len(item["question"]) > 50 else item["question"]
        for item in top10
    ]
    chart_counts = [item["count"] for item in top10]

    # Free-form causes (CauseRacine model) — agrégation par libellé
    from django.db.models import Count as _Count
    cause_qs = (
        CauseRacine.objects
        .filter(analyse__accident__in=accident_ids)
        .values("texte")
        .annotate(nb=_Count("id"))
        .order_by("-nb")
    )
    top_nodes = [(item["texte"], item["nb"]) for item in cause_qs[:15]]
    most_attributed = top_nodes[0][0] if top_nodes else None

    # Pour la visualisation en arbre : regrouper par libellé avec la liste des accidents
    cause_tree = []
    for item in cause_qs[:20]:
        accidents_avec_cause = list(
            CauseRacine.objects
            .filter(analyse__accident__in=accident_ids, texte=item["texte"])
            .select_related("analyse__accident")
            .values_list("analyse__accident__reference", "analyse__accident__pk")
            .distinct()
        )
        cause_tree.append({
            "texte": item["texte"],
            "nb": item["nb"],
            "accidents": accidents_avec_cause,
        })

    # Répartition par section (for pie chart)
    from collections import defaultdict
    section_totals = defaultdict(int)
    for c in causes_oui:
        if c["section"]:
            section_totals[c["section"]] += c["count"]
    section_labels = json.dumps([s.split("—")[-1].strip() if "—" in s else s for s in section_totals])
    section_counts = json.dumps(list(section_totals.values()))

    return render(request, "accident_travail/causes_globales.html", {
        "total_accidents": total_accidents,
        "total_avec_analyse": total_avec_analyse,
        "causes_oui": causes_oui,
        "max_count": max_count,
        "most_popular": most_popular,
        "top_nodes": top_nodes,
        "most_attributed": most_attributed,
        "cause_tree": cause_tree,
        "cause_tree_json": json.dumps([{"texte": n["texte"], "nb": n["nb"]} for n in cause_tree]),
        "chart_labels": json.dumps(chart_labels),
        "chart_counts": json.dumps(chart_counts),
        "section_labels": section_labels,
        "section_counts": section_counts,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Rapport complet (impression / PDF)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@at_roles_required
def accident_rapport(request, pk):
    """Rapport complet d'un accident — imprimable et téléchargeable."""
    accident = get_object_or_404(AccidentTravail, pk=pk)
    has_analyse = hasattr(accident, "analyse_48h")
    has_lap = hasattr(accident, "lap_8j")

    analyse = accident.analyse_48h if has_analyse else None
    lap = accident.lap_8j if has_lap else None

    questionnaire_sections = {}
    if analyse:
        for item in analyse.questionnaire_items.all().order_by("code"):
            questionnaire_sections.setdefault(item.section, []).append(item)

    actions_immediates = accident.actions_correctives_immediates.all()
    actions_correctives = lap.actions.all() if lap else []

    causes = analyse.causes_racines.all() if analyse else []

    return render(request, "accident_travail/rapport_complet.html", {
        "accident": accident,
        "has_analyse": has_analyse,
        "has_lap": has_lap,
        "analyse": analyse,
        "lap": lap,
        "questionnaire_sections": questionnaire_sections,
        "actions_immediates": actions_immediates,
        "actions_correctives": actions_correctives,
        "causes": causes,
        "now": timezone.now(),
    })


@login_required
@at_roles_required
def accident_export_excel(request, pk):
    """Génère et télécharge un fichier Excel (format template interne) rempli avec les données de l'accident."""
    import os
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    accident = get_object_or_404(AccidentTravail, pk=pk)
    has_analyse = hasattr(accident, "analyse_48h")
    has_lap = hasattr(accident, "lap_8j")
    analyse = accident.analyse_48h if has_analyse else None
    lap = accident.lap_8j if has_lap else None

    TEMPLATE_PATH = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
        "Fichier", "Accident", "Doc_56_Declaration_AT_du_10_11_2025_sans_macro.xlsx"
    )

    # Helper: format time field
    def fmt_time(t):
        return t.strftime("%H:%M") if t else ""

    # Helper: format date field
    def fmt_date(d):
        return d.strftime("%d/%m/%Y") if d else ""

    # Helper: boolean to Oui/Non
    def oui_non(val):
        if val is None:
            return ""
        return "OUI" if val else "NON"

    # ── Try to use the real template, fall back to building from scratch ──
    if os.path.exists(TEMPLATE_PATH):
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        # Strip column-dimension entries that extend to max=16384.
        # The template carries default col-style ranges covering all 16384
        # Excel columns; openpyxl re-writes them on save, and LibreOffice Calc
        # then warns "maximum number of columns per sheet was exceeded".
        # Removing entries whose max exceeds 200 is safe because no sheet has
        # real data beyond column 88.
        for _ws in wb.worksheets:
            inflated = [k for k, cd in _ws.column_dimensions.items() if cd.max > 200]
            for k in inflated:
                del _ws.column_dimensions[k]
    else:
        wb = openpyxl.Workbook()
        wb.active.title = "Doc.56 (24H)"
        wb.create_sheet("Doc 56.1 (48H)")
        wb.create_sheet("Doc 56.1 (8J)")

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 1: Doc.56 (24H)
    # ══════════════════════════════════════════════════════════════════════════
    ws24 = wb["Doc.56 (24H)"] if "Doc.56 (24H)" in wb.sheetnames else wb.active

    def _wc(ws, row, col, value):
        """Write value to cell, resolving merged-cell slaves to their master."""
        from openpyxl.cell import MergedCell
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            for mr in ws.merged_cells.ranges:
                if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                    cell = ws.cell(row=mr.min_row, column=mr.min_col)
                    break
        cell.value = value

    def w(row, col, value):
        _wc(ws24, row, col, value)

    # ── Row 2: main header data (labels are in row 3) ──
    w(2, 1, accident.victime_nom_prenom)                      # A2 → Victime
    w(2, 11, str(accident.societe))                           # K2 → Société
    w(2, 18, fmt_date(accident.date_accident))                # R2 → Date
    w(2, 21, fmt_time(accident.heure_accident))               # U2 → Heure
    # Horaires de travail  
    horaires = ""
    if accident.horaire_debut1:
        horaires = f"{fmt_time(accident.horaire_debut1)}-{fmt_time(accident.horaire_fin1)}"
        if accident.horaire_debut2:
            horaires += f" / {fmt_time(accident.horaire_debut2)}-{fmt_time(accident.horaire_fin2)}"
    w(2, 24, horaires)                                        # X2 → Horaires
    w(2, 33, accident.lieu_accident)                          # AG2 → Lieu

    # ── Row 6: site ──
    w(6, 12, accident.lieu_accident)                          # L6 → Site/Chantier

    # ── Row 10: circonstances détaillées ──
    w(10, 1, accident.circonstances_detaillees or accident.activite_victime)  # A10

    # ── Row 11: EPI Oui/Non (between AG11 and AM11) ──
    if accident.epi_endommage is True:
        w(11, 34+1, "X")   # AH11 — Oui
    elif accident.epi_endommage is False:
        w(11, 39+1, "X")   # AN11 — Non

    # ── Rows 14-17: element matériel checkboxes ──
    # Map labels to row/col positions: (B=2, F=6, J=10, N=14) × rows 14-17
    ELEM_MAP = {
        "Circulation de plain pied": (14, 1), "Chute de dénivellation": (14, 5),
        "Objet": (14, 9), "Transport manuel": (14, 13),
        "Particules corps étrangers": (15, 1), "Appareils levage": (15, 5),
        "Elingues, Chaînes": (15, 9), "Organes en mouvement": (15, 13),
        "Matériel à souder": (16, 1), "Outils mécaniques manuels": (16, 5),
        "Outil à main": (16, 9), "Produits chimiques": (16, 13),
        "Electricité": (17, 1), "Divers": (17, 5),
        "Manutention manuelle": (17, 13),
    }
    for elem in (accident.elements_materiels or []):
        pos = ELEM_MAP.get(elem)
        if pos:
            _wc(ws24, pos[0], pos[1], "X")

    # ── Nature lésions checkboxes (R=18, V=22, ... cols 18-21) ──
    LESION_MAP = {
        "Fractures": (14, 17), "Douleurs": (14, 21),
        "Brûlures / Coup d'arc": (15, 17), "Corps étrangers": (15, 21),
        "Entorse / Foulure": (16, 17), "Plaies Piqûres Coupures": (16, 21),
        "Contusions Hématomes": (17, 17), "Malaise": (17, 21),
    }
    for lesion in (accident.nature_lesions or []):
        pos = LESION_MAP.get(lesion)
        if pos:
            _wc(ws24, pos[0], pos[1], "X")

    # ── Siège lésions (Z=26, AD=30) ──
    siege_str = ", ".join(accident.siege_lesions or [])
    w(18, 26, siege_str)  # Z18

    # ── Row 19: Premiers soins ──
    w(19, 8, oui_non(accident.premiers_soins) if accident.premiers_soins else "")   # H19
    w(19, 11, oui_non(not accident.premiers_soins) if accident.premiers_soins is not None and not accident.premiers_soins else "")  # K19
    w(19, 15, accident.premiers_soins_par)           # O19 → Par qui
    w(19, 26, accident.premiers_soins_lesquels)      # Z19 → Lesquels

    # ── Row 20: Consultation médecin ──
    w(20, 8, "X" if accident.consultation_medecin else "")    # H20
    w(20, 11, "X" if accident.consultation_medecin is False else "")  # K20
    w(20, 15, accident.consultation_medecin_info)    # O20

    # ── Row 21: Transporté à ──
    w(21, 8, "X" if accident.transport_hopital else "")        # H21 → Oui
    w(21, 11, "X" if accident.transport_hopital is False else "")  # K21 → Non
    w(21, 23, accident.transport_hopital_nom)        # W21 → data area (W21:AR21)

    # ── Row 23: Tiers ──
    w(23, 12, accident.tiers_nom_adresse if accident.tiers_implique else "")  # L23 → data area (L23:U23)
    w(23, 34, accident.declarant_nom_prenom)         # AH23 → Etabli par (AH23:AR23 master)

    # ── Rows 26-27: Témoin ──
    w(26, 16, accident.temoin_nom_prenom)            # P26 → data area (P26:AG26 master)
    w(27, 16, accident.temoin_adresse)               # P27 → data area (P27:AR27 master)

    # ── Rows 31+: Actions correctives immédiates ──
    row = 31
    for i, action in enumerate(accident.actions_correctives_immediates.all()):
        _wc(ws24, row, 1, i + 1)
        _wc(ws24, row, 4, action.description)
        _wc(ws24, row, 36, action.pilote or "")
        row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 2: Doc 56.1 (48H)
    # ══════════════════════════════════════════════════════════════════════════
    if "Doc 56.1 (48H)" in wb.sheetnames and analyse:
        ws48 = wb["Doc 56.1 (48H)"]

        def w48(row, col, value):
            _wc(ws48, row, col, value)

        # ── Row 2: header data (labels in row 3 → I3, S3, Z3, AC3, AF3) ──
        w48(2, 9, accident.victime_nom_prenom)          # I2 → Victime
        w48(2, 19, str(accident.societe))               # S2 → Société
        w48(2, 26, fmt_date(accident.date_accident))    # Z2 → Date
        w48(2, 29, fmt_time(accident.heure_accident))   # AC2 → Heure
        w48(2, 32, horaires)                            # AF2 → Horaires

        # ── Participants (rows 7-11 — labels at col 5/27, data areas at col 12/31) ──
        w48(7, 12, analyse.participants_exploitation)   # L7 → Exploitation data area (L7:Z7)
        w48(7, 31, analyse.participants_finance)        # AE7 → Finance data area (AE7:AN7)
        w48(8, 12, analyse.participants_cssct)          # L8 → CSSCT data area (L8:Z8)
        w48(8, 31, analyse.participants_achat)          # AE8 → Achat data area
        w48(9, 12, analyse.participants_rh)             # L9 → RH data area
        w48(9, 31, analyse.participants_commercial)     # AE9 → Commercial data area
        w48(10, 12, analyse.participants_qse)           # L10 → QSE data area
        w48(10, 31, analyse.participants_autre)         # AE10 → Autre data area
        w48(11, 12, analyse.participants_direction)     # L11 → Direction data area
        w48(11, 31, analyse.participants_client)        # AE11 → Client data area

        # ── Victime compléments (row 43-44 — "2 - La victime" header at row 42) ──
        # Row 43: A43:B43=Age label, C43:E43=age data, F43:G43=Sexe label,
        #         H43:J43=sexe data, K43:M43=Date VM label, N43:P43=date VM data,
        #         Q43:V43=Restriction label, W43:Y43=restriction data, Z43:AC43=Précisions label, AD43:AN43=préc data
        # Row 44: A44:G44=Au moment label, H44:L44=Poste label, M44:Y44=poste data,
        #         Z44:AE44=Ancienneté label, AF44:AL44=ancienneté data
        w48(43, 3, analyse.victime_age or "")                       # C43 → age data area
        w48(43, 8, accident.victime_sexe or "")                     # H43 → sexe data area
        w48(43, 14, fmt_date(analyse.victime_date_visite_medicale)) # N43 → date VM data area
        w48(43, 23, oui_non(analyse.victime_restriction_aptitude))  # W43 → restriction data area
        w48(44, 13, analyse.victime_poste_occupe)                   # M44 → poste data area
        w48(44, 32, analyse.victime_anciennete_poste_analyse)       # AF44 → ancienneté data area

        # ── Questionnaire items ──
        for item in analyse.questionnaire_items.all().order_by("code"):
            # Match row by code prefix: template cells start with "1.1 : ", "1.2 : " etc.
            code_prefix = f"{item.code} :"
            for qrow in ws48.iter_rows(min_col=1, max_col=1):
                cell = qrow[0]
                if cell.value and str(cell.value).startswith(code_prefix):
                    r = cell.row
                    # Response cols: 38=NA, 39=O(Oui), 40=N(Non)
                    _wc(ws48, r, 38, "X" if item.reponse == "NA" else "")
                    _wc(ws48, r, 39, "X" if item.reponse == "O" else "")
                    _wc(ws48, r, 40, "X" if item.reponse == "N" else "")
                    # Précisions: label "Précisions :" in col 1 of r+1, write text to col 3
                    if item.precisions:
                        _wc(ws48, r + 1, 3, item.precisions)
                    break

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 3: Doc 56.1 (8J)
    # ══════════════════════════════════════════════════════════════════════════
    if "Doc 56.1 (8J)" in wb.sheetnames and lap:
        ws8j = wb["Doc 56.1 (8J)"]

        def w8j(row, col, value):
            _wc(ws8j, row, col, value)

        # ── Row 2: header data (labels in row 2 → I2, S2, Z2, AC2, AF2) ──
        w8j(2, 9, accident.victime_nom_prenom)
        w8j(2, 19, str(accident.societe))
        w8j(2, 26, fmt_date(accident.date_accident))
        w8j(2, 29, fmt_time(accident.heure_accident))
        w8j(2, 32, horaires)

        # ── Participants (rows 7-11 — same layout as 48H sheet) ──
        w8j(7, 12, lap.participants_exploitation)
        w8j(7, 31, lap.participants_finance)
        w8j(8, 12, lap.participants_cssct)
        w8j(8, 31, lap.participants_achat)
        w8j(9, 12, lap.participants_rh)
        w8j(9, 31, lap.participants_commercial)
        w8j(10, 12, lap.participants_qse)
        w8j(10, 31, lap.participants_autre)
        w8j(11, 12, lap.participants_direction)
        w8j(11, 31, lap.participants_client)

        # ── Actions correctives / préventives (header at row 15, data rows 16+) ──
        row = 16
        for action in lap.actions.all():
            _wc(ws8j, row, 1, action.cause_racine)
            _wc(ws8j, row, 8, action.description)
            _wc(ws8j, row, 25, action.pilote or "")
            _wc(ws8j, row, 29, fmt_date(action.delai))
            _wc(ws8j, row, 32, fmt_date(action.date_realisation) if hasattr(action, "date_realisation") else "")
            _wc(ws8j, row, 35, fmt_date(action.date_verification) if hasattr(action, "date_verification") else "")
            row += 1

        # ── Communication / DUER / Transversalisation (row 35 — Oui/Non checkboxes) ──
        # R34: labels, R35: Oui(col2/12/22) Non(col5/15/25), R36: dates
        w8j(35, 2, "X" if lap.communication_alerte_securite else "")
        w8j(35, 5, "X" if lap.communication_alerte_securite is False else "")
        w8j(35, 12, "X" if lap.mise_a_jour_duer else "")
        w8j(35, 15, "X" if lap.mise_a_jour_duer is False else "")
        w8j(35, 22, "X" if lap.transversalisation_groupe else "")
        w8j(35, 25, "X" if lap.transversalisation_groupe is False else "")

    # ── Save to BytesIO buffer and return as HTTP response ──
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"AT_{accident.reference}_{accident.victime_nom_prenom.replace(' ', '_')}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────────────────────────────────────────
# API dynamique sections / sites
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@at_roles_required
def api_sections_by_societe(request, societe_id):
    from accounts.models import Section
    sections = Section.objects.filter(societe_id=societe_id).values("id", "nom")
    return JsonResponse({"sections": list(sections)})


@login_required
@at_roles_required
def api_sites_by_section(request, section_id):
    from accounts.models import Site
    sites = Site.objects.filter(section_id=section_id).values("id", "nom")
    return JsonResponse({"sites": list(sites)})


# ─────────────────────────────────────────────────────────────────────────────
# Auto-sauvegarde brouillon
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@at_roles_required
def questionnaire_autosave(request, pk):
    """AJAX — sauvegarde partielle des réponses questionnaire (sans soumettre le formulaire)."""
    if request.method != "POST":
        return JsonResponse({"error": "POST requis"}, status=405)

    analyse = get_object_or_404(Analyse48h, accident__pk=pk)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "JSON invalide"}, status=400)

    updated = 0
    for item in analyse.questionnaire_items.all():
        # Correspond au filtre |cut:'.' dans le template
        key = item.code.replace(".", "")
        reponse    = data.get(f"q_{key}_reponse", "")
        precisions = data.get(f"q_{key}_precisions", "")
        if reponse in ("NA", "O", "N"):
            item.reponse    = reponse
            item.precisions = precisions
            item.save(update_fields=["reponse", "precisions"])
            updated += 1

    return JsonResponse({"saved": True, "updated": updated})


@login_required
@at_roles_required
def autosave_analyse_48h(request, pk):
    """AJAX — sauvegarde automatique complète de l'analyse 48h lors du changement d'onglet."""
    if request.method != "POST":
        return JsonResponse({"error": "POST requis"}, status=405)

    FORMATION_TYPES = [
        "Formation au poste de travail", "Formation renforcée intérimaires",
        "Habilitation électrique", "CACES", "PRAP", "SST", "Autres formations",
    ]
    NUISANCE_KEYS = [
        "bruit", "vibrations", "vapeurs", "climatiques",
        "eclairage", "poussieres", "thermiques", "autres",
    ]

    analyse = get_object_or_404(Analyse48h, accident__pk=pk)

    # 1. Champs Analyse48hForm
    form = Analyse48hForm(request.POST, instance=analyse)
    if form.is_valid():
        form.save()
        analyse.refresh_from_db()

    # 2. Questionnaire
    for item in analyse.questionnaire_items.all():
        key = item.code.replace(".", "")
        reponse = request.POST.get(f"q_{key}_reponse", "")
        precisions = request.POST.get(f"q_{key}_precisions", "")
        if reponse in ("NA", "O", "N"):
            item.reponse = reponse
            item.precisions = precisions
            item.save(update_fields=["reponse", "precisions"])

    # 3. Formations
    formations = []
    for i, nom in enumerate(FORMATION_TYPES):
        realise = request.POST.get(f"formation_{i}_realise", "NA")
        formations.append({
            "nom": nom,
            "realise": realise if realise in ("NA", "O", "N") else "NA",
            "date": request.POST.get(f"formation_{i}_date", ""),
            "commentaire": request.POST.get(f"formation_{i}_commentaire", ""),
        })
    analyse.formations = formations

    # 4. Nuisances
    analyse.nuisances = {k: (f"nuisance_{k}" in request.POST) for k in NUISANCE_KEYS}
    analyse.save(update_fields=["formations", "nuisances"])

    # 5. Données 5 Pourquoi
    pourquoi_json = request.POST.get("pourquoi_data_json", "")
    if pourquoi_json:
        try:
            pourquoi_data = json.loads(pourquoi_json)
            if isinstance(pourquoi_data, dict) and pourquoi_data:
                analyse.pourquoi_data = pourquoi_data
                analyse.save(update_fields=["pourquoi_data"])
                analyse.causes_racines.all().delete()
                ordre = 0
                for code, pdata in pourquoi_data.items():
                    if not isinstance(pdata, dict):
                        continue
                    p5 = ((pdata.get("pourquoi") or ["", "", "", "", ""])[4] or "").strip()
                    if p5:
                        CauseRacine.objects.create(analyse=analyse, texte=p5, ordre=ordre)
                        ordre += 1
        except (json.JSONDecodeError, ValueError):
            pass

    return JsonResponse({"saved": True})


@login_required
@at_roles_required
def autosave_lap_8j(request, pk):
    """AJAX — sauvegarde automatique du LAP 8 jours."""
    if request.method != "POST":
        return JsonResponse({"error": "POST requis"}, status=405)

    lap = get_object_or_404(LAP8Jours, accident__pk=pk)
    form = LAP8JForm(request.POST, instance=lap)
    if form.is_valid():
        form.save()
        return JsonResponse({"saved": True})
    return JsonResponse({"saved": False, "errors": str(form.errors)})


@login_required
@at_roles_required
def auto_save_draft(request, pk):
    """Endpoint de sauvegarde automatique pour le formulaire 24h."""
    if request.method != "POST":
        return JsonResponse({"error": "POST requis"}, status=405)
    accident = get_object_or_404(AccidentTravail, pk=pk)
    form = AccidentForm24h(request.POST, instance=accident, user=request.user)
    if form.is_valid():
        a = form.save(commit=False)
        a.updated_by = request.user
        a.save()
        return JsonResponse({"saved": True, "reference": accident.reference})
    return JsonResponse({"saved": False, "errors": form.errors})


# ─────────────────────────────────────────────────────────────────────────────
# CRUD Questionnaire Template — questions & sous-questions 48h
# ─────────────────────────────────────────────────────────────────────────────

def _require_questionnaire_admin(user):
    """Accès questionnaire réservé à superadmin, RS et RO."""
    return user_has_at_access(user)


@login_required
@at_roles_required
def questionnaire_list(request):
    """Liste des questions groupées par section, avec leurs sous-questions."""
    if not _require_questionnaire_admin(request.user):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied
    qs = QuestionnaireTemplate.objects.all().order_by("ordre", "code")
    # Build a grouped structure: section → [question, ...]
    # Each question carries its .sous_questions list
    sections = {}
    questions_by_code = {}
    for q in qs:
        questions_by_code[q.code] = q
        q.sous_questions = []

    for q in qs:
        if q.est_sous_question:
            parent = questions_by_code.get(q.parent_code)
            if parent:
                parent.sous_questions.append(q)
        else:
            sections.setdefault(q.section, []).append(q)

    return render(request, "accident_travail/questionnaire/list.html", {
        "sections": sections,
        "total": qs.count(),
    })


@login_required
@at_roles_required
def questionnaire_create(request):
    """Créer une nouvelle question ou sous-question."""
    if not _require_questionnaire_admin(request.user):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied
    if request.method == "POST":
        form = QuestionnaireTemplateForm(request.POST)
        if form.is_valid():
            q = form.save()
            messages.success(request, f"Question « {q.code} » créée avec succès.")
            if request.POST.get("save_and_add"):
                return redirect("at:questionnaire_create")
            return redirect("at:questionnaire_list")
    else:
        form = QuestionnaireTemplateForm()
    return render(request, "accident_travail/questionnaire/form.html", {
        "form": form,
        "action": "Créer",
        "is_create": True,
    })


@login_required
@at_roles_required
def questionnaire_update(request, pk):
    """Modifier une question ou sous-question existante."""
    if not _require_questionnaire_admin(request.user):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied
    question = get_object_or_404(QuestionnaireTemplate, pk=pk)
    if request.method == "POST":
        form = QuestionnaireTemplateForm(request.POST, instance=question)
        if form.is_valid():
            q = form.save()
            messages.success(request, f"Question « {q.code} » modifiée avec succès.")
            return redirect("at:questionnaire_list")
    else:
        form = QuestionnaireTemplateForm(instance=question)
    return render(request, "accident_travail/questionnaire/form.html", {
        "form": form,
        "action": "Modifier",
        "question": question,
        "is_create": False,
    })


@login_required
@at_roles_required
def questionnaire_delete(request, pk):
    """Supprimer une question (et ses sous-questions)."""
    if not _require_questionnaire_admin(request.user):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied
    question = get_object_or_404(QuestionnaireTemplate, pk=pk)
    sous_questions = QuestionnaireTemplate.objects.filter(parent_code=question.code)
    if request.method == "POST":
        sous_questions.delete()
        question.delete()
        messages.success(request, f"Question « {question.code} » supprimée.")
        return redirect("at:questionnaire_list")
    return render(request, "accident_travail/questionnaire/confirm_delete.html", {
        "question": question,
        "sous_questions": sous_questions,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Arbre des causes
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@at_roles_required
def arbre_causes_view(request, pk):
    """Affiche l'arbre des causes pour un accident de travail."""
    accident = get_object_or_404(AccidentTravail, pk=pk)
    
    # Vérifier les permissions
    can_edit = user_has_at_access(request.user)
    
    # Récupérer l'analyse 48h
    try:
        analyse = accident.analyse_48h
        arbre_data = analyse.arbre_causes or {}
    except Analyse48h.DoesNotExist:
        arbre_data = {}
    
    # S'assurer que la structure de base existe
    if not arbre_data:
        arbre_data = {"causes": []}
    
    import json
    return render(request, "accident_travail/arbre_causes.html", {
        "accident": accident,
        "can_edit": can_edit,
        "arbre_json": json.dumps(arbre_data),
    })


@login_required
@at_roles_required
def arbre_causes_api(request, pk):
    """API pour sauvegarder l'arbre des causes (JSON)."""
    import json
    from django.http import JsonResponse
    
    accident = get_object_or_404(AccidentTravail, pk=pk)
    
    # Vérifier les permissions
    if not user_has_at_access(request.user):
        return JsonResponse({"status": "error", "message": "Permission refusée"}, status=403)
    
    if request.method == "POST":
        try:
            # Récupérer ou créer l'analyse 48h
            analyse, created = Analyse48h.objects.get_or_create(accident=accident)
            
            # Sauvegarder les données de l'arbre
            data = json.loads(request.body)
            analyse.arbre_causes = data
            analyse.save()
            
            return JsonResponse({"status": "ok", "message": "Arbre sauvegardé avec succès"})
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)
    
    return JsonResponse({"status": "error", "message": "Méthode non autorisée"}, status=405)
